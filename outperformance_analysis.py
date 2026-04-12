#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║  POST-CUTOFF OUTPERFORMANCE ANALYSER                                 ║
║  Cutoff   : 02 Apr 2026  (pre-conditions measured ON this date)      ║
║  Period   : 03 Apr 2026 → today  (performance window)               ║
║  DB       : bhav  (bhav2025 + bhav2026 + indexbhav + sectors)        ║
║                                                                      ║
║  RS method    : stock_return% − index_return%  (excess return)       ║
║  Index        : Nifty Smallcap 250  (from indexbhav)                 ║
║  Volatility   : ATR% 21-day  = ATR_21d / close × 100                ║
╚══════════════════════════════════════════════════════════════════════╝
REQUIREMENTS:  pip install mysql-connector-python pandas openpyxl
RUN:           python outperformance_analysis.py
"""

import math
import sys
import warnings
from datetime import date
from pathlib import Path

import mysql.connector
import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ── Config ────────────────────────────────────────────────────────────
DB_HOST = "localhost"
DB_PORT = 3306
DB_USER = "root"
DB_PASS = "root"
DB_NAME = "bhav"

CUTOFF  = date(2026, 4, 2)
TODAY   = date.today()

ATR_PERIOD      = 21
RS_SLOPE_LOOKBACK = 21

STOCKS = [
    "PFOCUS","MTARTECH","STLTECH","NATIONALUM","GVT&D",
    "POWERINDIA","BLISSGVS","ATHERENERG","AEROFLEX","DEEDEV",
    "OMNI","QPOWER","APOLLOPIPE","AVANTIFEED","NETWEB",
    "TDPOWERSYS","DATAPATTNS","GESHIP","BAJAJCON","GMDCLTD",
]

OUT_DIR = Path(__file__).parent / "output"
OUT_DIR.mkdir(exist_ok=True)

# ── DB ────────────────────────────────────────────────────────────────
def get_conn():
    return mysql.connector.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=DB_PASS,
        database=DB_NAME
    )

# ── Load OHLCV (bhav2025 UNION bhav2026) ─────────────────────────────
def load_history(conn, symbols, from_date, to_date):
    sym_csv = ",".join(f"'{s}'" for s in symbols)
    tables  = []
    if from_date.year <= 2025: tables.append("bhav2025")
    if to_date.year   >= 2026: tables.append("bhav2026")
    if not tables: tables = ["bhav2025", "bhav2026"]

    parts = [f"""
        SELECT SYMBOL, MKTDATE,
               OPEN, HIGH, LOW, CLOSE, VOLUME,
               `20dma`, `50dma`, `200dma`
        FROM {t}
        WHERE SYMBOL IN ({sym_csv})
          AND MKTDATE BETWEEN %s AND %s
    """ for t in tables]

    sql    = " UNION ALL ".join(parts) + " ORDER BY SYMBOL, MKTDATE"
    params = [from_date, to_date] * len(tables)

    df = pd.read_sql(sql, conn, params=params)
    df["MKTDATE"] = pd.to_datetime(df["MKTDATE"]).dt.date
    for c in ["OPEN","HIGH","LOW","CLOSE","VOLUME","20dma","50dma","200dma"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df.sort_values(["SYMBOL","MKTDATE"]).reset_index(drop=True)

# ── Load Nifty Smallcap 250 from indexbhav ────────────────────────────
def load_index(conn, from_date, to_date):
    sql = """
        SELECT mktdate AS MKTDATE, close AS CLOSE
        FROM indexbhav
        WHERE UPPER(symbol) LIKE %s
          AND mktdate BETWEEN %s AND %s
        ORDER BY mktdate
    """
    for pat in ("%SMALLCAP%250%", "%NIFTY%SMALLCAP%", "%NIFTY%50%"):
        df = pd.read_sql(sql, conn, params=[pat, from_date, to_date])
        if not df.empty:
            df["MKTDATE"] = pd.to_datetime(df["MKTDATE"]).dt.date
            df["CLOSE"]   = pd.to_numeric(df["CLOSE"], errors="coerce")
            df = df.drop_duplicates("MKTDATE").sort_values("MKTDATE")
            print(f"  Index: {pat.replace('%','')} — {len(df)} rows")
            return df
    print("  WARNING: indexbhav data not found — RS vs index will be NaN")
    return pd.DataFrame(columns=["MKTDATE","CLOSE"])

# ── Load sectors ──────────────────────────────────────────────────────
def load_sectors(conn, symbols):
    sym_csv = ",".join(f"'{s.upper()}'" for s in symbols)
    try:
        df = pd.read_sql(
            f"SELECT UPPER(symbol) AS SYM, sector1 AS SECTOR "
            f"FROM sectors WHERE UPPER(symbol) IN ({sym_csv})", conn
        )
        return dict(zip(df["SYM"], df["SECTOR"]))
    except Exception:
        return {}

# ── Helpers ───────────────────────────────────────────────────────────
def safe_return(c_now, c_then):
    if pd.isna(c_now) or pd.isna(c_then) or c_then == 0:
        return np.nan
    return round((c_now / c_then - 1) * 100, 2)

def lookback_close(series, n):
    """Close n trading-days back from the last row."""
    if len(series) <= n:
        return np.nan
    return float(series.iloc[-(n + 1)])

def compute_atr_pct(high, low, close, period=21):
    """ATR% = rolling(period) mean of True Range / last close × 100."""
    prev_close = close.shift(1)
    tr = pd.concat([
        high - low,
        (high - prev_close).abs(),
        (low  - prev_close).abs(),
    ], axis=1).max(axis=1)
    atr = tr.rolling(period, min_periods=1).mean()
    last_close = float(close.iloc[-1])
    if last_close == 0:
        return np.nan
    return round(float(atr.iloc[-1] / last_close * 100), 4)

# ── Main ──────────────────────────────────────────────────────────────
def analyse():
    print("\n" + "═"*65)
    print("  POST-CUTOFF OUTPERFORMANCE ANALYSER")
    print(f"  Cutoff : {CUTOFF}   |   Today : {TODAY}")
    print("═"*65)

    conn = get_conn()
    print("\n  ✓ Connected to MySQL bhav database")

    from_hist = date(2025, 1, 1)   # 15 months → covers 252-td (12M) lookback
    df_all    = load_history(conn, STOCKS, from_hist, TODAY)
    print(f"  ✓ Loaded {len(df_all):,} rows  |  {df_all['SYMBOL'].nunique()} symbols")

    idx_all    = load_index(conn, from_hist, TODAY)
    sector_map = load_sectors(conn, STOCKS)
    conn.close()

    # ── Index returns on cutoff date ──────────────────────────────────
    idx_pre = idx_all[idx_all["MKTDATE"] <= CUTOFF].reset_index(drop=True)
    ic      = idx_pre["CLOSE"]
    idx_ret = {
        "1m":  safe_return(float(ic.iloc[-1]) if len(ic) else np.nan, lookback_close(ic, 21)),
        "3m":  safe_return(float(ic.iloc[-1]) if len(ic) else np.nan, lookback_close(ic, 63)),
        "6m":  safe_return(float(ic.iloc[-1]) if len(ic) else np.nan, lookback_close(ic, 126)),
        "12m": safe_return(float(ic.iloc[-1]) if len(ic) else np.nan, lookback_close(ic, 252)),
    }
    print(f"  Index returns on {CUTOFF}:  "
          f"1M={idx_ret['1m']:.1f}%  3M={idx_ret['3m']:.1f}%  "
          f"6M={idx_ret['6m']:.1f}%  12M={idx_ret['12m']:.1f}%")

    # ── Per-stock analysis ────────────────────────────────────────────
    records = []

    for sym in STOCKS:
        g = df_all[df_all["SYMBOL"] == sym].copy().reset_index(drop=True)
        if g.empty:
            print(f"  ✗  {sym:15s} — no data"); continue

        pre  = g[g["MKTDATE"] <= CUTOFF].reset_index(drop=True)
        post = g[g["MKTDATE"]  > CUTOFF].reset_index(drop=True)

        if pre.empty or post.empty:
            print(f"  ✗  {sym:15s} — missing pre or post data"); continue

        # ── Prices ───────────────────────────────────────────────────
        close_cut    = float(pre.iloc[-1]["CLOSE"])
        close_latest = float(post.iloc[-1]["CLOSE"])
        latest_date  = post.iloc[-1]["MKTDATE"]
        post_ret     = safe_return(close_latest, close_cut)

        n  = len(pre)
        pc = pre["CLOSE"]
        ph = pre["HIGH"]
        pl = pre["LOW"]
        pv = pre["VOLUME"]

        # ── 52-week high / low ────────────────────────────────────────
        w52      = min(252, n)
        high_52w = float(ph.iloc[-w52:].max())
        low_52w  = float(pl.iloc[-w52:].min())
        pct_52wh = safe_return(close_cut, high_52w)     # ≤ 0
        pct_52wl = safe_return(close_cut, low_52w)      # ≥ 0

        idx_52wh    = int(ph.iloc[-w52:].idxmax())
        days_52wh   = (n - 1) - idx_52wh   # trading days since 52wk high

        range_pos = round((close_cut - low_52w) / (high_52w - low_52w) * 100, 1) \
                    if high_52w != low_52w else 50.0

        # ── ATH (all available history) ───────────────────────────────
        ath     = float(ph.max())
        pct_ath = safe_return(close_cut, ath)

        # ── Momentum ─────────────────────────────────────────────────
        mom_1m  = safe_return(close_cut, lookback_close(pc, 21))
        mom_3m  = safe_return(close_cut, lookback_close(pc, 63))
        mom_6m  = safe_return(close_cut, lookback_close(pc, 126))
        mom_12m = safe_return(close_cut, lookback_close(pc, 252))

        # ── RS vs Nifty Smallcap 250  (excess return, %-pts) ─────────
        #   RS = stock_return% − index_return%
        def rs(stock_mom, key):
            im = idx_ret.get(key)
            if pd.isna(stock_mom) or im is None or pd.isna(im):
                return np.nan
            return round(stock_mom - im, 2)

        rs_1m  = rs(mom_1m,  "1m")
        rs_3m  = rs(mom_3m,  "3m")
        rs_6m  = rs(mom_6m,  "6m")
        rs_12m = rs(mom_12m, "12m")

        # ── RS line slope (21-day)  — stock_close / index_close ──────
        if not idx_pre.empty:
            merged = pre[["MKTDATE","CLOSE"]].merge(
                idx_pre[["MKTDATE","CLOSE"]].rename(columns={"CLOSE":"IDX"}),
                on="MKTDATE", how="inner"
            )
            if len(merged) >= RS_SLOPE_LOOKBACK and not (merged["IDX"] == 0).any():
                rs_line        = merged["CLOSE"] / merged["IDX"]
                rs_slope_21d   = round(float(rs_line.iloc[-1] - rs_line.iloc[-RS_SLOPE_LOOKBACK]), 6)
                rs_at_52w_high = bool(rs_line.iloc[-1] >= rs_line.tail(min(252,len(rs_line))).max())
            else:
                rs_slope_21d, rs_at_52w_high = np.nan, False
        else:
            rs_slope_21d, rs_at_52w_high = np.nan, False

        # ── ATR% 21-day (same as stock_rating.py) ────────────────────
        atr_pct = compute_atr_pct(ph, pl, pc, ATR_PERIOD)

        # ── Moving averages (use DB pre-computed; fallback to calc) ───
        row_cut = pre.iloc[-1]
        def get_ma(col, span):
            v = row_cut.get(col, np.nan)
            if not pd.isna(v) and v > 0:
                return round(float(v), 2)
            return round(float(pc.iloc[-min(span,n):].mean()), 2)

        dma20  = get_ma("20dma",  20)
        dma50  = get_ma("50dma",  50)
        dma200 = get_ma("200dma", 200)

        above_20  = close_cut > dma20
        above_50  = close_cut > dma50
        above_200 = close_cut > dma200

        # ── Volume trend ──────────────────────────────────────────────
        vol20     = float(pv.iloc[-min(20,n):].mean())
        vol60     = float(pv.iloc[-min(60,n):].mean())
        vol_ratio = round(vol20 / vol60, 2) if vol60 > 0 else np.nan

        # ── Consecutive up/down streak into cutoff ────────────────────
        streak = 0
        for i in range(n - 1, 0, -1):
            if   pc.iloc[i] > pc.iloc[i-1]: up = True
            elif pc.iloc[i] < pc.iloc[i-1]: up = False
            else: break
            if streak == 0:
                streak = 1 if up else -1
            elif (streak > 0 and up) or (streak < 0 and not up):
                streak += (1 if up else -1)
            else:
                break

        sector = sector_map.get(sym.upper(), "—")

        records.append({
            "Symbol"               : sym,
            "Sector"               : sector,
            "Cutoff Close"         : round(close_cut, 2),
            "Latest Close"         : round(close_latest, 2),
            "Latest Date"          : str(latest_date),
            "Post Return %"        : round(post_ret, 2),
            # 52-week
            "52wk High"            : round(high_52w, 2),
            "52wk Low"             : round(low_52w, 2),
            "% from 52wk High"     : pct_52wh,
            "% above 52wk Low"     : pct_52wl,
            "Days since 52wk High" : days_52wh,
            "52wk Range Pos %"     : range_pos,
            # ATH
            "ATH"                  : round(ath, 2),
            "% from ATH"           : pct_ath,
            # Momentum
            "1M Momentum %"        : mom_1m,
            "3M Momentum %"        : mom_3m,
            "6M Momentum %"        : mom_6m,
            "12M Momentum %"       : mom_12m,
            # RS vs Nifty Smallcap 250 (excess return %-pts)
            "RS 1M (exc.ret %)"    : rs_1m,
            "RS 3M (exc.ret %)"    : rs_3m,
            "RS 6M (exc.ret %)"    : rs_6m,
            "RS 12M (exc.ret %)"   : rs_12m,
            "RS Line Slope 21d"    : rs_slope_21d,
            "RS at 52wk High"      : rs_at_52w_high,
            # Volatility
            "ATR% 21d"             : atr_pct,
            # Moving averages
            "20 DMA"               : dma20,
            "50 DMA"               : dma50,
            "200 DMA"              : dma200,
            "Above 20 DMA"         : above_20,
            "Above 50 DMA"         : above_50,
            "Above 200 DMA"        : above_200,
            # Volume
            "Vol Ratio (20d/60d)"  : vol_ratio,
            # Streak
            "Streak (td)"          : streak,
        })

        print(f"  ✓  {sym:15s}  post={post_ret:+.1f}%  "
              f"3M={mom_3m:+.1f}%  RS3M={rs_3m:+.1f}pp  "
              f"52wkH={pct_52wh:.1f}%  ATR%={atr_pct:.2f}  "
              f"RSslope={rs_slope_21d:.5f}")

    if not records:
        sys.exit("No records — check symbol names.")

    df = pd.DataFrame(records)
    df.sort_values("Post Return %", ascending=False, inplace=True)
    df.reset_index(drop=True, inplace=True)
    df.insert(0, "Rank", range(1, len(df) + 1))

    # ── Peer-group percentile ranks ───────────────────────────────────
    for src, dst in [
        ("3M Momentum %",     "3M Mom Pctile"),
        ("6M Momentum %",     "6M Mom Pctile"),
        ("12M Momentum %",    "12M Mom Pctile"),
        ("RS 3M (exc.ret %)", "RS 3M Pctile"),
        ("RS 6M (exc.ret %)", "RS 6M Pctile"),
        ("52wk Range Pos %",  "Range Pctile"),
    ]:
        df[dst] = df[src].rank(pct=True, na_option="bottom").mul(100).round(1)

    # ── Composite score ───────────────────────────────────────────────
    for c in ["3M Momentum %","6M Momentum %","12M Momentum %",
              "RS 3M (exc.ret %)","52wk Range Pos %","Vol Ratio (20d/60d)"]:
        df[f"__r_{c}"] = df[c].rank(ascending=True, na_option="bottom")
    r_cols = [c for c in df.columns if c.startswith("__r_")]
    df["Composite Score"] = df[r_cols].mean(axis=1).round(2)
    df.drop(columns=r_cols, inplace=True)

    # ── Reason tags ───────────────────────────────────────────────────
    def build_reasons(row):
        r = []
        if row["% from 52wk High"] >= -3:
            r.append(f"Near 52wk High ({row['% from 52wk High']:+.1f}%)")
        elif row["% from 52wk High"] >= -8:
            r.append(f"Close to 52wk High ({row['% from 52wk High']:+.1f}%)")
        if row["Days since 52wk High"] <= 5:
            r.append(f"52wk High just {row['Days since 52wk High']}td ago")
        if row["52wk Range Pos %"] >= 80:
            r.append(f"Upper 52wk range ({row['52wk Range Pos %']:.0f}%)")
        m3 = row["3M Momentum %"]
        m6 = row["6M Momentum %"]
        if not pd.isna(m3) and m3 > 10:
            r.append(f"Strong 3M momentum (+{m3:.1f}%)")
        if not pd.isna(m6) and m6 > 15:
            r.append(f"Strong 6M momentum (+{m6:.1f}%)")
        rs3 = row["RS 3M (exc.ret %)"]
        rs6 = row["RS 6M (exc.ret %)"]
        if not pd.isna(rs3) and rs3 > 10:
            r.append(f"RS 3M +{rs3:.1f}pp vs index")
        if not pd.isna(rs6) and rs6 > 15:
            r.append(f"RS 6M +{rs6:.1f}pp vs index")
        if row.get("RS at 52wk High"):
            r.append("RS line at 52wk high")
        rsl = row["RS Line Slope 21d"]
        if not pd.isna(rsl) and rsl > 0:
            r.append(f"RS slope rising (+{rsl:.5f})")
        if row["Above 200 DMA"]:
            r.append("Above 200 DMA")
        if row["Above 50 DMA"] and not row["Above 200 DMA"]:
            r.append("Above 50 DMA (not 200)")
        vr = row["Vol Ratio (20d/60d)"]
        if not pd.isna(vr) and vr > 1.3:
            r.append(f"Volume expanding ({vr:.2f}x)")
        atr = row["ATR% 21d"]
        if not pd.isna(atr) and atr < 2.5:
            r.append(f"Low volatility (ATR% {atr:.2f})")
        if row["Streak (td)"] >= 3:
            r.append(f"{row['Streak (td)']}d up-streak at cutoff")
        return " | ".join(r) if r else "—"

    df["Reasons"] = df.apply(build_reasons, axis=1)

    # ── Terminal top-5 ────────────────────────────────────────────────
    print("\n" + "═"*65)
    print(f"  TOP PERFORMERS  (03 Apr → {TODAY})")
    print("═"*65)
    for _, row in df.head(5).iterrows():
        print(f"\n  #{int(row['Rank'])}  {row['Symbol']}  →  {row['Post Return %']:+.2f}%  [{row['Sector']}]")
        print(f"      Cutoff ₹{row['Cutoff Close']}  →  Latest ₹{row['Latest Close']}  ({row['Latest Date']})")
        print(f"      52wk High ₹{row['52wk High']} ({row['% from 52wk High']:+.1f}%), "
              f"{row['Days since 52wk High']}td ago  |  Range pos {row['52wk Range Pos %']:.0f}%")
        print(f"      Momentum  1M={row['1M Momentum %']:+.1f}%  3M={row['3M Momentum %']:+.1f}%  "
              f"6M={row['6M Momentum %']:+.1f}%  12M={row['12M Momentum %']:+.1f}%")
        print(f"      RS (exc.ret vs SC250)  "
              f"1M={row['RS 1M (exc.ret %)']: +.1f}pp  "
              f"3M={row['RS 3M (exc.ret %)']: +.1f}pp  "
              f"6M={row['RS 6M (exc.ret %)']: +.1f}pp  "
              f"12M={row['RS 12M (exc.ret %)']: +.1f}pp")
        print(f"      RS slope 21d={row['RS Line Slope 21d']:.5f}  "
              f"RS@52wkH={'YES' if row['RS at 52wk High'] else 'no'}")
        print(f"      ATR% 21d={row['ATR% 21d']:.2f}  |  "
              f"DMA 20/50/200: {'✓' if row['Above 20 DMA'] else '✗'}/"
              f"{'✓' if row['Above 50 DMA'] else '✗'}/"
              f"{'✓' if row['Above 200 DMA'] else '✗'}  |  "
              f"Vol ratio {row['Vol Ratio (20d/60d)']:.2f}x")
        print(f"      Reasons: {row['Reasons']}")

    # ── Outputs ───────────────────────────────────────────────────────
    xlsx_path = OUT_DIR / f"outperformance_{TODAY}.xlsx"
    html_path = OUT_DIR / f"outperformance_{TODAY}.html"

    _write_excel(df, xlsx_path)
    print(f"\n  ✓  Excel → {xlsx_path}")

    _write_html(df, html_path)
    print(f"  ✓  HTML  → {html_path}\n")


# ── Excel ─────────────────────────────────────────────────────────────
def _write_excel(df, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    SUMMARY_COLS = [
        "Rank","Symbol","Sector","Post Return %",
        "% from 52wk High","Days since 52wk High","52wk Range Pos %",
        "1M Momentum %","3M Momentum %","6M Momentum %","12M Momentum %",
        "RS 1M (exc.ret %)","RS 3M (exc.ret %)","RS 6M (exc.ret %)","RS 12M (exc.ret %)",
        "RS Line Slope 21d","RS at 52wk High",
        "ATR% 21d","Vol Ratio (20d/60d)","Composite Score",
        "Reasons",
    ]
    FULL_COLS = SUMMARY_COLS[:] + [
        "Cutoff Close","Latest Close","Latest Date",
        "52wk High","52wk Low","% above 52wk Low","ATH","% from ATH",
        "3M Mom Pctile","RS 3M Pctile","Range Pctile",
        "20 DMA","50 DMA","200 DMA",
        "Above 20 DMA","Above 50 DMA","Above 200 DMA",
        "Streak (td)",
    ]

    HDR  = PatternFill("solid", fgColor="1F4E79")
    HFNT = Font(bold=True, color="FFFFFF", size=10)
    GOLD = PatternFill("solid", fgColor="FFF2CC")
    GRN  = PatternFill("solid", fgColor="C6EFCE")
    RED  = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(border_style="thin", color="C0C0C0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, cols):
        avail = [c for c in cols if c in df.columns]
        ws.append(avail)
        for cell in ws[1]:
            cell.fill = HDR; cell.font = HFNT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = bdr
        ws.row_dimensions[1].height = 40

        for _, row in df[avail].iterrows():
            ws.append([row[c] for c in avail])

        pr_i = avail.index("Post Return %") + 1
        for ri in range(2, ws.max_row + 1):
            v = ws.cell(ri, pr_i).value or 0
            fill = GRN if v > 10 else GOLD if v >= 0 else RED
            for ci in range(1, ws.max_column + 1):
                ws.cell(ri, ci).fill = fill

        pr_col = get_column_letter(pr_i)
        ws.conditional_formatting.add(
            f"{pr_col}2:{pr_col}{ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="num",   mid_value=0, mid_color="FFEB84",
                end_type="max",   end_color="63BE7B",
            )
        )
        widths = {"Reasons": 62, "Symbol": 14, "Sector": 22, "Latest Date": 13}
        for i, h in enumerate(avail, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 13)
        ws.freeze_panes = "D2"

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb.create_sheet("Summary"), SUMMARY_COLS)
    write_sheet(wb.create_sheet("Full Detail"), FULL_COLS)
    wb.save(path)


# ── HTML ──────────────────────────────────────────────────────────────
def _write_html(df, path):
    rows = ""
    for _, r in df.iterrows():
        ret = r["Post Return %"]
        bg  = "#c6efce" if ret > 10 else "#fff2cc" if ret >= 0 else "#ffc7ce"
        fc  = "#375623" if ret >= 0 else "#9c0006"

        def f(v, pct=False, pp=False, dp=2):
            if pd.isna(v): return "—"
            if pct: return f"{float(v):+.{dp}f}%"
            if pp:  return f"{float(v):+.{dp}f}pp"
            return f"{float(v):.{dp}f}"

        rows += f"""<tr style="background:{bg}">
          <td style="text-align:center;font-weight:bold">{int(r['Rank'])}</td>
          <td style="font-weight:bold">{r['Symbol']}</td>
          <td style="font-size:11px">{r.get('Sector','—')}</td>
          <td style="text-align:right;font-weight:bold;color:{fc}">{f(ret,pct=True)}</td>
          <td style="text-align:right">₹{f(r['Cutoff Close'])}</td>
          <td style="text-align:right">{f(r['% from 52wk High'],pct=True)}</td>
          <td style="text-align:right">{int(r['Days since 52wk High'])}td</td>
          <td style="text-align:right">{f(r['52wk Range Pos %'],dp=1)}%</td>
          <td style="text-align:right">{f(r['1M Momentum %'],pct=True)}</td>
          <td style="text-align:right">{f(r['3M Momentum %'],pct=True)}</td>
          <td style="text-align:right">{f(r['6M Momentum %'],pct=True)}</td>
          <td style="text-align:right">{f(r['12M Momentum %'],pct=True)}</td>
          <td style="text-align:right">{f(r['RS 1M (exc.ret %)'],pp=True)}</td>
          <td style="text-align:right">{f(r['RS 3M (exc.ret %)'],pp=True)}</td>
          <td style="text-align:right">{f(r['RS 6M (exc.ret %)'],pp=True)}</td>
          <td style="text-align:right">{f(r['RS 12M (exc.ret %)'],pp=True)}</td>
          <td style="text-align:right">{f(r['RS Line Slope 21d'],dp=5)}</td>
          <td style="text-align:center">{'✓' if r['RS at 52wk High'] else '—'}</td>
          <td style="text-align:right">{f(r['ATR% 21d'],dp=2)}</td>
          <td style="text-align:center">{'✓' if r['Above 20 DMA'] else '✗'}/{'✓' if r['Above 50 DMA'] else '✗'}/{'✓' if r['Above 200 DMA'] else '✗'}</td>
          <td style="text-align:right">{f(r['Vol Ratio (20d/60d)'],dp=2)}x</td>
          <td style="text-align:right;font-weight:bold">{f(r['Composite Score'],dp=2)}</td>
          <td style="font-size:11px">{r['Reasons']}</td>
        </tr>"""

    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>Outperformance Report {TODAY}</title>
<style>
body{{font-family:'Segoe UI',Arial,sans-serif;margin:0;background:#f0f2f8}}
.hdr{{background:linear-gradient(135deg,#1F4E79,#2E75B6);color:#fff;padding:22px 32px}}
.hdr h1{{margin:0 0 4px;font-size:19px}}.hdr p{{margin:0;font-size:12px;opacity:.85}}
.sub{{background:#dce6f1;padding:8px 32px;font-size:12px;color:#1F4E79;font-weight:bold}}
.wrap{{padding:20px 32px}}
table{{border-collapse:collapse;width:100%;font-size:12px;background:#fff;
       box-shadow:0 2px 10px rgba(0,0,0,.1)}}
th{{background:#1F4E79;color:#fff;padding:8px 7px;text-align:center;
    font-size:11px;white-space:nowrap}}
td{{padding:6px 7px;border-bottom:1px solid #e0e4ee;vertical-align:middle}}
tr:hover td{{filter:brightness(.96)}}
.leg{{margin-top:14px;font-size:11px;color:#555}}
.dot{{display:inline-block;width:11px;height:11px;border-radius:2px;
      margin-right:4px;vertical-align:middle}}
</style></head><body>
<div class="hdr">
  <h1>📈 Post-Cutoff Outperformance Report</h1>
  <p>Cutoff: <b>02 Apr 2026</b> &nbsp;|&nbsp;
     Window: <b>03 Apr → {TODAY}</b> &nbsp;|&nbsp;
     Universe: <b>{len(df)} stocks</b></p>
</div>
<div class="sub">
  RS benchmark: Nifty Smallcap 250 (excess return = stock% − index%) &nbsp;|&nbsp;
  Volatility: ATR% 21-day = ATR₂₁ / close × 100
</div>
<div class="wrap">
<table>
<thead><tr>
  <th>#</th><th>Symbol</th><th>Sector</th><th>Post Return</th>
  <th>Cutoff ₹</th>
  <th>Δ 52wkH</th><th>Days 52wkH</th><th>52wk Pos</th>
  <th>1M Mom</th><th>3M Mom</th><th>6M Mom</th><th>12M Mom</th>
  <th>RS 1M</th><th>RS 3M</th><th>RS 6M</th><th>RS 12M</th>
  <th>RS Slope</th><th>RS@52H</th>
  <th>ATR%</th><th>DMA 20/50/200</th><th>Vol Ratio</th>
  <th>Score</th><th>Reasons</th>
</tr></thead>
<tbody>{rows}</tbody>
</table>
<div class="leg">
  <span class="dot" style="background:#c6efce"></span>Return &gt;10%&nbsp;
  <span class="dot" style="background:#fff2cc"></span>Return 0–10%&nbsp;
  <span class="dot" style="background:#ffc7ce"></span>Negative &nbsp;|&nbsp;
  <b>RS (exc.ret)</b>: stock% − Nifty SC250% (pp = percentage points) &nbsp;|&nbsp;
  <b>RS Slope</b>: 21-day change in RS line (stock/index) &nbsp;|&nbsp;
  <b>RS@52H</b>: RS line at its own 52-week high &nbsp;|&nbsp;
  <b>ATR%</b>: 21-day ATR / close × 100
</div></div></body></html>"""

    with open(path, "w", encoding="utf-8") as fh:
        fh.write(html)


if __name__ == "__main__":
    analyse()
