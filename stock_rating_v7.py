# -*- coding: utf-8 -*-
"""
stock_rating_v7.py
─────────────────────────────────────────────────────────────────────
Rates NSE stocks on technical criteria as of a given AS-OF date,
using ONLY data up to that date.

SCORING:
  +4 pts  — Within 10% of 52-week high
  +2 pts  — Within 15% of 52-week high  (only if NOT within 10%)
  +4 pts  — 52W high made in last 10 TDs
  +2 pts  — 52W high made in last 15 TDs  (only if NOT in last 10 TDs)
  +2 pts  — BONUS: both within 10% AND high in last 10 TDs
  +1 pt   — Above 50-SMA
  +1 pt   — Above 21-EMA
  +1 pt   — Above 8-EMA
  +3 pts  — 8-EMA uptrend (last 5 TDs)
  +2 pts  — 21-EMA uptrend (last 10 TDs)
  +2 pts  — 8-EMA > 21-EMA > 50-SMA stack
  +3 pts  — Return from reset-date low is in top 10th percentile
  +6 pts  — Return from reset-date low is in top 5th percentile  (overrides above)
  +2 pts  — 1-day return in top 10th percentile
  -2 pts  — 1-day return in bottom 10th percentile
  +4 pts  — Green candles (since reset) in top 1st percentile
  +2 pts  — Green candles (since reset) in top 10th percentile  (exclusive with above)
  +2 pts  — Stock listed < 30 trading days
  +1 pt   — Stock listed < 60 trading days  (exclusive with above)
  +2 pts  — 1+ other stocks from same sector in top 25th percentile score
  +3 pts  — 3+ other stocks from same sector in top 25th percentile score
  +4 pts  — 4+ other stocks from same sector in top 25th percentile score
  +6 pts  — 5+ other stocks from same sector in top 25th percentile score
  -4 pts  — 1-day return > 5% AND close > previous day's high × 1.01 (spike penalty)
  +10 pts — Avg 21-TD turnover (vol×close) in top 30th percentile
  -5 pts  — Avg 21-TD turnover (vol×close) in bottom 30th percentile
  -6 pts  — ATR-14 in bottom 10th percentile (low volatility penalty)
  -3 pts  — ATR-14 in bottom 20th percentile (only if NOT bottom 10th)
  +4 pts  — 8-EMA > 21-EMA > 50-SMA stack held on >= 80% of last 63 TDs (3 months)
  +4 pts  — IPO (< 3 months old): close >= 15% above listing-day open (proxy issue price)
  +6 pts  — IPO (< 3 months old): close >= 30% above listing-day open  (overrides +4)
  SECTOR COMPOSITE ONLY: +3 bonus if sector has more than 3 stocks
  +5 pts  — 1Y-high-vol spike in last 10 TDs  (top 5% of 1Y daily volume)
  +6 pts  — same spike + close >3% on that day  (overrides +5)
  +3 pts  — 1Y-high-vol spike only in last 30 TDs  (tiered: if no 10D spike)
  +2 pts  — 1Y-high-vol spike only in last 60 TDs  (tiered: if no 10/30D spike)

REQUIREMENTS:
    pip install kiteconnect pandas openpyxl

USAGE:
    python stock_rating_v7.py --as-of 2021-04-12 --reset 2021-03-01 --symbols stocks.txt

ARGUMENTS:
    --as-of     YYYY-MM-DD     Rating date. Default: parsed from symbols filename, then today
    --reset     YYYY-MM-DD     Return measured from DAY'S LOW on this date
    --symbols   PATH           Text file with NSE symbols (one per line, NSE: prefix optional)
    --out       PATH           Output directory. Default: E:\reports
    --token     PATH           kite_token.txt path. Default: kite_token.txt
─────────────────────────────────────────────────────────────────────
"""

import os
import sys
import time
import argparse
import subprocess
import webbrowser
import pandas as pd
from datetime import datetime, date, timedelta
from collections import defaultdict
from kiteconnect import KiteConnect

# ── Fixed parameters ──────────────────────────────────────────────
DAYS_52W          = 365
TRADING_DAYS_10   = 10
TRADING_DAYS_15   = 15
TRADING_DAYS_5    = 5      # kept for backward-compat imports
SMA_PERIOD        = 50
EMA_21_PERIOD     = 21
EMA_8_PERIOD      = 8
PCT_FROM_HIGH     = 10.0   # ≤10% proximity → +4
PCT_FROM_HIGH_15  = 15.0   # ≤15% proximity → +2
HIGH_BOTH_BONUS   = 2      # extra pts when ≤10% AND high in ≤10 TDs
EMA_8_UPTREND_LB  = 6     # 6 values = 5 consecutive rising pairs
EMA_21_UPTREND_LB = 11    # 11 values = 10 consecutive rising pairs
ATR_PERIOD        = 14    # Average True Range period

# Percentile thresholds
RET_LOW_P5        = 5.0    # top 5th percentile → 6 pts
RET_LOW_P10       = 10.0   # top 10th percentile → 3 pts
RET_1D_TOP_P10    = 10.0   # top 10th percentile 1D → +2 pts
RET_1D_BOT_P10    = 10.0   # bottom 10th percentile 1D → -2 pts
GREEN_TOP_P1      = 1.0    # top 1st percentile green candles → 4 pts
GREEN_TOP_P10     = 10.0   # top 10th percentile green candles → 2 pts
SECTOR_TOP_P25    = 25.0   # sector bonus uses top 25th percentile score

NEW_LISTING_30    = 30
NEW_LISTING_60    = 60

# Volume spike thresholds (1-year highest volume conditions)
VOL_1Y_DAYS       = 252  # ~1 year of trading days for baseline
VOL_WIN_10        = 10   # last 10 TDs window
VOL_WIN_30        = 30   # last 30 TDs window
VOL_WIN_60        = 60   # last 60 TDs window
VOL_SPIKE_PCT     = 5.0  # top 5% of 1Y daily volume = "one of the highest"
VOL_BIG_MOVE_PCT  = 3.0  # % close return threshold for "big price move" bonus
VOL_VBIG_MOVE_PCT = 6.0  # % close return threshold for larger move bonus

# Turnover percentile (avg daily value traded over last 21 TDs)
TURNOVER_DAYS    = 42    # trading days window for avg turnover
TURNOVER_TOP_P30 = 30.0  # top 30th percentile → +10 pts
TURNOVER_BOT_P30 = 30.0  # bottom 30th percentile → -5 pts

# ATR percentile (higher ATR preferred — low ATR penalised)
ATR_BOT_P10 = 10.0  # bottom 10th percentile → -6 pts
ATR_BOT_P20 = 20.0  # bottom 20th percentile → -3 pts (only if NOT bot 10th)

# Uptrend consistency (8EMA > 21EMA > 50SMA held on >= X% of days in last Y TDs)
UPTREND_CON_WINDOW  = 63    # ~3 months of trading days
UPTREND_CON_MIN_PCT = 80.0  # minimum % of days in uptrend stack → +4 pts

# Sector size bonus in composite score
SECTOR_SIZE_BONUS_MIN = 3    # sector must have MORE than this many stocks
SECTOR_SIZE_BONUS_PTS = 3    # bonus points added to composite

# IPO performance bonus (only for stocks < 3 months / ~63 TDs old)
# Issue price proxy = first available day's open price in the fetched data
IPO_AGE_TDS      = 63    # ~3 months of trading days
IPO_PERF_HIGH    = 30.0  # >= 30% above proxy → +6 pts
IPO_PERF_LOW     = 15.0  # >= 15% above proxy → +4 pts (exclusive with HIGH)
IPO_PERF_HIGH_PT = 6
IPO_PERF_LOW_PT  = 4

# ── BSE scrip master ──────────────────────────────────────────────
# BSE publishes a free CSV with columns:
#   Security Code | Security Id | Security Name | Status | Group |
#   Face Value | ISIN No | Industry | Instrument
# "Security Id" == BSE trading symbol (same as NSE symbol for most stocks).
# "Industry"    == Indian-market terminology, e.g. "AUTO ANCILLARIES",
#                  "NON FERROUS METALS", "IT-SOFTWARE".
#
# Download manually from:
#   https://www.bseindia.com/corporates/List_Scrips.html
#   (select Segment=Equity, Status=Active, click Download)
# Save as  bse_master.csv  next to this script, OR pass --bse-master <path>.
BSE_MASTER_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "bse_master.csv"
)
# Direct-download URLs tried in order (BSE may block without a session cookie;
# manual download is the reliable fallback).
BSE_MASTER_URLS = [
    "https://www.bseindia.com/corporates/List_Scrips.aspx?bFlag=&Group=",
    "https://www.bseindia.com/downloads1/List_of_companies.csv",
]
# ─────────────────────────────────────────────────────────────────


# ── Extract date from a file name ────────────────────────────────
import re as _re

def _date_from_filename(filepath):
    """Try to parse an as-of date from the base name of *filepath*.

    Patterns tried (first match wins):
      YYYY-MM-DD   e.g. stocks_2025-12-08.txt
      YYYY_MM_DD   e.g. stocks_2025_12_08.txt
      YYYYMMDD     e.g. stocks_20251208.txt
      DDMMMYYYY    e.g. stocks_08Dec2025.txt  (case-insensitive)
      DD-Mon-YYYY  e.g. 19-jun-2025.txt       (case-insensitive, with separators)
      DD-MM-YYYY   e.g. stocks_08-12-2025.txt
      DD_MM_YYYY   e.g. stocks_08_12_2025.txt

    Returns a date object on success, or None if no date is found.
    """
    name = os.path.splitext(os.path.basename(filepath))[0]
    patterns = [
        (r'(\d{4})[_-](\d{2})[_-](\d{2})',           lambda m: date(int(m[1]), int(m[2]), int(m[3]))),
        (r'(\d{8})',                                   lambda m: datetime.strptime(m[1], "%Y%m%d").date()),
        (r'(\d{2})[_-]([A-Za-z]{3})[_-](\d{4})',     lambda m: datetime.strptime(f"{m[1]}-{m[2].capitalize()}-{m[3]}", "%d-%b-%Y").date()),
        (r'(\d{2})([A-Za-z]{3})(\d{4})',              lambda m: datetime.strptime(f"{m[1]}{m[2].capitalize()}{m[3]}", "%d%b%Y").date()),
        (r'(\d{2})[_-](\d{2})[_-](\d{4})',           lambda m: date(int(m[3]), int(m[2]), int(m[1]))),
    ]
    for pat, parser in patterns:
        m = _re.search(pat, name)
        if m:
            try:
                return parser(m)
            except (ValueError, TypeError):
                continue
    return None


# ── CLI ───────────────────────────────────────────────────────────
def parse_args():
    parser = argparse.ArgumentParser(
        description="Rate NSE stocks on technical criteria."
    )
    parser.add_argument("--as-of", "-a",
        default=None,
        help="Rating date YYYY-MM-DD. Default: today")
    parser.add_argument("--reset", "-r", required=True,
        help="Reset date YYYY-MM-DD - return from DAY'S LOW, green candles counted from here")
    parser.add_argument("--symbols", "-s", default=None,
        help="Optional symbols file path. Default: auto-resolve gmlist_<as-of>.txt from reports folder")
    parser.add_argument("--out", "-o", default=r"E:\reports",
        help=r"Output directory. Default: E:\reports")
    parser.add_argument("--token", "-t", default="kite_token.txt",
        help="kite_token.txt path. Default: kite_token.txt")
    parser.add_argument("--bse-master", default=None,
        metavar="PATH",
        help="Path to BSE scrip master CSV (downloaded from bseindia.com). "
             "Default: bse_master.csv next to this script.")
    args = parser.parse_args()

    if args.as_of:
        try:
            as_of = datetime.strptime(args.as_of, "%Y-%m-%d").date()
        except ValueError:
            print(f"ERROR: Invalid --as-of '{args.as_of}'. Use YYYY-MM-DD.")
            sys.exit(1)
    else:
        as_of = date.today()
        print(f"  As-of date not passed - using today: {as_of.strftime('%Y-%m-%d')}")

    reset = None
    if args.reset:
        try:
            reset = datetime.strptime(args.reset, "%Y-%m-%d").date()
        except ValueError:
            print(f"ERROR: Invalid --reset '{args.reset}'. Use YYYY-MM-DD.")
            sys.exit(1)
        if reset >= as_of:
            print(f"ERROR: --reset ({reset}) must be before --as-of ({as_of}).")
            sys.exit(1)

    symbols_path = resolve_symbols_file(args.symbols, as_of)
    return as_of, reset, symbols_path, args.out, args.token, args.bse_master


def resolve_symbols_file(symbols_arg, as_of):
    """
    Resolve the symbols file path from the as-of date by default.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    reports_dir = os.path.join(script_dir, "reports")
    asof_name = f"gmlist_{as_of.strftime('%d%b%Y')}.txt"

    names_to_try = [asof_name]
    if symbols_arg:
        candidate = os.path.expanduser(str(symbols_arg).strip())
        if os.path.exists(candidate):
            return candidate
        candidate_name = os.path.basename(candidate)
        if candidate_name and candidate_name not in names_to_try:
            names_to_try.insert(0, candidate_name)

    paths_to_try = []
    for name in names_to_try:
        paths_to_try.extend([
            name,
            os.path.join(os.getcwd(), name),
            os.path.join(os.getcwd(), "reports", name),
            os.path.join(reports_dir, name),
        ])

    for path_candidate in paths_to_try:
        if os.path.exists(path_candidate):
            return path_candidate

    print("ERROR: Symbols file not found.")
    print(f"  Expected  : {asof_name}")
    if symbols_arg:
        print(f"  Requested : {symbols_arg}")
    print(f"  Tried     : {paths_to_try}")
    sys.exit(1)


def read_symbols(filepath):
    if not os.path.exists(filepath):
        print(f"ERROR: Symbols file not found: {filepath}")
        sys.exit(1)
    symbols = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            sym = line.upper()
            for prefix in ["NSE:", "BSE:", "NSE/", "BSE/"]:
                if sym.startswith(prefix):
                    sym = sym[len(prefix):]
                    break
            symbols.append(sym)
    # deduplicate preserving order
    seen, unique = set(), []
    for s in symbols:
        if s not in seen:
            seen.add(s)
            unique.append(s)
    print(f"  Loaded {len(unique)} symbols from {os.path.basename(filepath)}")
    return unique


# ── Kite helpers ─────────────────────────────────────────────────
def read_token_file(filepath):
    if not os.path.exists(filepath):
        print(f"ERROR: {filepath} not found. Run kite_get_access_token.py.")
        sys.exit(1)
    data = {}
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            data[k.strip()] = v.strip()
    if "API_KEY" not in data or "ACCESS_TOKEN" not in data:
        print(f"ERROR: {filepath} missing API_KEY or ACCESS_TOKEN.")
        sys.exit(1)
    if "GENERATED" in data:
        gen = data["GENERATED"][:10]
        if gen != datetime.now().strftime("%Y-%m-%d"):
            print(f"WARNING: Token from {gen} — may be expired.")
    return data


def get_kite_session(token_file):
    creds = read_token_file(token_file)
    kite  = KiteConnect(api_key=creds["API_KEY"])
    kite.set_access_token(creds["ACCESS_TOKEN"])
    print(f"  Kite OK  |  Generated: {creds.get('GENERATED','?')}")
    return kite


def build_token_map(kite, symbols):
    print("  Loading NSE instruments...")
    df        = pd.DataFrame(kite.instruments("NSE"))
    token_map = {}
    missing   = []
    for sym in symbols:
        row = df[df["tradingsymbol"] == sym]
        if not row.empty:
            token_map[sym] = {
                "token":        int(row.iloc[0]["instrument_token"]),
                "listing_date": row.iloc[0].get("listing_date", None),
            }
        else:
            missing.append(sym)
    if missing:
        print(f"  WARNING — not found on NSE: {missing}")
    print(f"  {len(token_map)} tokens mapped")
    return token_map


def fetch_daily(kite, token, sym, from_date, to_date):
    try:
        data = kite.historical_data(
            instrument_token=token,
            from_date=datetime.combine(from_date, datetime.min.time()),
            to_date=datetime.combine(to_date,   datetime.min.time()),
            interval="day", continuous=False, oi=False,
        )
        if not data:
            return None
        df = pd.DataFrame(data)
        df["date"] = pd.to_datetime(df["date"]).dt.date
        df.set_index("date", inplace=True)
        return df.sort_index()
    except Exception as e:
        print(f"    ERROR {sym}: {e}")
        return None


# ── Indicators ────────────────────────────────────────────────────
def calc_ema(series, period):
    return series.ewm(span=period, adjust=False).mean()


def is_uptrend(series, lookback):
    vals = series.dropna().tail(lookback).values
    if len(vals) < lookback:
        return False
    return all(vals[i] > vals[i - 1] for i in range(1, len(vals)))


def count_green_candles(df, from_date):
    subset = df[df.index >= from_date]
    if subset.empty:
        return 0, 0
    total = len(subset)
    green = int((subset["close"] > subset["open"]).sum())
    return green, total


# ── Percentile helper ─────────────────────────────────────────────
def percentile_threshold(values, pct, top=True):
    """
    Returns the threshold value for the given percentile.
    top=True  → top X% (descending)
    top=False → bottom X% (ascending)
    """
    if not values:
        return None
    n = len(values)
    k = max(1, int(n * pct / 100))
    if top:
        return sorted(values, reverse=True)[k - 1]
    else:
        return sorted(values)[k - 1]


# ── Rate one stock (base criteria; relative scores applied later) ─
def rate_stock(sym, df, as_of, reset_date, sector_map, listing_date=None):
    df = df[df.index <= as_of].copy()
    if df.empty or len(df) < 2:
        return None

    close  = float(df.iloc[-1]["close"])
    n_days = len(df)

    # 3-month, 6-month, 12-month returns from as_of (~91 / ~182 / ~365 calendar days back)
    _cutoff_3m  = as_of - timedelta(days=91)
    _cutoff_6m  = as_of - timedelta(days=182)
    _cutoff_12m = as_of - timedelta(days=365)
    _df_3m  = df[df.index <= _cutoff_3m]
    _df_6m  = df[df.index <= _cutoff_6m]
    _df_12m = df[df.index <= _cutoff_12m]
    ret_3m  = round((close / float(_df_3m.iloc[-1]["close"])  - 1) * 100, 2) if not _df_3m.empty  else None
    ret_6m  = round((close / float(_df_6m.iloc[-1]["close"])  - 1) * 100, 2) if not _df_6m.empty  else None
    ret_12m = round((close / float(_df_12m.iloc[-1]["close"]) - 1) * 100, 2) if not _df_12m.empty else None

    # For stocks with limited history we cannot compute all MAs.
    # Conditions that require more data than available are ASSUMED favourable
    # (user instruction: new stocks are treated as above all MAs/trends).
    is_limited = n_days < SMA_PERIOD   # True  → fewer than 50 TDs of data

    # 1. 52W-high proximity: ≤10% → +4, ≤15% → +2
    cutoff_52w = as_of - timedelta(days=DAYS_52W)
    high_52w   = float(df[df.index >= cutoff_52w]["high"].max())
    pct_hi     = ((high_52w - close) / high_52w) * 100
    within_10  = pct_hi <= PCT_FROM_HIGH
    within_15  = pct_hi <= PCT_FROM_HIGH_15
    if within_10:
        s1 = 4; prox_lbl = "≤10%"
    elif within_15:
        s1 = 2; prox_lbl = "≤15%"
    else:
        s1 = 0; prox_lbl = ">15%"

    # 2. 52W-high recency: ≤10 TDs → +4, ≤15 TDs → +2 (mutually exclusive)
    hi10td = float(df.tail(TRADING_DAYS_10)["high"].max()) >= high_52w
    hi15td = float(df.tail(TRADING_DAYS_15)["high"].max()) >= high_52w
    if hi10td:
        s2 = 4; hi_label = "≤10 TDs"
    elif hi15td:
        s2 = 2; hi_label = "≤15 TDs"
    else:
        s2 = 0; hi_label = "NO"

    # Bonus: both within 10% proximity AND high in last 10 TDs → +2
    s_hi_bonus   = HIGH_BOTH_BONUS if (within_10 and hi10td) else 0
    hi_bonus_lbl = "YES" if s_hi_bonus else "NO"

    # 3. Above 50-SMA (+1)
    sma50_s = df["close"].rolling(SMA_PERIOD).mean()   # full series needed for consistency check
    if n_days >= SMA_PERIOD:
        sma50          = float(sma50_s.iloc[-1])
        s3             = 1 if close > sma50 else 0
        above_sma_lbl  = "YES" if s3 else "NO"
    else:
        sma50          = None          # not computable
        s3             = 1             # assumed above (limited history)
        above_sma_lbl  = "ASSUMED"

    # 4. Above 21-EMA (+1)
    ema21_s = calc_ema(df["close"], EMA_21_PERIOD)
    ema21   = float(ema21_s.iloc[-1])
    if n_days >= EMA_21_PERIOD:
        s4             = 1 if close > ema21 else 0
        above_ema21_lbl = "YES" if s4 else "NO"
    else:
        s4             = 1             # assumed above
        above_ema21_lbl = "ASSUMED"

    # 5. Above 8-EMA (+1)
    ema8_s = calc_ema(df["close"], EMA_8_PERIOD)
    ema8   = float(ema8_s.iloc[-1])
    if n_days >= EMA_8_PERIOD:
        s5             = 1 if close > ema8 else 0
        above_ema8_lbl  = "YES" if s5 else "NO"
    else:
        s5             = 1             # assumed above
        above_ema8_lbl  = "ASSUMED"

    # 6. 8-EMA uptrend last 5 TDs (+3)
    if n_days >= EMA_8_UPTREND_LB:
        s6              = 3 if is_uptrend(ema8_s, EMA_8_UPTREND_LB) else 0
        ema8_trend_lbl  = "YES" if s6 else "NO"
    else:
        s6              = 3            # assumed uptrend
        ema8_trend_lbl  = "ASSUMED"

    # 7. 21-EMA uptrend last 10 TDs (+2)
    if n_days >= EMA_21_UPTREND_LB:
        s7              = 2 if is_uptrend(ema21_s, EMA_21_UPTREND_LB) else 0
        ema21_trend_lbl = "YES" if s7 else "NO"
    else:
        s7              = 2            # assumed uptrend
        ema21_trend_lbl = "ASSUMED"

    # 8. EMA stack 8 > 21 > 50 (+2)
    if n_days >= SMA_PERIOD and sma50 is not None:
        s8              = 2 if (ema8 > ema21 > sma50) else 0
        stack_lbl       = "YES" if s8 else "NO"
    else:
        s8              = 2            # assumed stacked (limited history)
        stack_lbl       = "ASSUMED"

    # 9. Return from LOW of reset date (score applied later)
    ret_from_low = None
    reset_low    = None
    reset_actual = None
    reset_cands  = df[df.index >= reset_date]
    if not reset_cands.empty:
        reset_actual = reset_cands.index[0]
        reset_low    = float(reset_cands.iloc[0]["low"])
        ret_from_low = ((close - reset_low) / reset_low) * 100

    # 10. 1-day return (score applied later)
    ret_1d = None
    if len(df) >= 2:
        prev_close = float(df.iloc[-2]["close"])
        ret_1d     = ((close - prev_close) / prev_close) * 100

    # 13. Spike penalty: 1D return > 5% AND close > prev day high + 1%
    prev_high     = float(df.iloc[-2]["high"]) if len(df) >= 2 else None
    spike_penalty = False
    s_spike       = 0
    if ret_1d is not None and prev_high is not None:
        if ret_1d > 5.0 and close > prev_high * 1.01:
            spike_penalty = True
            s_spike       = -4

    # 14. Volume spike: compare window peak vs 1-year top-5% threshold
    #   Tiered / exclusive: highest-priority window wins.
    #   10D window: +6 if spike + big price move (>3%), +5 if spike alone.
    #   30D window (only if no 10D spike): +3.
    #   60D window (only if no 10D/30D spike): +2.
    vol_enough_1y = len(df) >= VOL_1Y_DAYS

    # 1-year volume 95th-percentile threshold ("one of the highest")
    if vol_enough_1y:
        vols_1y  = df["volume"].tail(VOL_1Y_DAYS)
        k        = max(1, int(len(vols_1y) * VOL_SPIKE_PCT / 100))
        vol_1y_p95 = float(sorted(vols_1y, reverse=True)[k - 1])
    else:
        vol_1y_p95 = None

    def _window_spike(n):
        """Return (spike_flag, peak_vol, move_pct) for last n TDs."""
        if vol_1y_p95 is None or len(df) < n:
            return False, None, None
        window = df.tail(n)
        peak_vol = float(window["volume"].max())
        if peak_vol < vol_1y_p95:
            return False, peak_vol, None
        # Spike confirmed – check if the peak day also had a big price move
        peak_idx  = window["volume"].idxmax()
        peak_pos  = df.index.get_loc(peak_idx)
        if peak_pos > 0:
            prev_c = float(df.iloc[peak_pos - 1]["close"])
            peak_c = float(df.loc[peak_idx, "close"])
            move_pct = ((peak_c - prev_c) / prev_c * 100) if prev_c else None
        else:
            move_pct = None
        return True, peak_vol, move_pct

    spike_10d, peak10, mv_10d = _window_spike(VOL_WIN_10)
    spike_30d, peak30, mv_30d = _window_spike(VOL_WIN_30)
    spike_60d, peak60, mv_60d = _window_spike(VOL_WIN_60)

    # Build labels (show peak volume for context)
    def _spike_label(spike, peak, move_pct=None):
        if vol_1y_p95 is None:
            return "N/A"
        if not spike:
            return "NO"
        if move_pct is not None and move_pct > VOL_VBIG_MOVE_PCT:
            return f"YES+>6%Move ({peak:,.0f})"
        if move_pct is not None and move_pct >= VOL_BIG_MOVE_PCT:
            return f"YES+>=3%Move ({peak:,.0f})"
        return f"YES ({peak:,.0f})"

    vol_10d_label = _spike_label(spike_10d, peak10, mv_10d)
    vol_30d_label = _spike_label(spike_30d, peak30, mv_30d)
    vol_60d_label = _spike_label(spike_60d, peak60, mv_60d)

    # Scoring: tiered/exclusive — most recent window wins
    if spike_10d:
        s_vol10 = 5
        if mv_10d is not None and mv_10d > VOL_VBIG_MOVE_PCT:
            s_vol10 += 4
        elif mv_10d is not None and mv_10d >= VOL_BIG_MOVE_PCT:
            s_vol10 += 2
        s_vol30, s_vol60 = 0, 0
    elif spike_30d:
        s_vol10, s_vol30, s_vol60 = 0, 4, 0
        if mv_30d is not None and mv_30d > VOL_VBIG_MOVE_PCT:
            s_vol30 += 4
        elif mv_30d is not None and mv_30d >= VOL_BIG_MOVE_PCT:
            s_vol30 += 2
    elif spike_60d:
        s_vol10, s_vol30, s_vol60 = 0, 0, 2
        if mv_60d is not None and mv_60d >= VOL_BIG_MOVE_PCT:
            s_vol60 += 2
    else:
        s_vol10, s_vol30, s_vol60 = 0, 0, 0

    s_vol = s_vol10 + s_vol30 + s_vol60

    # 15. Avg / median 42-TD turnover = value traded over last 42 TDs (in Crore)
    turnover_window = df.tail(TURNOVER_DAYS)
    avg_turnover_42d = float((turnover_window["volume"] * turnover_window["close"]).mean()) / 1e7 \
                       if len(turnover_window) >= 1 else None   # Crore rupees
    median_turnover_42d = float((turnover_window["volume"] * turnover_window["close"]).median()) / 1e7 \
                          if len(turnover_window) >= 1 else None   # Crore rupees

    # 11. Green candles since reset date
    green_count, total_count = count_green_candles(df, reset_date)
    green_pct = round(green_count / total_count * 100, 1) if total_count > 0 else 0.0

    # 12. New listing bonus / IPO age.
    # Kite's NSE instrument dump often has no listing_date, so fall back to
    # observed history length to avoid missing recent IPOs like OMNI while
    # also keeping older stocks like GESHIP out of IPO scoring.
    td_since_list_observed = n_days
    if listing_date and isinstance(listing_date, date):
        td_since_list_meta = max(0, (as_of - listing_date).days * 5 // 7)
        td_since_list = min(td_since_list_meta, td_since_list_observed)
    else:
        td_since_list = td_since_list_observed
    if td_since_list < NEW_LISTING_30:
        s_listing = 2; listing_cat = f"< {NEW_LISTING_30} TDs"
    elif td_since_list < NEW_LISTING_60:
        s_listing = 1; listing_cat = f"< {NEW_LISTING_60} TDs"
    else:
        s_listing = 0; listing_cat = "Established"

    # 17. IPO performance bonus (< 3 months old only)
    #     Issue price proxy = first available open in the observed history.
    s_ipo           = 0
    ipo_proxy       = None
    ipo_perf_pct    = None
    if td_since_list < IPO_AGE_TDS and not df.empty:
        ipo_proxy = float(df.iloc[0]["open"])
        if ipo_proxy and ipo_proxy > 0:
            ipo_perf_pct = round((close - ipo_proxy) / ipo_proxy * 100, 2)
            if ipo_perf_pct >= IPO_PERF_HIGH:
                s_ipo = IPO_PERF_HIGH_PT          # +6
            elif ipo_perf_pct >= IPO_PERF_LOW:
                s_ipo = IPO_PERF_LOW_PT           # +4

    # 16. Uptrend consistency: 8EMA > 21EMA > 50SMA on >= 80% of last 63 TDs (+4)
    uptrend_con_pct = None
    s_uptrend_con   = 0
    if n_days >= SMA_PERIOD:
        w8   = ema8_s.tail(UPTREND_CON_WINDOW)
        w21  = ema21_s.tail(UPTREND_CON_WINDOW)
        w50  = sma50_s.tail(UPTREND_CON_WINDOW)
        mask = w8.notna() & w21.notna() & w50.notna()
        n_valid = int(mask.sum())
        if n_valid > 0:
            n_up = int(((w8[mask] > w21[mask]) & (w21[mask] > w50[mask])).sum())
            uptrend_con_pct = round(n_up / n_valid * 100, 1)
            s_uptrend_con   = 4 if uptrend_con_pct >= UPTREND_CON_MIN_PCT else 0
    uptrend_con_lbl = f"{uptrend_con_pct:.1f}%" if uptrend_con_pct is not None else "N/A"
    uptrend_con_yn  = ("YES" if s_uptrend_con else "NO") if uptrend_con_pct is not None else "N/A"

    # 18. Average True Range (ATR-14) — informational, no score impact
    atr_val = None
    atr_pct = None
    if len(df) >= ATR_PERIOD + 1:
        prev_close = df["close"].shift(1)
        tr = pd.concat([
            df["high"] - df["low"],
            (df["high"] - prev_close).abs(),
            (df["low"]  - prev_close).abs(),
        ], axis=1).max(axis=1)
        atr_series = tr.rolling(ATR_PERIOD).mean()
        last_atr   = atr_series.iloc[-1]
        if pd.notna(last_atr):
            atr_val = round(float(last_atr), 2)
            atr_pct = round(atr_val / close * 100, 2) if close else None

    base = s1 + s2 + s_hi_bonus + s3 + s4 + s5 + s6 + s7 + s8 + s_listing + s_spike + s_vol + s_uptrend_con + s_ipo

    return {
        # Identifiers
        "_sym":                 sym,
        "Symbol":               f"NSE:{sym}",
        "Sector":               sector_map.get(sym, "Unknown"),
        "Close":                round(close, 2),
        # 52W
        "52W High":             round(high_52w, 2),
        "% From 52W High":      round(pct_hi, 2),
        "Within 10% of 52W":   prox_lbl,       # "≤10%", "≤15%", or ">15%"
        "Score: Within 10%":    s1,             # 4 / 2 / 0
        "52W Hi 5/10 TDs":      hi_label,       # "≤10 TDs", "≤15 TDs", or "NO"
        "Score: 52W Hi":        s2,             # 4 / 2 / 0
        "52W Hi Both Bonus":   hi_bonus_lbl,   # "YES" / "NO"
        "Score: 52W Bonus":    s_hi_bonus,      # 2 / 0
        # MAs
        "50-SMA":               round(sma50, 2) if sma50 is not None else None,
        "Above 50-SMA":         above_sma_lbl,
        "Score: 50-SMA":        s3,
        "21-EMA":               round(ema21, 2),
        "Above 21-EMA":         above_ema21_lbl,
        "Score: 21-EMA":        s4,
        "8-EMA":                round(ema8, 2),
        "Above 8-EMA":          above_ema8_lbl,
        "Score: 8-EMA":         s5,
        "8-EMA Uptrend 5TD":    ema8_trend_lbl,
        "Score: 8EMA Trend":    s6,
        "21-EMA Uptrend 10TD":  ema21_trend_lbl,
        "Score: 21EMA Trend":   s7,
        "EMA Stack 8>21>50":    stack_lbl,
        "Score: EMA Stack":     s8,
        # Reset return (relative, filled later)
        "Reset Date Used":      str(reset_actual) if reset_actual else "N/A",
        "Reset Day Low":        round(reset_low, 2) if reset_low else None,
        "Ret From Low %":       round(ret_from_low, 2) if ret_from_low is not None else None,
        "Ret Top 5th Pct":     "TBD",
        "Ret Top 10th Pct":    "TBD",
        "Score: Low Return":    0,
        # 1D return (relative, filled later)
        "1D Return %":          round(ret_1d, 2) if ret_1d is not None else None,
        "1D Top 10th Pct":     "TBD",
        "1D Bot 10th Pct":     "TBD",
        "Score: 1D Return":     0,
        # Spike penalty: 1D > 5% and close > prev high + 1%
        "Prev Day High":        round(prev_high, 2) if prev_high is not None else None,
        "1D>5%&Close>PrevHi+1%": "YES" if spike_penalty else "NO",
        "Score: Spike":         s_spike,
        # Volume spike vs 1-year highest (top 5th percentile)
        "1Y Vol P95":           round(vol_1y_p95) if vol_1y_p95 else None,
        "10D Vol Spike":        vol_10d_label,
        "Score: Vol 10D":       s_vol10,
        "30D Vol Spike":        vol_30d_label,
        "Score: Vol 30D":       s_vol30,
        "60D Vol Spike":        vol_60d_label,
        "Score: Vol 60D":       s_vol60,
        # Green candles (relative, filled later)
        "Green Candles":        green_count,
        "Total Candles":        total_count,
        "Green %":              green_pct,
        "Green Top 1st Pct":   "TBD",
        "Green Top 10th Pct":  "TBD",
        "Score: Green":         0,
        "Green Candle Rank":    0,
        # Listing
        "Listing Category":     listing_cat,
        "Score: Listing":       s_listing,
        # Sector bonus (filled later)
        "Sector Top25 Count":   0,
        "Score: Sector":        0,
        # Uptrend consistency: 8EMA>21EMA>50SMA for >= 80% of last 63 TDs
        "Uptrend Con %":        uptrend_con_pct,
        "Uptrend Con Lbl":      uptrend_con_lbl,
        "Uptrend >=80% 3M":     uptrend_con_yn,
        "Score: Uptrend Con":   s_uptrend_con,
        # IPO performance (only for stocks < 63 TDs old; N/A otherwise)
        "IPO Issue Proxy":      round(ipo_proxy, 2)    if ipo_proxy    is not None else None,
        "IPO Perf %":           ipo_perf_pct,
        "Score: IPO":           s_ipo,
        # Avg 42D turnover (score filled later via percentile)
        "Avg Turnover 42D":     round(avg_turnover_42d, 2) if avg_turnover_42d is not None else None,
        "Median Turnover 42D":  round(median_turnover_42d, 2) if median_turnover_42d is not None else None,
        "Turnover Top30Pct":    "TBD",
        "Turnover Bot30Pct":    "TBD",
        "Score: Turnover":      0,
        # 3M / 6M / 12M absolute returns (used by sector_discovery & fno_scanner for percentile scoring)
        "3M Return %":          ret_3m,
        "6M Return %":          ret_6m,
        "12M Return %":         ret_12m,
        # ATR (higher is better; relative score applied later)
        "ATR 14D":              atr_val,
        "ATR %":                atr_pct,
        "ATR Bot10Pct":         "TBD",
        "ATR Bot20Pct":         "TBD",
        "Score: ATR":           0,
        # Total
        "Limited History":      "YES" if is_limited else "NO",
        "TOTAL SCORE":          base,
        "RATING":              "",
    }


def get_rating(score):
    if score >= 25: return "STRONG BUY ★★★"
    if score >= 18: return "BUY ★★"
    if score >= 11: return "WEAK BUY ★"
    if score >= 5:  return "NEUTRAL"
    return "WEAK"


# ── Apply all relative/percentile scores ─────────────────────────
def apply_relative_scores(results):
    n = len(results)

    # ── Return from low: top 5th → 6pts, top 10th → 3pts ────────
    valid_low = [(i, r) for i, r in enumerate(results)
                 if r["Ret From Low %"] is not None]
    if valid_low:
        rets   = [r["Ret From Low %"] for _, r in valid_low]
        th_5   = percentile_threshold(rets, RET_LOW_P5,  top=True)
        th_10  = percentile_threshold(rets, RET_LOW_P10, top=True)
        print(f"  Low-ret  top-5th: {th_5:.2f}%  top-10th: {th_10:.2f}%")
        for i, r in valid_low:
            rv = r["Ret From Low %"]
            if rv >= th_5:
                results[i]["Ret Top 5th Pct"]  = "YES"
                results[i]["Ret Top 10th Pct"] = "YES (via top5)"
                results[i]["Score: Low Return"] = 6
            elif rv >= th_10:
                results[i]["Ret Top 5th Pct"]  = "NO"
                results[i]["Ret Top 10th Pct"] = "YES"
                results[i]["Score: Low Return"] = 3
            else:
                results[i]["Ret Top 5th Pct"]  = "NO"
                results[i]["Ret Top 10th Pct"] = "NO"
                results[i]["Score: Low Return"] = 0
            results[i]["TOTAL SCORE"] += results[i]["Score: Low Return"]
    for i, r in enumerate(results):
        if r["Ret From Low %"] is None:
            results[i]["Ret Top 5th Pct"]  = "N/A"
            results[i]["Ret Top 10th Pct"] = "N/A"

    # ── 1D return: top 10th → +2, bottom 10th → -2 ───────────────
    valid_1d = [(i, r) for i, r in enumerate(results)
                if r["1D Return %"] is not None]
    if valid_1d:
        rets_1d  = [r["1D Return %"] for _, r in valid_1d]
        th_top   = percentile_threshold(rets_1d, RET_1D_TOP_P10, top=True)
        th_bot   = percentile_threshold(rets_1d, RET_1D_BOT_P10, top=False)
        print(f"  1D-ret   top-10th: {th_top:.2f}%  bot-10th: {th_bot:.2f}%")
        for i, r in valid_1d:
            rv = r["1D Return %"]
            if rv >= th_top:
                results[i]["1D Top 10th Pct"] = "YES"
                results[i]["1D Bot 10th Pct"] = "NO"
                results[i]["Score: 1D Return"] = 2
            elif rv <= th_bot:
                results[i]["1D Top 10th Pct"] = "NO"
                results[i]["1D Bot 10th Pct"] = "YES"
                results[i]["Score: 1D Return"] = -2
            else:
                results[i]["1D Top 10th Pct"] = "NO"
                results[i]["1D Bot 10th Pct"] = "NO"
                results[i]["Score: 1D Return"] = 0
            results[i]["TOTAL SCORE"] += results[i]["Score: 1D Return"]
    for i, r in enumerate(results):
        if r["1D Return %"] is None:
            results[i]["1D Top 10th Pct"] = "N/A"
            results[i]["1D Bot 10th Pct"] = "N/A"

    # ── Green candles: rank + percentile bonus ────────────────────
    valid_gc = [(i, r) for i, r in enumerate(results)]
    gc_sorted = sorted(valid_gc, key=lambda x: -x[1]["Green Candles"])
    for rank, (i, _) in enumerate(gc_sorted, 1):
        results[i]["Green Candle Rank"] = rank

    gc_vals = [r["Green Candles"] for r in results]
    th_gc1  = percentile_threshold(gc_vals, GREEN_TOP_P1,  top=True)
    th_gc10 = percentile_threshold(gc_vals, GREEN_TOP_P10, top=True)
    print(f"  GreenC   top-1st: {th_gc1}  top-10th: {th_gc10}")
    for i, r in enumerate(results):
        gc = r["Green Candles"]
        if gc >= th_gc1:
            results[i]["Green Top 1st Pct"]  = "YES"
            results[i]["Green Top 10th Pct"] = "YES (via top1)"
            results[i]["Score: Green"]        = 4
        elif gc >= th_gc10:
            results[i]["Green Top 1st Pct"]  = "NO"
            results[i]["Green Top 10th Pct"] = "YES"
            results[i]["Score: Green"]        = 2
        else:
            results[i]["Green Top 1st Pct"]  = "NO"
            results[i]["Green Top 10th Pct"] = "NO"
            results[i]["Score: Green"]        = 0
        results[i]["TOTAL SCORE"] += results[i]["Score: Green"]

    # ── Sector bonus: based on top-25th-percentile final scores ──
    # Determine top-25th-percentile threshold of TOTAL SCORE
    all_scores = [r["TOTAL SCORE"] for r in results]
    th_score_p25 = percentile_threshold(all_scores, SECTOR_TOP_P25, top=True)
    print(f"  Score    top-25th threshold: {th_score_p25}")

    # Tag each result as top-25th-percentile
    for i, r in enumerate(results):
        results[i]["_in_top25"] = r["TOTAL SCORE"] >= th_score_p25

    # Group by sector and count top-25th-percentile stocks per sector
    sector_top25 = defaultdict(int)
    for r in results:
        if r["_in_top25"] and is_known_sector(r["Sector"]):
            sector_top25[r["Sector"]] += 1

    for i, r in enumerate(results):
        sector  = r["Sector"]
        if not is_known_sector(sector):
            results[i]["Sector Top25 Count"] = 0
            results[i]["Score: Sector"] = 0
            continue
        n_top25 = sector_top25[sector]
        # Number of OTHER stocks in same sector in top 25th percentile
        # (exclude self if self is also in top 25th)
        n_others = n_top25 - (1 if r["_in_top25"] else 0)

        if   n_others >= 5: s_sec = 6   # >4 others
        elif n_others >= 4: s_sec = 4   # >3 others
        elif n_others >= 3: s_sec = 3   # >2 others
        elif n_others >= 1: s_sec = 2   # at least 1 other
        else:               s_sec = 0

        results[i]["Sector Top25 Count"] = n_others
        results[i]["Score: Sector"]      = s_sec
        results[i]["TOTAL SCORE"]       += s_sec

    # Median 42D Turnover: top 30% -> +10 pts, bottom 30% -> -5 pts.
    # Fallback to Avg Turnover 42D for older callers that don't provide median.
    def _turnover_metric(row):
        if row.get("Median Turnover 42D") is not None:
            return row.get("Median Turnover 42D")
        return row.get("Avg Turnover 42D")

    valid_to = [(i, r) for i, r in enumerate(results)
                if _turnover_metric(r) is not None]
    if valid_to:
        tvals  = [_turnover_metric(r) for _, r in valid_to]
        th_top = percentile_threshold(tvals, TURNOVER_TOP_P30, top=True)
        th_bot = percentile_threshold(tvals, TURNOVER_BOT_P30, top=False)
        using_median = any(r.get("Median Turnover 42D") is not None for _, r in valid_to)
        label = "Median turnover" if using_median else "Avg turnover"
        print(f"  {label} top-30th: {th_top:.2f} Cr  bot-30th: {th_bot:.2f} Cr")
        for i, r in valid_to:
            tv = _turnover_metric(r)
            if tv >= th_top:
                results[i]["Turnover Top30Pct"] = "YES"
                results[i]["Turnover Bot30Pct"] = "NO"
                results[i]["Score: Turnover"]   = 10
            elif tv <= th_bot:
                results[i]["Turnover Top30Pct"] = "NO"
                results[i]["Turnover Bot30Pct"] = "YES"
                results[i]["Score: Turnover"]   = -5
            else:
                results[i]["Turnover Top30Pct"] = "NO"
                results[i]["Turnover Bot30Pct"] = "NO"
                results[i]["Score: Turnover"]   = 0
            results[i]["TOTAL SCORE"] += results[i]["Score: Turnover"]
    for i, r in enumerate(results):
        if _turnover_metric(r) is None:
            results[i]["Turnover Top30Pct"] = "N/A"
            results[i]["Turnover Bot30Pct"] = "N/A"

    # ── ATR: bottom 10th → -6 pts, bottom 20th → -3 pts ──────────
    # New listings < 30 TDs are excluded from ATR scoring — their ATR is
    # computed on too few candles to be comparable with established stocks.
    def _is_new_listing(r):
        cat = r.get("Listing Category", "Established")
        return isinstance(cat, str) and cat.startswith(f"< {NEW_LISTING_30}")

    # Use ATR % (ATR / Close × 100) so scores are price-normalised.
    # A ₹1,300 stock and a ₹500 stock are judged on relative volatility,
    # not absolute ₹ swing size.
    valid_atr = [(i, r) for i, r in enumerate(results)
                 if r.get("ATR %") is not None and not _is_new_listing(r)]
    if valid_atr:
        avals   = [r["ATR %"] for _, r in valid_atr]
        th_b10  = percentile_threshold(avals, ATR_BOT_P10,  top=False)
        th_b20  = percentile_threshold(avals, ATR_BOT_P20,  top=False)
        print(f"  ATR% bot-10th: {th_b10:.2f}%  bot-20th: {th_b20:.2f}%")
        for i, r in valid_atr:
            av = r["ATR %"]
            if av <= th_b10:
                results[i]["ATR Bot10Pct"] = "YES"
                results[i]["ATR Bot20Pct"] = "YES"
                results[i]["Score: ATR"]   = -6
            elif av <= th_b20:
                results[i]["ATR Bot10Pct"] = "NO"
                results[i]["ATR Bot20Pct"] = "YES"
                results[i]["Score: ATR"]   = -3
            else:
                results[i]["ATR Bot10Pct"] = "NO"
                results[i]["ATR Bot20Pct"] = "NO"
                results[i]["Score: ATR"]   = 0
            results[i]["TOTAL SCORE"] += results[i]["Score: ATR"]
    # Stocks with no ATR data or new listings — mark as exempt, score 0
    for i, r in enumerate(results):
        if r.get("ATR %") is None:
            results[i]["ATR Bot10Pct"] = "N/A"
            results[i]["ATR Bot20Pct"] = "N/A"
        elif _is_new_listing(r):
            results[i]["ATR Bot10Pct"] = "NEW"
            results[i]["ATR Bot20Pct"] = "NEW"
            results[i]["Score: ATR"]   = 0

    # ── Final ratings ─────────────────────────────────────────────
    for i in range(len(results)):
        results[i]["RATING"] = get_rating(results[i]["TOTAL SCORE"])
        del results[i]["_in_top25"]

    return results


# ── BSE master helpers ───────────────────────────────────────────

def _try_download_bse_master(dest=BSE_MASTER_FILE):
    """
    Attempt a direct HTTP download of the BSE scrip master CSV.
    Returns True if the file was saved successfully, False otherwise.
    BSE often blocks requests without a browser session, so this may
    fail — in which case the user should download manually.
    """
    import urllib.request
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Referer": "https://www.bseindia.com/",
        "Accept":  "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    for url in BSE_MASTER_URLS:
        try:
            req  = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
            # Sanity-check: real CSV starts with "Security Code"
            if not raw.lstrip().startswith("Security Code"):
                continue
            with open(dest, "w", encoding="utf-8") as fh:
                fh.write(raw)
            print(f"  BSE master downloaded → {dest}")
            return True
        except Exception as exc:
            print(f"  Auto-download failed ({url}): {exc}")
    return False


def _load_bse_master(path=None):
    """
    Load the BSE scrip master CSV and return a dict
        { NSE_symbol_upper: industry_title_case }

    Resolution order for the file:
      1. path argument (--bse-master CLI flag)
      2. bse_master.csv next to this script
      3. Attempt auto-download from BSE website
      4. Return empty dict (graceful degradation)

    Column mapping used:
      Security Id  →  symbol key   (BSE symbol ≈ NSE symbol for most stocks)
      Industry     →  sector value
      Status       →  filter to 'Active' only
    """
    filepath = path or BSE_MASTER_FILE

    if not os.path.exists(filepath):
        print(f"  BSE master not found at {filepath}.")
        print("  Attempting auto-download …")
        if not _try_download_bse_master(filepath):
            print("  Auto-download failed.  Download manually from:")
            print("    https://www.bseindia.com/corporates/List_Scrips.html")
            print("  (Segment=Equity, Status=Active → Download)")
            print(f"  Save as: {filepath}")
            print("  Or pass:  --bse-master <path>")
            return {}

    # Try common encodings
    df = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            df = pd.read_csv(filepath, dtype=str, encoding=encoding)
            break
        except Exception:
            df = None
    if df is None:
        print(f"  ERROR: Could not read {filepath} with any encoding.")
        return {}

    # Normalise column names: strip whitespace, lowercase for matching
    df.columns = [c.strip() for c in df.columns]
    cols_lower  = {c.lower(): c for c in df.columns}

    print(f"  Sector file: {len(df):,} rows | columns: {list(df.columns)}")

    # ── Auto-detect column layout ─────────────────────────────────
    # Format A (from build_sector_csv.py):  symbol | sector
    # Format B (BSE scrip master):          Security Id | Industry  [+ Status]
    sym_col = None
    ind_col = None

    for candidate in ("symbol", "security id", "securityid",
                       "scrip id", "scripid", "tradingsymbol"):
        if candidate in cols_lower:
            sym_col = cols_lower[candidate]
            break

    for candidate in ("sector", "industry", "sector1",
                       "industry name", "industry group"):
        if candidate in cols_lower:
            ind_col = cols_lower[candidate]
            break

    if not sym_col or not ind_col:
        print(f"  ERROR: Cannot locate symbol/sector columns.")
        print(f"  Columns found: {list(df.columns)}")
        return {}

    print(f"  Using columns: symbol='{sym_col}'  sector='{ind_col}'")

    # ── Optional Status filter (BSE master only) ──────────────────
    if "status" in cols_lower:
        status_col = cols_lower["status"]
        before = len(df)
        df = df[df[status_col].str.strip().str.lower() == "active"]
        print(f"  Active filter: {len(df):,} rows kept  "
              f"({before - len(df):,} removed)")

    # ── Build symbol → sector mapping ────────────────────────────
    mapping = {}
    for _, row in df.iterrows():
        sym    = str(row[sym_col]).strip().upper()
        sector = str(row[ind_col]).strip()
        if sym and sector and sector.lower() not in ("nan", "", "none", "-"):
            mapping[sym] = sector.title()

    print(f"  Loaded: {len(mapping):,} symbols with sector data")
    for sym, sec in list(mapping.items())[:5]:
        print(f"    {sym:<20} → {sec}")
    return mapping


# ── Build sector map from symbols ─────────────────────────────────

def build_sector_map(symbols, bse_master_path=None):
    """
    Return a {symbol: sector} dict for every symbol in the list.
    All sector data comes directly from the BSE master CSV — no cache,
    no hardcoded fallback.  Symbols absent from the BSE master are
    mapped to "Unknown".
    """
    bse = _load_bse_master(bse_master_path)

    sector_map = {sym: bse.get(sym, "Unknown") for sym in symbols}

    unknowns = [s for s in symbols if sector_map[s] == "Unknown"]
    resolved = len(symbols) - len(unknowns)
    print(f"  Sector map: {resolved}/{len(symbols)} resolved from BSE master"
          + (f"  |  Unknown: {unknowns}" if unknowns else ""))
    return sector_map


def is_known_sector(sector):
    sec = str(sector or "").strip()
    return sec.lower() not in ("", "unknown", "none", "-", "nan")


# ── Force-delete helper (closes file in Excel if open) ───────────
def _force_delete_xlsx(fname):
    """
    Deletes fname.  If the file is locked by Excel (PermissionError):
      1. Tries win32com to close just that workbook.
      2. Falls back to PowerShell COM automation.
      3. Retries the delete.
      4. Hard-exits with a clear message if still locked.
    """
    if not os.path.exists(fname):
        return
    try:
        os.remove(fname)
        print(f"  Deleted old file: {os.path.basename(fname)}")
        return
    except PermissionError:
        pass

    print(f"  File is open in Excel — attempting to close it: {os.path.basename(fname)}")

    # ── Attempt 1: win32com ───────────────────────────────────────
    closed = False
    try:
        import win32com.client
        xl = win32com.client.GetActiveObject("Excel.Application")
        for wb in list(xl.Workbooks):
            if os.path.normcase(wb.FullName) == os.path.normcase(fname):
                wb.Close(SaveChanges=False)
                closed = True
                print("  Closed workbook via Excel COM.")
                break
    except Exception:
        pass

    # ── Attempt 2: PowerShell COM fallback ────────────────────────
    if not closed:
        try:
            import subprocess
            ps = (
                "$xl = [Runtime.InteropServices.Marshal]"
                "::GetActiveObject('Excel.Application'); "
                f"$xl.Workbooks | Where-Object {{ $_.FullName -eq '{fname}' }}"
                " | ForEach-Object { $_.Close($false) }"
            )
            subprocess.run(["powershell", "-Command", ps],
                           capture_output=True, timeout=10)
            closed = True
            print("  Closed workbook via PowerShell.")
        except Exception:
            pass

    # ── Retry delete ──────────────────────────────────────────────
    time.sleep(0.5)
    try:
        os.remove(fname)
        print(f"  Deleted old file: {os.path.basename(fname)}")
    except PermissionError:
        print(f"\n  ERROR: Still cannot delete {fname}.")
        print("  Please close Excel completely and re-run.")
        sys.exit(1)


# ── Excel export ─────────────────────────────────────────────────
def export_excel(ratings, as_of, reset_date, symbols_file, output_dir):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(output_dir, exist_ok=True)
    fname = os.path.join(
        output_dir,
        f"StockRating_v7_{as_of.strftime('%d%b%Y')}"
        f"_Reset{reset_date.strftime('%d%b%Y')}.xlsx"
    )
    _force_delete_xlsx(fname)
    wb = Workbook()

    t_fill   = PatternFill("solid", start_color="1F3864")
    h_fill   = PatternFill("solid", start_color="2E4B8F")
    sb_fill  = PatternFill("solid", start_color="1E6B2E")
    b_fill   = PatternFill("solid", start_color="375623")
    wbuy_fill= PatternFill("solid", start_color="E2EFDA")
    n_fill   = PatternFill("solid", start_color="FFF2CC")
    w_fill   = PatternFill("solid", start_color="FCE4D6")
    yes_fill = PatternFill("solid", start_color="E2EFDA")
    no_fill  = PatternFill("solid", start_color="FCE4D6")
    neg_fill = PatternFill("solid", start_color="FF9999")
    alt_fill = PatternFill("solid", start_color="EBF3FB")
    grn_fill = PatternFill("solid", start_color="CCFFCC")
    gold_fill= PatternFill("solid", start_color="FFD700")
    sec_fill = PatternFill("solid", start_color="FFF9E6")

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sc(cell, bold=False, color="000000", fill=None,
           align="center", size=9, wrap=False):
        cell.font = Font(bold=bold, color=color, name="Arial", size=size)
        if fill: cell.fill = fill
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   wrap_text=wrap)
        cell.border = bdr

    def rf(score):
        if score >= 25: return sb_fill,   "FFFFFF"
        if score >= 18: return b_fill,    "FFFFFF"
        if score >= 11: return wbuy_fill, "000000"
        if score >= 5:  return n_fill,    "000000"
        return w_fill, "000000"

    def yn(val):
        v = str(val).upper()
        if v.startswith("YES"):    return yes_fill,  "1F6B3A", False
        if v == "NO":              return no_fill,   "8B0000", False
        if v == "NEG":             return neg_fill,  "8B0000", True
        if v == "NEW":             return n_fill,    "7B6000", False  # new listing exempt
        return alt_fill, "555555", False

    # ── Sheet 1: Full Rating ──────────────────────────────────────
    ws = wb.active
    ws.title = "Full Rating"
    rd = reset_date.strftime('%d-%b-%y')

    headers = [
        "#","Symbol","Sector",f"Close\n{as_of.strftime('%d-%b')}",
        "52W Hi","% From\n52W Hi","52W\nProx","Sc\nProx",
        "52W Hi\nRecency","Sc\nRec",
        "Both\n≤10%+≤10TD?","Sc\nBon",
        "50-SMA","Above\n50SMA?","Sc",
        "21-EMA","Above\n21EMA?","Sc",
        "8-EMA","Above\n8EMA?","Sc",
        "8EMA Up\n5TD?","Sc",
        "21EMA Up\n10TD?","Sc",
        "8>21>50\nStack?","Sc",
        "Reset\nDate","Reset\nLow",
        f"Ret\nFrom Low%","Top5th\nPct?","Top10th\nPct?","Sc",
        "1D Ret\n%","Top10th\nPct?","Bot10th\nPct?","Sc",
        "Prev Day\nHigh","1D>5%&\nPrevHi+1%?","Sc\nSpike",
        "1Y Vol\nP95","10D Vol\nSpike","Sc\nV10",
        "30D Vol\nSpike","Sc\nV30",
        "60D Vol\nSpike","Sc\nV60",
        f"Green C\nsince {rd}","Total C",
        "Green\n%","Top1st\nPct?","Top10th\nPct?","Sc","GC\nRank",
        "Listing\nCat","Sc",
        "Sector\nTop25 Peers","Sc",
        "Avg TO\n42D (Cr)","Median TO\n42D (Cr)","Top30%\nTO?","Bot30%\nTO?","Sc\nTO",
        "Uptrend\nCon%",">=80%\n3M Uptrnd?","Sc\nUC",
        "IPO\nProxy","IPO\nGain%","Sc\nIPO",
        "ATR\n14D","ATR\n%","ATR\nBot10%?","ATR\nBot20%?","Sc\nATR",
        "Ltd\nHist?",
        "TOTAL\nSCORE","RATING",
    ]
    ncols = len(headers)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"] = (
        f"NSE STOCK TECHNICAL RATING v7  |  As of {as_of.strftime('%d-%b-%Y')}  |  "
        f"Reset (Low): {reset_date.strftime('%d-%b-%Y')}  |  "
        f"Symbols: {os.path.basename(symbols_file)}  |  {len(ratings)} stocks"
    )
    ws["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ws["A1"].fill = t_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws["A2"] = (
        "52WProx≤10%=+4/≤15%=+2 | 52WRec≤10TD=+4/≤15TD=+2 | BothBonus≤10%+≤10TD=+2 | "
        "Above50SMA=1 | Above21EMA=1 | Above8EMA=1 | "
        "8EMAUp5TD=3 | 21EMAUp10TD=2 | EMAStack=2 | Top5%LowRet=6/Top10%LowRet=3 | "
        "Top10%1D=+2/Bot10%1D=-2 | Spike1D>5%+PrevHi=-4 | "
        "Vol1YSpike10D+Move=+6/10D=+5/30D=+3/60D=+2 | "
        "GreenTop1%=4/Top10%=2 | NewList<30TD=2/<60TD=1 | "
        "SectorTop25Peers: 1+=+2, 3+=+3, 4+=+4, 5+=+6 | "
        "MedianTurnover42D Top30%=+10/Bot30%=-5 | "
        "Uptrend8>21>50 >=80% of 63TDs=+4 | SectorSize>3=+3(composite) | "
        "IPO(<3M) >=15%AboveProxy=+4/>=30%=+6"
    )
    ws["A2"].font = Font(italic=True, color="444444", name="Arial", size=8)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        sc(cell, bold=True, color="FFFFFF", fill=h_fill, size=8, wrap=True)
    ws.row_dimensions[3].height = 46

    # NEW cols 11=Both≤10%+≤10TD?(yn), 12=ScBon(score)
    # All subsequent cols shift +2 vs prior version:
    # col 59=Avg TO 42D, 60=Median TO 42D, 61=Top30%TO?(yn), 62=Bot30%TO?(yn), 63=ScTO(score)
    # col 64=Uptrend Con%, 65=>=80% 3M?(yn), 66=ScUC(score)
    # col 67=IPO Proxy, 68=IPO Gain%, 69=ScIPO(score)
    # col 70=ATR 14D, 71=ATR %, 72=ATR Bot10%?(yn), 73=ATR Bot20%?(yn)
    # col 74=Sc ATR(score), 75=Ltd Hist?(yn), 76=TOTAL SCORE, 77=RATING
    yn_cols    = {11, 14, 17, 20, 22, 24, 29, 30, 33, 34, 39, 49, 50, 61, 62, 65, 72, 73, 75}
    score_cols = {8, 10, 12, 15, 18, 21, 23, 25, 31, 35, 40, 43, 45, 47, 51, 53, 55, 63, 66, 69, 74}

    for i, r in enumerate(ratings, 1):
        row   = i + 3
        score = r["TOTAL SCORE"]
        fill_r, col_r = rf(score)
        afill = alt_fill if i % 2 == 0 else PatternFill("solid",start_color="FFFFFF")
        gc_rank = r["Green Candle Rank"]

        vals = [
            i, r["Symbol"], r["Sector"], r["Close"],
            r["52W High"], r["% From 52W High"],
            r["Within 10% of 52W"], r["Score: Within 10%"],   # cols 7,8
            r["52W Hi 5/10 TDs"],   r["Score: 52W Hi"],        # cols 9,10
            r["52W Hi Both Bonus"], r["Score: 52W Bonus"],     # cols 11,12
            r["50-SMA"], r["Above 50-SMA"], r["Score: 50-SMA"],
            r["21-EMA"], r["Above 21-EMA"], r["Score: 21-EMA"],
            r["8-EMA"],  r["Above 8-EMA"],  r["Score: 8-EMA"],
            r["8-EMA Uptrend 5TD"],  r["Score: 8EMA Trend"],
            r["21-EMA Uptrend 10TD"],r["Score: 21EMA Trend"],
            r["EMA Stack 8>21>50"],  r["Score: EMA Stack"],
            r["Reset Date Used"],
            r["Reset Day Low"] if r["Reset Day Low"] else "N/A",
            r["Ret From Low %"] if r["Ret From Low %"] is not None else "N/A",
            r["Ret Top 5th Pct"], r["Ret Top 10th Pct"], r["Score: Low Return"],
            r["1D Return %"] if r["1D Return %"] is not None else "N/A",
            r["1D Top 10th Pct"], r["1D Bot 10th Pct"], r["Score: 1D Return"],
            r["Prev Day High"] if r["Prev Day High"] is not None else "N/A",
            r["1D>5%&Close>PrevHi+1%"], r["Score: Spike"],
            r["1Y Vol P95"] if r["1Y Vol P95"] is not None else "N/A",
            r["10D Vol Spike"], r["Score: Vol 10D"],
            r["30D Vol Spike"], r["Score: Vol 30D"],
            r["60D Vol Spike"], r["Score: Vol 60D"],
            r["Green Candles"], r["Total Candles"], r["Green %"],
            r["Green Top 1st Pct"], r["Green Top 10th Pct"],
            r["Score: Green"], gc_rank,
            r["Listing Category"], r["Score: Listing"],
            r["Sector Top25 Count"], r["Score: Sector"],
            r["Avg Turnover 42D"] if r["Avg Turnover 42D"] is not None else "N/A",
            r["Median Turnover 42D"] if r["Median Turnover 42D"] is not None else "N/A",
            r["Turnover Top30Pct"], r["Turnover Bot30Pct"], r["Score: Turnover"],
            r["Uptrend Con Lbl"], r["Uptrend >=80% 3M"], r["Score: Uptrend Con"],
            r["IPO Issue Proxy"] if r["IPO Issue Proxy"] is not None else "N/A",
            r["IPO Perf %"]      if r["IPO Perf %"]      is not None else "N/A",
            r["Score: IPO"],
            r["ATR 14D"] if r.get("ATR 14D") is not None else "N/A",
            r["ATR %"]   if r.get("ATR %")   is not None else "N/A",
            r.get("ATR Bot10Pct", "N/A"),
            r.get("ATR Bot20Pct", "N/A"),
            r.get("Score: ATR",   0),
            r["Limited History"],
            score, r["RATING"],
        ]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col in [76, 77]:   # TOTAL SCORE, RATING
                sc(cell, bold=True, fill=fill_r, color=col_r,
                   align="left" if col==77 else "center", size=9)
            elif col in score_cols:
                if isinstance(val, (int,float)) and val < 0:
                    sc(cell, bold=True, fill=neg_fill, color="8B0000", size=8)
                else:
                    sf = yes_fill if (isinstance(val,(int,float)) and val > 0) else no_fill
                    sc(cell, bold=True,
                       fill=sf,
                       color="1F6B3A" if (isinstance(val,(int,float)) and val>0) else "8B0000",
                       size=8)
            elif col == 7:   # 52W Prox text: ≤10%=green, ≤15%=yellow, >15%=neutral
                lbl = str(val)
                if lbl == "≤10%":
                    sc(cell, bold=True, fill=yes_fill, color="1F6B3A", size=8)
                elif lbl == "≤15%":
                    sc(cell, bold=True, fill=n_fill, color="7B6000", size=8)
                else:
                    sc(cell, fill=afill, size=8)
            elif col == 9:   # 52W Hi Recency: ≤10 TDs=green, ≤15 TDs=yellow, NO=red
                lbl = str(val)
                if lbl == "≤10 TDs":
                    sc(cell, bold=True, fill=yes_fill, color="1F6B3A", size=8)
                elif lbl == "≤15 TDs":
                    sc(cell, bold=True, fill=n_fill, color="7B6000", size=8)
                else:
                    sc(cell, fill=no_fill, size=8)
            elif col in yn_cols:
                yf, yc, bld = yn(str(val))
                sc(cell, bold=bld, fill=yf, color=yc, size=8)
            elif col == 42:  # 10D Vol Spike label (shifted from 40)
                lbl = str(val)
                if lbl.startswith("YES+Move"):
                    sc(cell, bold=True, fill=gold_fill, color="8B4513", size=8)
                elif lbl.startswith("YES"):
                    sc(cell, bold=True, fill=yes_fill, color="1F6B3A", size=8)
                else:
                    sc(cell, fill=afill, size=8)
            elif col in (44, 46):  # 30D / 60D Vol Spike labels (shifted from 42,44)
                lbl = str(val)
                if lbl.startswith("YES"):
                    sc(cell, bold=True, fill=yes_fill, color="1F6B3A", size=8)
                else:
                    sc(cell, fill=afill, size=8)
            elif col in [48, 49, 50]:  # Green candle raw cols (shifted from 46,47,48)
                sc(cell, fill=grn_fill, size=9)
            elif col == 54:  # GC rank (shifted from 52)
                gf = gold_fill if isinstance(val,int) and val<=int(max(1,len(ratings)*0.01)) \
                     else grn_fill if isinstance(val,int) and val<=int(max(1,len(ratings)*0.10)) \
                     else afill
                sc(cell, bold=(isinstance(val,int) and val<=3), fill=gf, size=9)
            elif col == 3:
                sc(cell, fill=sec_fill, align="left", size=8)
            elif col == 2:
                sc(cell, fill=afill, align="left", size=9)
            else:
                sc(cell, fill=afill, size=8)
                if col in [4, 5, 6, 13, 16, 19, 29, 30, 34, 38]:  # number cols (shifted)
                    cell.number_format = "#,##0.00"
                elif col == 41:  # 1Y Vol P95 (shifted from 39)
                    cell.number_format = "#,##0"
                elif col in [59, 60]:  # Avg / Median Turnover 42D (Cr)
                    cell.number_format = "#,##0.00"
                elif col == 66:  # IPO Issue Proxy
                    cell.number_format = "#,##0.00"
                elif col == 67:  # IPO Gain %
                    cell.number_format = "#,##0.00"
                elif col == 69:  # ATR 14D
                    cell.number_format = "#,##0.00"
                elif col == 70:  # ATR %
                    cell.number_format = "#,##0.00"

        ws.row_dimensions[row].height = 16

    col_widths = [
        4,18,20,9, 9,8,7,4, 12,4, 10,4, 9,7,4, 9,7,4, 9,7,4,
        #                   ^^^^  new Both Bonus? + ScBon cols
        9,4, 9,4, 7,4, 12,9,10, 8,8,4, 9,8,8,4,
        9,10,4, 10,12,4,12,4,12,4,
        9,9,7, 8,8,4,6, 12,4, 10,4,
        11,11,8,8,4,   # Avg TO 42D (Cr), Median TO 42D (Cr), Top30% TO?, Bot30% TO?, Sc TO
        9,10,4,     # Uptrend Con%, >=80% 3M Uptrnd?, Sc UC
        9,9,4,          # IPO Proxy, IPO Gain%, Sc IPO
        8,6, 6,6,4,     # ATR 14D, ATR %, ATR Bot10%?, ATR Bot20%?, Sc ATR
        6, 8,18,        # Ltd Hist?, TOTAL SCORE, RATING
    ]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "D4"

    # ── Sheet 2: Ranked Summary ───────────────────────────────────
    ws2 = wb.create_sheet("Ranked Summary")
    ws2.merge_cells("A1:L1")
    ws2["A1"] = (
        f"RANKED SUMMARY  |  As of {as_of.strftime('%d-%b-%Y')}  |  "
        f"Reset: {reset_date.strftime('%d-%b-%Y')}  |  {len(ratings)} stocks"
    )
    ws2["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ws2["A1"].fill = t_fill
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for col, h in enumerate(
        ["Rank","Symbol","Sector","Close","Score","Rating",
         "Ret From Low%","1D Ret%","Green Candles","GC Rank",
         "Avg TO 42D (Cr)","Median TO 42D (Cr)",
         "Criteria Met (key)"], 1
    ):
        cell = ws2.cell(row=2, column=col, value=h)
        sc(cell, bold=True, color="FFFFFF", fill=h_fill, size=10, wrap=True)
    ws2.row_dimensions[2].height = 20

    cmap = [
        ("Score: Within 10%",  "52W Prox ≤10%(+4)/≤15%(+2)"),
        ("Score: 52W Hi",      "52W Recency ≤10TD(+4)/≤15TD(+2)"),
        ("Score: 52W Bonus",   "Both ≤10%+≤10TD bonus (+2)"),
        ("Score: 50-SMA",      "Above 50-SMA (+1)"),
        ("Score: 21-EMA",      "Above 21-EMA (+1)"),
        ("Score: 8-EMA",       "Above 8-EMA (+1)"),
        ("Score: 8EMA Trend",  "8-EMA Uptrend (+3)"),
        ("Score: 21EMA Trend", "21-EMA Uptrend (+2)"),
        ("Score: EMA Stack",   "8>21>50 Stack (+2)"),
        ("Score: Low Return",  "Low-Ret Top5%(+6)/Top10%(+3)"),
        ("Score: 1D Return",   "1D Top10%(+2)/Bot10%(-2)"),
        ("Score: Spike",       "1D>5%&Close>PrevHi+1% (-4)"),
        ("Score: Vol 10D",     "1Y-Vol Spike 10D(+5) + >=3%(+2) / >6%(+4)"),
        ("Score: Vol 30D",     "1Y-Vol Spike 30D(+4) + >=3%(+2) / >6%(+4)"),
        ("Score: Vol 60D",     "1Y-Vol Spike 60D(+2) + >=3%(+2) / >6%(+2)"),
        ("Score: Green",       "GreenC Top1%(+4)/Top10%(+2)"),
        ("Score: Listing",     "New Listing (+1/+2)"),
        ("Score: Sector",      "Sector Top25 Peers 1+=+2/3+=+3/4+=+4/5+=+6"),
        ("Score: Turnover",    "Median Turnover 42D Top30%(+10)/Bot30%(-5)"),
        ("Score: ATR",         "ATR-14 Bot10%(-6)/Bot20%(-3)"),
        ("Score: Uptrend Con", "Uptrend 8>21>50 >=80% of 63TDs (+4)"),
        ("Score: IPO",         "IPO <3M: >=15% above proxy (+4) / >=30% (+6)"),
    ]

    for i, r in enumerate(ratings, 1):
        row   = i + 2
        score = r["TOTAL SCORE"]
        fill_r, col_r = rf(score)
        afill = alt_fill if i % 2 == 0 else PatternFill("solid",start_color="FFFFFF")
        met   = [lbl for key, lbl in cmap if r.get(key, 0) > 0]
        vals = [
            i, r["Symbol"], r["Sector"], r["Close"], score, r["RATING"],
            r["Ret From Low %"] if r["Ret From Low %"] is not None else "N/A",
            r["1D Return %"]    if r["1D Return %"]    is not None else "N/A",
            r["Green Candles"], r["Green Candle Rank"],
            r["Avg Turnover 42D"] if r["Avg Turnover 42D"] is not None else "N/A",
            " | ".join(met) if met else "—",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            if col in [5, 6]:
                sc(cell, bold=True, fill=fill_r, color=col_r, size=10)
            elif col == 9:
                sc(cell, fill=grn_fill, size=10)
            elif col == 10:
                gf = gold_fill if isinstance(val,int) and val<=int(max(1,len(ratings)*0.01)) \
                     else grn_fill
                sc(cell, fill=gf, size=10)
            elif col in [2, 3, 13]:
                sc(cell, fill=afill, align="left", size=10)
            else:
                sc(cell, fill=afill, size=10)
                if col in [4, 7, 8]:
                    cell.number_format = "#,##0.00"
                elif col in [11, 12]:   # Avg / Median TO 42D
                    cell.number_format = "#,##0.00"
        ws2.row_dimensions[row].height = 16

    for col, w in zip(range(1,14),[6,20,22,10,8,18,13,10,12,8,12,12,70]):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A3"

    # ── Sheet 3: Green Candle Ranking ─────────────────────────────
    ws3 = wb.create_sheet("Green Candle Rank")
    ws3.merge_cells("A1:G1")
    ws3["A1"] = (
        f"GREEN CANDLE RANKING  |  Since {reset_date.strftime('%d-%b-%Y')} "
        f"to {as_of.strftime('%d-%b-%Y')}  |  Rank 1 = Most Green Candles"
    )
    ws3["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ws3["A1"].fill = t_fill
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    for col, h in enumerate(
        ["GC Rank","Symbol","Green Candles","Total Candles","Green %",
         "GC Score","Total Score"], 1
    ):
        cell = ws3.cell(row=2, column=col, value=h)
        sc(cell, bold=True, color="FFFFFF", fill=h_fill, size=10)
    ws3.row_dimensions[2].height = 18

    gc_sorted = sorted(ratings, key=lambda x: x["Green Candle Rank"])
    p1_cutoff  = max(1, int(len(ratings) * GREEN_TOP_P1  / 100))
    p10_cutoff = max(1, int(len(ratings) * GREEN_TOP_P10 / 100))
    for i, r in enumerate(gc_sorted, 1):
        row   = i + 2
        score = r["TOTAL SCORE"]
        fill_r, col_r = rf(score)
        rank  = r["Green Candle Rank"]
        gf    = gold_fill if rank <= p1_cutoff \
                else grn_fill if rank <= p10_cutoff \
                else (alt_fill if i % 2 == 0 else PatternFill("solid",start_color="FFFFFF"))
        vals  = [rank, r["Symbol"], r["Green Candles"], r["Total Candles"],
                 r["Green %"], r["Score: Green"], score]
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            if col == 7:
                sc(cell, bold=True, fill=fill_r, color=col_r, size=10)
            elif col == 2:
                sc(cell, fill=gf, align="left", size=10)
            else:
                sc(cell, fill=gf, size=10)
                if col == 5:
                    cell.number_format = "0.0"
        ws3.row_dimensions[row].height = 16

    for col, w in zip(range(1,8),[8,20,14,14,10,10,10]):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = "A3"

    # ── Sheet 4: Sector View ──────────────────────────────────────
    ws4 = wb.create_sheet("Sector View")
    ws4.merge_cells("A1:G1")
    ws4["A1"] = "SECTOR VIEW — Stocks ranked by score within each sector"
    ws4["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ws4["A1"].fill = t_fill
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 24

    sector_groups = defaultdict(list)
    for r in ratings:
        sector_groups[r["Sector"]].append(r)

    row4 = 2
    for sector in sorted(sector_groups.keys()):
        stocks = sorted(sector_groups[sector], key=lambda x: -x["TOTAL SCORE"])
        ws4.merge_cells(f"A{row4}:G{row4}")
        cell = ws4.cell(row=row4, column=1,
                        value=f"  {sector.upper()}  ({len(stocks)} stocks)")
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="2E4B8F")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws4.row_dimensions[row4].height = 18
        row4 += 1
        for r in stocks:
            score = r["TOTAL SCORE"]
            fill_r, col_r = rf(score)
            vals = [r["Symbol"], r["Close"], score, r["RATING"],
                    r["Ret From Low %"] if r["Ret From Low %"] is not None else "N/A",
                    r["1D Return %"]    if r["1D Return %"]    is not None else "N/A",
                    r["Green Candle Rank"]]
            for col, val in enumerate(vals, 1):
                cell = ws4.cell(row=row4, column=col, value=val)
                if col in [3, 4]:
                    sc(cell, bold=True, fill=fill_r, color=col_r, size=10)
                else:
                    sc(cell, fill=alt_fill,
                       align="left" if col==1 else "center", size=10)
                    if col in [2, 5, 6]:
                        cell.number_format = "#,##0.00"
            ws4.row_dimensions[row4].height = 16
            row4 += 1

    for col, w in zip(range(1,8),[20,10,8,18,13,10,8]):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # ── Sector Summary sheet ──────────────────────────────────────
    sec_stocks = defaultdict(list)
    for r in ratings:
        if is_known_sector(r["Sector"]):
            sec_stocks[r["Sector"]].append(r)

    sec_rows = []
    total_rated = len(ratings)
    top25_threshold = percentile_threshold(
        [r["TOTAL SCORE"] for r in ratings], SECTOR_TOP_P25, top=True)

    for sec, stocks in sec_stocks.items():
        n = len(stocks)
        scores      = [s["TOTAL SCORE"] for s in stocks]
        avg_score   = round(sum(scores) / n, 1)
        med_score   = round(sorted(scores)[n // 2], 1)

        above_8ema  = sum(1 for s in stocks if s.get("Above 8-EMA")  in ("YES", "ASSUMED"))
        above_21ema = sum(1 for s in stocks if s.get("Above 21-EMA") in ("YES", "ASSUMED"))
        above_50sma = sum(1 for s in stocks if s.get("Above 50-SMA") in ("YES", "ASSUMED"))
        breadth_8   = round(above_8ema  / n * 100, 1)
        breadth_21  = round(above_21ema / n * 100, 1)
        breadth_50  = round(above_50sma / n * 100, 1)
        avg_breadth = round((breadth_8 + breadth_21 + breadth_50) / 3, 1)

        rets = [s["Ret From Low %"] for s in stocks if s["Ret From Low %"] is not None]
        avg_ret = round(sum(rets) / len(rets), 1) if rets else None

        d1s = [s["1D Return %"] for s in stocks if s["1D Return %"] is not None]
        avg_1d = round(sum(d1s) / len(d1s), 2) if d1s else None

        n_top25  = sum(1 for s in stocks if s["TOTAL SCORE"] >= top25_threshold)
        pct_top25 = round(n_top25 / n * 100, 1)

        vol_spikes = sum(1 for s in stocks
                         if (s.get("Score: Vol 10D", 0) or 0) > 0
                         or (s.get("Score: Vol 30D", 0) or 0) > 0
                         or (s.get("Score: Vol 60D", 0) or 0) > 0)

        # Composite rank score: avg_score (40%) + avg_breadth (35%) + avg_ret (25%)
        #   + size bonus (+3 if sector has more than 3 stocks)
        ret_norm   = min(avg_ret or 0, 100) / 100 * 20   # cap at 100%, scale to 20
        size_bonus = SECTOR_SIZE_BONUS_PTS if n > SECTOR_SIZE_BONUS_MIN else 0
        composite  = round(avg_score * 0.40 + avg_breadth * 0.35 + ret_norm + size_bonus, 2)

        sec_rows.append({
            "Sector":        sec,
            "# Stocks":      n,
            "Avg Score":     avg_score,
            "Med Score":     med_score,
            "# Top25%":      n_top25,
            "% Top25%":      pct_top25,
            "Breadth 8EMA%": breadth_8,
            "Breadth 21EMA%":breadth_21,
            "Breadth 50SMA%":breadth_50,
            "Avg Breadth%":  avg_breadth,
            "Avg Ret/Low%":  avg_ret,
            "Avg 1D Ret%":   avg_1d,
            "Vol Spikes":    vol_spikes,
            "Size Bonus":    size_bonus,
            "Composite":     composite,
        })

    sec_rows = [s for s in sec_rows if s["# Stocks"] >= 3]
    sec_rows.sort(key=lambda x: -x["Composite"])

    ws5 = wb.create_sheet("Sector Summary")

    # Title
    tc = ws5.cell(row=1, column=1,
                  value=f"Sector Summary — as of {as_of.strftime('%d-%b-%Y')}")
    tc.font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    tc.fill      = PatternFill("solid", start_color="1F3864")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    ws5.row_dimensions[1].height = 22

    # Headers
    sec_hdrs = ["Rank", "Sector", "# Stocks",
                "Avg Score", "Med Score", "# Top25%", "% Top25%",
                "Breadth\n8EMA%", "Breadth\n21EMA%", "Breadth\n50SMA%", "Avg Breadth%",
                "Avg Ret/Low%", "Avg 1D Ret%", "Vol Spikes", "Size\nBonus", "Composite"]
    for ci, h in enumerate(sec_hdrs, 1):
        c = ws5.cell(row=2, column=ci, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c.fill      = PatternFill("solid", start_color="2E4B8F")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws5.row_dimensions[2].height = 30

    # Top-3 gold/silver/bronze fills
    medal_fills = [
        PatternFill("solid", start_color="FFD700"),   # gold
        PatternFill("solid", start_color="C0C0C0"),   # silver
        PatternFill("solid", start_color="CD7F32"),   # bronze
    ]

    for ri, row in enumerate(sec_rows, 3):
        rank = ri - 2
        is_medal = rank <= 3
        row_bg = medal_fills[rank - 1] if is_medal else (
            PatternFill("solid", start_color="FFFFFF") if ri % 2 == 1 else alt_fill)

        vals = [
            rank,
            row["Sector"],
            row["# Stocks"],
            row["Avg Score"],
            row["Med Score"],
            row["# Top25%"],
            row["% Top25%"],
            row["Breadth 8EMA%"],
            row["Breadth 21EMA%"],
            row["Breadth 50SMA%"],
            row["Avg Breadth%"],
            row["Avg Ret/Low%"] if row["Avg Ret/Low%"] is not None else "N/A",
            row["Avg 1D Ret%"]  if row["Avg 1D Ret%"]  is not None else "N/A",
            row["Vol Spikes"],
            row["Size Bonus"],
            row["Composite"],
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws5.cell(row=ri, column=ci, value=val)
            cell.fill      = row_bg
            cell.font      = Font(bold=is_medal, name="Arial", size=9,
                                  color="000000")
            cell.alignment = Alignment(horizontal="left"  if ci == 2 else "center",
                                       vertical="center")
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
        ws5.row_dimensions[ri].height = 16

    sec_col_widths = [5, 26, 7, 8, 8, 7, 7, 9, 10, 10, 11, 11, 10, 9, 7, 9]
    for ci, w in enumerate(sec_col_widths, 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    ws5.freeze_panes = ws5["A3"]

    wb.save(fname)
    return fname


# ── ELITE WATCHLIST GENERATOR ────────────────────────────────────
def generate_elite_watchlist(sd_excel_path, as_of, output_dir):
    """
    Read the 'Top 30 Combined' sheet from sector_discovery Excel output,
    group symbols by sector in rank order, and write a TradingView-compatible
    watchlist file:  elite_DDMMMYYYY.txt

    Format:
        ###SectorName
        NSE:SYMBOL
        NSE:SYMBOL
        ...
    """
    try:
        import openpyxl
    except ImportError:
        print("  [elite] openpyxl not available — cannot read sector_discovery Excel.")
        return None

    if not os.path.exists(sd_excel_path):
        print(f"  [elite] Sector discovery Excel not found: {sd_excel_path}")
        return None

    try:
        wb = openpyxl.load_workbook(sd_excel_path, read_only=True, data_only=True)
        if "Top 30 Combined" not in wb.sheetnames:
            print("  [elite] 'Top 30 Combined' sheet not found in sector discovery Excel.")
            wb.close()
            return None

        ws = wb["Top 30 Combined"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        print(f"  [elite] Failed to read sector discovery Excel: {exc}")
        return None

    # Row 1 = title, Row 2 = headers, Row 3+ = data
    if len(rows) < 3:
        print("  [elite] Top 30 Combined sheet has no data rows.")
        return None

    headers = [str(c).strip() if c is not None else "" for c in rows[1]]

    # Locate Symbol and Sector columns (case-insensitive)
    sym_col = next((i for i, h in enumerate(headers) if h.lower() == "symbol"), None)
    sec_col = next((i for i, h in enumerate(headers) if h.lower() == "sector"), None)
    if sym_col is None or sec_col is None:
        print(f"  [elite] Could not find Symbol/Sector columns in Top 30 Combined. Headers: {headers}")
        return None

    # Collect (sector, symbol) pairs in rank order (rows already sorted by score)
    from collections import OrderedDict
    sectors = OrderedDict()
    for row in rows[2:]:
        if not any(row):
            continue
        sym = row[sym_col]
        sec = row[sec_col]
        if not sym or sym == "Symbol":
            continue
        sym = str(sym).strip().upper().replace("NSE:", "")
        sec = str(sec).strip() if sec else "Unknown"
        if sec not in sectors:
            sectors[sec] = []
        tv_sym = f"NSE:{sym}"
        if tv_sym not in sectors[sec]:
            sectors[sec].append(tv_sym)

    if not sectors:
        print("  [elite] No symbols found in Top 30 Combined sheet.")
        return None

    # Build file content
    lines = []
    for sec, syms in sectors.items():
        lines.append(f"###{sec}")
        lines.extend(syms)

    elite_name = f"elite_{as_of.strftime('%d%b%Y')}.txt"
    elite_path = os.path.join(output_dir, elite_name)
    os.makedirs(output_dir, exist_ok=True)
    with open(elite_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print(f"  [elite] Wrote {len(lines) - len(sectors)} symbols across "
          f"{len(sectors)} sectors → {elite_path}")
    return elite_path


# ── MAIN ─────────────────────────────────────────────────────────
def main():
    as_of, reset_date, symbols_file, output_dir, token_file, bse_master = parse_args()

    print("=" * 70)
    print("  NSE STOCK TECHNICAL RATING v7")
    print(f"  As-of date   : {as_of.strftime('%d-%b-%Y')}")
    print(f"  Reset date   : {reset_date.strftime('%d-%b-%Y')}")
    print(f"  Symbols file : {symbols_file}")
    print(f"  Output       : {output_dir}")
    print("=" * 70)

    symbols    = read_symbols(symbols_file)
    sector_map = build_sector_map(symbols, bse_master_path=bse_master)

    kite       = get_kite_session(token_file)
    token_map  = build_token_map(kite, symbols)

    fetch_from = min(reset_date, as_of - timedelta(days=DAYS_52W)) - timedelta(days=30)
    fetch_to   = as_of

    print(f"\n  Fetching data: {fetch_from} → {fetch_to}\n")

    ratings  = []
    skipped  = []
    total    = len(token_map)

    for idx, (sym, info) in enumerate(token_map.items(), 1):
        print(f"  [{idx:>2}/{total}] {sym:<18}", end=" ")

        df = fetch_daily(kite, info["token"], sym, fetch_from, fetch_to)
        if df is None or df.empty:
            print("NO DATA — skipped")
            skipped.append(sym)
            time.sleep(0.35)
            continue

        df = df[df.index <= as_of]
        if df.empty:
            print("No data before as-of — skipped")
            skipped.append(sym)
            time.sleep(0.35)
            continue

        ld = info.get("listing_date")
        if ld and isinstance(ld, str):
            try: ld = datetime.strptime(ld[:10], "%Y-%m-%d").date()
            except: ld = None

        result = rate_stock(sym, df, df.index[-1], reset_date, sector_map, ld)
        if result:
            ratings.append(result)
            ret_str  = f"{result['Ret From Low %']:>8.2f}%" \
                       if result["Ret From Low %"] is not None else "     N/A"
            lim_flag = " [LIMITED]" if result.get("Limited History") else ""
            d1_str   = f"{result['1D Return %']:>6.2f}%" \
                       if result["1D Return %"] is not None else "   N/A"
            print(f"Ret={ret_str}  1D={d1_str}  "
                  f"GreenC={result['Green Candles']:>3}  "
                  f"Base={result['TOTAL SCORE']}{lim_flag}")
        else:
            print("No data — skipped")
            skipped.append(sym)

        time.sleep(0.35)

    print(f"\n  Applying percentile/relative scores...")
    ratings = apply_relative_scores(ratings)
    ratings.sort(key=lambda x: -x["TOTAL SCORE"])

    print(f"\n{'='*70}")
    print(f"  Rated: {len(ratings)}  |  Skipped: {len(skipped)}")
    if skipped: print(f"  Skipped: {skipped}")
    print(f"{'='*70}")
    print(f"\n  {'RK':<4}{'SYMBOL':<20}{'SC':>5}  "
          f"{'RET_LOW%':>10}  {'1D%':>7}  {'GRN':>4}  {'GRNRK':>5}  RATING")
    print(f"  {'─'*72}")
    for i, r in enumerate(ratings, 1):
        rl  = f"{r['Ret From Low %']:>9.2f}%" if r["Ret From Low %"] is not None else "      N/A"
        d1  = f"{r['1D Return %']:>6.2f}%"    if r["1D Return %"]    is not None else "   N/A"
        star = "★★" if r["Ret Top 5th Pct"] == "YES" else \
               "★"  if r["Ret Top 10th Pct"] == "YES" else " "
        print(f"  {i:<4}{r['Symbol']:<20}{r['TOTAL SCORE']:>5}  "
              f"{rl}  {d1}  {r['Green Candles']:>4}  "
              f"{r['Green Candle Rank']:>5}  {star} {r['RATING']}")

    if ratings:
        fp = export_excel(ratings, as_of, reset_date, symbols_file, output_dir)
        print(f"\n  Saved → {fp}")

  
    # ── Auto-run Sector Discovery 2 (new — top-5 sectors, filtered) ──
    sd2_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sector_discovery2.py")
    if os.path.exists(sd2_script) and ratings:
        print("\n" + "=" * 70)
        print("  AUTO-RUNNING SECTOR DISCOVERY 2 ...")
        print("=" * 70)
        sd2_cmd = [
            sys.executable, sd2_script,
            "--as-of",   as_of.strftime("%Y-%m-%d"),
            "--reset",   reset_date.strftime("%Y-%m-%d"),
            "--symbols", symbols_file,
            "--out",     output_dir,
            "--token",   token_file,
        ]
        if bse_master:
            sd2_cmd += ["--bse-master", bse_master]
        subprocess.run(sd2_cmd)
    elif not os.path.exists(sd2_script):
        print(f"\n  [!] sector_discovery2.py not found — skipping.")

    print("\n  Done!")


if __name__ == "__main__":
    main()
