from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT_DIR = Path(__file__).resolve().parent
REPORTS_DIR = ROOT_DIR / "reports"
OUTPUT_DIR = ROOT_DIR / "output"

TITLE_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FILL = PatternFill("solid", fgColor="2E4B8F")
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge neo_liquid_momentum_scanner and stock_rating Excel outputs."
    )
    parser.add_argument("cutoff_date", help="As-of date in YYYY-MM-DD format")
    parser.add_argument("reset_date", help="Reset date in YYYY-MM-DD format")
    parser.add_argument("--neo-file", help="Optional explicit neo liquid scanner workbook path")
    parser.add_argument("--rating-file", help="Optional explicit stock rating workbook path")
    parser.add_argument("--output-dir", default=str(REPORTS_DIR), help="Output folder for consolidated workbook")
    return parser.parse_args()


def parse_iso_date(raw: str) -> date:
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError as exc:
        raise SystemExit(f"Invalid date '{raw}'. Use YYYY-MM-DD.") from exc


def locate_file(explicit_path: str | None, candidates: Iterable[Path], label: str) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if not path.exists():
            raise SystemExit(f"{label} file not found: {path}")
        return path
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    tried = "\n".join(f"  - {p}" for p in candidates)
    raise SystemExit(f"Could not locate {label} file. Tried:\n{tried}")


def normalize_column_name(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").replace("₹", "rs").replace("%", "pct")
    cleaned = "".join(ch.lower() if ch.isalnum() else "_" for ch in text)
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_")


def load_stock_rating_sheet(path: Path) -> pd.DataFrame:
    excel = pd.ExcelFile(path, engine="openpyxl")
    sheet_name = next((name for name in excel.sheet_names if "Full Ratings" in name), None)
    if not sheet_name:
        raise SystemExit(f"Could not find Full Ratings sheet in {path}")
    df = pd.read_excel(path, sheet_name=sheet_name, header=2, engine="openpyxl")
    df.columns = [normalize_column_name(col) for col in df.columns]
    if df.columns.tolist():
        first_col = df.columns[0]
        if first_col == "":
            df = df.rename(columns={first_col: "rating_rank"})
    df = df.rename(
        columns={
            "symbol": "symbol",
            "sector": "rating_sector",
            "total_score": "rating_total_score",
            "pre_sec_score": "rating_pre_sector_score",
            "close": "rating_close",
            "avg_to_42d_cr": "rating_avg_to_42d_cr",
            "med_to_42d_cr": "rating_med_to_42d_cr",
            "ret_12m": "rating_ret_12m",
            "ret_6m": "rating_ret_6m",
            "ret_3m": "rating_ret_3m",
            "perf_total": "rating_perf_total",
            "rs_total": "rating_rs_total",
            "uptrnd_con_pct": "rating_uptrend_con_pct",
            "days_50dma": "rating_daysbelow50dma",
            "spike_total": "rating_spike_total",
            "sc_gapup": "rating_gapup_score",
            "atr_21d": "rating_atr_pct_21d",
            "sc_trend": "rating_trend_score",
            "sc_sect": "rating_sector_score",
        }
    )
    keep_cols = [
        "symbol",
        "rating_rank",
        "rating_sector",
        "rating_total_score",
        "rating_pre_sector_score",
        "rating_close",
        "rating_avg_to_42d_cr",
        "rating_med_to_42d_cr",
        "rating_ret_12m",
        "rating_ret_6m",
        "rating_ret_3m",
        "rating_perf_total",
        "rating_rs_total",
        "rating_uptrend_con_pct",
        "rating_daysbelow50dma",
        "rating_spike_total",
        "rating_gapup_score",
        "rating_atr_pct_21d",
        "rating_trend_score",
        "rating_sector_score",
    ]
    available = [col for col in keep_cols if col in df.columns]
    out = df[available].copy()
    out["symbol"] = out["symbol"].astype(str).str.strip().str.upper()
    out = out[out["symbol"].ne("") & out["symbol"].ne("NAN")]
    for col in ["rating_ret_12m", "rating_ret_6m", "rating_ret_3m"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
            if out[col].dropna().abs().max() <= 10:
                out[col] = out[col] * 100.0
    return out


def load_neo_sheet(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="All Scored Stocks", header=0, engine="openpyxl")
    df.columns = [normalize_column_name(col) for col in df.columns]
    df = df.rename(
        columns={
            "rank": "neo_rank",
            "symbol": "symbol",
            "source": "neo_source",
            "sector": "neo_sector",
            "close_rs": "neo_close",
            "total_score": "neo_total_score",
            "rating": "neo_rating",
            "avg_to_21d_rs_cr": "neo_avg_to_21d_cr",
            "median_to_21d_rs_cr": "neo_med_to_21d_cr",
            "ret_resetpct": "neo_ret_reset",
            "1d_retpct": "neo_ret_1d",
            "12m_retpct": "neo_ret_12m",
            "6m_retpct": "neo_ret_6m",
            "3m_retpct": "neo_ret_3m",
            "uptrend_con_pct": "neo_uptrend_con_pct",
            "pct_from_52w_high": "neo_pct_from_52w_high",
            "tds_since_52w_high": "neo_days_since_52w_high",
            "vol_period": "neo_vol_period",
            "vol_day_movepct": "neo_vol_day_move_pct",
            "atrpct": "neo_atr_pct",
            "rs_compositepct": "neo_rs_composite_pct",
            "rs_percentile": "neo_rs_percentile",
        }
    )
    keep_cols = [
        "symbol",
        "neo_rank",
        "neo_source",
        "neo_sector",
        "neo_close",
        "neo_total_score",
        "neo_rating",
        "neo_avg_to_21d_cr",
        "neo_med_to_21d_cr",
        "neo_ret_reset",
        "neo_ret_1d",
        "neo_ret_12m",
        "neo_ret_6m",
        "neo_ret_3m",
        "neo_uptrend_con_pct",
        "neo_pct_from_52w_high",
        "neo_days_since_52w_high",
        "neo_vol_period",
        "neo_vol_day_move_pct",
        "neo_atr_pct",
        "neo_rs_composite_pct",
        "neo_rs_percentile",
    ]
    available = [col for col in keep_cols if col in df.columns]
    out = df[available].copy()
    out["symbol"] = out["symbol"].astype(str).str.strip().str.upper()
    out = out[out["symbol"].ne("") & out["symbol"].ne("NAN")]
    return out


def build_consolidated(neo_df: pd.DataFrame, rating_df: pd.DataFrame) -> pd.DataFrame:
    merged = neo_df.merge(rating_df, on="symbol", how="outer")
    merged["sector"] = merged["rating_sector"].combine_first(merged["neo_sector"])
    merged["in_neo"] = merged["neo_rank"].notna()
    merged["in_rating"] = merged["rating_rank"].notna()
    merged["presence"] = "Both"
    merged.loc[merged["in_neo"] & ~merged["in_rating"], "presence"] = "Neo Only"
    merged.loc[~merged["in_neo"] & merged["in_rating"], "presence"] = "Stock Rating Only"

    merged["consensus_rank"] = pd.concat(
        [merged["neo_rank"], merged["rating_rank"]], axis=1
    ).mean(axis=1, skipna=True)
    merged["score_delta_rating_minus_neo"] = merged["rating_total_score"] - merged["neo_total_score"]
    merged = merged.sort_values(
        ["presence", "consensus_rank", "rating_total_score", "neo_total_score", "symbol"],
        ascending=[True, True, False, False, True],
        na_position="last",
    ).reset_index(drop=True)
    merged.insert(0, "combined_rank", range(1, len(merged) + 1))
    return merged


def style_sheet(ws, title: str, percent_columns: set[str] | None = None) -> None:
    percent_columns = percent_columns or set()
    ws.insert_rows(1)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=ws.max_column).column_letter}{ws.max_row}"

    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(3, ws.max_row + 1):
        row_fill = ALT_FILL if row % 2 == 1 else WHITE_FILL
        for cell in ws[row]:
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    header_map = {ws.cell(row=2, column=col).value: col for col in range(1, ws.max_column + 1)}
    for header, col_idx in header_map.items():
        letter = ws.cell(row=2, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(12, min(24, len(str(header)) + 2))
        if header in percent_columns:
            for row in range(3, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "0.0%"


def write_dataframe_sheet(wb: Workbook, sheet_name: str, title: str, df: pd.DataFrame, percent_cols: set[str]) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    style_sheet(ws, title, percent_cols)


def write_summary_sheet(
    wb: Workbook,
    cutoff_date: date,
    reset_date: date,
    neo_path: Path,
    rating_path: Path,
    consolidated: pd.DataFrame,
) -> None:
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Cutoff Date", cutoff_date.isoformat()),
        ("Reset Date", reset_date.isoformat()),
        ("Neo Liquid Workbook", neo_path.name),
        ("Stock Rating Workbook", rating_path.name),
        ("Consolidated Symbols", int(len(consolidated))),
        ("Present In Both", int((consolidated["presence"] == "Both").sum())),
        ("Only In Neo", int((consolidated["presence"] == "Neo Only").sum())),
        ("Only In Stock Rating", int((consolidated["presence"] == "Stock Rating Only").sum())),
    ]
    ws.append(["Metric", "Value"])
    for row in rows:
        ws.append(list(row))
    style_sheet(
        ws,
        f"Consolidated Scanner Summary | As of {cutoff_date.strftime('%d-%b-%Y')} | Reset {reset_date.strftime('%d-%b-%Y')}",
        set(),
    )
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45


def main() -> None:
    args = parse_args()
    cutoff_date = parse_iso_date(args.cutoff_date)
    reset_date = parse_iso_date(args.reset_date)

    neo_candidates = [
        REPORTS_DIR / f"LiquidCandidates_{cutoff_date.strftime('%d%b%Y').upper()}_Reset{reset_date.strftime('%d%b%Y').upper()}.xlsx",
    ]
    rating_candidates = [
        OUTPUT_DIR / f"stock_rating_{cutoff_date.strftime('%d%b%Y')}.xlsx",
    ]

    neo_path = locate_file(args.neo_file, neo_candidates, "Neo liquid scanner")
    rating_path = locate_file(args.rating_file, rating_candidates, "Stock rating")

    print(f"Using neo workbook    : {neo_path}")
    print(f"Using rating workbook : {rating_path}")

    neo_df = load_neo_sheet(neo_path)
    rating_df = load_stock_rating_sheet(rating_path)
    consolidated = build_consolidated(neo_df, rating_df)

    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / (
        f"consolidated_{cutoff_date.strftime('%d%b%Y')}_reset_{reset_date.strftime('%d%b%Y')}.xlsx"
    )

    wb = Workbook()
    write_summary_sheet(wb, cutoff_date, reset_date, neo_path, rating_path, consolidated)

    display_columns = [
        "combined_rank",
        "symbol",
        "sector",
        "presence",
        "consensus_rank",
        "neo_rank",
        "rating_rank",
        "neo_total_score",
        "rating_total_score",
        "score_delta_rating_minus_neo",
        "neo_rating",
        "neo_avg_to_21d_cr",
        "rating_avg_to_42d_cr",
        "rating_med_to_42d_cr",
        "neo_ret_12m",
        "neo_ret_6m",
        "neo_ret_3m",
        "rating_ret_12m",
        "rating_ret_6m",
        "rating_ret_3m",
        "neo_rs_composite_pct",
        "rating_rs_total",
        "rating_spike_total",
        "rating_gapup_score",
    ]
    display_df = consolidated[[col for col in display_columns if col in consolidated.columns]].copy()

    percent_cols: set[str] = set()

    write_dataframe_sheet(
        wb,
        "Consolidated",
        f"Consolidated Rankings | As of {cutoff_date.strftime('%d-%b-%Y')}",
        display_df,
        percent_cols,
    )
    write_dataframe_sheet(
        wb,
        "Overlap Only",
        "Symbols Present In Both Reports",
        display_df[display_df["presence"] == "Both"].copy(),
        percent_cols,
    )
    write_dataframe_sheet(
        wb,
        "Neo Only",
        "Symbols Only In Neo Liquid Scanner",
        display_df[display_df["presence"] == "Neo Only"].copy(),
        percent_cols,
    )
    write_dataframe_sheet(
        wb,
        "Rating Only",
        "Symbols Only In Stock Rating",
        display_df[display_df["presence"] == "Stock Rating Only"].copy(),
        percent_cols,
    )

    wb.save(out_path)
    print(f"Consolidated workbook : {out_path}")


if __name__ == "__main__":
    main()
