#!/usr/bin/env python3
"""
Liquid Momentum Screener v2.0
Scans NSE universe, scores stocks on composite momentum + relative-strength,
identifies top 5 sectors, and produces an Excel report of the top 20 stocks.

Usage:
    python liquid_momentum_scanner.py <as_of_date> <reset_date> [options]

Arguments:
    as_of_date      Date to score as of (YYYY-MM-DD)
    reset_date      Portfolio reset/entry date (YYYY-MM-DD)

Options:
    --source        kite|bhav  (default: kite; stock OHLC from Kite, turnover/index from bhav)
    --extended      Include extended universe scan (stocks outside top 150)
    --run-fundamentals  Run screener_fundamentals.py on final output
"""

import argparse
import json
import logging
import os
import subprocess
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
KITE_RATE_LIMIT = 3          # requests per second
KITE_BATCH_PAUSE = 0.34      # seconds between calls
TOP_N = 125
TOP_N_REPORT = 20
MIN_CLOSE = 50.0             # ₹
MIN_EXTENDED_TURNOVER_CR = 20.0   # ₹ Crore
ATR_PERIOD = 14
TURNOVER_DAYS = 21
EXTENDED_UNIVERSE_TURNOVER_CRORE = 20.0
TOP30_MIN_ATR_PCT = 2.5

# 52W high gate
MAX_DIST_FROM_52W_HIGH_PCT = 15.0  # ≤15%
MAX_TDS_SINCE_52W_HIGH = 15        # ≤15 trading days for recency override
MAX_RECENCY_GATE_DIST_PCT = 20.0   # hard cap for recent-high inclusion

# 12M return gate (60th percentile)
RETURN_GATE_PERCENTILE = 50

# Nifty Smallcap 250 benchmark from indexbhav
NIFTY_SMALLCAP_250_SYMBOL = "NIFTY SMALLCAP 250"
NIFTY_50_SYMBOL = "NIFTY 50"

# RS weights
RS_WEIGHTS = {"3m": 0.40, "6m": 0.30, "12m": 0.20, "1m": 0.10}
UPTREND_CON_WINDOW = 63
UPTREND_CON_MIN_PCT = 80.0
SECTOR_W_SCORE = 0.40
SECTOR_W_BREADTH = 0.25
SECTOR_W_RS = 0.20
SECTOR_W_COUNT = 0.15

# Report output directories
OUTPUT_DIR = Path("output")
REPORTS_DIR = Path("reports")
GMLIST_DIR = Path("gmlist")
TOKEN_FILE = Path("kite_token.txt")
ROOT_DIR = Path(__file__).resolve().parent
LOGS_DIR = ROOT_DIR / "logs"


def setup_run_logging(as_of: datetime, reset: datetime) -> Path:
    """Attach a persistent file log for this scanner run."""
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    log_path = LOGS_DIR / (
        f"neo_liquid_momentum_{as_of.strftime('%Y%m%d')}_{reset.strftime('%Y%m%d')}.log"
    )

    root_logger = logging.getLogger()
    resolved_log_path = log_path.resolve()
    for handler in root_logger.handlers:
        if isinstance(handler, logging.FileHandler):
            try:
                if Path(handler.baseFilename).resolve() == resolved_log_path:
                    return log_path
            except Exception:
                continue

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    root_logger.addHandler(file_handler)
    return log_path

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db_connection():
    try:
        import pymysql
        conn = pymysql.connect(
            host="localhost",
            port=3306,
            user="root",
            password="root",
            database="bhav",
            charset="utf8mb4",
            autocommit=True,
        )
        return conn
    except ImportError:
        log.error("pymysql not installed. Run: pip install pymysql")
        sys.exit(1)
    except Exception as e:
        log.error(f"DB connection failed: {e}")
        sys.exit(1)


def query_df(conn, sql: str, params=None) -> pd.DataFrame:
    """Run a query through a DB cursor and return a DataFrame."""
    cursor = conn.cursor()
    try:
        cursor.execute(sql, params or [])
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description] if cursor.description else []
    finally:
        cursor.close()
    return pd.DataFrame(rows, columns=columns)


def fetch_bhav_turnover(conn, as_of_date: datetime, lookback_days: int = 30) -> pd.DataFrame:
    """Return avg 21-trading-day turnover per symbol from bhav tables."""
    year = as_of_date.year
    prev_year = year - 1
    tables = [f"bhav{prev_year}", f"bhav{year}"]

    frames = []
    for tbl in tables:
        try:
            sql = f"""
                SELECT symbol, mktdate AS date, close, volume
                FROM {tbl}
                WHERE mktdate <= %s
                ORDER BY mktdate
            """
            df = query_df(conn, sql, [as_of_date.strftime("%Y-%m-%d")])
            frames.append(df)
        except Exception as e:
            log.warning(f"Could not read {tbl}: {e}")

    if not frames:
        log.error("No bhav data found.")
        return pd.DataFrame(columns=[
            "symbol", "avg_turnover_21d", "median_turnover_21d",
            "latest_close", "avg_turnover_cr", "median_turnover_cr"
        ])

    raw = pd.concat(frames, ignore_index=True)
    raw.sort_values(["symbol", "date"], inplace=True)
    raw["turnover"] = raw["close"] * raw["volume"]

    # Keep last 21 trading days per symbol
    raw = raw.groupby("symbol").tail(TURNOVER_DAYS)
    summary = raw.groupby("symbol").agg(
        avg_turnover_21d=("turnover", "mean"),
        median_turnover_21d=("turnover", "median"),
        latest_close=("close", "last"),
    ).reset_index()
    summary["avg_turnover_cr"] = summary["avg_turnover_21d"] / 1e7  # to Crore
    summary["median_turnover_cr"] = summary["median_turnover_21d"] / 1e7  # to Crore
    return summary


def fetch_sectors(conn) -> pd.DataFrame:
    """Return symbol → sector mapping from bhav.sectors."""
    try:
        return query_df(conn, "SELECT symbol, sector1 AS sector FROM sectors")
    except Exception as e:
        log.warning(f"Could not read sectors table: {e}")
        return pd.DataFrame(columns=["symbol", "sector"])


def read_token_file(filepath: Path) -> dict:
    path = Path(filepath)
    if not path.exists():
        log.error(f"{path} not found. Run kite_get_access_token.py.")
        sys.exit(1)

    data = {}
    for raw_line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        data[key.strip().upper()] = value.strip()

    if "API_KEY" not in data or "ACCESS_TOKEN" not in data:
        log.error(f"{path} missing API_KEY or ACCESS_TOKEN.")
        sys.exit(1)
    return data


# ---------------------------------------------------------------------------
# Kite API helpers
# ---------------------------------------------------------------------------

def get_kite_client():
    try:
        from kiteconnect import KiteConnect
    except ImportError:
        log.error("kiteconnect not installed. Run: pip install kiteconnect")
        sys.exit(1)

    creds = read_token_file(TOKEN_FILE)
    kite = KiteConnect(api_key=creds["API_KEY"])
    kite.set_access_token(creds["ACCESS_TOKEN"])
    return kite


def is_stock_instrument(inst: dict) -> bool:
    """Keep only real NSE cash-equity stocks; exclude indices and ETFs."""
    if inst.get("exchange") != "NSE":
        return False
    if inst.get("instrument_type") != "EQ":
        return False
    if str(inst.get("segment") or "").upper() == "INDICES":
        return False

    tradingsymbol = str(inst.get("tradingsymbol") or "").upper()
    name = str(inst.get("name") or "").upper()

    if any(marker in name for marker in ("ETF", "INDEX ETF", "EXCHANGE TRADED FUND")):
        return False
    if tradingsymbol.endswith("BEES") or tradingsymbol.endswith("ETF"):
        return False
    if any(marker in tradingsymbol for marker in ("ETF",)):
        return False
    return True


def get_nse_eq_symbols(kite) -> set:
    """Return set of valid NSE EQ symbols from Kite instruments."""
    instruments = kite.instruments("NSE")
    return {i["tradingsymbol"] for i in instruments if is_stock_instrument(i)}


def build_nse_token_map(kite, symbols: list[str]) -> dict[str, int]:
    """Return {symbol: instrument_token} for NSE EQ symbols."""
    instruments = pd.DataFrame(kite.instruments("NSE"))
    if instruments.empty:
        return {}

    eq = instruments[instruments.apply(
        lambda row: is_stock_instrument(row.to_dict()), axis=1
    )].copy()
    eq = eq[eq["tradingsymbol"].isin(symbols)]
    return {
        str(row["tradingsymbol"]): int(row["instrument_token"])
        for _, row in eq.iterrows()
    }


def get_index_instrument_token(kite, symbol: str):
    """Return instrument_token for a Nifty index symbol."""
    instruments = kite.instruments("NSE")
    for inst in instruments:
        if inst.get("tradingsymbol") == symbol or inst.get("name", "").upper() == symbol.upper():
            return inst["instrument_token"]
    return None


def fetch_ohlcv_kite(kite, symbol: str, from_date: datetime, to_date: datetime,
                     is_index: bool = False, failed_symbols: list = None,
                     instrument_token: int | None = None) -> pd.DataFrame:
    """Fetch daily OHLCV from Kite for a single symbol. Returns DataFrame."""
    if is_index:
        token = get_index_instrument_token(kite, symbol)
        if token is None:
            log.warning(f"Index token not found for {symbol}")
            return pd.DataFrame()
        instrument_token = token
    elif instrument_token is None:
        log.warning(f"Kite instrument token missing for {symbol}")
        if failed_symbols is not None:
            failed_symbols.append(symbol)
        return pd.DataFrame()

    time.sleep(KITE_BATCH_PAUSE)
    try:
        data = kite.historical_data(
            instrument_token,
            from_date.strftime("%Y-%m-%d"),
            to_date.strftime("%Y-%m-%d"),
            "day",
        )
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame(data)
        df.rename(columns={"date": "date", "open": "open", "high": "high",
                            "low": "low", "close": "close", "volume": "volume"}, inplace=True)
        df["date"] = pd.to_datetime(df["date"]).dt.tz_localize(None).dt.normalize()
        return df.sort_values("date").reset_index(drop=True)
    except Exception as e:
        log.warning(f"Kite fetch failed for {symbol}: {e}")
        if failed_symbols is not None:
            failed_symbols.append(symbol)
        return pd.DataFrame()


def fetch_ohlcv_bhav(conn, symbol: str, from_date: datetime, to_date: datetime) -> pd.DataFrame:
    """Fallback: fetch OHLCV from bhav DB."""
    frames = []
    for year in range(from_date.year, to_date.year + 1):
        try:
            sql = f"""
                SELECT mktdate AS date, open, high, low, close, volume
                FROM bhav{year}
                WHERE symbol = %s AND mktdate BETWEEN %s AND %s
                ORDER BY mktdate
            """
            df = query_df(conn, sql, [
                symbol,
                from_date.strftime("%Y-%m-%d"),
                to_date.strftime("%Y-%m-%d"),
            ])
            frames.append(df)
        except Exception:
            pass
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    return df.sort_values("date").reset_index(drop=True)


def merge_ohlcv_prefer_primary(primary_df: pd.DataFrame, secondary_df: pd.DataFrame) -> pd.DataFrame:
    """Merge OHLCV series by date, preferring primary rows and backfilling missing dates from secondary."""
    if primary_df.empty:
        return secondary_df.sort_values("date").reset_index(drop=True) if not secondary_df.empty else pd.DataFrame()
    if secondary_df.empty:
        return primary_df.sort_values("date").reset_index(drop=True)

    primary = primary_df.copy()
    secondary = secondary_df.copy()
    primary["date"] = pd.to_datetime(primary["date"]).dt.normalize()
    secondary["date"] = pd.to_datetime(secondary["date"]).dt.normalize()
    merged = pd.concat([primary, secondary], ignore_index=True)
    merged = merged.drop_duplicates(subset=["date"], keep="first")
    return merged.sort_values("date").reset_index(drop=True)


def fetch_index_ohlcv_indexbhav(conn, symbol: str, from_date: datetime, to_date: datetime) -> pd.DataFrame:
    """Fetch index OHLC from bhav.indexbhav."""
    if conn is None:
        return pd.DataFrame()
    start_date = pd.Timestamp(from_date).date()
    end_date = pd.Timestamp(to_date).date()
    cursor = conn.cursor()
    try:
        exact_sql = """
            SELECT mktdate, open, high, low, close, symbol
            FROM indexbhav
            WHERE mktdate BETWEEN %s AND %s
              AND UPPER(symbol) = UPPER(%s)
            ORDER BY mktdate
        """
        cursor.execute(exact_sql, (start_date, end_date, symbol))
        rows = cursor.fetchall()

        if not rows and "SMALLCAP" in str(symbol).upper() and "250" in str(symbol):
            like_sql = """
                SELECT mktdate, open, high, low, close, symbol
                FROM indexbhav
                WHERE mktdate BETWEEN %s AND %s
                  AND UPPER(symbol) LIKE %s
                ORDER BY mktdate
            """
            cursor.execute(like_sql, (start_date, end_date, "%SMALLCAP%250%"))
            rows = cursor.fetchall()
    except Exception as e:
        log.warning(f"indexbhav fetch failed for {symbol}: {e}")
        cursor.close()
        return pd.DataFrame()
    finally:
        try:
            cursor.close()
        except Exception:
            pass

    if not rows:
        return pd.DataFrame(columns=["date", "open", "high", "low", "close", "symbol"])
    df = pd.DataFrame(rows, columns=["date", "open", "high", "low", "close", "symbol"])
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    return df.drop_duplicates(subset=["date"], keep="last").sort_values("date").reset_index(drop=True)


def fetch_ohlcv(kite, conn, symbol: str, from_date: datetime, to_date: datetime,
                source: str = "kite", failed_symbols: list = None,
                token_map: dict[str, int] | None = None) -> pd.DataFrame:
    """Fetch OHLCV preferring source, falling back to other."""
    if source == "kite":
        token = token_map.get(symbol) if token_map else None
        kite_df = fetch_ohlcv_kite(
            kite, symbol, from_date, to_date,
            failed_symbols=failed_symbols,
            instrument_token=token,
        )
        if conn:
            bhav_df = fetch_ohlcv_bhav(conn, symbol, from_date, to_date)
            if kite_df.empty:
                df = bhav_df
            else:
                df = merge_ohlcv_prefer_primary(kite_df, bhav_df)
                backfilled = max(0, len(df) - len(kite_df))
                if backfilled > 0:
                    log.info(f"Backfilled {backfilled} missing bhav candle(s) into Kite OHLCV for {symbol}")
        else:
            df = kite_df
    else:
        df = fetch_ohlcv_bhav(conn, symbol, from_date, to_date) if conn else pd.DataFrame()
        if df.empty and kite:
            token = token_map.get(symbol) if token_map else None
            df = fetch_ohlcv_kite(
                kite, symbol, from_date, to_date,
                failed_symbols=failed_symbols,
                instrument_token=token,
            )
    return df


# ---------------------------------------------------------------------------
# Metric computation
# ---------------------------------------------------------------------------

def calendar_anchor(as_of: datetime, *, years: int = 0, months: int = 0, days: int = 0) -> pd.Timestamp:
    ts = pd.Timestamp(as_of)
    if years or months:
        ts = ts - pd.DateOffset(years=years, months=months)
    if days:
        ts = ts - pd.Timedelta(days=days)
    return ts


def compute_return_from_anchor(df: pd.DataFrame, as_of: datetime, start_date: datetime | pd.Timestamp,
                               fallback_to_listing: bool = False) -> float:
    """Return % gain from a calendar anchor date to as_of."""
    if df.empty:
        return np.nan
    df = df[df["date"] <= pd.Timestamp(as_of)]
    if df.empty:
        return np.nan
    end_price = df.iloc[-1]["close"]
    start_date = pd.Timestamp(start_date)
    past = df[df["date"] <= start_date]
    if past.empty:
        if not fallback_to_listing:
            return np.nan
        start_price = df.iloc[0].get("open", np.nan)
        if pd.isna(start_price) or start_price == 0:
            start_price = df.iloc[0].get("close", np.nan)
        if pd.isna(start_price) or start_price == 0:
            return np.nan
        return (end_price - start_price) / start_price * 100
    start_price = past.iloc[-1]["close"]
    if start_price == 0:
        return np.nan
    return (end_price - start_price) / start_price * 100


def compute_return(df: pd.DataFrame, as_of: datetime, days: int, fallback_to_listing: bool = False) -> float:
    """Return % gain from `days` calendar days ago to as_of."""
    return compute_return_from_anchor(
        df,
        as_of,
        calendar_anchor(as_of, days=days),
        fallback_to_listing=fallback_to_listing,
    )


def compute_reset_return(df: pd.DataFrame, as_of: datetime, reset_date: datetime) -> float:
    """Return % gain from reset_date to as_of."""
    if df.empty:
        return np.nan
    df_sorted = df[df["date"] <= pd.Timestamp(as_of)].copy()
    if df_sorted.empty:
        return np.nan
    end_price = df_sorted.iloc[-1]["close"]
    past = df_sorted[df_sorted["date"] <= pd.Timestamp(reset_date)]
    if past.empty:
        return np.nan
    start_price = past.iloc[-1]["close"]
    if start_price == 0:
        return np.nan
    return (end_price - start_price) / start_price * 100


def compute_1d_return(df: pd.DataFrame, as_of: datetime) -> float:
    """Return % change on as_of day (close-to-close)."""
    if df.empty:
        return np.nan
    df_sorted = df[df["date"] <= pd.Timestamp(as_of)].copy()
    if len(df_sorted) < 2:
        return np.nan
    return (df_sorted.iloc[-1]["close"] - df_sorted.iloc[-2]["close"]) / df_sorted.iloc[-2]["close"] * 100


def compute_atr_pct(df: pd.DataFrame, as_of: datetime, period: int = ATR_PERIOD) -> float:
    """Compute ATR% = ATR(14) / current close * 100."""
    if df.empty:
        return np.nan
    df_sorted = df[df["date"] <= pd.Timestamp(as_of)].copy().tail(period + 1)
    if len(df_sorted) < 2:
        return np.nan
    highs = df_sorted["high"].values
    lows = df_sorted["low"].values
    closes = df_sorted["close"].values
    trs = []
    for i in range(1, len(df_sorted)):
        tr = max(highs[i] - lows[i], abs(highs[i] - closes[i - 1]), abs(lows[i] - closes[i - 1]))
        trs.append(tr)
    atr = np.mean(trs[-period:])
    current_close = closes[-1]
    if current_close == 0:
        return np.nan
    return atr / current_close * 100


def compute_52w_metrics(df: pd.DataFrame, as_of: datetime):
    """Return (high_52w, pct_from_high, tds_since_high)."""
    if df.empty:
        return np.nan, np.nan, np.nan
    window_start = calendar_anchor(as_of, years=1)
    df_window = df[(df["date"] >= window_start) & (df["date"] <= pd.Timestamp(as_of))].copy()
    if df_window.empty:
        return np.nan, np.nan, np.nan
    idx_high = df_window["high"].idxmax()
    high_52w = df_window.loc[idx_high, "high"]
    current_close = df_window.iloc[-1]["close"]
    if high_52w == 0:
        return np.nan, np.nan, np.nan
    pct_from_high = (high_52w - current_close) / high_52w * 100
    # trading days since high
    high_date = df_window.loc[idx_high, "date"]
    tds_since_high = len(df_window[df_window["date"] > high_date])
    return high_52w, pct_from_high, tds_since_high


def compute_volume_metrics(df: pd.DataFrame, as_of: datetime):
    """
    Find the most recent top-1%-volume event in the past six months and
    which recency window it falls in.
    Returns (vol_period, vol_day_move_pct).
    vol_period: 'last_10' | 'last_30' | 'last_60' | 'none'
    """
    if df.empty:
        return "none", np.nan
    window_start = pd.Timestamp(as_of) - timedelta(days=183)
    df_window = df[(df["date"] > window_start) & (df["date"] <= pd.Timestamp(as_of))].copy()
    if df_window.empty:
        return "none", np.nan
    df_window = df_window.reset_index(drop=True)
    n = len(df_window)

    volume_cutoff = df_window["volume"].quantile(0.99)
    qualifying = df_window[df_window["volume"] >= volume_cutoff].copy()
    if qualifying.empty:
        return "none", np.nan

    idx_max_vol = int(qualifying.index.max())

    # Position from end (0 = last trading day)
    pos_from_end = n - 1 - idx_max_vol

    if pos_from_end < 10:
        vol_period = "last_10"
    elif pos_from_end < 30:
        vol_period = "last_30"
    elif pos_from_end < 60:
        vol_period = "last_60"
    else:
        vol_period = "none"

    # Move on the selected qualifying day (close vs prev close).
    # Only positive expansion should qualify for a bullish momentum score.
    row = df_window.iloc[idx_max_vol]
    if idx_max_vol > 0:
        prev_close = df_window.iloc[idx_max_vol - 1]["close"]
        if prev_close > 0:
            vol_day_move_pct = (row["close"] - prev_close) / prev_close * 100
        else:
            vol_day_move_pct = np.nan
    else:
        vol_day_move_pct = np.nan

    if pd.isna(vol_day_move_pct) or vol_day_move_pct <= 0:
        return "none", np.nan

    return vol_period, vol_day_move_pct


def compute_trend_metrics(df: pd.DataFrame, as_of: datetime, reset_date: datetime | None = None) -> dict:
    """Compute moving-average state and uptrend consistency."""
    if df.empty:
        return {
            "sma50": np.nan,
            "ema21": np.nan,
            "ema8": np.nan,
            "above_50dma": False,
            "stack_8_21_50": False,
            "uptrend_consistency_pct": np.nan,
            "green_candle_count": np.nan,
        }

    df_sorted = df[df["date"] <= pd.Timestamp(as_of)].copy()
    if df_sorted.empty:
        return {
            "sma50": np.nan,
            "ema21": np.nan,
            "ema8": np.nan,
            "above_50dma": False,
            "stack_8_21_50": False,
            "uptrend_consistency_pct": np.nan,
            "green_candle_count": np.nan,
        }

    close_s = df_sorted["close"].astype(float)
    sma50_s = close_s.rolling(50).mean()
    ema21_s = close_s.ewm(span=21, adjust=False).mean()
    ema8_s = close_s.ewm(span=8, adjust=False).mean()

    current_close = float(close_s.iloc[-1])
    sma50 = float(sma50_s.iloc[-1]) if not pd.isna(sma50_s.iloc[-1]) else np.nan
    ema21 = float(ema21_s.iloc[-1]) if not pd.isna(ema21_s.iloc[-1]) else np.nan
    ema8 = float(ema8_s.iloc[-1]) if not pd.isna(ema8_s.iloc[-1]) else np.nan

    above_50dma = pd.notna(sma50) and current_close > sma50
    stack_8_21_50 = pd.notna(ema8) and pd.notna(ema21) and pd.notna(sma50) and (ema8 > ema21 > sma50)

    uptrend_pct = np.nan
    green_candle_count = np.nan
    green_window_start = pd.Timestamp(reset_date) if reset_date is not None else None
    if green_window_start is not None:
        green_window = df_sorted[df_sorted["date"] >= green_window_start].copy()
    else:
        green_window = df_sorted.tail(UPTREND_CON_WINDOW).copy()

    if not green_window.empty:
        green_candle_count = int(
            (green_window["close"].astype(float) > green_window["open"].astype(float)).sum()
        )

    window = df_sorted.tail(UPTREND_CON_WINDOW).copy()
    if len(window) >= 50:
        w_close = window["close"].astype(float)
        w_sma50 = w_close.rolling(50).mean()
        w_ema21 = w_close.ewm(span=21, adjust=False).mean()
        w_ema8 = w_close.ewm(span=8, adjust=False).mean()
        valid = pd.DataFrame({"ema8": w_ema8, "ema21": w_ema21, "sma50": w_sma50}).dropna()
        if not valid.empty:
            n_up = ((valid["ema8"] > valid["ema21"]) & (valid["ema21"] > valid["sma50"])).sum()
            uptrend_pct = round(n_up / len(valid) * 100.0, 2)

    return {
        "sma50": sma50,
        "ema21": ema21,
        "ema8": ema8,
        "above_50dma": bool(above_50dma),
        "stack_8_21_50": bool(stack_8_21_50),
        "uptrend_consistency_pct": uptrend_pct,
        "green_candle_count": green_candle_count,
    }


def compute_rs_composite(stock_df: pd.DataFrame, index_df: pd.DataFrame, as_of: datetime) -> dict:
    """
    Compute RS excess returns and weighted composite vs benchmark index.
    Returns dict: rs_1m, rs_3m, rs_6m, rs_12m, rs_composite
    """
    timeframes = {
        "1m": {"months": 1},
        "3m": {"months": 3},
        "6m": {"months": 6},
        "12m": {"years": 1},
    }
    rs_values = {}
    for key, offset_kwargs in timeframes.items():
        anchor = calendar_anchor(as_of, **offset_kwargs)
        stock_ret = compute_return_from_anchor(stock_df, as_of, anchor)
        idx_ret = compute_return_from_anchor(index_df, as_of, anchor)
        if pd.isna(stock_ret) or pd.isna(idx_ret):
            rs_values[key] = np.nan
        else:
            rs_values[key] = stock_ret - idx_ret

    available = {k: v for k, v in rs_values.items() if not pd.isna(v)}

    # Need at least 2 timeframes
    if len(available) < 2:
        rs_composite = np.nan
    else:
        total_weight = sum(RS_WEIGHTS[k] for k in available)
        rs_composite = sum(RS_WEIGHTS[k] * v for k, v in available.items()) / total_weight

    return {
        "rs_1m": rs_values.get("1m", np.nan),
        "rs_3m": rs_values.get("3m", np.nan),
        "rs_6m": rs_values.get("6m", np.nan),
        "rs_12m": rs_values.get("12m", np.nan),
        "rs_composite": rs_composite,
    }


# ---------------------------------------------------------------------------
# Scoring helpers
# ---------------------------------------------------------------------------

def _percentile_score(series: pd.Series, tiers: list, bottom_penalty_pct: float = None,
                      bottom_penalty_pts: int = 0) -> pd.Series:
    """
    Assign points based on percentile rank.
    tiers: list of (percentile_threshold, points).
    Lower-point tiers are applied first so higher-point tiers overwrite them.
    """
    pct = series.rank(pct=True, na_option="keep") * 100
    scores = pd.Series(0, index=series.index, dtype=int)
    for threshold, pts in sorted(tiers, key=lambda x: x[1]):
        scores[pct >= threshold] = pts
    if bottom_penalty_pct is not None:
        scores[pct <= bottom_penalty_pct] = bottom_penalty_pts
    return scores


def score_52w_high(pct_from_high: pd.Series, tds_since_high: pd.Series) -> pd.Series:
    """SC-1: 52-week high composite ladder with distance-aware recency."""
    scores = pd.Series(0, index=pct_from_high.index, dtype=int)

    within_10 = pct_from_high <= 10.0
    within_15 = (pct_from_high > 10.0) & (pct_from_high <= 15.0)
    within_20 = (pct_from_high > 15.0) & (pct_from_high <= 20.0)
    off_20 = (pct_from_high > 20.0) & (pct_from_high < 25.0)
    off_25 = pct_from_high >= 25.0

    last_10 = tds_since_high < 10
    last_15 = (tds_since_high >= 10) & (tds_since_high < 15)

    # Single mutually-exclusive ladder: recency is only rewarded if the stock is still
    # reasonably close to the 52-week high, avoiding stacked proximity+recency bonuses.
    scores[within_10 & last_10] = 12
    scores[within_10 & last_15] = 10
    scores[within_15 & last_10] = 8
    scores[within_15 & last_15] = 6
    scores[within_20 & last_10] = 4
    scores[within_20 & last_15] = 2

    scores[off_20] = -4
    scores[off_25] = -8

    return scores


def score_atr(atr_pct: pd.Series) -> pd.Series:
    """SC-2: ATR% Penalty — bottom 30th percentile and ATR% < 3.5 → -6 pts."""
    pct = atr_pct.rank(pct=True, na_option="keep") * 100
    scores = pd.Series(0, index=atr_pct.index, dtype=int)
    scores[(pct <= 30) & (atr_pct < 3.5)] = -6
    return scores


def score_volume(vol_period: pd.Series, vol_day_move_pct: pd.Series) -> pd.Series:
    """SC-3: Highest-Volume-in-Year."""
    scores = pd.Series(0, index=vol_period.index, dtype=int)

    scores[vol_period == "last_10"] += 6
    scores[vol_period == "last_30"] += 4
    scores[vol_period == "last_60"] += 3

    # Move bonus
    move_bonus = pd.Series(0, index=vol_day_move_pct.index, dtype=int)
    valid = vol_period.isin(["last_10", "last_30", "last_60"])
    move_bonus[valid & (vol_day_move_pct >= 10.0)] = 6
    move_bonus[valid & (vol_day_move_pct >= 6.0) & (vol_day_move_pct < 10.0)] = 4
    move_bonus[valid & (vol_day_move_pct >= 4.0) & (vol_day_move_pct < 6.0)] = 2

    return scores + move_bonus


def score_reset_return(reset_ret: pd.Series) -> pd.Series:
    """SC-4: Reset-Date Return."""
    return _percentile_score(
        reset_ret,
        tiers=[(90, 4), (80, 3), (70, 2)],
        bottom_penalty_pct=10,
        bottom_penalty_pts=-2,
    )


def score_day_return(day_ret: pd.Series) -> pd.Series:
    """SC-5: 1-Day Return (0 pts if negative)."""
    scores = _percentile_score(day_ret, tiers=[(90, 3), (80, 2), (70, 1)])
    # Zero out negative returns
    scores[day_ret < 0] = 0
    return scores


def score_rs(rs_composite: pd.Series) -> pd.Series:
    """SC-8: Relative Strength scoring."""
    scores = _percentile_score(
        rs_composite,
        tiers=[(90, 6), (80, 4), (70, 2)],
        bottom_penalty_pct=10,
        bottom_penalty_pts=-2,
    )
    # NaN RS → 0
    scores[rs_composite.isna()] = 0
    return scores


def score_turnover(avg_turnover_cr: pd.Series) -> pd.Series:
    """No turnover penalty; liquidity is already enforced by the universe filter."""
    return pd.Series(0, index=avg_turnover_cr.index, dtype=int)


def score_ret12_bonus(ret_12m: pd.Series) -> pd.Series:
    scores = pd.Series(0, index=ret_12m.index, dtype=int)
    pct = ret_12m.rank(pct=True, na_option="keep") * 100
    scores[pct >= 80] = 4
    scores[ret_12m.isna()] = 0
    return scores


def score_ret6_bonus(ret_6m: pd.Series) -> pd.Series:
    scores = pd.Series(0, index=ret_6m.index, dtype=int)
    pct = ret_6m.rank(pct=True, na_option="keep") * 100
    scores[pct >= 80] = 2
    scores[ret_6m.isna()] = 0
    return scores


def score_ret3_bonus(ret_3m: pd.Series) -> pd.Series:
    scores = pd.Series(0, index=ret_3m.index, dtype=int)
    pct = ret_3m.rank(pct=True, na_option="keep") * 100
    scores[pct >= 80] = 4
    scores[ret_3m.isna()] = 0
    return scores


def score_ma_trend(above_50dma: pd.Series, stack_8_21_50: pd.Series,
                   uptrend_consistency_pct: pd.Series,
                   green_candle_count: pd.Series) -> tuple[pd.Series, pd.Series, pd.Series, pd.Series, pd.Series]:
    score_50dma = pd.Series(0, index=above_50dma.index, dtype=int)
    score_stack = pd.Series(0, index=stack_8_21_50.index, dtype=int)
    score_consistency = pd.Series(0, index=uptrend_consistency_pct.index, dtype=int)
    score_green_candles = pd.Series(0, index=green_candle_count.index, dtype=int)

    score_50dma[above_50dma.fillna(False)] = 2
    score_stack[stack_8_21_50.fillna(False)] = 2
    score_consistency[uptrend_consistency_pct >= UPTREND_CON_MIN_PCT] = 2
    green_pct = green_candle_count.rank(pct=True, na_option="keep") * 100
    score_green_candles[green_pct >= 80] = 2
    score_green_candles[green_pct >= 90] = 4

    total = score_50dma + score_stack + score_consistency + score_green_candles
    return score_50dma, score_stack, score_consistency, score_green_candles, total


def assign_rating(scores: pd.Series) -> pd.Series:
    """SC-7: Rating Assignment based on total score percentile."""
    pct = scores.rank(pct=True) * 100
    rating = pd.Series("WEAK BUY ★", index=scores.index)
    rating[pct > 40] = "BUY ★★"
    rating[pct > 70] = "STRONG BUY ★★★"
    return rating


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------

def compute_all_metrics(symbols: list, ohlcv_map: dict, index_df: pd.DataFrame,
                        as_of: datetime, reset_date: datetime) -> pd.DataFrame:
    """Compute per-symbol metrics for a set of symbols."""
    rows = []
    for sym in symbols:
        df = ohlcv_map.get(sym, pd.DataFrame())
        if df.empty:
            continue
        ret_12m = compute_return_from_anchor(df, as_of, calendar_anchor(as_of, years=1), fallback_to_listing=True)
        ret_6m = compute_return_from_anchor(df, as_of, calendar_anchor(as_of, months=6))
        ret_3m = compute_return_from_anchor(df, as_of, calendar_anchor(as_of, months=3))
        ret_1m = compute_return_from_anchor(df, as_of, calendar_anchor(as_of, months=1))
        ret_1d = compute_1d_return(df, as_of)
        ret_reset = compute_reset_return(df, as_of, reset_date)
        atr_pct_val = compute_atr_pct(df, as_of)
        high_52w, pct_from_high, tds_since_high = compute_52w_metrics(df, as_of)
        vol_period, vol_day_move = compute_volume_metrics(df, as_of)
        trend = compute_trend_metrics(df, as_of, reset_date)
        current_close = df[df["date"] <= pd.Timestamp(as_of)].iloc[-1]["close"] if not df.empty else np.nan

        rs = compute_rs_composite(df, index_df, as_of) if not index_df.empty else {
            "rs_1m": np.nan, "rs_3m": np.nan, "rs_6m": np.nan, "rs_12m": np.nan, "rs_composite": np.nan
        }

        rows.append({
            "symbol": sym,
            "close": current_close,
            "ret_12m": ret_12m,
            "ret_6m": ret_6m,
            "ret_3m": ret_3m,
            "ret_1m": ret_1m,
            "ret_reset": ret_reset,
            "ret_1d": ret_1d,
            "atr_pct": atr_pct_val,
            "high_52w": high_52w,
            "pct_from_52w_high": pct_from_high,
            "tds_since_52w_high": tds_since_high,
            "vol_period": vol_period,
            "vol_day_move_pct": vol_day_move,
            **trend,
            **rs,
        })

    return pd.DataFrame(rows)


def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    """Apply pre-score gates: 52W proximity + 12M return percentile."""
    if df.empty:
        return df.copy()

    start_count = len(df)

    # 52W high gate: pass if near the 52W high, or made a fresh high recently,
    # but recent-high inclusion is still capped at 20% away from the high.
    df = df[
        (df["pct_from_52w_high"] <= MAX_RECENCY_GATE_DIST_PCT) &
        (
            (df["pct_from_52w_high"] <= MAX_DIST_FROM_52W_HIGH_PCT) |
            (df["tds_since_52w_high"] <= MAX_TDS_SINCE_52W_HIGH)
        )
    ].copy()
    log.info(f"Filter pass - 52W gate: {len(df)}/{start_count}")

    # 12M return gate: threshold is computed on the post-52W survivors.
    df = df[df["ret_12m"].notna()].copy()
    if df.empty:
        log.warning("Filter pass - 12M gate: 0 stocks with valid 12M return")
        return df

    threshold_70 = df["ret_12m"].quantile(RETURN_GATE_PERCENTILE / 100)
    df = df[df["ret_12m"] >= threshold_70].copy()
    log.info(
        f"Filter pass - 12M gate: {len(df)} stocks >= {RETURN_GATE_PERCENTILE}th percentile "
        f"({threshold_70:.2f})"
    )

    return df


def score_universe(df: pd.DataFrame) -> pd.DataFrame:
    """Apply all scoring components and total score."""
    df = df.copy()
    df["score_52w"] = score_52w_high(df["pct_from_52w_high"], df["tds_since_52w_high"])
    df["score_atr"] = score_atr(df["atr_pct"])
    df["score_volume"] = score_volume(df["vol_period"], df["vol_day_move_pct"])
    df["score_reset"] = score_reset_return(df["ret_reset"])
    df["score_1d"] = score_day_return(df["ret_1d"])
    df["score_rs"] = score_rs(df["rs_composite"])
    df["score_turnover"] = score_turnover(df["avg_turnover_cr"])
    df["score_12m_bonus"] = score_ret12_bonus(df["ret_12m"])
    df["score_6m_bonus"] = score_ret6_bonus(df["ret_6m"])
    df["score_3m_bonus"] = score_ret3_bonus(df["ret_3m"])
    (
        df["score_50dma"],
        df["score_stack"],
        df["score_uptrend_consistency"],
        df["score_green_candles"],
        df["score_trend"],
    ) = score_ma_trend(
        df["above_50dma"],
        df["stack_8_21_50"],
        df["uptrend_consistency_pct"],
        df["green_candle_count"],
    )
    df["total_score"] = (
        df["score_52w"] + df["score_atr"] + df["score_volume"] +
        df["score_reset"] + df["score_1d"] + df["score_rs"] + df["score_turnover"] +
        df["score_12m_bonus"] + df["score_6m_bonus"] + df["score_3m_bonus"] +
        df["score_trend"]
    )
    # RS percentile within this universe
    df["rs_percentile"] = df["rs_composite"].rank(pct=True, na_option="keep") * 100
    df["rating"] = assign_rating(df["total_score"])
    return df.sort_values("total_score", ascending=False).reset_index(drop=True)


def compute_sector_summary(scored_df: pd.DataFrame, sectors_df: pd.DataFrame) -> pd.DataFrame:
    """Compute top-5 sector summary from scored stocks."""
    merged = scored_df.merge(sectors_df, on="symbol", how="left")
    merged["sector"] = merged["sector"].fillna("Unknown")

    sector_groups = merged.groupby("sector")
    sector_rows = []
    for sector, grp in sector_groups:
        if len(grp) < 2:
            continue
        top_stock = grp.nsmallest(1, "rank")["symbol"].values[0] if "rank" in grp.columns else grp.iloc[0]["symbol"]
        best_rank = grp["rank"].min() if "rank" in grp.columns else None
        breadth_pct = (
            grp["rating"].astype(str).str.contains("STRONG", na=False).mean() * 100
            if "rating" in grp.columns else 0.0
        )
        sector_rows.append({
            "sector": sector,
            "stock_count": len(grp),
            "avg_score": grp["total_score"].mean(),
            "breadth_pct": breadth_pct,
            "avg_rs_composite": grp["rs_composite"].mean() if "rs_composite" in grp.columns else np.nan,
            "avg_12m": grp["ret_12m"].mean(),
            "avg_6m": grp["ret_6m"].mean(),
            "avg_3m": grp["ret_3m"].mean(),
            "avg_reset": grp["ret_reset"].mean(),
            "avg_1d": grp["ret_1d"].mean(),
            "avg_turnover_cr": grp["avg_turnover_cr"].mean() if "avg_turnover_cr" in grp.columns else np.nan,
            "median_turnover_cr": grp["median_turnover_cr"].mean() if "median_turnover_cr" in grp.columns else np.nan,
            "avg_atr_pct": grp["atr_pct"].mean(),
            "top_stock": top_stock,
            "best_rank": best_rank,
            "constituent_symbols": ", ".join(grp["symbol"].tolist()),
        })

    if not sector_rows:
        return pd.DataFrame()

    sector_df = pd.DataFrame(sector_rows)
    def _norm(s: pd.Series) -> pd.Series:
        lo, hi = s.min(), s.max()
        return (s - lo) / (hi - lo) if hi > lo else pd.Series(0.5, index=s.index)

    n_score = _norm(sector_df["avg_score"].astype(float))
    n_breadth = _norm(sector_df["breadth_pct"].astype(float))
    n_count = _norm(sector_df["stock_count"].astype(float))
    n_rs = _norm(sector_df["avg_rs_composite"].fillna(0.0).astype(float))

    sector_df["n_avg_score"] = n_score.round(4)
    sector_df["n_breadth"] = n_breadth.round(4)
    sector_df["n_rs"] = n_rs.round(4)
    sector_df["n_count"] = n_count.round(4)
    sector_df["sector_composite_score"] = (
        SECTOR_W_SCORE * n_score
        + SECTOR_W_BREADTH * n_breadth
        + SECTOR_W_RS * n_rs
        + SECTOR_W_COUNT * n_count
    ).round(4)
    sector_df = sector_df.sort_values(
        ["sector_composite_score", "avg_score", "avg_12m"], ascending=[False, False, False]
    ).head(5).reset_index(drop=True)
    sector_df.insert(0, "rank", range(1, len(sector_df) + 1))
    return sector_df


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def _col_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter(s)."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _force_delete_xlsx(fname: Path) -> None:
    """
    Delete fname. If Excel has it open, try closing just that workbook first.
    """
    if not fname.exists():
        return
    try:
        fname.unlink()
        log.info(f"Deleted old file: {fname.name}")
        return
    except PermissionError:
        pass

    log.warning(f"File is open in Excel - attempting to close it: {fname.name}")

    closed = False
    try:
        import win32com.client
        xl = win32com.client.GetActiveObject("Excel.Application")
        for wb in list(xl.Workbooks):
            if os.path.normcase(wb.FullName) == os.path.normcase(str(fname.resolve())):
                wb.Close(SaveChanges=False)
                closed = True
                log.info("Closed workbook via Excel COM.")
                break
    except Exception:
        pass

    if not closed:
        try:
            ps = (
                "$xl = [Runtime.InteropServices.Marshal]"
                "::GetActiveObject('Excel.Application'); "
                f"$xl.Workbooks | Where-Object {{ $_.FullName -eq '{str(fname.resolve())}' }}"
                " | ForEach-Object { $_.Close($false) }"
            )
            subprocess.run(["powershell", "-Command", ps], capture_output=True, timeout=10)
            closed = True
            log.info("Closed workbook via PowerShell.")
        except Exception:
            pass

    time.sleep(0.5)
    try:
        fname.unlink()
        log.info(f"Deleted old file: {fname.name}")
    except PermissionError:
        log.error(f"Still cannot delete {fname}. Please close Excel completely and re-run.")
        sys.exit(1)


def write_excel_report(
    top30: pd.DataFrame,
    excluded_top30_atr: pd.DataFrame,
    top5_sectors: pd.DataFrame,
    all_scored: pd.DataFrame,
    pipeline_summary: dict,
    extended_scored: pd.DataFrame,
    as_of_date: datetime,
    reset_date: datetime,
    failed_symbols: list,
) -> Path:
    """Generate the Excel report with a ready reckoner and analysis sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Alignment, Font, PatternFill,
                                     numbers, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    OUTPUT_DIR.mkdir(exist_ok=True)
    filename = f"LiquidCandidates_{as_of_date.strftime('%d%b%Y').upper()}_Reset{reset_date.strftime('%d%b%Y').upper()}.xlsx"
    filepath = OUTPUT_DIR / filename
    _force_delete_xlsx(filepath)

    wb = Workbook()

    # ----- colour palette -----
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    ALT_ROW_FILL = PatternFill("solid", start_color="EBF3FB")
    SCORE_FILL = PatternFill("solid", start_color="FFFF99")
    FONT_NORMAL = Font(name="Arial", size=9)
    FONT_BOLD = Font(name="Arial", bold=True, size=9)
    GREEN_FONT = Font(name="Arial", color="006100", size=9)
    RED_FONT = Font(name="Arial", color="9C0006", size=9)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row_num: int, ncols: int):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    def style_data_row(ws, row_num: int, ncols: int, alt: bool):
        fill = ALT_ROW_FILL if alt else PatternFill("solid", start_color="FFFFFF")
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def colour_return_cell(cell, value):
        if pd.isna(value):
            return
        if value > 0:
            cell.font = GREEN_FONT
        elif value < 0:
            cell.font = RED_FONT

    def excel_value(value):
        if pd.isna(value):
            return ""
        if isinstance(value, (np.floating, float)):
            return round(float(value), 2)
        return value

    def threshold_or_blank(series: pd.Series, q: float):
        clean = pd.to_numeric(series, errors="coerce").dropna()
        if clean.empty:
            return ""
        return round(float(clean.quantile(q)), 2)

    # ================================================================
    # Sheet 1: Top 20 Stocks
    # ================================================================
    ws1 = wb.active
    ws1.title = "Top 20 Stocks"

    top30_cols = [
        ("Rank", "rank"), ("Symbol", "symbol"), ("Source", "source"), ("Sector", "sector"),
        ("Close ₹", "close"), ("Total Score", "total_score"), ("Rating", "rating"),
        ("Avg TO 21D (₹ Cr)", "avg_turnover_cr"),
        ("Median TO 21D (??? Cr)", "median_turnover_cr"),
        ("Ret Reset%", "ret_reset"), ("1D Ret%", "ret_1d"),
        ("12M Ret%", "ret_12m"), ("6M Ret%", "ret_6m"), ("3M Ret%", "ret_3m"),
        ("50DMA", "sma50"), ("21EMA", "ema21"), ("8EMA", "ema8"),
        ("Above 50DMA", "above_50dma"), ("8EMA>21EMA>50DMA", "stack_8_21_50"),
        ("Uptrend Con %", "uptrend_consistency_pct"), ("Green Candles", "green_candle_count"),
        ("52W High", "high_52w"), ("% From 52W High", "pct_from_52w_high"),
        ("TDs Since 52W High", "tds_since_52w_high"),
        ("Vol Period", "vol_period"), ("Vol Day Move%", "vol_day_move_pct"), ("ATR%", "atr_pct"),
        ("RS Composite%", "rs_composite"), ("RS Percentile", "rs_percentile"),
        ("RS 3M vs Index%", "rs_3m"), ("Score: RS", "score_rs"),
        ("Score: TO", "score_turnover"),
        ("Score: 12M Bonus", "score_12m_bonus"), ("Score: 6M Bonus", "score_6m_bonus"),
        ("Score: 3M Bonus", "score_3m_bonus"),
        ("Score: 50DMA", "score_50dma"), ("Score: Stack", "score_stack"),
        ("Score: Uptrend Con", "score_uptrend_consistency"), ("Score: Green Candles", "score_green_candles"),
        ("Score: 52W", "score_52w"), ("Score: ATR", "score_atr"), ("Score: Vol", "score_volume"),
        ("Score: Reset", "score_reset"), ("Score: 1D", "score_1d"),
    ]
    headers = [c[0] for c in top30_cols]
    keys = [c[1] for c in top30_cols]
    score_col_indices = {i + 1 for i, k in enumerate(keys) if k.startswith("score_") or k == "total_score"}

    ws1.append(headers)
    style_header_row(ws1, 1, len(headers))
    ws1.row_dimensions[1].height = 30

    for i, (_, row) in enumerate(top30.iterrows()):
        data = [excel_value(row.get(k, "")) for k in keys]
        ws1.append(data)
        r = i + 2
        style_data_row(ws1, r, len(headers), alt=(i % 2 == 1))
        # score columns → yellow bg
        for col in score_col_indices:
            ws1.cell(row=r, column=col).fill = SCORE_FILL
            ws1.cell(row=r, column=col).font = FONT_BOLD
        # colour return columns
        for ret_key in ["ret_reset", "ret_1d", "ret_12m", "ret_6m", "ret_3m",
                         "rs_composite", "rs_3m"]:
            if ret_key in keys:
                col_idx = keys.index(ret_key) + 1
                colour_return_cell(ws1.cell(row=r, column=col_idx), row.get(ret_key, np.nan))

    # Column widths
    for col_idx, header in enumerate(headers, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(12, len(header) + 2)

    ws1.freeze_panes = "A2"

    # ================================================================
    # Sheet 2: Top 5 Sectors
    # ================================================================
    ws2 = wb.create_sheet("Top 5 Sectors")
    sec_cols = [
        ("Rank", "rank"), ("Sector", "sector"), ("# Stocks", "stock_count"),
        ("Top Stock", "top_stock"), ("Best Rank", "best_rank"),
        ("Sector Composite", "sector_composite_score"),
        ("Avg Score", "avg_score"), ("Breadth %", "breadth_pct"), ("Avg RS", "avg_rs_composite"),
        ("Norm Score", "n_avg_score"), ("Norm Breadth", "n_breadth"),
        ("Norm RS", "n_rs"), ("Norm Count", "n_count"),
        ("Avg 12M%", "avg_12m"),
        ("Avg 6M%", "avg_6m"), ("Avg 3M%", "avg_3m"),
        ("Avg Reset%", "avg_reset"), ("Avg 1D%", "avg_1d"),
        ("Avg TO 21D (₹ Cr)", "avg_turnover_cr"), ("Avg ATR%", "avg_atr_pct"),
        ("Constituent Symbols", "constituent_symbols"),
    ]
    sec_headers = [c[0] for c in sec_cols]
    sec_keys = [c[1] for c in sec_cols]
    ws2.append(sec_headers)
    style_header_row(ws2, 1, len(sec_headers))
    ws2.row_dimensions[1].height = 25

    for i, (_, row) in enumerate(top5_sectors.iterrows()):
        data = [excel_value(row.get(k, "")) for k in sec_keys]
        ws2.append(data)
        style_data_row(ws2, i + 2, len(sec_headers), alt=(i % 2 == 1))

    for col_idx, header in enumerate(sec_headers, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)
    ws2.column_dimensions[get_column_letter(len(sec_headers))].width = 60
    ws2.freeze_panes = "A2"

    # ================================================================
    # Sheet 3: All Scored Stocks
    # ================================================================
    ws3 = wb.create_sheet("All Scored Stocks")
    ws3.append(headers)
    style_header_row(ws3, 1, len(headers))
    ws3.row_dimensions[1].height = 30
    for i, (_, row) in enumerate(all_scored.iterrows()):
        data = [excel_value(row.get(k, "")) for k in keys]
        ws3.append(data)
        style_data_row(ws3, i + 2, len(headers), alt=(i % 2 == 1))
    for col_idx in range(1, len(headers) + 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 13
    ws3.freeze_panes = "A2"

    # ================================================================
    # Sheet 4: Pipeline Summary
    # ================================================================
    ws4 = wb.create_sheet("Pipeline Summary")
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 18

    ws4["A1"] = "Pipeline Summary"
    ws4["A1"].font = Font(name="Arial", bold=True, size=12)

    summary_rows = [
        ("Stage", "Count"),
        ("Total NSE EQ symbols", pipeline_summary.get("total_eq", "")),
        ("After close ≥ ₹50 filter", pipeline_summary.get("after_close_filter", "")),
        (f"Top {TOP_N} by avg turnover", pipeline_summary.get("topn_count", "")),
        (f"After ATR% >= {TOP30_MIN_ATR_PCT:.1f} pre-filter", pipeline_summary.get("after_atr_filter", "")),
        ("After 52W high gate", pipeline_summary.get("after_52w_gate", "")),
        ("After 12M return gate (70th pctile)", pipeline_summary.get("after_12m_gate", "")),
        (f"Scored Top {TOP_N} stocks", pipeline_summary.get("scored_topn", "")),
        ("Extended universe candidates", pipeline_summary.get("extended_candidates", "")),
        ("Scored Extended stocks", pipeline_summary.get("scored_extended", "")),
        ("Total scored (merged)", pipeline_summary.get("total_scored", "")),
        ("Initial Top 20 shortlist", pipeline_summary.get("top30_initial_count", "")),
        (f"Top 20 excluded: ATR% < {TOP30_MIN_ATR_PCT}", pipeline_summary.get("top30_atr_excluded_count", "")),
        ("Top 20 final report", pipeline_summary.get("top30_count", "")),
        ("Failed API calls", len(failed_symbols)),
    ]
    for r_idx, (label, val) in enumerate(summary_rows, start=2):
        ws4.cell(row=r_idx, column=1, value=label).font = FONT_BOLD if r_idx == 2 else FONT_NORMAL
        ws4.cell(row=r_idx, column=2, value=val).font = FONT_NORMAL

    # RS stats
    ws4["A14"] = "RS Stats"
    ws4["A14"].font = Font(name="Arial", bold=True, size=10)
    if not all_scored.empty and "rs_composite" in all_scored.columns:
        rs_valid = all_scored["rs_composite"].dropna()
        ws4["A15"] = "RS Composite — Mean"
        ws4["B15"] = round(rs_valid.mean(), 2) if len(rs_valid) else ""
        ws4["A16"] = "RS Composite — Median"
        ws4["B16"] = round(rs_valid.median(), 2) if len(rs_valid) else ""
        ws4["A17"] = "Stocks with RS >= 0"
        ws4["B17"] = int((rs_valid >= 0).sum())

    if failed_symbols:
        ws4["A19"] = "Failed Symbols"
        ws4["A19"].font = FONT_BOLD
        for i, sym in enumerate(failed_symbols, start=20):
            ws4.cell(row=i, column=1, value=sym).font = FONT_NORMAL

    # ================================================================
    # Sheet 5: Ready Reckoner
    # ================================================================
    ws_rr = wb.create_sheet("Ready Reckoner")
    ws_rr.merge_cells("A1:E1")
    ws_rr["A1"] = (
        f"READY RECKONER | As of {as_of_date.strftime('%d-%b-%Y')} | "
        "Filters, scoring logic, and live thresholds used by this run"
    )
    ws_rr["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    ws_rr["A1"].fill = HEADER_FILL
    ws_rr["A1"].alignment = Alignment(horizontal="center", vertical="center")

    rr_headers = ["Category", "Metric", "Rule", "Points", "Current Threshold / Notes"]
    ws_rr.append(rr_headers)
    style_header_row(ws_rr, 2, len(rr_headers))
    ws_rr.row_dimensions[2].height = 28

    top30_turnover_floor = threshold_or_blank(top30["avg_turnover_cr"], 0.0) if not top30.empty and "avg_turnover_cr" in top30.columns else ""
    atr_bottom30 = threshold_or_blank(all_scored["atr_pct"], 0.30) if not all_scored.empty and "atr_pct" in all_scored.columns else ""
    turnover_bottom20 = threshold_or_blank(all_scored["avg_turnover_cr"], 0.20) if not all_scored.empty and "avg_turnover_cr" in all_scored.columns else ""
    ret12_top80 = threshold_or_blank(all_scored["ret_12m"], 0.80) if not all_scored.empty and "ret_12m" in all_scored.columns else ""
    ret6_top80 = threshold_or_blank(all_scored["ret_6m"], 0.80) if not all_scored.empty and "ret_6m" in all_scored.columns else ""
    ret3_top80 = threshold_or_blank(all_scored["ret_3m"], 0.80) if not all_scored.empty and "ret_3m" in all_scored.columns else ""
    reset_top90 = threshold_or_blank(all_scored["ret_reset"], 0.90) if not all_scored.empty and "ret_reset" in all_scored.columns else ""
    reset_top80 = threshold_or_blank(all_scored["ret_reset"], 0.80) if not all_scored.empty and "ret_reset" in all_scored.columns else ""
    reset_top70 = threshold_or_blank(all_scored["ret_reset"], 0.70) if not all_scored.empty and "ret_reset" in all_scored.columns else ""
    reset_bottom10 = threshold_or_blank(all_scored["ret_reset"], 0.10) if not all_scored.empty and "ret_reset" in all_scored.columns else ""
    day_top90 = threshold_or_blank(all_scored["ret_1d"], 0.90) if not all_scored.empty and "ret_1d" in all_scored.columns else ""
    day_top80 = threshold_or_blank(all_scored["ret_1d"], 0.80) if not all_scored.empty and "ret_1d" in all_scored.columns else ""
    day_top70 = threshold_or_blank(all_scored["ret_1d"], 0.70) if not all_scored.empty and "ret_1d" in all_scored.columns else ""
    rs_top90 = threshold_or_blank(all_scored["rs_composite"], 0.90) if not all_scored.empty and "rs_composite" in all_scored.columns else ""
    rs_top80 = threshold_or_blank(all_scored["rs_composite"], 0.80) if not all_scored.empty and "rs_composite" in all_scored.columns else ""
    rs_top70 = threshold_or_blank(all_scored["rs_composite"], 0.70) if not all_scored.empty and "rs_composite" in all_scored.columns else ""
    rs_bottom10 = threshold_or_blank(all_scored["rs_composite"], 0.10) if not all_scored.empty and "rs_composite" in all_scored.columns else ""
    green_top80 = threshold_or_blank(all_scored["green_candle_count"], 0.80) if not all_scored.empty and "green_candle_count" in all_scored.columns else ""
    green_top90 = threshold_or_blank(all_scored["green_candle_count"], 0.90) if not all_scored.empty and "green_candle_count" in all_scored.columns else ""

    rr_rows = [
        ("Universe Filter", "Close", "Latest close must be >= ₹50", "Required", "Pre-universe gate"),
        ("Universe Filter", "Turnover rank", f"Top {TOP_N} by average turnover", "Required", f"Current shortlist floor ≈ {top30_turnover_floor} Cr"),
        ("Universe Filter", "52W gate", f"Within {MAX_DIST_FROM_52W_HIGH_PCT:.0f}% of 52W high OR made 52W high within {MAX_TDS_SINCE_52W_HIGH} trading days, and in all cases within {MAX_RECENCY_GATE_DIST_PCT:.0f}% of 52W high", "Required", "Pre-score gate with 20% hard cap"),
        ("Universe Filter", "12M return gate", f"12M return >= {RETURN_GATE_PERCENTILE}th percentile", "Required", f"Current cutoff ≈ {ret12_top80 if RETURN_GATE_PERCENTILE == 80 else threshold_or_blank(all_scored['ret_12m'], RETURN_GATE_PERCENTILE/100) if not all_scored.empty and 'ret_12m' in all_scored.columns else ''}%"),
        ("Scoring", "52W composite", "<=10% & <10D / <=10% & <15D", "+12 / +10", "Distance-aware recency ladder"),
        ("Scoring", "52W composite", ">10-15% & <10D / >10-15% & <15D", "+8 / +6", "Distance-aware recency ladder"),
        ("Scoring", "52W composite", ">15-20% & <10D / >15-20% & <15D", "+4 / +2", "No 52W recency points beyond 20% away"),
        ("Scoring", "52W penalty", ">20-25% / >=25% from 52W high", "-4 / -8", "Distance penalty"),
        ("Scoring", "ATR%", "Bottom 30th percentile ATR% and ATR% < 3.5", "-6", f"Current percentile cutoff ≈ {atr_bottom30}"),
        ("Scoring", "Volume event", "Most recent day in top 1% volume of last 6 months, within last 10D / 30D / 60D", "+6 / +4 / +3", "Recent top-1%-volume event score"),
        ("Scoring", "Volume move bonus", "Qualifying volume event with move >= 10% / 6% / 4%", "+6 / +4 / +2", "Adds to volume score"),
        ("Scoring", "Reset-date return", "Top 90 / 80 / 70 percentile", "+4 / +3 / +2", f"Current cutoffs ≈ {reset_top90} / {reset_top80} / {reset_top70}"),
        ("Scoring", "Reset-date return", "Bottom 10th percentile", "-2", f"Current cutoff ≈ {reset_bottom10}"),
        ("Scoring", "1-day return", "Top 90 / 80 / 70 percentile with non-negative return", "+3 / +2 / +1", f"Current cutoffs ≈ {day_top90} / {day_top80} / {day_top70}"),
        ("Scoring", "Relative strength composite", "Top 90 / 80 / 70 percentile", "+6 / +4 / +2", f"Current cutoffs ≈ {rs_top90} / {rs_top80} / {rs_top70}"),
        ("Scoring", "Relative strength composite", "Bottom 10th percentile", "-2", f"Current cutoff ≈ {rs_bottom10}"),
        ("Scoring", "Turnover", "No turnover penalty inside shortlisted liquid universe", "0", f"Top {TOP_N} liquidity filter already applied"),
        ("Scoring", "12M bonus", "12M return >= 80th percentile", "+4", f"Current cutoff ≈ {ret12_top80}"),
        ("Scoring", "6M bonus", "6M return >= 80th percentile", "+2", f"Current cutoff ≈ {ret6_top80}"),
        ("Scoring", "3M bonus", "3M return >= 80th percentile", "+4", f"Current cutoff ≈ {ret3_top80}"),
        ("Scoring", "Trend", "Above 50DMA", "+2", "Current close > 50DMA"),
        ("Scoring", "Trend stack", "8EMA > 21EMA > 50DMA", "+2", "Stacked trend"),
        ("Scoring", "Uptrend consistency", f"Uptrend consistency >= {UPTREND_CON_MIN_PCT:.0f}%", "+2", f"Window = {UPTREND_CON_WINDOW} trading days"),
        ("Scoring", "Green candles", "Top 80 / 90 percentile", "+2 / +4", f"Current cutoffs ≈ {green_top80} / {green_top90}"),
        ("Universe Filter", "ATR% pre-filter", f"ATR% must be >= {TOP30_MIN_ATR_PCT:.1f}", "Required", "Applied before 52W / 12M gates and scoring"),
        ("Sector", "Top 5 sectors", "Composite = 40% avg score + 25% breadth + 20% RS + 15% count", "Reference", "Sector ranking logic"),
        ("Output", "ATR guard", f"No additional Top 20 ATR exclusion beyond ATR% >= {TOP30_MIN_ATR_PCT:.1f}", "Reference", "Top 20 exclusion sheet should normally be empty"),
    ]

    for row in rr_rows:
        ws_rr.append(list(row))
    for r in range(3, ws_rr.max_row + 1):
        style_data_row(ws_rr, r, len(rr_headers), alt=((r - 3) % 2 == 1))
        ws_rr.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws_rr.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_rr.cell(row=r, column=5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_rr.column_dimensions["A"].width = 18
    ws_rr.column_dimensions["B"].width = 24
    ws_rr.column_dimensions["C"].width = 48
    ws_rr.column_dimensions["D"].width = 14
    ws_rr.column_dimensions["E"].width = 34
    ws_rr.freeze_panes = "A3"

    # ================================================================
    # Sheet 6: Extended Universe Stocks
    # ================================================================
    ws5 = wb.create_sheet("Extended Universe Stocks")
    ws5.append(headers)
    style_header_row(ws5, 1, len(headers))
    ws5.row_dimensions[1].height = 30
    if not extended_scored.empty:
        for i, (_, row) in enumerate(extended_scored.iterrows()):
            data = [excel_value(row.get(k, "")) for k in keys]
            ws5.append(data)
            style_data_row(ws5, i + 2, len(headers), alt=(i % 2 == 1))
    for col_idx in range(1, len(headers) + 1):
        ws5.column_dimensions[get_column_letter(col_idx)].width = 13
    ws5.freeze_panes = "A2"

    # ================================================================
    # Sheet 7: ATR Excluded Top 20
    # ================================================================
    ws6 = wb.create_sheet("ATR Excluded Top20")
    ws6.append(headers)
    style_header_row(ws6, 1, len(headers))
    ws6.row_dimensions[1].height = 30
    if not excluded_top30_atr.empty:
        for i, (_, row) in enumerate(excluded_top30_atr.iterrows()):
            data = [excel_value(row.get(k, "")) for k in keys]
            ws6.append(data)
            style_data_row(ws6, i + 2, len(headers), alt=(i % 2 == 1))
            for col in score_col_indices:
                ws6.cell(row=i + 2, column=col).fill = SCORE_FILL
                ws6.cell(row=i + 2, column=col).font = FONT_BOLD
    for col_idx in range(1, len(headers) + 1):
        ws6.column_dimensions[get_column_letter(col_idx)].width = 13
    ws6.freeze_panes = "A2"

    wb.save(filepath)
    log.info(f"Excel report saved: {filepath}")
    return filepath


def write_tradingview_watchlist(top30: pd.DataFrame, sectors_df: pd.DataFrame,
                                filepath: Path, as_of_date: datetime):
    """Generate TradingView watchlist .txt in final_wl format grouped by sector."""
    merged = top30.copy()
    if "sector" not in merged.columns:
        if not sectors_df.empty:
            merged = merged.merge(sectors_df[["symbol", "sector"]], on="symbol", how="left")
        else:
            merged["sector"] = "Unknown"
    elif not sectors_df.empty and merged["sector"].isna().any():
        sector_map = sectors_df.set_index("symbol")["sector"]
        merged["sector"] = merged["sector"].fillna(merged["symbol"].map(sector_map))

    merged["sector"] = (
        merged["sector"]
        .astype(str)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
        .fillna("Unknown")
    )

    sectors: dict[str, list[str]] = {}
    for _, row in merged.iterrows():
        sector = str(row.get("sector", "Unknown")).strip() or "Unknown"
        sym = str(row.get("symbol", "")).strip().replace("-", "_")
        if not sym:
            continue
        tv_sym = sym if ":" in sym else f"NSE:{sym}"
        sectors.setdefault(sector, []).append(tv_sym)

    lines = [
        f"### Final Watchlist — {as_of_date.strftime('%d %b %Y')} "
        f"(Top {len(top30)} by Momentum Score)",
        "",
    ]
    for sector in sorted(sectors.keys()):
        lines.append(f"###{sector}")
        lines.extend(sectors[sector])
        lines.append("")

    filepath.write_text("\n".join(lines), encoding="utf-8")
    log.info(f"TradingView watchlist saved: {filepath}")


def read_symbol_list_file(filepath: Path) -> list[str]:
    if not filepath.exists():
        return []

    symbols: list[str] = []
    seen: set[str] = set()
    for raw_line in filepath.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        symbol = line.split(":", 1)[-1].strip().upper().replace("-", "_")
        if not symbol or symbol in seen:
            continue
        seen.add(symbol)
        symbols.append(symbol)
    return symbols


def write_updated_gmlist(top30: pd.DataFrame, as_of_date: datetime) -> Path:
    GMLIST_DIR.mkdir(parents=True, exist_ok=True)
    base_name = as_of_date.strftime("%d%b%Y")
    original_path = GMLIST_DIR / f"gmlist_{base_name}.txt"
    updated_path = GMLIST_DIR / f"updated_gmlist_{base_name}.txt"

    merged_symbols: list[str] = []
    seen: set[str] = set()

    for symbol in read_symbol_list_file(original_path):
        if symbol not in seen:
            seen.add(symbol)
            merged_symbols.append(symbol)

    if "symbol" in top30.columns:
        for raw_symbol in top30["symbol"].tolist():
            symbol = str(raw_symbol).strip().upper().replace("NSE:", "").replace("-", "_")
            if not symbol or symbol == "NAN" or symbol in seen:
                continue
            seen.add(symbol)
            merged_symbols.append(symbol)

    updated_path.write_text("\n".join(merged_symbols) + ("\n" if merged_symbols else ""), encoding="utf-8")
    log.info(f"Updated gmlist saved: {updated_path}")
    return updated_path


def build_liquid_leader_bonus_payload(liquid_leaders: pd.DataFrame) -> dict[str, int]:
    payload: dict[str, int] = {}
    if "symbol" not in liquid_leaders.columns or liquid_leaders.empty:
        return payload
    for index, raw_symbol in enumerate(liquid_leaders.head(20)["symbol"].tolist(), start=1):
        symbol = str(raw_symbol).strip().upper().replace("NSE:", "").replace("-", "_")
        if not symbol or symbol == "NAN":
            continue
        payload[symbol] = 8 if index <= 10 else 6
    return payload


def debug_symbol_row(label: str, symbol: str, df: pd.DataFrame, cols: list[str] | None = None) -> None:
    """Log a single-symbol snapshot from a dataframe for debugging."""
    if not symbol or df is None or df.empty or "symbol" not in df.columns:
        log.info(f"[DEBUG] {label}: {symbol} not present")
        return

    rows = df[df["symbol"].astype(str).str.upper() == symbol.upper()]
    if rows.empty:
        log.info(f"[DEBUG] {label}: {symbol} not present")
        return

    row = rows.iloc[0]
    if cols is None:
        cols = [c for c in ["symbol", "latest_close", "avg_turnover_cr", "median_turnover_cr",
                            "close", "ret_12m", "pct_from_52w_high", "tds_since_52w_high",
                            "total_score", "rating", "source"] if c in rows.columns]
    payload = {c: row.get(c) for c in cols if c in rows.columns}
    log.info(f"[DEBUG] {label}: {payload}")


def debug_gate_result(symbol: str, stage: str, passed: bool, details: str) -> None:
    status = "PASS" if passed else "FAIL"
    log.info(f"[DEBUG] {symbol} {stage}: {status} | {details}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(description="Liquid Momentum Screener v2.0")
    parser.add_argument("as_of_date", help="As-of date (YYYY-MM-DD)")
    parser.add_argument("reset_date", help="Reset/entry date (YYYY-MM-DD)")
    parser.add_argument("mode", nargs="?", choices=["full"], help="Use 'full' to also run stock_rating.py after this scanner.")
    parser.add_argument("--source", choices=["kite", "bhav"], default="kite")
    parser.add_argument("--extended", action="store_true", help="Enable extended universe scan")
    parser.add_argument("--run-fundamentals", action="store_true")
    parser.add_argument("--debug", metavar="SYMBOL", help="Show detailed diagnostics for one symbol")
    return parser.parse_args()


def main():
    args = parse_args()
    as_of = datetime.strptime(args.as_of_date, "%Y-%m-%d")
    reset = datetime.strptime(args.reset_date, "%Y-%m-%d")
    run_log_path = setup_run_logging(as_of, reset)
    ohlcv_from = min(as_of - timedelta(days=375), reset - timedelta(days=5))
    debug_symbol = args.debug.strip().upper() if args.debug else None

    pipeline_summary = {}
    failed_symbols = []

    log.info(f"Persistent log file: {run_log_path}")
    log.info(f"Stock OHLC source: {args.source}")
    if debug_symbol:
        log.info(f"[DEBUG] Trace mode enabled for symbol: {debug_symbol}")

    # ------------------------------------------------------------------
    # 1. Database + Kite setup
    # ------------------------------------------------------------------
    conn = get_db_connection()
    kite = get_kite_client() if args.source == "kite" else None

    # ------------------------------------------------------------------
    # 2. Fetch bhav turnover data for all symbols
    # ------------------------------------------------------------------
    log.info("Fetching bhav turnover data...")
    turnover_df = fetch_bhav_turnover(conn, as_of)
    if turnover_df.empty or "symbol" not in turnover_df.columns:
        log.error("No turnover data available from bhav DB. Check bhav tables and column names.")
        sys.exit(1)
    if debug_symbol:
        debug_symbol_row("Raw turnover universe", debug_symbol, turnover_df)
    pipeline_summary["total_eq"] = len(turnover_df)

    # ------------------------------------------------------------------
    # 3. Validate against Kite NSE EQ instruments
    # ------------------------------------------------------------------
    if kite:
        log.info("Fetching Kite NSE EQ instruments...")
        valid_eq = get_nse_eq_symbols(kite)
        turnover_df = turnover_df[turnover_df["symbol"].isin(valid_eq)]
        token_map = build_nse_token_map(kite, turnover_df["symbol"].tolist())
        if debug_symbol:
            debug_gate_result(
                debug_symbol,
                "EQ validation",
                debug_symbol in valid_eq,
                f"valid_eq={debug_symbol in valid_eq}, token_mapped={debug_symbol in token_map}",
            )
    else:
        token_map = {}

    # ------------------------------------------------------------------
    # 4. Filter and build Top 150
    # ------------------------------------------------------------------
    turnover_df = turnover_df[turnover_df["latest_close"] >= MIN_CLOSE].copy()
    if debug_symbol:
        debug_symbol_row("After close filter", debug_symbol, turnover_df)
        raw_match = "symbol" in turnover_df.columns and any(turnover_df["symbol"].astype(str).str.upper() == debug_symbol)
        debug_gate_result(debug_symbol, "Close filter", raw_match, f"min_close={MIN_CLOSE}")
    pipeline_summary["after_close_filter"] = len(turnover_df)
    turnover_df = turnover_df[
        (turnover_df["avg_turnover_cr"] >= MIN_EXTENDED_TURNOVER_CR) &
        (turnover_df["median_turnover_cr"] >= MIN_EXTENDED_TURNOVER_CR)
    ].copy()
    if debug_symbol:
        debug_symbol_row("After turnover filter", debug_symbol, turnover_df)
        turn_match = "symbol" in turnover_df.columns and any(turnover_df["symbol"].astype(str).str.upper() == debug_symbol)
        debug_gate_result(
            debug_symbol,
            "Liquidity filter",
            turn_match,
            f"avg_turnover>={MIN_EXTENDED_TURNOVER_CR}Cr and median_turnover>={MIN_EXTENDED_TURNOVER_CR}Cr",
        )
    pipeline_summary["after_turnover_filter"] = len(turnover_df)

    topn_df = turnover_df.nlargest(TOP_N, "avg_turnover_cr").reset_index(drop=True)
    topn_symbols = topn_df["symbol"].tolist()
    if debug_symbol:
        rank_df = turnover_df.sort_values("avg_turnover_cr", ascending=False).reset_index(drop=True)
        matches = rank_df[rank_df["symbol"].astype(str).str.upper() == debug_symbol]
        if not matches.empty:
            rank = matches.index[0] + 1
            log.info(f"[DEBUG] Turnover rank: {debug_symbol} rank_by_avg_to={rank} top{TOP_N}={rank <= TOP_N}")
            debug_gate_result(debug_symbol, f"Top {TOP_N} turnover shortlist", rank <= TOP_N, f"rank_by_avg_to={rank}")
        else:
            log.info(f"[DEBUG] Turnover rank: {debug_symbol} not present after liquidity filters")
            debug_gate_result(debug_symbol, f"Top {TOP_N} turnover shortlist", False, "not present after liquidity filters")
    pipeline_summary["topn_count"] = len(topn_symbols)
    log.info(f"Top {TOP_N} symbols selected. Fetching OHLCV...")

    # ------------------------------------------------------------------
    # 5. Fetch benchmark index OHLCV (DS-8)
    # ------------------------------------------------------------------
    log.info("Fetching Nifty Smallcap 250 index data from indexbhav...")
    index_df = fetch_index_ohlcv_indexbhav(conn, NIFTY_SMALLCAP_250_SYMBOL, ohlcv_from, as_of)
    if index_df.empty:
        log.warning("Nifty Smallcap 250 unavailable in indexbhav, falling back to Nifty 50.")
        index_df = fetch_index_ohlcv_indexbhav(conn, NIFTY_50_SYMBOL, ohlcv_from, as_of)
    if index_df.empty:
        log.warning("Benchmark index data unavailable in indexbhav; RS metrics will be NaN.")

    # ------------------------------------------------------------------
    # 6. Fetch OHLCV for Top 150
    # ------------------------------------------------------------------
    ohlcv_map = {}
    for i, sym in enumerate(topn_symbols):
        if i > 0 and i % 50 == 0:
            log.info(f"  Fetched {i}/{len(topn_symbols)} symbols...")
        ohlcv_map[sym] = fetch_ohlcv(
            kite, conn, sym, ohlcv_from, as_of,
            args.source, failed_symbols, token_map=token_map
        )

    # ------------------------------------------------------------------
    # 7. Compute metrics and score Top 150
    # ------------------------------------------------------------------
    log.info(f"Computing metrics for Top {TOP_N}...")
    metrics_df = compute_all_metrics(topn_symbols, ohlcv_map, index_df, as_of, reset)
    metrics_df = metrics_df.merge(
        topn_df[["symbol", "avg_turnover_cr", "median_turnover_cr"]],
        on="symbol", how="left"
    )
    metrics_df = metrics_df[metrics_df["atr_pct"].notna()].copy()
    metrics_df = metrics_df[metrics_df["atr_pct"] >= TOP30_MIN_ATR_PCT].copy()
    pipeline_summary["after_atr_filter"] = len(metrics_df)

    filtered_df = apply_filters(metrics_df)
    if debug_symbol:
        debug_symbol_row(
            f"Top{TOP_N} metrics", debug_symbol, metrics_df,
            cols=["symbol", "avg_turnover_cr", "median_turnover_cr", "ret_12m",
                  "pct_from_52w_high", "tds_since_52w_high", "atr_pct", "ret_reset", "ret_1d"]
        )
        topn_debug = metrics_df[metrics_df["symbol"].astype(str).str.upper() == debug_symbol]
        if not topn_debug.empty:
            row = topn_debug.iloc[0]
            ret12_valid = pd.notna(row.get("ret_12m"))
            gate_base_df = metrics_df[
                (metrics_df["pct_from_52w_high"] <= MAX_RECENCY_GATE_DIST_PCT) &
                (
                    (metrics_df["pct_from_52w_high"] <= MAX_DIST_FROM_52W_HIGH_PCT) |
                    (metrics_df["tds_since_52w_high"] <= MAX_TDS_SINCE_52W_HIGH)
                )
            ].copy()
            threshold_70 = (
                gate_base_df["ret_12m"].dropna().quantile(RETURN_GATE_PERCENTILE / 100)
                if gate_base_df["ret_12m"].notna().any()
                else np.nan
            )
            pass_52w = (
                pd.notna(row.get("pct_from_52w_high")) and pd.notna(row.get("tds_since_52w_high")) and
                (
                    row["pct_from_52w_high"] <= MAX_RECENCY_GATE_DIST_PCT and
                    (
                        row["pct_from_52w_high"] <= MAX_DIST_FROM_52W_HIGH_PCT or
                        row["tds_since_52w_high"] <= MAX_TDS_SINCE_52W_HIGH
                    )
                )
            )
            pass_12m = ret12_valid and (row["ret_12m"] >= threshold_70 if pd.notna(threshold_70) else False)
            log.info(
                f"[DEBUG] Top{TOP_N} gates: pass_52w={pass_52w} "
                f"(pct_from_high={row.get('pct_from_52w_high')}, tds_since_high={row.get('tds_since_52w_high')}) "
                f"pass_12m={pass_12m} (ret_12m={row.get('ret_12m')}, "
                f"post_52w_threshold={threshold_70}, gate_base_count={len(gate_base_df)})"
            )
            debug_gate_result(
                debug_symbol,
                "52W gate",
                pass_52w,
                f"pct_from_high={row.get('pct_from_52w_high')}, tds_since_high={row.get('tds_since_52w_high')}, "
                f"rule=((<= {MAX_DIST_FROM_52W_HIGH_PCT}% OR <= {MAX_TDS_SINCE_52W_HIGH} trading days) "
                f"AND <= {MAX_RECENCY_GATE_DIST_PCT}% from high)",
            )
            debug_gate_result(
                debug_symbol,
                "12M gate",
                pass_12m,
                f"ret_12m={row.get('ret_12m')}, post_52w_threshold={threshold_70}, gate_base_count={len(gate_base_df)}",
            )
    pipeline_summary["after_52w_gate"] = len(metrics_df[
        (metrics_df["pct_from_52w_high"] <= MAX_RECENCY_GATE_DIST_PCT) &
        (
            (metrics_df["pct_from_52w_high"] <= MAX_DIST_FROM_52W_HIGH_PCT) |
            (metrics_df["tds_since_52w_high"] <= MAX_TDS_SINCE_52W_HIGH)
        )
    ])
    pipeline_summary["after_12m_gate"] = len(filtered_df)

    scored_topn = score_universe(filtered_df)
    scored_topn["source"] = f"Top{TOP_N}"
    scored_topn["rank"] = range(1, len(scored_topn) + 1)
    if debug_symbol:
        debug_symbol_row(
            f"Top{TOP_N} scored", debug_symbol, scored_topn,
            cols=["symbol", "total_score", "rating", "score_52w", "score_atr", "score_volume",
                  "score_reset", "score_1d", "score_rs", "score_turnover", "rank"]
        )
        top_scored = "symbol" in scored_topn.columns and any(scored_topn["symbol"].astype(str).str.upper() == debug_symbol)
        debug_gate_result(debug_symbol, "Scored main universe", top_scored, "present in final scored shortlist after all gates")
    pipeline_summary["scored_topn"] = len(scored_topn)

    # ------------------------------------------------------------------
    # 8. Sector data
    # ------------------------------------------------------------------
    sectors_df = fetch_sectors(conn)

    # ------------------------------------------------------------------
    # 9. Sector analysis — Top 5 sectors
    # ------------------------------------------------------------------
    log.info("Computing top 5 sectors...")
    top5_sectors = compute_sector_summary(scored_topn, sectors_df)

    # ------------------------------------------------------------------
    # 10. Extended universe (if --extended)
    # ------------------------------------------------------------------
    scored_extended = pd.DataFrame()
    extended_scored_df = pd.DataFrame()
    pipeline_summary["extended_candidates"] = 0
    pipeline_summary["scored_extended"] = 0

    if args.extended:
        log.info("Building extended universe...")
        topn_set = set(topn_symbols)
        ext_candidates = turnover_df[
            (~turnover_df["symbol"].isin(topn_set)) &
            (turnover_df["avg_turnover_cr"] >= MIN_EXTENDED_TURNOVER_CR) &
            (turnover_df["median_turnover_cr"] >= MIN_EXTENDED_TURNOVER_CR)
        ]["symbol"].tolist()
        if debug_symbol:
            log.info(f"[DEBUG] Extended universe: {debug_symbol} in_extended_candidates={debug_symbol in ext_candidates}")
            debug_gate_result(debug_symbol, "Extended universe candidate", debug_symbol in ext_candidates, "outside main shortlist but passed liquidity filters")
        pipeline_summary["extended_candidates"] = len(ext_candidates)
        log.info(f"Extended universe: {len(ext_candidates)} candidates. Fetching OHLCV...")

        ext_ohlcv_map = {}
        for i, sym in enumerate(ext_candidates):
            if i > 0 and i % 50 == 0:
                log.info(f"  Extended fetch {i}/{len(ext_candidates)}...")
            ext_ohlcv_map[sym] = fetch_ohlcv(
                kite, conn, sym, ohlcv_from, as_of,
                args.source, failed_symbols, token_map=token_map
            )

        ext_turnover = turnover_df[
            turnover_df["symbol"].isin(ext_candidates)
        ][["symbol", "avg_turnover_cr", "median_turnover_cr"]]
        ext_metrics = compute_all_metrics(ext_candidates, ext_ohlcv_map, index_df, as_of, reset)
        ext_metrics = ext_metrics.merge(ext_turnover, on="symbol", how="left")
        ext_metrics = ext_metrics[ext_metrics["atr_pct"].notna()].copy()
        ext_metrics = ext_metrics[ext_metrics["atr_pct"] >= TOP30_MIN_ATR_PCT].copy()
        ext_filtered = apply_filters(ext_metrics)
        if debug_symbol:
            debug_symbol_row(
                "Extended metrics", debug_symbol, ext_metrics,
                cols=["symbol", "avg_turnover_cr", "median_turnover_cr", "ret_12m",
                      "pct_from_52w_high", "tds_since_52w_high", "atr_pct", "ret_reset", "ret_1d"]
            )
            ext_debug = ext_metrics[ext_metrics["symbol"].astype(str).str.upper() == debug_symbol]
            if not ext_debug.empty:
                row = ext_debug.iloc[0]
                ext_gate_base_df = ext_metrics[
                    (ext_metrics["pct_from_52w_high"] <= MAX_RECENCY_GATE_DIST_PCT) &
                    (
                        (ext_metrics["pct_from_52w_high"] <= MAX_DIST_FROM_52W_HIGH_PCT) |
                        (ext_metrics["tds_since_52w_high"] <= MAX_TDS_SINCE_52W_HIGH)
                    )
                ].copy()
                threshold_70 = ext_gate_base_df["ret_12m"].dropna().quantile(RETURN_GATE_PERCENTILE / 100) if ext_gate_base_df["ret_12m"].notna().any() else np.nan
                pass_52w = (
                    pd.notna(row.get("pct_from_52w_high")) and pd.notna(row.get("tds_since_52w_high")) and
                    (
                        row["pct_from_52w_high"] <= MAX_RECENCY_GATE_DIST_PCT and
                        (
                            row["pct_from_52w_high"] <= MAX_DIST_FROM_52W_HIGH_PCT or
                            row["tds_since_52w_high"] <= MAX_TDS_SINCE_52W_HIGH
                        )
                    )
                )
                pass_12m = pd.notna(row.get("ret_12m")) and (row["ret_12m"] >= threshold_70 if pd.notna(threshold_70) else False)
                log.info(
                    f"[DEBUG] Extended gates: pass_52w={pass_52w} "
                    f"(pct_from_high={row.get('pct_from_52w_high')}, tds_since_high={row.get('tds_since_52w_high')}) "
                    f"pass_12m={pass_12m} (ret_12m={row.get('ret_12m')}, post_52w_threshold={threshold_70})"
                )
                debug_gate_result(
                    debug_symbol,
                    "Extended 52W gate",
                    pass_52w,
                    f"pct_from_high={row.get('pct_from_52w_high')}, tds_since_high={row.get('tds_since_52w_high')}, "
                    f"rule=((<= {MAX_DIST_FROM_52W_HIGH_PCT}% OR <= {MAX_TDS_SINCE_52W_HIGH} trading days) "
                    f"AND <= {MAX_RECENCY_GATE_DIST_PCT}% from high)",
                )
                debug_gate_result(
                    debug_symbol,
                    "Extended 12M gate",
                    pass_12m,
                    f"ret_12m={row.get('ret_12m')}, post_52w_threshold={threshold_70}",
                )
        scored_extended = score_universe(ext_filtered)
        scored_extended["source"] = "Extended"
        scored_extended["rank"] = range(1, len(scored_extended) + 1)
        if debug_symbol:
            debug_symbol_row(
                "Extended scored", debug_symbol, scored_extended,
                cols=["symbol", "total_score", "rating", "score_52w", "score_atr", "score_volume",
                      "score_reset", "score_1d", "score_rs", "score_turnover", "rank"]
            )
            ext_scored = "symbol" in scored_extended.columns and any(scored_extended["symbol"].astype(str).str.upper() == debug_symbol)
            debug_gate_result(debug_symbol, "Scored extended universe", ext_scored, "present in extended scored list")
        pipeline_summary["scored_extended"] = len(scored_extended)
        extended_scored_df = scored_extended.copy()

    # ------------------------------------------------------------------
    # 11. Merge and select Top 20
    # ------------------------------------------------------------------
    scored_frames = [df for df in [scored_topn, scored_extended] if not df.empty]
    if scored_frames:
        all_scored = pd.concat(scored_frames, ignore_index=True)
    else:
        log.warning("No stocks qualified after scoring. Writing an empty report.")
        all_scored = pd.DataFrame(columns=list(scored_topn.columns))
        top5_sectors = pd.DataFrame()
    # Add sector info
    if not sectors_df.empty:
        all_scored = all_scored.merge(sectors_df, on="symbol", how="left")
        all_scored["sector"] = all_scored["sector"].fillna("Unknown")

    if not all_scored.empty:
        all_scored = all_scored.sort_values(
            ["total_score", "ret_12m"], ascending=[False, False]
        ).reset_index(drop=True)
        all_scored["rank"] = range(1, len(all_scored) + 1)
    if debug_symbol:
        final_present = "symbol" in all_scored.columns and any(all_scored["symbol"].astype(str).str.upper() == debug_symbol)
        if final_present:
            row = all_scored[all_scored["symbol"].astype(str).str.upper() == debug_symbol].iloc[0]
            debug_gate_result(debug_symbol, "Final merged output", True, f"rank={row.get('rank')}, total_score={row.get('total_score')}, source={row.get('source')}")
        else:
            debug_gate_result(debug_symbol, "Final merged output", False, "not present after gating/scoring")
    pipeline_summary["total_scored"] = len(all_scored)

    top30_initial = all_scored.head(TOP_N_REPORT).copy()
    excluded_top30_atr = top30_initial.iloc[0:0].copy()
    top30 = top30_initial.copy()
    pipeline_summary["top30_initial_count"] = len(top30_initial)
    pipeline_summary["top30_atr_excluded_count"] = len(excluded_top30_atr)
    pipeline_summary["top30_count"] = len(top30)

    # ------------------------------------------------------------------
    # 12. Write Excel report
    # ------------------------------------------------------------------
    log.info("Writing Excel report...")
    excel_path = write_excel_report(
        top30, excluded_top30_atr, top5_sectors, all_scored, pipeline_summary,
        extended_scored_df, as_of, reset, failed_symbols,
    )

    # ------------------------------------------------------------------
    # 13. Write TradingView watchlist
    # ------------------------------------------------------------------
    tv_path = REPORTS_DIR / f"final_wl_{as_of.strftime('%d%b%Y')}.txt"
    write_tradingview_watchlist(top30, sectors_df, tv_path, as_of)
    liquid_leader_bonus_payload = build_liquid_leader_bonus_payload(top30)
    updated_gmlist_path = write_updated_gmlist(top30, as_of)

    # ------------------------------------------------------------------
    # 14. Optionally run fundamentals
    # ------------------------------------------------------------------
    log.info(f"  Report: {excel_path}")
    log.info(f"  Watchlist: {tv_path}")
    if liquid_leader_bonus_payload:
        preview = ", ".join(
            f"{symbol}:{bonus}" for symbol, bonus in list(liquid_leader_bonus_payload.items())[:5]
        )
        log.info(f"  Liquid leader bonus map: {preview}{' ...' if len(liquid_leader_bonus_payload) > 5 else ''}")
    log.info(f"  Updated gmlist: {updated_gmlist_path}")
    if failed_symbols:
        log.warning(f"  {len(failed_symbols)} symbols failed API calls: {failed_symbols[:10]}{'...' if len(failed_symbols) > 10 else ''}")

    if args.run_fundamentals:
        log.info("Running screener_fundamentals.py...")
        os.system(
            f'python screener_fundamentals.py --date {as_of.strftime("%Y-%m-%d")} '
            f'--reports-dir "{REPORTS_DIR.resolve()}"'
        )

    if args.mode == "full":
        stock_rating_path = Path(__file__).resolve().parent / "stock_rating.py"
        cmd = [
            sys.executable,
            str(stock_rating_path),
            as_of.strftime("%Y-%m-%d"),
            reset.strftime("%Y-%m-%d"),
            "--liquid-leader-map",
            json.dumps(liquid_leader_bonus_payload, separators=(",", ":")),
        ]
        log.info("Running stock_rating.py after neo reports and updated gmlist with inline liquid leader bonus map...")
        result = subprocess.run(cmd, cwd=str(Path(__file__).resolve().parent))
        if result.returncode != 0:
            log.error(f"stock_rating.py failed with exit code {result.returncode}")
            sys.exit(result.returncode)

    log.info("Done.")


if __name__ == "__main__":
    main()
