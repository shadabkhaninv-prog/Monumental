"""
ATR 21-Day % Report for all NSE stocks
---------------------------------------
- Reads symbol list from bhav2026 (latest date) via local MySQL bhav database
- Fetches daily OHLC via Kite Connect API
- Calculates 21-day ATR and ATR% for each symbol
- Outputs ranked Excel report

Usage:
    python atr_report.py
    python atr_report.py --date 2026-04-10          # specific date for symbol list
    python atr_report.py --out C:\reports            # custom output folder
    python atr_report.py --token kite_token.txt      # custom token file path
"""

import argparse
import os
import sys
import traceback
from datetime import date, datetime, timedelta
from pathlib import Path

import mysql.connector
import pandas as pd
import numpy as np
from kiteconnect import KiteConnect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DB_HOST     = "localhost"
DB_PORT     = 3306
DB_USER     = "root"
DB_PASSWORD = "root"
DB_NAME     = "bhav"

SCRIPT_DIR  = Path(__file__).parent
TOKEN_FILE  = SCRIPT_DIR / "kite_token.txt"
OUT_DIR     = SCRIPT_DIR / "output"

ATR_PERIOD      = 21          # days for ATR calculation
DATA_DAYS       = ATR_PERIOD + 10   # extra buffer for warm-up

# ---------------------------------------------------------------------------
# Token helpers
# ---------------------------------------------------------------------------
def read_token(filepath):
    creds = {}
    with open(filepath, encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                creds[k.strip()] = v.strip()
    if "API_KEY" not in creds or "ACCESS_TOKEN" not in creds:
        sys.exit(f"ERROR: kite_token.txt must contain API_KEY and ACCESS_TOKEN")
    gen = creds.get("GENERATED", "")
    try:
        gen_dt = datetime.strptime(gen, "%Y-%m-%d %H:%M:%S")
        if gen_dt.date() < date.today():
            print(f"WARNING: Token was generated on {gen_dt.date()}, may be stale. Re-generate if errors occur.")
    except Exception:
        pass
    return creds

def get_kite(token_file):
    creds = read_token(token_file)
    kite = KiteConnect(api_key=creds["API_KEY"])
    kite.set_access_token(creds["ACCESS_TOKEN"])
    return kite

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------
def get_symbols(as_of_date):
    """Return list of distinct symbols from bhav2026 for the given date.
    If no data on that date, use the most recent available date."""
    conn = mysql.connector.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=DB_PASSWORD,
        database=DB_NAME
    )
    cur = conn.cursor()

    # Check if data exists for as_of_date
    cur.execute("SELECT COUNT(*) FROM bhav2026 WHERE MKTDATE = %s", (as_of_date,))
    count = cur.fetchone()[0]

    if count == 0:
        cur.execute("SELECT MAX(MKTDATE) FROM bhav2026 WHERE MKTDATE <= %s", (as_of_date,))
        row = cur.fetchone()
        if not row or not row[0]:
            conn.close()
            sys.exit(f"ERROR: No data in bhav2026 on or before {as_of_date}")
        actual_date = row[0]
        if hasattr(actual_date, 'date'):
            actual_date = actual_date.date()
        print(f"No data on {as_of_date}, using latest available: {actual_date}")
    else:
        actual_date = as_of_date

    cur.execute(
        "SELECT DISTINCT SYMBOL FROM bhav2026 WHERE MKTDATE = %s ORDER BY SYMBOL",
        (actual_date,)
    )
    symbols = [r[0] for r in cur.fetchall()]
    conn.close()
    print(f"Loaded {len(symbols)} symbols from bhav2026 for {actual_date}")
    return symbols, actual_date

# ---------------------------------------------------------------------------
# Instrument token map
# ---------------------------------------------------------------------------
def build_token_map(kite, symbols):
    print("Fetching NSE instrument list from Kite...")
    instruments = pd.DataFrame(kite.instruments("NSE"))
    instruments = instruments[instruments["segment"] == "NSE"]
    token_map = {}
    sym_set = set(symbols)
    for _, row in instruments.iterrows():
        if row["tradingsymbol"] in sym_set:
            token_map[row["tradingsymbol"]] = int(row["instrument_token"])
    found = len(token_map)
    missing = len(sym_set) - found
    print(f"Mapped {found} symbols to instrument tokens  ({missing} not found in NSE instruments)")
    return token_map

# ---------------------------------------------------------------------------
# OHLC fetch
# ---------------------------------------------------------------------------
def fetch_ohlc(kite, token, symbol, from_date, to_date):
    try:
        data = kite.historical_data(
            instrument_token=token,
            from_date=datetime.combine(from_date, datetime.min.time()),
            to_date=datetime.combine(to_date, datetime.min.time()),
            interval="day",
            continuous=False,
            oi=False
        )
        if not data:
            return None
        df = pd.DataFrame(data)
        df["date"] = pd.to_datetime(df["date"]).dt.date
        df = df.set_index("date").sort_index()
        return df
    except Exception as e:
        print(f"  WARN: {symbol} fetch failed — {e}")
        return None

# ---------------------------------------------------------------------------
# ATR calculation
# ---------------------------------------------------------------------------
def calc_atr(df, period=21):
    """Returns ATR and ATR% for the last row of df."""
    if len(df) < period + 1:
        return None, None

    high  = df["high"]
    low   = df["low"]
    close = df["close"]
    prev_close = close.shift(1)

    tr = pd.concat([
        high - low,
        (high - prev_close).abs(),
        (low  - prev_close).abs()
    ], axis=1).max(axis=1)

    # Drop the first row (NaN prev_close)
    tr = tr.iloc[1:]

    if len(tr) < period:
        return None, None

    atr = tr.iloc[-period:].mean()
    latest_close = close.iloc[-1]

    if latest_close <= 0:
        return None, None

    atr_pct = (atr / latest_close) * 100
    return round(atr, 4), round(atr_pct, 4)

# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------
NAVY       = "1F4E79"
BLUE       = "2E75B6"
WHITE      = "FFFFFF"
ALT_ROW    = "F2F7FB"
GREEN_BG   = "E2EFDA"
RED_BG     = "FCE4EC"
YELLOW_BG  = "FFF9C4"

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_report(results, as_of_date, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = out_dir / f"ATR21_Report_{as_of_date}.xlsx"

    df = pd.DataFrame(results, columns=["Symbol", "Close", "ATR_21d", "ATR_Pct", "Status"])
    df = df[df["ATR_Pct"].notna()].copy()
    df = df.sort_values("ATR_Pct", ascending=False).reset_index(drop=True)

    # Percentile rank (0-100, higher = more volatile)
    df["Percentile"] = df["ATR_Pct"].rank(pct=True).mul(100).round(1)
    df["Rank"] = range(1, len(df) + 1)

    # Bottom 30th percentile flag
    p30_threshold = df["ATR_Pct"].quantile(0.30)
    df["Low_Vol_Flag"] = df["ATR_Pct"] <= p30_threshold

    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Full ranked list
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "ATR 21d Report"
    ws.sheet_properties.tabColor = NAVY

    hdr_font  = Font(name="Arial", size=10, bold=True, color=WHITE)
    hdr_fill  = PatternFill("solid", fgColor=NAVY)
    body_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    wrap_al   = Alignment(vertical="center", wrap_text=False)
    center_al = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"21-Day ATR% Report — NSE Stocks — As of {as_of_date}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Total stocks: {len(df)}   |   Bottom 30th percentile ATR% threshold: {p30_threshold:.2f}%   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Rank", "Symbol", "Close", "ATR 21d (₹)", "ATR% 21d", "Percentile", "Low Vol Flag", "Status"]
    col_widths = [7, 16, 12, 14, 12, 12, 14, 12]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_al
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[4].height = 20

    for ri, row in df.iterrows():
        excel_row = ri + 5
        alt = ri % 2 == 1
        fill = PatternFill("solid", fgColor=ALT_ROW) if alt else None

        vals = [
            row["Rank"],
            row["Symbol"],
            row["Close"],
            row["ATR_21d"],
            row["ATR_Pct"],
            row["Percentile"],
            "YES" if row["Low_Vol_Flag"] else "",
            row["Status"],
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=excel_row, column=ci, value=val)
            cell.font = body_font
            cell.border = _border()
            cell.alignment = center_al if ci in (1, 7, 8) else Alignment(horizontal="right", vertical="center")
            if fill:
                cell.fill = fill

        # Highlight low-vol stocks
        if row["Low_Vol_Flag"]:
            ws.cell(row=excel_row, column=2).fill = PatternFill("solid", fgColor="FDEBD0")
            ws.cell(row=excel_row, column=5).fill = PatternFill("solid", fgColor=RED_BG)
            ws.cell(row=excel_row, column=7).font = Font(name="Arial", size=10, bold=True, color="C0392B")

    # Number formatting
    for row_cells in ws.iter_rows(min_row=5, max_row=4+len(df), min_col=3, max_col=6):
        for cell in row_cells:
            cell.number_format = "#,##0.00"

    # Freeze header
    ws.freeze_panes = "A5"

    # Color scale on ATR% column (col 5)
    last_data_row = 4 + len(df)
    atr_range = f"E5:E{last_data_row}"
    ws.conditional_formatting.add(atr_range, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"
    ))

    # -----------------------------------------------------------------------
    # Sheet 2: Bottom 30th percentile (low volatility / penalised stocks)
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Low Volatility Stocks")
    ws2.sheet_properties.tabColor = "C0392B"

    low_vol_df = df[df["Low_Vol_Flag"]].copy()

    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"Low Volatility Stocks — Bottom 30th Percentile ATR%  (ATR% ≤ {p30_threshold:.2f}%)  —  {len(low_vol_df)} stocks"
    ws2["A1"].font = Font(name="Arial", size=12, bold=True, color="C0392B")
    ws2["A1"].alignment = Alignment(horizontal="center")

    h2 = ["Rank", "Symbol", "Close", "ATR 21d (₹)", "ATR% 21d", "Percentile", "Rating Impact"]
    cw2 = [7, 16, 12, 14, 12, 12, 15]
    for ci, (h, w) in enumerate(zip(h2, cw2), 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="C0392B")
        cell.alignment = center_al
        cell.border = _border()
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for ri, (_, row) in enumerate(low_vol_df.iterrows()):
        er = ri + 4
        alt = ri % 2 == 1
        fill = PatternFill("solid", fgColor="FDEDEC") if alt else PatternFill("solid", fgColor="FEF9E7")
        vals = [row["Rank"], row["Symbol"], row["Close"], row["ATR_21d"], row["ATR_Pct"], row["Percentile"], "-6 pts"]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(row=er, column=ci, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.border = _border()
            cell.alignment = center_al if ci in (1, 2, 7) else Alignment(horizontal="right", vertical="center")
        ws2.cell(row=er, column=7).font = Font(name="Arial", size=10, bold=True, color="C0392B")

    for row_cells in ws2.iter_rows(min_row=4, max_row=3+len(low_vol_df), min_col=3, max_col=6):
        for cell in row_cells:
            cell.number_format = "#,##0.00"

    ws2.freeze_panes = "A4"

    # -----------------------------------------------------------------------
    # Sheet 3: Summary stats
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_properties.tabColor = "27AE60"

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 18

    ws3.merge_cells("A1:B1")
    ws3["A1"] = "ATR 21d% Report — Summary Statistics"
    ws3["A1"].font = Font(name="Arial", size=13, bold=True, color=NAVY)
    ws3["A1"].alignment = Alignment(horizontal="center")

    stats = [
        ("As-of Date", str(as_of_date)),
        ("Total Stocks Rated", len(df)),
        ("Stocks Skipped (no data)", len([r for r in results if r[2] is None])),
        ("", ""),
        ("ATR% — Mean", f"{df['ATR_Pct'].mean():.2f}%"),
        ("ATR% — Median", f"{df['ATR_Pct'].median():.2f}%"),
        ("ATR% — Std Dev", f"{df['ATR_Pct'].std():.2f}%"),
        ("ATR% — Min", f"{df['ATR_Pct'].min():.2f}%"),
        ("ATR% — Max", f"{df['ATR_Pct'].max():.2f}%"),
        ("", ""),
        ("10th Percentile threshold", f"{df['ATR_Pct'].quantile(0.10):.2f}%"),
        ("20th Percentile threshold", f"{df['ATR_Pct'].quantile(0.20):.2f}%"),
        ("30th Percentile threshold (penalty)", f"{p30_threshold:.2f}%"),
        ("50th Percentile (median)", f"{df['ATR_Pct'].quantile(0.50):.2f}%"),
        ("70th Percentile threshold", f"{df['ATR_Pct'].quantile(0.70):.2f}%"),
        ("90th Percentile threshold", f"{df['ATR_Pct'].quantile(0.90):.2f}%"),
        ("", ""),
        ("Low volatility stocks (bottom 30%)", len(low_vol_df)),
        ("% of universe penalised (-6 pts)", f"{len(low_vol_df)/len(df)*100:.1f}%"),
    ]

    sf = Font(name="Arial", size=10, bold=True, color=NAVY)
    bf = Font(name="Arial", size=10)
    for ri, (label, val) in enumerate(stats, 3):
        ws3.cell(row=ri, column=1, value=label).font = sf if label else bf
        ws3.cell(row=ri, column=2, value=val).font = bf
        if label:
            ws3.cell(row=ri, column=1).border = _border()
            ws3.cell(row=ri, column=2).border = _border()
            if ri % 2 == 0:
                for c in (1, 2):
                    ws3.cell(row=ri, column=c).fill = PatternFill("solid", fgColor=ALT_ROW)

    wb.save(fname)
    return fname

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Generate 21-day ATR% report for NSE stocks")
    parser.add_argument("--date",  default=str(date.today() - timedelta(days=1)),
                        help="As-of date for symbol list (YYYY-MM-DD). Default: yesterday")
    parser.add_argument("--out",   default=str(OUT_DIR),
                        help="Output directory. Default: ./output")
    parser.add_argument("--token", default=str(TOKEN_FILE),
                        help="Path to kite_token.txt")
    args = parser.parse_args()

    as_of_date = datetime.strptime(args.date, "%Y-%m-%d").date()
    print(f"\n{'='*60}")
    print(f"  ATR 21d% Report  |  As-of: {as_of_date}")
    print(f"{'='*60}\n")

    # 1. Get symbols from DB
    symbols, actual_date = get_symbols(as_of_date)

    # 2. Init Kite
    kite = get_kite(args.token)
    token_map = build_token_map(kite, symbols)

    # 3. Date range for OHLC fetch
    fetch_to   = actual_date
    fetch_from = fetch_to - timedelta(days=DATA_DAYS + 15)   # extra calendar buffer for weekends/holidays

    # 4. Fetch OHLC and compute ATR
    results = []
    total = len(token_map)
    print(f"\nFetching OHLC data and computing 21d ATR% for {total} symbols...\n")

    for i, (symbol, token) in enumerate(token_map.items(), 1):
        if i % 100 == 0 or i == total:
            print(f"  [{i}/{total}] processed...")

        df = fetch_ohlc(kite, token, symbol, fetch_from, fetch_to)
        if df is None or len(df) == 0:
            results.append((symbol, None, None, None, "No data"))
            continue

        # Use only up to as_of_date
        df = df[df.index <= actual_date]
        if len(df) == 0:
            results.append((symbol, None, None, None, "No data"))
            continue

        latest_close = df["close"].iloc[-1]
        atr, atr_pct = calc_atr(df, ATR_PERIOD)

        if atr is None:
            results.append((symbol, latest_close, None, None, "Insufficient history"))
        else:
            results.append((symbol, latest_close, atr, atr_pct, "OK"))

    # 5. Write report
    ok_count   = sum(1 for r in results if r[4] == "OK")
    skip_count = len(results) - ok_count
    print(f"\nComputed ATR for {ok_count} stocks. Skipped {skip_count} (no data / insufficient history).")
    print("Writing Excel report...")

    fname = write_report(results, actual_date, args.out)
    print(f"\nReport saved to: {fname}\n")

if __name__ == "__main__":
    main()
