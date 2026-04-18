#!/usr/bin/env python3
"""
Institutional Picks — Fire Status Report Generator
====================================================
Fetches daily close-to-close % changes via Kite API for the institutional
picks list, then classifies each stock as Fired / Extended / Yet-to-Fire
and produces a colour-coded multi-sheet Excel workbook.

Usage
-----
    python ip_fire_report.py <cutoff_date> <end_date> [options]

    cutoff_date  : DD-MM-YYYY  — date the picks list was generated
                                 (base close used; tracking starts next trading day)
    end_date     : DD-MM-YYYY  — last date to include in the report
                                 (use today's date for a live report)

Options
-------
    --reports-dir PATH   Folder containing institutional_picks_*.txt
                         and where output files are written  [default: reports]
    --token-file  PATH   kite_token.txt path                 [default: kite_token.txt]
    --big-day-pct FLOAT  Single-day % threshold for a "big day"  [default: 5.0]
    --fired-total FLOAT  Total return threshold to be "Fired"     [default: 10.0]
    --extended-total FLOAT  Total return threshold for "Extended" [default: 20.0]
    --sleep-seconds FLOAT   Delay between Kite API calls         [default: 0.35]

Output
------
    reports/ip_fire_report_<cutoff_tag>_to_<end_tag>.xlsx
    (5 sheets: Fire Status · Today's Focus · Daily Heatmap · Extended-Avoid · Summary)

Examples
--------
    python ip_fire_report.py 02-04-2026 16-04-2026
    python ip_fire_report.py 02-04-2026 16-04-2026 --reports-dir C:/Users/shada/Monumental/reports
"""

from __future__ import annotations

import argparse
import re
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def parse_date(s: str, label: str) -> date:
    s = s.strip()
    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", s)
    if not m:
        raise SystemExit(f"{label} must be DD-MM-YYYY, e.g. 02-04-2026")
    dd, mm, yyyy = map(int, m.groups())
    return date(yyyy, mm, dd)


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Generate Institutional Picks Fire-Status Excel report via Kite API."
    )
    p.add_argument("cutoff_date", help="Picks generation date  DD-MM-YYYY")
    p.add_argument("end_date",    help="Last date to include   DD-MM-YYYY")
    p.add_argument("--reports-dir",    default="reports",       help="Reports directory  [reports]")
    p.add_argument("--token-file",     default="kite_token.txt", help="Kite token file   [kite_token.txt]")
    p.add_argument("--big-day-pct",    type=float, default=5.0,  help="Big-day threshold %%  [5.0]")
    p.add_argument("--fired-total",    type=float, default=10.0, help="Fired total %%        [10.0]")
    p.add_argument("--extended-total", type=float, default=20.0, help="Extended total %%     [20.0]")
    p.add_argument("--laggard-total",  type=float, default=6.0,  help="Laggard threshold %%  [6.0]")
    p.add_argument("--sleep-seconds",  type=float, default=0.35, help="Inter-call delay s   [0.35]")
    return p


# ─────────────────────────────────────────────────────────────────────────────
# Kite helpers (copied from institutional_picks_daily_pct_change_report.py)
# ─────────────────────────────────────────────────────────────────────────────

def read_kite_token_file(token_file: Path) -> Dict[str, str]:
    if not token_file.exists():
        raise SystemExit(f"Missing token file: {token_file}")
    values: Dict[str, str] = {}
    for raw_line in token_file.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        values[k.strip().upper()] = v.strip()
    if "API_KEY" not in values or "ACCESS_TOKEN" not in values:
        raise SystemExit(f"{token_file} must contain API_KEY and ACCESS_TOKEN")
    return values


def get_kite_client(token_file: Path):
    try:
        from kiteconnect import KiteConnect
    except ImportError:
        raise SystemExit("kiteconnect not installed.  pip install kiteconnect")
    creds = read_kite_token_file(token_file)
    kite = KiteConnect(api_key=creds["API_KEY"])
    kite.set_access_token(creds["ACCESS_TOKEN"])
    return kite


def normalize_symbol(symbol: str) -> str:
    s = symbol.upper().strip().replace("&", "AND")
    return re.sub(r"[^A-Z0-9]", "", s)


def parse_picks_file(txt_path: Path) -> List[Tuple[str, str]]:
    """Return list of (exchange, symbol) from institutional_picks_*.txt"""
    if not txt_path.exists():
        raise SystemExit(f"Picks file not found: {txt_path}")
    pairs: List[Tuple[str, str]] = []
    seen: set[Tuple[str, str]] = set()
    for raw in txt_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if ":" in line:
            exch, sym = line.split(":", 1)
        else:
            exch, sym = "NSE", line
        key = (exch.strip().upper(), sym.strip().upper())
        if key[1] and key not in seen:
            pairs.append(key)
            seen.add(key)
    if not pairs:
        raise SystemExit(f"No symbols found in {txt_path}")
    return pairs


def build_token_lookups(kite, exchange: str):
    instruments = pd.DataFrame(kite.instruments(exchange))
    if instruments.empty:
        raise SystemExit(f"kite.instruments({exchange}) returned nothing")
    if "segment" in instruments.columns:
        instruments = instruments[instruments["segment"] == exchange]
    if "instrument_type" in instruments.columns:
        instruments = instruments[instruments["instrument_type"] == "EQ"]
    instruments["ts_upper"] = instruments["tradingsymbol"].astype(str).str.upper()
    instruments["ts_norm"]  = instruments["ts_upper"].map(normalize_symbol)
    by_upper: Dict[str, int] = {}
    by_norm:  Dict[str, int] = {}
    for _, row in instruments.iterrows():
        token = int(row["instrument_token"])
        by_upper.setdefault(row["ts_upper"], token)
        by_norm.setdefault(row["ts_norm"],  token)
    return by_upper, by_norm


def resolve_token(by_upper, by_norm, raw_symbol: str) -> Tuple[Optional[int], Optional[str]]:
    up = raw_symbol.upper().strip()
    if up in by_upper:
        return by_upper[up], up
    nm = normalize_symbol(up)
    if nm in by_norm:
        return by_norm[nm], nm
    return None, None


def fetch_daily_closes(kite, token: int, from_dt: date, to_dt: date) -> pd.DataFrame:
    rows = kite.historical_data(
        instrument_token=token,
        from_date=datetime.combine(from_dt, datetime.min.time()),
        to_date=datetime.combine(to_dt,   datetime.min.time()),
        interval="day",
        continuous=False,
        oi=False,
    )
    if not rows:
        return pd.DataFrame(columns=["date", "close"])
    df = pd.DataFrame(rows)
    df["date"]  = pd.to_datetime(df["date"]).dt.date
    df["close"] = pd.to_numeric(df["close"], errors="coerce")
    return df[["date", "close"]].dropna().sort_values("date")


# ─────────────────────────────────────────────────────────────────────────────
# Fire-status classification
# ─────────────────────────────────────────────────────────────────────────────

def classify_stock(sym: str, daily: List[float], big_day_pct: float,
                   fired_total: float, extended_total: float,
                   laggard_total: float = 6.0) -> dict:
    """
    Classify a stock based on its daily % changes since the cutoff date.

    LAGGARD criteria (either condition triggers it):
      • total return ≤ -laggard_total  (default -6%): significantly underwater
      • total < 0  AND  fallen ≥ laggard_total% from its intra-period peak
        (e.g. stock reached +5% then slid to -2% — a 7% drawdown from peak)

    Stocks with small negative returns (-0.1% to -5.9%) that haven't had a
    meaningful drawdown fall through to YET TO FIRE or STEADY RUNNER.
    """
    total    = sum(daily)                       # simple sum (not compounded — fast scan)
    big_days = [d for d in daily if d >= big_day_pct]
    n_big    = len(big_days)
    max_day  = max(daily) if daily else 0.0
    recent3  = sum(daily[-3:]) if len(daily) >= 3 else sum(daily)
    today    = daily[-1] if daily else 0.0

    # ── Compute intra-period peak cumulative return ───────────────────────────
    peak_total = 0.0
    running    = 0.0
    for d in daily:
        running += d
        if running > peak_total:
            peak_total = running
    drawdown_from_peak = peak_total - total     # how far stock has fallen from its high

    # ── True laggard: significantly down OR big drawdown from peak ────────────
    is_laggard = (total <= -laggard_total) or (total < 0 and drawdown_from_peak >= laggard_total)

    if total >= extended_total * 1.5 or n_big >= 3:
        status, emoji, color, order = "HIGHLY EXTENDED", "🚀", "FF4444", 1
    elif today >= big_day_pct and total < extended_total * 1.5:
        status, emoji, color, order = "JUST FIRED TODAY", "⚡", "FF8C00", 2
    elif total >= extended_total and n_big >= 2:
        status, emoji, color, order = "EXTENDED", "⚠️", "FFA500", 3
    elif n_big >= 1 and total >= fired_total:
        # Check if recent 3 days are negative → retreating
        if recent3 < 0:
            status, emoji, color, order = "FIRED & RETREATING", "🔄", "9370DB", 6
        else:
            status, emoji, color, order = "FIRED", "🔥", "FFD700", 4
    elif is_laggard:
        status, emoji, color, order = "LAGGARD", "❌", "C0C0C0", 8
    elif n_big == 0 and total < fired_total:
        status, emoji, color, order = "YET TO FIRE", "⏳", "00B050", 5
    else:
        status, emoji, color, order = "STEADY RUNNER", "📈", "70AD47", 7

    return dict(symbol=sym, daily=daily, total=round(total, 2),
                max_day=round(max_day, 2), n_big=n_big,
                recent3=round(recent3, 2), today=round(today, 2),
                peak_total=round(peak_total, 2),
                drawdown_from_peak=round(drawdown_from_peak, 2),
                status=status, emoji=emoji, color=color, order=order)


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", start_color=hex6, end_color=hex6)


def _font(size=10, bold=False, color="000000", name="Arial") -> Font:
    return Font(name=name, size=size, bold=bold, color=color)


def _border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _ctr(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def write_header_row(ws, row: int, cols, texts, bg="2E75B6", height=32):
    for c, txt in zip(cols, texts):
        cell = ws.cell(row=row, column=c, value=txt)
        cell.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill      = _fill(bg)
        cell.alignment = _ctr(wrap=True)
        cell.border    = _border()
    ws.row_dimensions[row].height = height


def pct_num_fmt() -> str:
    return '+0.00%;-0.00%;"0.00%"'


# heatmap colour for a single day's %
_HM_PALETTE = [
    (8.0,  "005C00", "FFFFFF"),   # very dark green, white text
    (5.0,  "00B050", "FFFFFF"),   # green
    (3.0,  "70AD47", "000000"),
    (1.0,  "C6EFCE", "000000"),
    (0.0,  "EBF7EC", "000000"),
    (-2.0, "FFCCCC", "000000"),
    (-99,  "FF4444", "FFFFFF"),
]

def hm_color(val: float) -> Tuple[str, str]:
    for thresh, bg, fg in _HM_PALETTE:
        if val >= thresh:
            return bg, fg
    return "FF4444", "FFFFFF"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def sheet_fire_status(wb: Workbook, stocks: List[dict], end_date: date):
    ws = wb.active
    ws.title = "🎯 Fire Status"
    ws.freeze_panes = "A3"

    ed = end_date.strftime("%d-%b-%Y")
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = f"📊 Institutional Picks — Fire Status Dashboard  |  As of {ed}"
    t.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _ctr()
    ws.row_dimensions[1].height = 28

    hdrs = ["#", "Symbol", "Status",
            "Total\nReturn%", f"Today\n{end_date.strftime('%d-%b')}%",
            "Max Single\nDay%", "# Big Days\n(≥5%)", "Recent 3D\nTrend%",
            "Wk1 Perf%", "Wk2+ Perf%", "Action Signal"]
    write_header_row(ws, 2, range(1, 12), hdrs)

    STATUS_TEXT_COLOR = {
        "FF4444": "FFFFFF", "FF8C00": "FFFFFF", "FFA500": "000000",
        "FFD700": "000000", "00B050": "FFFFFF", "9370DB": "FFFFFF",
        "70AD47": "FFFFFF", "C0C0C0": "000000",
    }
    ROW_BG = {
        "FF4444": "FFE5E5", "FF8C00": "FFF3E0", "FFA500": "FFF8E1",
        "FFD700": "FFFDE7", "00B050": "E8F5E9", "9370DB": "F3E5F5",
        "70AD47": "F1F8E9", "C0C0C0": "F5F5F5",
    }
    ACTIONS = {
        "YET TO FIRE":        "✅ Focus Today — maximum upside potential",
        "HIGHLY EXTENDED":    "🚫 Avoid — likely to pull back",
        "JUST FIRED TODAY":   "⚡ Breakout today — watch for follow-through",
        "FIRED & RETREATING": "🔄 Wait for reset / base formation",
        "EXTENDED":           "⚠️ Partial exit / raise stop",
        "FIRED":              "🔥 Hold / trail stop",
        "LAGGARD":            "❌ Skip / remove from list",
        "STEADY RUNNER":      "📈 Monitor for acceleration",
    }

    sorted_stocks = sorted(stocks, key=lambda x: x["order"])
    for i, st in enumerate(sorted_stocks, 1):
        r = i + 2
        half = len(st["daily"]) // 2
        wk1 = sum(st["daily"][:half])
        wk2 = sum(st["daily"][half:])
        row_bg  = ROW_BG.get(st["color"], "FFFFFF")
        stat_bg = st["color"]
        stat_fc = STATUS_TEXT_COLOR.get(st["color"], "000000")

        vals = [i, st["symbol"], f"{st['emoji']} {st['status']}",
                st["total"]/100, st["today"]/100, st["max_day"]/100,
                st["n_big"],     st["recent3"]/100,
                wk1/100,         wk2/100,
                ACTIONS.get(st["status"], "—")]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border    = _border()
            cell.alignment = _ctr() if c <= 10 else _left()
            cell.fill      = _fill(row_bg)
            cell.font      = _font(bold=(c == 2))
            if c == 3:
                cell.fill = _fill(stat_bg)
                cell.font = _font(bold=True, color=stat_fc)
            if isinstance(v, float):
                cell.number_format = pct_num_fmt()
        ws.row_dimensions[r].height = 20

    for i, w in enumerate([4, 14, 24, 12, 12, 12, 12, 12, 12, 12, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_todays_focus(wb: Workbook, stocks: List[dict], end_date: date):
    ws = wb.create_sheet("⏳ Today's Focus")
    ed = end_date.strftime("%d-%b-%Y")

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"⏳ Today's Focus — Stocks That Have NOT Yet Made Their Big Move  |  {ed}"
    t.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    t.fill      = _fill("00703C")
    t.alignment = _ctr()
    ws.row_dimensions[1].height = 28

    focus = ([s for s in stocks if s["status"] == "YET TO FIRE"]
           + [s for s in stocks if s["status"] == "FIRED & RETREATING"])
    if not focus:
        ws.cell(row=3, column=1,
                value="No 'Yet to Fire' or 'Fired & Retreating' stocks found.").font = _font(bold=True)
        return

    write_header_row(ws, 2, range(1, 10),
        ["#", "Symbol", "Status", "Total %\n(So Far)",
         f"Today\n{end_date.strftime('%d-%b')}%",
         "Max Day%", "# Big\nDays", "Recent 3D%", "Notes"],
        bg="00B050")

    for i, st in enumerate(focus, 1):
        r = i + 2
        bg = "E8F5E9" if st["status"] == "YET TO FIRE" else "F3E5F5"
        stat_bg = "00B050" if st["status"] == "YET TO FIRE" else "9370DB"
        vals = [i, st["symbol"], f"{st['emoji']} {st['status']}",
                st["total"]/100, st["today"]/100, st["max_day"]/100,
                st["n_big"],     st["recent3"]/100,
                "No big day yet — watch for breakout" if st["status"] == "YET TO FIRE"
                else "Pulled back after firing — wait for re-base"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border    = _border()
            cell.alignment = _ctr() if c <= 8 else _left(wrap=True)
            cell.fill      = _fill(bg)
            cell.font      = _font(bold=(c == 2))
            if c == 3:
                cell.fill = _fill(stat_bg)
                cell.font = _font(bold=True, color="FFFFFF")
            if isinstance(v, float):
                cell.number_format = pct_num_fmt()
        ws.row_dimensions[r].height = 22

    for i, w in enumerate([4, 14, 24, 12, 12, 12, 10, 12, 42], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_daily_heatmap(wb: Workbook, stocks: List[dict],
                        date_labels: List[str]):
    ws = wb.create_sheet("📊 Daily Heatmap")
    ws.freeze_panes = "C3"

    ncols = 2 + len(date_labels) + 2  # # + sym + dates + total + status
    last_col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = f"📊 Daily % Change Heatmap  |  {date_labels[0]} → {date_labels[-1]}"
    t.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _ctr()
    ws.row_dimensions[1].height = 26

    hdrs = ["#", "Symbol"] + date_labels + ["Total %", "Status"]
    write_header_row(ws, 2, range(1, len(hdrs)+1), hdrs)

    sorted_by_total = sorted(stocks, key=lambda x: x["total"], reverse=True)
    for i, st in enumerate(sorted_by_total, 1):
        r = i + 2
        ws.cell(row=r, column=1, value=i).alignment = _ctr()
        ws.cell(row=r, column=2, value=st["symbol"]).font = _font(bold=True)
        ws.cell(row=r, column=2).alignment = _ctr()
        ws.cell(row=r, column=2).border = _border()

        n = min(len(st["daily"]), len(date_labels))
        for j in range(n):
            d   = st["daily"][j]
            col = j + 3
            bg, fc = hm_color(d)
            cell = ws.cell(row=r, column=col, value=d/100)
            cell.number_format = pct_num_fmt()
            cell.alignment = _ctr()
            cell.border    = _border()
            cell.fill      = _fill(bg)
            cell.font      = _font(bold=(abs(d) >= 5), color=fc)

        # Total
        tc = len(date_labels) + 3
        tot_cell = ws.cell(row=r, column=tc, value=st["total"]/100)
        tot_cell.number_format = pct_num_fmt()
        tot_cell.alignment = _ctr()
        tot_cell.border = _border()
        t_bg = "00B050" if st["total"] >= 20 else ("70AD47" if st["total"] > 0 else "FF4444")
        t_fc = "FFFFFF" if st["total"] >= 15 or st["total"] < 0 else "000000"
        tot_cell.fill = _fill(t_bg)
        tot_cell.font = _font(bold=True, color=t_fc)

        # Status
        sc = tc + 1
        stat_cell = ws.cell(row=r, column=sc, value=f"{st['emoji']} {st['status']}")
        stat_cell.fill      = _fill(st["color"])
        stat_cell.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        stat_cell.alignment = _left()
        stat_cell.border    = _border()
        ws.row_dimensions[r].height = 20

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    for j in range(len(date_labels)):
        ws.column_dimensions[get_column_letter(j+3)].width = 10
    ws.column_dimensions[get_column_letter(len(date_labels)+3)].width = 10
    ws.column_dimensions[get_column_letter(len(date_labels)+4)].width = 22


def sheet_extended_avoid(wb: Workbook, stocks: List[dict]):
    ws = wb.create_sheet("🚫 Extended — Avoid")

    ext = [s for s in stocks
           if s["status"] in ("HIGHLY EXTENDED", "EXTENDED", "JUST FIRED TODAY")]
    ext_sorted = sorted(ext, key=lambda x: x["total"], reverse=True)

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "🚫 Extended Stocks — Risk of Pullback — Avoid Chasing"
    t.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    t.fill      = _fill("C00000")
    t.alignment = _ctr()
    ws.row_dimensions[1].height = 28

    write_header_row(ws, 2, range(1, 9),
        ["#", "Symbol", "Status", "Total %", "Max Day%",
         "# Big Days", "Wk1 Perf%", "Risk Notes"],
        bg="FF4444")

    RISK_NOTES = {
        "HIGHLY EXTENDED": "Multiple big days — parabolic. Likely to consolidate or pull back.",
        "EXTENDED":        "Strong run. Risk of profit-taking. Partial exit advised.",
        "JUST FIRED TODAY":"Today's breakout. Monitor closely — don't chase gaps.",
    }

    for i, st in enumerate(ext_sorted, 1):
        r = i + 2
        half = len(st["daily"]) // 2
        wk1  = sum(st["daily"][:half])
        bg = {"HIGHLY EXTENDED": "FFE5E5",
              "JUST FIRED TODAY": "FFF3E0"}.get(st["status"], "FFF8E1")
        vals = [i, st["symbol"], f"{st['emoji']} {st['status']}",
                st["total"]/100, st["max_day"]/100, st["n_big"],
                wk1/100, RISK_NOTES.get(st["status"], "—")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border    = _border()
            cell.alignment = _ctr() if c <= 7 else _left(wrap=True)
            cell.fill      = _fill(bg)
            cell.font      = _font(bold=(c == 2))
            if c == 3:
                cell.fill = _fill(st["color"])
                cell.font = _font(bold=True, color="FFFFFF")
            if isinstance(v, float):
                cell.number_format = pct_num_fmt()
        ws.row_dimensions[r].height = 22

    for i, w in enumerate([4, 14, 24, 10, 12, 10, 10, 44], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_summary(wb: Workbook, stocks: List[dict], cutoff_date: date, end_date: date):
    ws = wb.create_sheet("📋 Summary")

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = (f"📋 Portfolio Summary — Institutional Picks Fire Status\n"
               f"Cutoff: {cutoff_date.strftime('%d-%b-%Y')}  →  "
               f"End: {end_date.strftime('%d-%b-%Y')}  "
               f"({(end_date - cutoff_date).days} calendar days)")
    t.font      = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    t.fill      = _fill("1F4E79")
    t.alignment = _ctr(wrap=True)
    ws.row_dimensions[1].height = 38

    STATUS_GROUPS = [
        ("YET TO FIRE",        "⏳", "00B050", "Focus these today — maximum upside remaining"),
        ("FIRED & RETREATING", "🔄", "9370DB", "Fired before — wait for re-base / re-entry"),
        ("JUST FIRED TODAY",   "⚡", "FF8C00", "Breakout in progress — watch for follow-through"),
        ("FIRED",              "🔥", "FFD700", "In motion — hold / raise trail stop"),
        ("STEADY RUNNER",      "📈", "70AD47", "Grinding higher — monitor for acceleration"),
        ("EXTENDED",           "⚠️", "FFA500", "Big move already made — reduce / raise stop"),
        ("HIGHLY EXTENDED",    "🚀", "FF4444", "Avoid — parabolic / likely to pull back"),
        ("LAGGARD",            "❌", "C0C0C0", "No momentum — remove from focus"),
    ]

    write_header_row(ws, 2, range(1, 5),
        ["Status", "Count", "Symbols", "Trading Guidance"],
        bg="2E75B6")

    for i, (stat, emo, clr, guide) in enumerate(STATUS_GROUPS, 1):
        r = i + 2
        syms = [s["symbol"] for s in stocks if s["status"] == stat]

        c1 = ws.cell(row=r, column=1, value=f"{emo} {stat}")
        c1.fill = _fill(clr); c1.font = _font(bold=True, color="FFFFFF")
        c1.alignment = _ctr(); c1.border = _border()

        c2 = ws.cell(row=r, column=2, value=len(syms))
        c2.alignment = _ctr(); c2.border = _border()
        c2.font = Font(name="Arial", size=12, bold=True)
        c2.fill = _fill("F5F5F5")

        c3 = ws.cell(row=r, column=3, value=", ".join(syms) if syms else "—")
        c3.alignment = _left(wrap=True); c3.border = _border()
        c3.font = _font(); c3.fill = _fill("F5F5F5")

        c4 = ws.cell(row=r, column=4, value=guide)
        c4.alignment = _left(wrap=True); c4.border = _border()
        c4.font = _font(); c4.fill = _fill("FAFAFA")
        ws.row_dimensions[r].height = 28

    # Stats block
    total_stocks = len(stocks)
    ytf_count    = sum(1 for s in stocks if s["status"] == "YET TO FIRE")
    ext_count    = sum(1 for s in stocks if s["status"] in ("HIGHLY EXTENDED", "EXTENDED"))
    r_stats = len(STATUS_GROUPS) + 4
    ws.cell(row=r_stats, column=1, value="Total stocks tracked").font = _font(bold=True)
    ws.cell(row=r_stats, column=2, value=total_stocks).font = _font(bold=True)
    ws.cell(row=r_stats+1, column=1, value="Yet to Fire (focus)").font = _font(bold=True, color="00703C")
    ws.cell(row=r_stats+1, column=2, value=ytf_count).font = _font(bold=True, color="00703C")
    ws.cell(row=r_stats+2, column=1, value="Extended (caution)").font = _font(bold=True, color="C00000")
    ws.cell(row=r_stats+2, column=2, value=ext_count).font = _font(bold=True, color="C00000")

    for i, w in enumerate([26, 8, 42, 48], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Original daily-pct-change sheet (preserves existing report format)
# ─────────────────────────────────────────────────────────────────────────────

def sheet_daily_raw(wb: Workbook, pivot: pd.DataFrame, ref_dates: List[date]):
    """Replicate the original institutional_picks_daily_pct_change_report sheet."""
    ws = wb.create_sheet("📋 Raw Daily Data")

    light_fill = _fill("C6EFCE")
    dark_fill  = _fill("2E7D32")
    wf = Font(color="FFFFFF")
    nf = Font(color="000000")

    ws.cell(row=1, column=1, value="Symbol").font = _font(bold=True)
    for j, d in enumerate(ref_dates, 2):
        ws.cell(row=1, column=j, value=str(d)).font = _font(bold=True)
    tc = len(ref_dates) + 2
    ws.cell(row=1, column=tc, value="Total Return %").font = _font(bold=True)

    compounded = (1.0 + pivot / 100.0).prod(axis=1, skipna=True) - 1.0
    total_series = (compounded * 100.0).round(2)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, 2 + len(ref_dates)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    ws.column_dimensions[get_column_letter(tc)].width = 16

    for i, sym in enumerate(pivot.index, 2):
        ws.cell(row=i, column=1, value=str(sym)).alignment = _left()
        for j, d in enumerate(ref_dates, 2):
            val  = pivot.loc[sym, d]
            cell = ws.cell(row=i, column=j,
                           value=None if pd.isna(val) else round(float(val), 2))
            cell.alignment = _ctr()
            cell.number_format = "0.00"
            if pd.notna(val):
                v = float(val)
                if v > 5.0:   cell.fill, cell.font = dark_fill,  wf
                elif v > 3.0: cell.fill, cell.font = light_fill, nf

        tot_val = total_series.loc[sym]
        tc_cell = ws.cell(row=i, column=tc,
                          value=None if pd.isna(tot_val) else round(float(tot_val), 2))
        tc_cell.alignment = _ctr()
        tc_cell.number_format = "0.00"
        if pd.notna(tot_val):
            tv = float(tot_val)
            if tv > 5.0:
                tc_cell.fill = _fill("1B5E20"); tc_cell.font = Font(color="FFFFFF", bold=True)
            elif tv > 3.0:
                tc_cell.fill = _fill("9BE5A5"); tc_cell.font = Font(bold=True)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    args = build_arg_parser().parse_args()

    cutoff_date = parse_date(args.cutoff_date, "cutoff_date")
    end_date    = parse_date(args.end_date,    "end_date")

    if end_date <= cutoff_date:
        raise SystemExit("end_date must be after cutoff_date")

    reports_dir  = Path(args.reports_dir)
    reports_dir.mkdir(parents=True, exist_ok=True)

    cutoff_tag   = cutoff_date.strftime("%d%b%Y").lower()   # 02apr2026
    end_tag      = end_date.strftime("%d%b%Y").lower()      # 16apr2026
    input_txt    = reports_dir / f"institutional_picks_{cutoff_tag}.txt"

    # ── Load picks ────────────────────────────────────────────────────────────
    pairs    = parse_picks_file(input_txt)
    exchanges = sorted({exch for exch, _ in pairs})
    print(f"Loaded {len(pairs)} symbols from {input_txt.name}")

    # ── Connect to Kite ───────────────────────────────────────────────────────
    kite = get_kite_client(Path(args.token_file))

    token_lookups: Dict[str, tuple] = {}
    for exch in exchanges:
        print(f"  Building instrument lookup: {exch} …")
        token_lookups[exch] = build_token_lookups(kite, exch)

    # ── Fetch data: from day BEFORE start (for first pct-change) to end_date ─
    start_calendar = cutoff_date + timedelta(days=1)
    fetch_from     = cutoff_date - timedelta(days=5)  # ensure we have base close
    fetch_to       = end_date

    ref_dates: Optional[List[date]] = None
    symbol_data: Dict[str, Dict[date, float]] = {}
    missing: List[str] = []

    for exch, raw_sym in pairs:
        by_upper, by_norm = token_lookups[exch]
        token, resolved   = resolve_token(by_upper, by_norm, raw_sym)
        if token is None:
            missing.append(f"{exch}:{raw_sym}")
            print(f"  ⚠  Not found in Kite instruments: {exch}:{raw_sym}")
            continue

        print(f"  Fetching {exch}:{raw_sym} (token={token}) …")
        df = fetch_daily_closes(kite, token, fetch_from, fetch_to)
        if df.empty or len(df) < 2:
            print(f"     ↳ Insufficient data, skipping.")
            continue

        df["pct_change"] = df["close"].pct_change() * 100.0

        # Keep only rows from start_calendar onward (and up to end_date)
        mask = (df["date"] >= start_calendar) & (df["date"] <= end_date)
        df2  = df[mask]
        if df2.empty:
            continue

        symbol_data[raw_sym] = {
            row_d: float(row_p)
            for row_d, row_p in zip(df2["date"].tolist(), df2["pct_change"].tolist())
            if pd.notna(row_p)
        }

        # Build reference trading-date list from the richest symbol seen so far
        candidates = sorted(df2["date"].tolist())
        if ref_dates is None or len(candidates) > len(ref_dates):
            ref_dates = candidates

        time.sleep(args.sleep_seconds)

    if not ref_dates:
        raise SystemExit("No trading dates found between cutoff and end date.")

    print(f"\n  Reference trading days ({len(ref_dates)}): {ref_dates[0]} → {ref_dates[-1]}")

    # ── Build pivot ───────────────────────────────────────────────────────────
    rows = []
    for sym, d2p in symbol_data.items():
        for d in ref_dates:
            if d in d2p:
                rows.append({"symbol": sym, "date": d, "pct_change": round(d2p[d], 2)})
    long_df = pd.DataFrame(rows)
    if long_df.empty:
        raise SystemExit("No data to report.")

    pivot = (long_df
             .pivot_table(index="symbol", columns="date",
                          values="pct_change", aggfunc="first")
             .sort_index()
             .reindex(columns=ref_dates))

    # ── Classify each stock ───────────────────────────────────────────────────
    stocks: List[dict] = []
    for sym in pivot.index:
        daily = [float(v) for v in pivot.loc[sym].values if pd.notna(v)]
        if not daily:
            continue
        st = classify_stock(
            sym, daily,
            big_day_pct=args.big_day_pct,
            fired_total=args.fired_total,
            extended_total=args.extended_total,
            laggard_total=args.laggard_total,
        )
        stocks.append(st)

    # ── Build workbook ────────────────────────────────────────────────────────
    date_labels = [d.strftime("%d-%b") for d in ref_dates]

    wb = Workbook()
    sheet_fire_status(wb, stocks, end_date)
    sheet_todays_focus(wb, stocks, end_date)
    sheet_daily_heatmap(wb, stocks, date_labels)
    sheet_extended_avoid(wb, stocks)
    sheet_summary(wb, stocks, cutoff_date, end_date)
    sheet_daily_raw(wb, pivot, ref_dates)

    # ── Save ──────────────────────────────────────────────────────────────────
    out_name = f"ip_fire_report_{cutoff_tag}_to_{end_tag}.xlsx"
    out_path = reports_dir / out_name
    if out_path.exists():
        for i in range(2, 50):
            cand = reports_dir / f"ip_fire_report_{cutoff_tag}_to_{end_tag}_v{i}.xlsx"
            if not cand.exists():
                out_path = cand
                break

    wb.save(out_path)

    # ── Summary ───────────────────────────────────────────────────────────────
    ytf  = [s["symbol"] for s in stocks if s["status"] == "YET TO FIRE"]
    ext  = [s["symbol"] for s in stocks if s["status"] in ("HIGHLY EXTENDED", "EXTENDED")]
    jf   = [s["symbol"] for s in stocks if s["status"] == "JUST FIRED TODAY"]

    print(f"\n{'─'*60}")
    print(f"  Report saved → {out_path}")
    print(f"  Trading days : {len(ref_dates)}")
    print(f"  Stocks       : {len(stocks)}")
    print(f"  ⏳ Yet to Fire (TODAY'S FOCUS) : {', '.join(ytf) or 'none'}")
    if jf:
        print(f"  ⚡ Just Fired Today            : {', '.join(jf)}")
    if ext:
        print(f"  🚫 Extended (avoid)            : {', '.join(ext)}")
    if missing:
        print(f"  ⚠  Not found in Kite           : {', '.join(missing)}")
    print(f"{'─'*60}\n")


if __name__ == "__main__":
    main()
