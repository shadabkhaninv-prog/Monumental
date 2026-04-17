#!/usr/bin/env python3
"""
One-shot report generator:
  - Argument: cutoff date in DD-MM-YYYY (e.g., 02-04-2026)
  - Loads: reports/institutional_picks_<ddmonyyyy>.txt  (e.g., 02apr2026)
  - Computes: daily close-to-close % change for the first 10 trading days
               starting from (cutoff_date + 1 calendar day)
  - Output: transposed & formatted report with green shading in XLSX

Dependencies:
  pip install kiteconnect pandas openpyxl
"""

from __future__ import annotations

import argparse
import re
import time
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def parse_cutoff_date(s: str) -> date:
    """
    Parse DD-MM-YYYY (user format: 02-04-2026).
    """
    s = s.strip()
    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", s)
    if not m:
        raise SystemExit("Date must be in DD-MM-YYYY format, e.g. 02-04-2026")
    dd, mm, yyyy = map(int, m.groups())
    return date(yyyy, mm, dd)


def read_kite_token_file(token_file: Path) -> Dict[str, str]:
    if not token_file.exists():
        raise SystemExit(f"Missing token file: {token_file}")

    values: Dict[str, str] = {}
    for raw_line in token_file.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        values[k.strip().upper()] = v.strip()

    if "API_KEY" not in values or "ACCESS_TOKEN" not in values:
        raise SystemExit(f"{token_file} must contain API_KEY and ACCESS_TOKEN (key=value lines).")
    return values


def get_kite_client(token_file: Path):
    try:
        from kiteconnect import KiteConnect
    except ImportError:
        raise SystemExit("kiteconnect not installed. Run: pip install kiteconnect")

    creds = read_kite_token_file(token_file)
    kite = KiteConnect(api_key=creds["API_KEY"])
    kite.set_access_token(creds["ACCESS_TOKEN"])
    return kite


def normalize_symbol_for_matching(symbol: str) -> str:
    """
    Best-effort normalization to match Kite tradingsymbol variants.
    Example: GVT&D -> GVTANDD
    """
    s = symbol.upper().strip()
    s = s.replace("&", "AND")
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s


def parse_institutional_picks(txt_path: Path) -> List[Tuple[str, str]]:
    """
    Returns (exchange, symbol) pairs from your institutional_picks_*.txt file.
    """
    if not txt_path.exists():
        raise SystemExit(f"Input file not found: {txt_path}")

    pairs: List[Tuple[str, str]] = []
    seen: set[Tuple[str, str]] = set()

    for raw_line in txt_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or line.startswith("###"):
            continue

        if ":" in line:
            exch, sym = line.split(":", 1)
        else:
            exch, sym = "NSE", line

        exch = exch.strip().upper()
        sym = sym.strip().upper()
        if not sym:
            continue
        key = (exch, sym)
        if key not in seen:
            pairs.append(key)
            seen.add(key)

    if not pairs:
        raise SystemExit(f"No symbols found in: {txt_path}")
    return pairs


def build_instrument_token_lookup(kite, exchange: str) -> Tuple[Dict[str, int], Dict[str, int]]:
    """
    Returns lookup maps:
      - {TRADINGSYMBOL_UPPER -> instrument_token}
      - {normalized_TRADINGSYMBOL -> instrument_token}
    """
    instruments = pd.DataFrame(kite.instruments(exchange))
    if instruments.empty:
        raise SystemExit(f"kite.instruments({exchange!r}) returned no rows.")

    if "segment" in instruments.columns:
        instruments = instruments[instruments["segment"] == exchange]
    if "instrument_type" in instruments.columns:
        instruments = instruments[instruments["instrument_type"] == "EQ"]

    instruments["tradingsymbol"] = instruments["tradingsymbol"].astype(str)
    instruments["tradingsymbol_upper"] = instruments["tradingsymbol"].str.upper()
    instruments["tradingsymbol_norm"] = instruments["tradingsymbol_upper"].map(normalize_symbol_for_matching)

    token_by_ts_upper: Dict[str, int] = {}
    token_by_norm: Dict[str, int] = {}
    for _, row in instruments.iterrows():
        tsu = row["tradingsymbol_upper"]
        token = int(row["instrument_token"])
        token_by_ts_upper.setdefault(tsu, token)
        token_by_norm.setdefault(row["tradingsymbol_norm"], token)

    return token_by_ts_upper, token_by_norm


def resolve_instrument_token(
    token_by_ts_upper: Dict[str, int],
    token_by_norm: Dict[str, int],
    raw_symbol: str,
) -> Tuple[Optional[int], Optional[str]]:
    sym_upper = raw_symbol.upper().strip()
    if sym_upper in token_by_ts_upper:
        return token_by_ts_upper[sym_upper], sym_upper
    sym_norm = normalize_symbol_for_matching(sym_upper)
    if sym_norm in token_by_norm:
        return token_by_norm[sym_norm], sym_norm
    return None, None


def fetch_close_history(
    kite,
    instrument_token: int,
    fetch_from: date,
    fetch_to: date,
) -> pd.DataFrame:
    """
    Fetch daily close history, return DataFrame with columns: date, close.
    """
    rows = kite.historical_data(
        instrument_token=instrument_token,
        from_date=datetime.combine(fetch_from, datetime.min.time()),
        to_date=datetime.combine(fetch_to, datetime.min.time()),
        interval="day",
        continuous=False,
        oi=False,
    )
    if not rows:
        return pd.DataFrame(columns=["date", "close"])

    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"]).dt.date
    df = df.sort_values("date")
    df["close"] = pd.to_numeric(df["close"], errors="coerce")
    df = df[["date", "close"]].dropna(subset=["close"])
    return df


def build_transposed_xlsx(
    daily_pivot: pd.DataFrame,
    pivot_with_total: pd.DataFrame,
    ref_dates: List[date],
    move_order_df: pd.DataFrame,
    out_xlsx: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "daily_pct_change"

    # Daily columns (dates) + final total return column
    dates = list(ref_dates)
    total_col_name = "Total Return (Till Date) %"

    # Header row
    ws.cell(row=1, column=1, value="Symbol").font = Font(bold=True)
    for j, d in enumerate(dates, start=2):
        ws.cell(row=1, column=j, value=str(d)).font = Font(bold=True)
    ws.cell(row=1, column=2 + len(dates), value=total_col_name).font = Font(bold=True)

    # Styles for conditional green shading
    # Brighter than before to make >3% highlight more visible.
    light_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    dark_fill = PatternFill(start_color="FF2E7D32", end_color="FF2E7D32", fill_type="solid")
    # Distinct shades for the total-return column so it doesn't look identical to daily cells.
    light_total_fill = PatternFill(start_color="FF9BE5A5", end_color="FF9BE5A5", fill_type="solid")
    dark_total_fill = PatternFill(start_color="FF1B5E20", end_color="FF1B5E20", fill_type="solid")
    normal_font = Font(color="FF000000")
    white_font = Font(color="FFFFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    from openpyxl.styles import Border, Side
    thin_gray_border = Border(
        left=Side(style="thin", color="FFBDBDBD"),
        right=Side(style="thin", color="FFBDBDBD"),
        top=Side(style="thin", color="FFBDBDBD"),
        bottom=Side(style="thin", color="FFBDBDBD"),
    )

    # Body
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, 2 + len(dates)):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12
    ws.column_dimensions[ws.cell(row=1, column=2 + len(dates)).column_letter].width = 22

    for i, sym in enumerate(daily_pivot.index, start=2):
        ws.cell(row=i, column=1, value=str(sym)).alignment = left
        for j, d in enumerate(dates, start=2):
            val = daily_pivot.loc[sym, d]
            cell = ws.cell(row=i, column=j, value=None if pd.isna(val) else float(val))
            cell.alignment = center
            cell.number_format = "0.00"
            if pd.notna(val):
                v = float(val)
                # Highlight only positive moves per your request.
                if v > 5.0:
                    cell.fill = dark_fill
                    cell.font = white_font
                elif v > 3.0:
                    cell.fill = light_fill
                    cell.font = normal_font
                else:
                    cell.font = normal_font

        # Total return column (compounded from available daily values)
        total_val = pivot_with_total.loc[sym, total_col_name]
        total_cell = ws.cell(row=i, column=2 + len(dates), value=None if pd.isna(total_val) else float(total_val))
        total_cell.alignment = center
        total_cell.number_format = "0.00"
        if pd.notna(total_val):
            tv = float(total_val)
            if tv > 5.0:
                total_cell.fill = dark_total_fill
                total_cell.font = Font(color="FFFFFFFF", bold=True)
            elif tv > 3.0:
                total_cell.fill = light_total_fill
                total_cell.font = Font(color="FF000000", bold=True)
            else:
                total_cell.font = Font(color="FF000000", bold=False)
            total_cell.border = thin_gray_border

    # Second sheet: first date each stock crossed >3%
    ws2 = wb.create_sheet("move_order_over_3pct")
    ws2.cell(row=1, column=1, value="Order").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Symbol").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="First date >3%").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="First move >3% (%)").font = Font(bold=True)
    ws2.cell(row=1, column=5, value="Trading Day #").font = Font(bold=True)

    if move_order_df.empty:
        ws2.cell(row=2, column=1, value="No moves > 3% found in this window.")
    else:
        for r_idx, row in enumerate(move_order_df.itertuples(index=False), start=2):
            # row: (order, symbol, first_date, first_move_pct)
            ws2.cell(row=r_idx, column=1, value=int(row[0]))
            ws2.cell(row=r_idx, column=2, value=str(row[1]))
            ws2.cell(row=r_idx, column=3, value=str(row[2]) if pd.notna(row[2]) else "")
            first_move = None if pd.isna(row[3]) else float(row[3])
            move_cell = ws2.cell(row=r_idx, column=4, value=first_move)

            # Day-number within the window for quick "who was first" scanning.
            first_date_val = row[2]
            first_date_norm = None
            if pd.notna(first_date_val):
                if isinstance(first_date_val, pd.Timestamp):
                    first_date_norm = first_date_val.date()
                else:
                    first_date_norm = first_date_val
            if first_date_norm in ref_dates:
                ws2.cell(row=r_idx, column=5, value=ref_dates.index(first_date_norm) + 1)

            # Conditional coloring on the first move column.
            if first_move is not None:
                if first_move > 5.0:
                    move_cell.fill = dark_fill
                    move_cell.font = white_font
                elif first_move > 3.0:
                    move_cell.fill = light_fill
                    move_cell.font = normal_font

            # Also color the row's symbol cell lightly (same rule as move).
            sym_cell = ws2.cell(row=r_idx, column=2)
            if first_move is not None:
                if first_move > 5.0:
                    sym_cell.fill = dark_fill
                    sym_cell.font = white_font
                elif first_move > 3.0:
                    sym_cell.fill = light_fill
                    sym_cell.font = normal_font
        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 20
        ws2.column_dimensions["E"].width = 14

    wb.save(out_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate institutional picks daily pct change report.")
    parser.add_argument("cutoff_date", help="Cutoff date in DD-MM-YYYY format (e.g., 02-04-2026)")
    parser.add_argument(
        "--reports-dir",
        default=str(Path("reports")),
        help="Reports directory containing institutional_picks_*.txt (default: reports)",
    )
    parser.add_argument(
        "--token-file",
        default="kite_token.txt",
        help="kite_token.txt path (default: kite_token.txt in current folder)",
    )
    parser.add_argument("--sleep-seconds", type=float, default=0.35, help="Delay between Kite calls.")
    args = parser.parse_args()

    cutoff_date = parse_cutoff_date(args.cutoff_date)
    start_calendar = cutoff_date + timedelta(days=1)

    cutoff_tag = cutoff_date.strftime("%d%b%Y").lower()  # 02apr2026
    reports_dir = Path(args.reports_dir)
    input_txt = reports_dir / f"institutional_picks_{cutoff_tag}.txt"

    pairs = parse_institutional_picks(input_txt)
    exchanges = sorted({exch for exch, _ in pairs})

    kite = get_kite_client(Path(args.token_file))

    token_lookups: Dict[str, Tuple[Dict[str, int], Dict[str, int]]] = {}
    for exch in exchanges:
        print(f"Building instrument lookup for {exch} ...")
        token_by_ts_upper, token_by_norm = build_instrument_token_lookup(kite, exch)
        token_lookups[exch] = (token_by_ts_upper, token_by_norm)

    # We need prior trading day's close for the first output day.
    # Also need enough calendar room to reach 10 trading days.
    fetch_from = start_calendar - timedelta(days=45)
    fetch_to = min(date.today(), start_calendar + timedelta(days=180))

    ref_dates: Optional[List[date]] = None
    symbol_date_pct: Dict[str, Dict[date, float]] = {}
    missing: List[str] = []

    for exch, raw_sym in pairs:
        token_by_ts_upper, token_by_norm = token_lookups[exch]
        token, resolved = resolve_instrument_token(token_by_ts_upper, token_by_norm, raw_sym)

        if token is None:
            missing.append(f"{exch}:{raw_sym}")
            continue

        print(f"Fetching {exch}:{raw_sym} (token={token}) ...")
        df = fetch_close_history(kite, token, fetch_from, fetch_to)
        if df.empty or len(df) < 2:
            continue

        df = df.copy()
        df["pct_change"] = df["close"].pct_change() * 100.0

        # Candidate rows on/after start_calendar
        df2 = df[df["date"] >= start_calendar]
        if df2.empty:
            continue

        sym_out = raw_sym  # strip NSE: by using symbol part only
        symbol_date_pct[sym_out] = {
            row_date: float(row_pct)
            for row_date, row_pct in zip(df2["date"].tolist(), df2["pct_change"].tolist())
            if pd.notna(row_pct)
        }

        # Determine the reference trading-date list as:
        #   - first 10 trading days after start_calendar, BUT
        #   - if fewer than 10 exist up to fetch_to, we still take what exists.
        # Choose the candidate with the largest length (helps if first symbol is thin/new).
        candidate_dates = list(df2["date"].iloc[:10].values)
        if ref_dates is None or len(candidate_dates) > len(ref_dates):
            ref_dates = candidate_dates

        time.sleep(args.sleep_seconds)

    if ref_dates is None or not ref_dates:
        raise SystemExit("Could not determine any trading days. Check Kite availability/data.")

    # Pivot: symbols rows, dates columns.
    # Build the long dataframe from the final ref_dates
    rows: List[Dict[str, object]] = []
    for sym, d2p in symbol_date_pct.items():
        for d in ref_dates:
            if d in d2p:
                rows.append({"symbol": sym, "date": d, "pct_change": d2p[d]})
    long_df = pd.DataFrame(rows)
    if long_df.empty:
        raise SystemExit("No data returned for any symbol on/after the start date.")

    long_df["pct_change"] = pd.to_numeric(long_df["pct_change"], errors="coerce").round(2)
    pivot = long_df.pivot_table(index="symbol", columns="date", values="pct_change", aggfunc="first").sort_index()
    pivot = pivot.reindex(columns=ref_dates)

    # Output
    start_tag = start_calendar.strftime("%d%b%Y").lower()
    end_tag = ref_dates[-1].strftime("%d%b%Y").lower()
    td_count = len(ref_dates)
    if td_count < 10:
        print(f"NOTE: Only {td_count} trading day(s) available after start date up to {fetch_to}.")
    out_stem = (
        f"institutional_picks_daily_pct_change_{cutoff_tag}_start_{start_tag}"
        f"_{td_count}td_to_{end_tag}_enh"
    )

    out_xlsx = reports_dir / f"{out_stem}.xlsx"
    out_csv = reports_dir / f"{out_stem}.csv"
    # If Excel has the old file open, re-use a new versioned filename.
    if out_xlsx.exists():
        for i in range(2, 50):
            cand = reports_dir / f"{out_stem}_v{i}.xlsx"
            if not cand.exists():
                out_xlsx = cand
                break
    if out_csv.exists():
        for i in range(2, 50):
            cand = reports_dir / f"{out_stem}_v{i}.csv"
            if not cand.exists():
                out_csv = cand
                break

    # Add total return till date (compounded from available daily values)
    total_col_name = "Total Return (Till Date) %"
    # pivot has index=symbol, columns=dates
    compounded = (1.0 + (pivot / 100.0)).prod(axis=1, skipna=True) - 1.0
    pivot_with_total = pivot.copy()
    pivot_with_total[total_col_name] = (compounded * 100.0).round(2)

    # CSV with the total column
    pivot_with_total.to_csv(out_csv, index=True)

    # Build move order sheet: first day each stock crosses >3%
    move_rows: List[Dict[str, object]] = []
    daily_index = list(pivot.index)
    for sym in daily_index:
        s = pivot.loc[sym]
        over = s[s > 3.0]
        if over.empty:
            continue
        first_date = over.index[0]
        first_move_pct = float(over.iloc[0])
        move_rows.append(
            {
                "order_sort_date": first_date,
                "symbol": sym,
                "first_date": first_date,
                "first_move_pct": first_move_pct,
            }
        )
    move_order_df = pd.DataFrame(move_rows)
    if not move_order_df.empty:
        move_order_df = move_order_df.sort_values(
            ["order_sort_date", "first_move_pct"],
            ascending=[True, False],
        ).reset_index(drop=True)
        move_order_df.insert(0, "order", range(1, len(move_order_df) + 1))
        move_order_df = move_order_df[["order", "symbol", "first_date", "first_move_pct"]]
    else:
        move_order_df = pd.DataFrame(columns=["order", "symbol", "first_date", "first_move_pct"])

    build_transposed_xlsx(daily_pivot=pivot, pivot_with_total=pivot_with_total, ref_dates=ref_dates, move_order_df=move_order_df, out_xlsx=out_xlsx)

    print("Done.")
    print(f"CSV  -> {out_csv}")
    print(f"XLSX -> {out_xlsx}")
    if missing:
        print(f"Missing (not found in Kite instruments): {', '.join(missing)}")


if __name__ == "__main__":
    main()

