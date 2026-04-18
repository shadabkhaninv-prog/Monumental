"""
Stock rating application based on Stock_Rating_Spec v1.0 / v2.

Required positional arguments:
    python stock_rating.py 2026-04-11 2026-03-04

Optional flags:
    --symbols PATH      Explicit symbol list file
    --token PATH        Kite token file path (default: kite_token.txt)
    --output-dir PATH   Output directory (default: ./output)
    --index-symbol STR  Preferred index symbol in indexbhav
    --liquid-leader-map JSON  Optional neo liquid bonus map passed inline

Dependencies:
    pip install kiteconnect mysql-connector-python pandas numpy openpyxl
"""

from __future__ import annotations

import argparse
import json
import math
import os
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import numpy as np
import pandas as pd

try:
    import mysql.connector
except Exception:  # pragma: no cover - import failure handled at runtime
    mysql = None

try:
    from kiteconnect import KiteConnect
except Exception:  # pragma: no cover - import failure handled at runtime
    KiteConnect = None

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NSE_YEAR_TRADING_DAYS = 248   # NSE trading days in a 365-calendar-day fetch window
LOOKBACK_52W = NSE_YEAR_TRADING_DAYS
LOOKBACK_12M = NSE_YEAR_TRADING_DAYS
LOOKBACK_6M = 125
LOOKBACK_3M = 60
TURNOVER_LOOKBACK = 42
TURNOVER_LOOKBACK_SHORT = 21
ATR_LOOKBACK = 21
RS_SLOPE_LOOKBACK = 21
SPIKE_LOOKBACK = 125
SPIKE_WINDOW_10 = 10
SPIKE_WINDOW_30 = 30
SPIKE_WINDOW_60 = 60
TURNOVER_CRORE_DIVISOR = 10_000_000.0
MIN_MEDIAN_TURNOVER_CRORES = 15.0      # universe hard gate: median 42D turnover must be ≥ 15 Cr
MIN_AVG_TURNOVER_42D_CRORES = 15.0
INST_PICKS_MIN_MEDIAN_TURNOVER_CRORES = 10.0  # additional gate for institutional picks only
MEDIAN_TURNOVER_LOW_THRESHOLD_CRORES = 20.0  # absolute gate for percentile-based median-turnover penalty
GAPUP_LOOKBACK = 60
UPTREND_CONSISTENCY_MIN_LOOKBACK = 30

DB_CONFIG = {
    "host": "localhost",
    "port": 3306,
    "user": "root",
    "password": "root",
    "database": "bhav",
}

# The sector section of the spec defines the methodology but not the points.
# Assumption used here:
#   - Sector strength = share of sector stocks that landed in the top quartile
#     of pre-sector composite score.
#   - Stocks in sectors within the top 10% of sector-strength distribution get +4.
#   - Stocks in sectors within the top 30% (but not top 10%) get +2.
SECTOR_TOP10_POINTS = 8
SECTOR_TOP20_POINTS = 6
SECTOR_TOP30_POINTS = 4

# Top-20 concentration bonuses — added to each sector's leadership score before
# the top-10% / top-30% threshold comparison is made.
# Breadth bonus: how many of the sector's stocks appear in the final Top-20 list.
SECTOR_TOP20_BREADTH_BONUS_3PLUS = 6   # 3 or more sector stocks in Top-20
SECTOR_TOP20_BREADTH_BONUS_2     = 4   # exactly 2 sector stocks in Top-20
SECTOR_TOP20_BREADTH_BONUS_1     = 2   # exactly 1 sector stock in Top-20
# Penetration bonus: what fraction of the sector's stocks made the Top-20 list.
SECTOR_TOP20_PENETRATION_BONUS_100 = 3  # 100% of sector stocks in Top-20 (min 2 stocks)
SECTOR_TOP20_PENETRATION_BONUS_67  = 2  # >= 67% of sector stocks in Top-20
SECTOR_TOP20_PENETRATION_BONUS_50  = 1  # >= 50% of sector stocks in Top-20

# 12M return + 52W high proximity concentration bonus for sectors (min 2 stocks in sector).
# Counts stocks that are BOTH in the top-30% of 12M return AND within 20% of their 52W high.
# Base points + 1 per qualifying stock are added to the sector leadership score.
SECTOR_12M_52W_GT3_BASE = 6   # >3 qualifying stocks → base 6 + 1 per qualifying stock
SECTOR_12M_52W_GT2_BASE = 4   # >2 qualifying stocks → base 4 + 1 per qualifying stock

# IPO performance bonus (applies to stocks listed < 6 months with CMP > issue price)
IPO_TOP10_POINTS = 6
IPO_TOP20_POINTS = 4
IPO_TOP40_POINTS = 3
IPO_MIN_GAIN = 0.20


@dataclass
class WarningLog:
    symbol: str
    message: str
    avg_turnover_21d: Optional[float] = None
    avg_turnover_42d: Optional[float] = None
    median_turnover_42d: Optional[float] = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Rate NSE stocks using the Stock_Rating_Spec scoring model."
    )
    parser.add_argument("cutoff_date", help="As-of date in YYYY-MM-DD format")
    parser.add_argument("reset_date", help="Reset date in YYYY-MM-DD format")
    parser.add_argument(
        "--symbols",
        help="Optional explicit symbol file path. Default: auto-resolve gmlist_<DDMMMYYYY>.txt.",
    )
    parser.add_argument(
        "--symbols-dir",
        default="gmlist",
        help="Directory to search for gmlist files when --symbols is not provided.",
    )
    parser.add_argument(
        "--token",
        default="kite_token.txt",
        help="Path to kite_token.txt (default: kite_token.txt).",
    )
    parser.add_argument(
        "--output-dir",
        default="output",
        help="Directory for the generated workbook (default: ./output).",
    )
    parser.add_argument(
        "--index-symbol",
        default="Nifty Smallcap 250",
        help="Preferred index symbol label in indexbhav (default: Nifty Smallcap 250).",
    )
    parser.add_argument(
        "--liquid-leader-map",
        default="",
        help="Optional JSON object mapping top neo liquid leaders to bonus points.",
    )
    args = parser.parse_args()
    args.cutoff_date = parse_iso_date(args.cutoff_date, "cutoff_date")
    args.reset_date = parse_iso_date(args.reset_date, "reset_date")
    if args.reset_date > args.cutoff_date:
        raise SystemExit("reset_date must be on or before cutoff_date")
    return args


def parse_iso_date(value: str, field_name: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise SystemExit(f"Invalid {field_name}: {value}. Expected YYYY-MM-DD.") from exc


def locate_symbol_file(cutoff_date: date, explicit_path: Optional[str], symbols_dir: str) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if not path.exists():
            raise SystemExit(f"Symbol file not found: {path}")
        return path

    cutoff_stamp = cutoff_date.strftime('%d%b%Y')
    preferred_name = f"updated_gmlist_{cutoff_stamp}.txt"
    candidates = [
        Path.cwd() / symbols_dir / preferred_name,
        Path(__file__).resolve().parent / symbols_dir / preferred_name,
        Path.cwd() / preferred_name,
        Path(__file__).resolve().parent / preferred_name,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    tried = "\n".join(f"  - {p}" for p in candidates)
    raise SystemExit(f"Could not locate symbol file {preferred_name}. Tried:\n{tried}")


def read_symbols(path: Path) -> List[str]:
    symbols: List[str] = []
    seen = set()
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        symbol = line.upper()
        if symbol.startswith("NSE:"):
            symbol = symbol[4:]
        if symbol and symbol not in seen:
            symbols.append(symbol)
            seen.add(symbol)
    if not symbols:
        raise SystemExit(f"Symbol file is empty: {path}")
    return symbols


def parse_liquid_leader_bonus_map(raw_value: str, replacements: Dict[str, str]) -> Dict[str, int]:
    if not raw_value:
        return {}
    try:
        payload = json.loads(raw_value)
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Invalid --liquid-leader-map JSON: {exc}") from exc
    if not isinstance(payload, dict):
        raise SystemExit("--liquid-leader-map must be a JSON object")

    bonus_map: Dict[str, int] = {}
    for raw_symbol, raw_bonus in payload.items():
        symbol = str(raw_symbol).strip().upper()
        mapped = replacements.get(symbol, symbol)
        try:
            bonus = int(raw_bonus)
        except (TypeError, ValueError):
            continue
        if bonus <= 0 or not symbol:
            continue
        bonus_map[symbol] = max(bonus_map.get(symbol, 0), bonus)
        bonus_map[mapped] = max(bonus_map.get(mapped, 0), bonus)
    return bonus_map


def read_kite_token_file(path: Path) -> Dict[str, str]:
    if not path.exists():
        raise SystemExit(f"Token file not found: {path}")
    text = None
    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
    for encoding in encodings:
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise SystemExit(f"Unable to decode token file: {path}")
    values: Dict[str, str] = {}
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip()
    if "API_KEY" not in values or "ACCESS_TOKEN" not in values:
        raise SystemExit("kite_token.txt must contain API_KEY and ACCESS_TOKEN")
    generated = values.get("GENERATED")
    if generated:
        token_day = generated[:10]
        today = datetime.now().strftime("%Y-%m-%d")
        if token_day != today:
            print(f"WARNING: kite_token.txt GENERATED is {token_day}; today is {today}.")
    return values


def get_kite_client(token_file: Path):
    if KiteConnect is None:
        raise SystemExit("kiteconnect is not installed. Install it with pip.")
    creds = read_kite_token_file(token_file)
    kite = KiteConnect(api_key=creds["API_KEY"])
    kite.set_access_token(creds["ACCESS_TOKEN"])
    print("Kite session initialized")
    return kite


def get_db_connection():
    if mysql is None:
        raise SystemExit("mysql-connector-python is not installed. Install it with pip.")
    try:
        return mysql.connector.connect(**DB_CONFIG)
    except Exception as exc:
        raise SystemExit(f"Failed to connect to MySQL bhav database: {exc}") from exc


def ensure_inactive_symbols_table(conn) -> None:
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS inactive_symbols (
            symbol VARCHAR(50) NOT NULL,
            new_symbol VARCHAR(50) NULL,
            PRIMARY KEY (symbol)
        )
        """
    )
    conn.commit()
    cursor.close()


def load_symbol_replacements(conn) -> Dict[str, str]:
    ensure_inactive_symbols_table(conn)
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT UPPER(symbol) AS symbol, UPPER(TRIM(new_symbol)) AS new_symbol
        FROM inactive_symbols
        WHERE new_symbol IS NOT NULL
          AND TRIM(new_symbol) <> ''
        """
    )
    rows = cursor.fetchall()
    cursor.close()
    return {row[0]: row[1] for row in rows}


def record_inactive_symbol(conn, symbol: str) -> None:
    ensure_inactive_symbols_table(conn)
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO inactive_symbols (symbol, new_symbol)
        VALUES (%s, NULL)
        ON DUPLICATE KEY UPDATE symbol = VALUES(symbol)
        """,
        (symbol.upper(),),
    )
    conn.commit()
    cursor.close()


def remap_symbols(symbols: Iterable[str], replacements: Dict[str, str]) -> List[str]:
    remapped: List[str] = []
    seen: set[str] = set()
    for symbol in symbols:
        effective = replacements.get(symbol.upper(), symbol.upper())
        if effective not in seen:
            remapped.append(effective)
            seen.add(effective)
    return remapped


def load_sector_map(conn, symbols: Iterable[str]) -> Dict[str, str]:
    symbol_list = list(symbols)
    placeholders = ",".join(["%s"] * len(symbol_list))
    sql = (
        f"SELECT UPPER(symbol) AS symbol, sector1 "
        f"FROM sectors WHERE UPPER(symbol) IN ({placeholders})"
    )
    cursor = conn.cursor()
    cursor.execute(sql, [s.upper() for s in symbol_list])
    mapping = {row[0]: (row[1] or "Unknown") for row in cursor.fetchall()}
    cursor.close()
    return {sym: mapping.get(sym.upper(), "Unknown") for sym in symbol_list}


def load_index_history(conn, start_date: date, cutoff_date: date, preferred_symbol: str) -> pd.DataFrame:
    print(f"Loading index data from indexbhav for {start_date} to {cutoff_date}")
    cursor = conn.cursor()
    exact_sql = """
        SELECT mktdate, open, high, low, close, symbol
        FROM indexbhav
        WHERE mktdate BETWEEN %s AND %s
          AND UPPER(symbol) = UPPER(%s)
        ORDER BY mktdate
    """
    cursor.execute(exact_sql, (start_date, cutoff_date, preferred_symbol))
    rows = cursor.fetchall()
    if not rows:
        like_sql = """
            SELECT mktdate, open, high, low, close, symbol
            FROM indexbhav
            WHERE mktdate BETWEEN %s AND %s
              AND UPPER(symbol) LIKE %s
            ORDER BY mktdate
        """
        cursor.execute(like_sql, (start_date, cutoff_date, "%SMALLCAP%250%"))
        rows = cursor.fetchall()
    cursor.close()
    if not rows:
        raise SystemExit(
            f"Could not load Nifty Smallcap 250 data from indexbhav between {start_date} and {cutoff_date}."
        )
    df = pd.DataFrame(rows, columns=["date", "open", "high", "low", "close", "symbol"])
    df["date"] = pd.to_datetime(df["date"]).dt.date
    df = df.drop_duplicates(subset=["date"], keep="last").sort_values("date").set_index("date")
    return df


def get_existing_yearly_bhav_tables(conn, start_date: date, cutoff_date: date) -> List[str]:
    table_names = [f"bhav{year}" for year in range(start_date.year, cutoff_date.year + 1)]
    cursor = conn.cursor()
    existing: List[str] = []
    for table_name in table_names:
        cursor.execute("SHOW TABLES LIKE %s", (table_name,))
        if cursor.fetchone():
            existing.append(table_name)
    cursor.close()
    return existing


def load_turnover_map(
    conn,
    symbols: Iterable[str],
    start_date: date,
    cutoff_date: date,
) -> Dict[str, Dict[str, float]]:
    symbol_list = list(symbols)
    if not symbol_list:
        return {}

    tables = get_existing_yearly_bhav_tables(conn, start_date, cutoff_date)
    if not tables:
        return {}
    print(f"Loading turnover filter data from tables: {', '.join(tables)}")

    symbol_placeholders = ",".join(["%s"] * len(symbol_list))
    union_parts = []
    params: List[object] = []
    for table_name in tables:
        union_parts.append(
            f"""
            SELECT UPPER(SYMBOL) AS symbol, MKTDATE AS trade_date, CLOSE, VOLUME
            FROM {table_name}
            WHERE MKTDATE BETWEEN %s AND %s
              AND UPPER(SYMBOL) IN ({symbol_placeholders})
            """
        )
        params.extend([start_date, cutoff_date, *[s.upper() for s in symbol_list]])

    sql = f"""
        SELECT symbol, trade_date, CLOSE, VOLUME
        FROM (
            {" UNION ALL ".join(union_parts)}
        ) turnover_rows
        ORDER BY symbol, trade_date
    """

    df = pd.read_sql(sql, conn, params=params)
    if df.empty:
        return {}

    df["turnover"] = pd.to_numeric(df["CLOSE"], errors="coerce") * pd.to_numeric(df["VOLUME"], errors="coerce")
    turnover_map: Dict[str, Dict[str, float]] = {}
    for symbol, group in df.groupby("symbol"):
        sorted_group = group.sort_values("trade_date")
        recent = sorted_group.tail(TURNOVER_LOOKBACK)
        recent_21d = sorted_group.tail(TURNOVER_LOOKBACK_SHORT)
        turnover_map[str(symbol).upper()] = {
            "avg_turnover_21d": float(recent_21d["turnover"].mean()) / TURNOVER_CRORE_DIVISOR if not recent_21d.empty else math.nan,
            "avg_turnover_42d": float(recent["turnover"].mean()) / TURNOVER_CRORE_DIVISOR if not recent.empty else math.nan,
            "median_turnover_42d": float(recent["turnover"].median()) / TURNOVER_CRORE_DIVISOR if not recent.empty else math.nan,
        }
    return turnover_map


def load_ipo_issue_price_map(
    conn,
    symbols: Iterable[str],
    cutoff_date: date,
) -> Dict[str, float]:
    """Return {symbol: issue_price} for symbols present in ipobhav."""
    sym_list = [s.upper() for s in symbols]
    if not sym_list:
        return {}
    placeholders = ", ".join(["%s"] * len(sym_list))
    sql = (
        f"SELECT SYMBOL, ISSUE_PRICE FROM ipobhav "
        f"WHERE UPPER(SYMBOL) IN ({placeholders}) "
        f"AND LISTING_DATE <= %s AND ISSUE_PRICE > 0"
    )
    cursor = conn.cursor(dictionary=True)
    cursor.execute(sql, sym_list + [cutoff_date.strftime('%Y-%m-%d')])
    rows = cursor.fetchall()
    cursor.close()
    return {
        str(r["SYMBOL"]).upper(): float(r["ISSUE_PRICE"])
        for r in rows
        if r["ISSUE_PRICE"] and float(r["ISSUE_PRICE"]) > 0
    }


def filter_symbols_by_turnover(
    symbols: Iterable[str],
    turnover_map: Dict[str, Dict[str, float]],
    min_median_turnover_crores: float,
    min_avg_turnover_42d_crores: float,
    warnings: List[WarningLog],
) -> List[str]:
    eligible: List[str] = []
    for symbol in symbols:
        turnover_info = turnover_map.get(symbol.upper())
        avg_turnover_21d = turnover_info.get("avg_turnover_21d", math.nan) if turnover_info else math.nan
        median_turnover = turnover_info.get("median_turnover_42d", math.nan) if turnover_info else math.nan
        avg_turnover = turnover_info.get("avg_turnover_42d", math.nan) if turnover_info else math.nan
        if pd.isna(median_turnover):
            warnings.append(
                WarningLog(
                    symbol,
                    "Median turnover unavailable; excluded from rating universe.",
                    avg_turnover_21d=None if pd.isna(avg_turnover_21d) else float(avg_turnover_21d),
                    avg_turnover_42d=None if pd.isna(avg_turnover) else float(avg_turnover),
                    median_turnover_42d=None if pd.isna(median_turnover) else float(median_turnover),
                )
            )
            continue
        if pd.isna(avg_turnover):
            warnings.append(
                WarningLog(
                    symbol,
                    "Average 42D turnover unavailable; excluded from rating universe.",
                    avg_turnover_21d=None if pd.isna(avg_turnover_21d) else float(avg_turnover_21d),
                    avg_turnover_42d=None,
                    median_turnover_42d=None if pd.isna(median_turnover) else float(median_turnover),
                )
            )
            continue
        # Hard gate: excluded if BOTH median 42D < min_median AND avg 42D < min_avg.
        # (Either metric above the floor is sufficient to stay in the rated universe.)
        if float(median_turnover) < min_median_turnover_crores and float(avg_turnover) < min_avg_turnover_42d_crores:
            warnings.append(
                WarningLog(
                    symbol,
                    f"Turnover too low: median {round(float(median_turnover), 2)} Cr < {min_median_turnover_crores:.0f} Cr AND avg {round(float(avg_turnover), 2)} Cr < {min_avg_turnover_42d_crores:.0f} Cr; excluded from rating universe.",
                    avg_turnover_21d=None if pd.isna(avg_turnover_21d) else float(avg_turnover_21d),
                    avg_turnover_42d=None if pd.isna(avg_turnover) else float(avg_turnover),
                    median_turnover_42d=float(median_turnover),
                )
            )
            continue
        eligible.append(symbol)
    return eligible


def build_instrument_map(kite, symbols: Iterable[str]) -> Dict[str, Dict[str, object]]:
    print("Calling Kite: instruments('NSE')")
    instruments = pd.DataFrame(kite.instruments("NSE"))
    if instruments.empty:
        raise SystemExit("kite.instruments('NSE') returned no rows.")
    if "segment" in instruments.columns:
        instruments = instruments[instruments["segment"] == "NSE"]
    filtered = instruments[instruments["tradingsymbol"].isin(list(symbols))]
    print(f"Kite instruments mapped: {len(filtered)}/{len(list(symbols))}")
    instrument_map: Dict[str, Dict[str, object]] = {}
    for _, row in filtered.iterrows():
        listing_date = row.get("listing_date")
        parsed_listing_date = None
        if pd.notna(listing_date):
            try:
                parsed_listing_date = pd.to_datetime(listing_date).date()
            except Exception:
                parsed_listing_date = None
        instrument_map[row["tradingsymbol"]] = {
            "instrument_token": int(row["instrument_token"]),
            "listing_date": parsed_listing_date,
        }
    return instrument_map


def fetch_history_with_retry(
    kite,
    instrument_token: int,
    symbol: str,
    start_date: date,
    cutoff_date: date,
    retries: int = 3,
) -> Optional[pd.DataFrame]:
    for attempt in range(1, retries + 1):
        try:
            print(f"Calling Kite: historical_data('{symbol}') attempt {attempt}")
            rows = kite.historical_data(
                instrument_token=instrument_token,
                from_date=datetime.combine(start_date, datetime.min.time()),
                to_date=datetime.combine(cutoff_date, datetime.min.time()),
                interval="day",
                continuous=False,
                oi=False,
            )
            if not rows:
                return None
            df = pd.DataFrame(rows)
            df["date"] = pd.to_datetime(df["date"]).dt.date
            df = df.sort_values("date").set_index("date")
            return df[["open", "high", "low", "close", "volume"]]
        except Exception as exc:
            err_text = str(exc).lower()
            if "invalid token" in err_text or "token is invalid" in err_text:
                print(f"WARNING: historical_data failed for {symbol}: {exc}")
                return None
            if attempt == retries:
                print(f"WARNING: historical_data failed for {symbol}: {exc}")
                return None
            sleep_seconds = attempt * 2
            print(f"WARNING: retrying {symbol} after error: {exc} (sleep {sleep_seconds}s)")
            time.sleep(sleep_seconds)
    return None


def trading_lookback_value(series: pd.Series, lookback: int) -> float:
    clean = series.dropna()
    if clean.empty:
        return math.nan
    if len(clean) > lookback:
        return float(clean.iloc[-(lookback + 1)])
    return float(clean.iloc[0])


def safe_return(current_value: float, prior_value: float) -> float:
    if pd.isna(current_value) or pd.isna(prior_value) or prior_value == 0:
        return math.nan
    return (current_value - prior_value) / prior_value


def compute_atr_percent(df: pd.DataFrame, period: int = ATR_LOOKBACK) -> float:
    if df.empty:
        return math.nan
    prev_close = df["close"].shift(1)
    tr = pd.concat(
        [
            df["high"] - df["low"],
            (df["high"] - prev_close).abs(),
            (df["low"] - prev_close).abs(),
        ],
        axis=1,
    ).max(axis=1)
    atr = tr.rolling(period, min_periods=1).mean()
    current_close = float(df["close"].iloc[-1])
    if current_close == 0:
        return math.nan
    return float((atr.iloc[-1] / current_close) * 100.0)


def first_row_on_or_after(df: pd.DataFrame, target_date: date) -> Optional[pd.Series]:
    subset = df[df.index >= target_date]
    if subset.empty:
        return None
    return subset.iloc[0]


def compute_stock_metrics(
    symbol: str,
    df: pd.DataFrame,
    index_df: pd.DataFrame,
    sector: str,
    cutoff_date: date,
    reset_date: date,
    listing_date: Optional[date],
    turnover_override: Optional[Dict[str, float]],
    liquid_leader_bonus: int,
    issue_price: Optional[float],
    warnings: List[WarningLog],
) -> Optional[Dict[str, object]]:
    history = df[df.index <= cutoff_date].copy()
    if history.empty:
        warnings.append(WarningLog(symbol, "No price history returned by Kite; skipped."))
        return None

    if listing_date is not None:
        listed_days = len(history[history.index >= listing_date])
    else:
        listed_days = len(history)

    current_close = float(history["close"].iloc[-1])
    price_window = history.tail(LOOKBACK_52W)
    if len(history) < LOOKBACK_52W:
        warnings.append(WarningLog(symbol, "Insufficient history (<250 trading days); used available history for 52W high."))

    high_52w = float(price_window["high"].max())
    high_rows = price_window[price_window["high"] == high_52w]
    high_date = high_rows.index[-1]
    days_since_high = len(price_window.loc[high_date:]) - 1
    pct_from_high = ((high_52w - current_close) / high_52w) if high_52w else math.nan

    turnover_42d = (history["close"] * history["volume"]).tail(TURNOVER_LOOKBACK)
    avg_turnover_42d = (
        float(turnover_42d.mean()) / TURNOVER_CRORE_DIVISOR if not turnover_42d.empty else math.nan
    )
    median_turnover_42d = (
        float(turnover_42d.median()) / TURNOVER_CRORE_DIVISOR if not turnover_42d.empty else math.nan
    )
    if turnover_override:
        avg_value = turnover_override.get("avg_turnover_42d", math.nan)
        median_value = turnover_override.get("median_turnover_42d", math.nan)
        if not pd.isna(avg_value):
            avg_turnover_42d = float(avg_value)
        if not pd.isna(median_value):
            median_turnover_42d = float(median_value)

    stock_close_12m = trading_lookback_value(history["close"], LOOKBACK_12M)
    stock_close_6m = trading_lookback_value(history["close"], LOOKBACK_6M)
    stock_close_3m = trading_lookback_value(history["close"], LOOKBACK_3M)
    prev_close = float(history["close"].iloc[-2]) if len(history) >= 2 else math.nan
    ret_12m = safe_return(current_close, stock_close_12m)
    ret_6m = safe_return(current_close, stock_close_6m)
    ret_3m = safe_return(current_close, stock_close_3m)
    ret_1d = safe_return(current_close, prev_close)

    reset_row = first_row_on_or_after(history, reset_date)
    reset_low = float(reset_row["low"]) if reset_row is not None else math.nan
    reset_date_used = reset_row.name if reset_row is not None else None
    reset_recovery = safe_return(current_close, reset_low)

    aligned = history[["close"]].join(index_df[["close"]].rename(columns={"close": "index_close"}), how="inner")
    if aligned.empty:
        warnings.append(WarningLog(symbol, "No overlapping stock/index history; skipped."))
        return None

    index_current_close = float(aligned["index_close"].iloc[-1])
    index_close_12m = trading_lookback_value(aligned["index_close"], LOOKBACK_12M)
    index_close_6m = trading_lookback_value(aligned["index_close"], LOOKBACK_6M)
    index_close_3m = trading_lookback_value(aligned["index_close"], LOOKBACK_3M)
    index_ret_12m = safe_return(index_current_close, index_close_12m)
    index_ret_6m = safe_return(index_current_close, index_close_6m)
    index_ret_3m = safe_return(index_current_close, index_close_3m)

    rs_12m = ret_12m - index_ret_12m if not (pd.isna(ret_12m) or pd.isna(index_ret_12m)) else math.nan
    rs_6m = ret_6m - index_ret_6m if not (pd.isna(ret_6m) or pd.isna(index_ret_6m)) else math.nan
    rs_3m = ret_3m - index_ret_3m if not (pd.isna(ret_3m) or pd.isna(index_ret_3m)) else math.nan

    rs_line = aligned["close"] / aligned["index_close"]
    rs_line_high_52w = float(rs_line.tail(LOOKBACK_52W).max()) if not rs_line.empty else math.nan
    rs_line_at_high = bool(not rs_line.empty and rs_line.iloc[-1] >= rs_line_high_52w)
    if len(rs_line) >= RS_SLOPE_LOOKBACK:
        rs_line_slope_21d = float(rs_line.iloc[-1] - rs_line.iloc[-RS_SLOPE_LOOKBACK])
    else:
        rs_line_slope_21d = math.nan

    atr_percent = compute_atr_percent(history, ATR_LOOKBACK)

    ema8 = history["close"].ewm(span=8, adjust=False).mean()
    ema21 = history["close"].ewm(span=21, adjust=False).mean()
    dma50 = history["close"].rolling(50, min_periods=1).mean()
    ema8_value = float(ema8.iloc[-1]) if not ema8.empty and not pd.isna(ema8.iloc[-1]) else math.nan
    ema21_value = float(ema21.iloc[-1]) if not ema21.empty and not pd.isna(ema21.iloc[-1]) else math.nan
    uptrend_stack = (ema8 > ema21) & (ema21 > dma50)
    reset_history = history[history.index >= reset_date]
    uptrend_lookback = max(UPTREND_CONSISTENCY_MIN_LOOKBACK, len(reset_history))
    uptrend_consistency_pct = float(uptrend_stack.tail(uptrend_lookback).mean()) if not uptrend_stack.empty else math.nan
    ema21_window = history.tail(uptrend_lookback).copy()
    ema21_window_series = ema21.loc[ema21_window.index] if not ema21_window.empty else ema21.iloc[0:0]
    ema21_uptrend_mask = (
        (ema21_window["close"] > ema21_window_series)
        & (ema21_window_series.diff() > 0)
    ).fillna(False) if not ema21_window.empty else pd.Series(dtype=bool)
    ema21_uptrend_pct = float(ema21_uptrend_mask.mean()) if not ema21_uptrend_mask.empty else math.nan
    ema21_uptrend_days = 0
    ema21_uptrend_since = None
    if not ema21_uptrend_mask.empty:
        for is_uptrend in reversed(ema21_uptrend_mask.tolist()):
            if not is_uptrend:
                break
            ema21_uptrend_days += 1
        if ema21_uptrend_days > 0:
            ema21_uptrend_since = ema21_window.index[-ema21_uptrend_days]
    above_50dma = bool(not dma50.empty and not pd.isna(dma50.iloc[-1]) and current_close > float(dma50.iloc[-1]))
    above_21ema = bool(not pd.isna(ema21_value) and current_close > ema21_value)
    above_8ema = bool(not pd.isna(ema8_value) and current_close > ema8_value)
    ema8_slope_5d = (
        float(ema8.iloc[-1] - ema8.iloc[-6])
        if len(ema8) >= 6 and not pd.isna(ema8.iloc[-1]) and not pd.isna(ema8.iloc[-6])
        else math.nan
    )
    ema21_slope_5d = (
        float(ema21.iloc[-1] - ema21.iloc[-6])
        if len(ema21) >= 6 and not pd.isna(ema21.iloc[-1]) and not pd.isna(ema21.iloc[-6])
        else math.nan
    )
    short_trend_bearish = bool(
        not pd.isna(ema8_slope_5d)
        and not pd.isna(ema21_slope_5d)
        and (ema8_slope_5d < 0)
        and (ema21_slope_5d < 0)
        and (not above_21ema)
    )
    short_trend_breakdown = bool(
        short_trend_bearish
        and (not above_8ema)
        and (not pd.isna(ema8_value))
        and (not pd.isna(ema21_value))
        and (ema8_value <= ema21_value)
    )
    reset_dma50 = dma50.loc[reset_history.index] if not reset_history.empty else dma50.iloc[0:0]
    days_below_50dma = int((reset_history["close"] < reset_dma50).sum()) if not reset_history.empty else 0
    green_candle_count = int((reset_history["close"] > reset_history["open"]).sum()) if not reset_history.empty else 0

    spike_window = history.tail(SPIKE_LOOKBACK).copy()
    spike_window["prev_close"] = spike_window["close"].shift(1)
    spike_window["price_change_pct"] = np.where(
        spike_window["prev_close"].fillna(0) != 0,
        ((spike_window["close"] - spike_window["prev_close"]) / spike_window["prev_close"]) * 100.0,
        np.nan,
    )
    spike_base_points = 0
    spike_bonus_points = 0
    spike_total_points = 0
    spike_label = ""
    spike_date = None
    spike_volume = math.nan
    spike_price_change_pct = math.nan
    spike_window_days = math.nan
    spike_top1_threshold = math.nan
    if not spike_window.empty:
        spike_top1_threshold = top_n_threshold(spike_window["volume"], 1)
        spike_candidates = spike_window[spike_window["volume"] >= spike_top1_threshold].copy()
        if not spike_candidates.empty:
            spike_candidates["window_days"] = spike_candidates.index.map(lambda d: len(spike_window.loc[d:]))
            spike_candidates = spike_candidates[spike_candidates["window_days"] <= SPIKE_WINDOW_60]
        if not spike_candidates.empty:
            spike_candidates = spike_candidates.sort_values(["window_days", "volume"], ascending=[True, False])
            spike_row = spike_candidates.iloc[0]
            spike_date = spike_row.name
            spike_volume = float(spike_row["volume"])
            spike_price_change_pct = float(spike_row["price_change_pct"]) if not pd.isna(spike_row["price_change_pct"]) else math.nan
            spike_window_days = int(spike_row["window_days"])
            if spike_window_days <= SPIKE_WINDOW_10:
                spike_base_points = 10
                spike_label = "10D"
            elif spike_window_days <= SPIKE_WINDOW_30:
                spike_base_points = 8
                spike_label = "30D"
            elif spike_window_days <= SPIKE_WINDOW_60:
                spike_base_points = 6
                spike_label = "60D"

            if spike_base_points:
                if not pd.isna(spike_price_change_pct) and spike_price_change_pct > 6:
                    spike_bonus_points = 4
                elif not pd.isna(spike_price_change_pct) and spike_price_change_pct > 3:
                    spike_bonus_points = 2
                spike_total_points = spike_base_points + spike_bonus_points
    if listed_days < 15:
        spike_base_points = 0
        spike_bonus_points = 0
        spike_total_points = 0
        spike_label = ""
        spike_date = None
        spike_volume = math.nan
        spike_price_change_pct = math.nan
        spike_window_days = math.nan

    gapup_window = history.tail(GAPUP_LOOKBACK).copy()
    gapup_window["prev_close"] = gapup_window["close"].shift(1)
    gapup_window["gap_up_pct"] = np.where(
        gapup_window["prev_close"].fillna(0) != 0,
        ((gapup_window["open"] - gapup_window["prev_close"]) / gapup_window["prev_close"]) * 100.0,
        np.nan,
    )
    gapup_window["close_gain_pct"] = np.where(
        gapup_window["prev_close"].fillna(0) != 0,
        ((gapup_window["close"] - gapup_window["prev_close"]) / gapup_window["prev_close"]) * 100.0,
        np.nan,
    )
    gapup_date = None
    gapup_volume = math.nan
    gapup_pct = math.nan
    gapup_close_pct = math.nan
    gapup_top1_threshold = math.nan
    score_gapup = 0
    if not gapup_window.empty:
        gapup_top1_threshold = top_n_threshold(gapup_window["volume"], 1)
        qualifying_gapups = gapup_window[
            (gapup_window["gap_up_pct"] > 3.0)
            & (gapup_window["close_gain_pct"] > 5.0)
            & (
                (gapup_window["close_gain_pct"] >= 9.0)
                | (gapup_window["volume"] >= gapup_top1_threshold)
            )
        ]
        if not qualifying_gapups.empty:
            gapup_row = qualifying_gapups.iloc[-1]
            gapup_date = gapup_row.name
            gapup_volume = float(gapup_row["volume"])
            gapup_pct = float(gapup_row["gap_up_pct"])
            gapup_close_pct = float(gapup_row["close_gain_pct"])
            score_gapup = 6

    score_new_listing = 0
    if listed_days < 30:
        score_new_listing = 4
    elif listed_days < 60:
        score_new_listing = 2

    return {
        "symbol": symbol,
        "sector": sector or "Unknown",
        "listing_date": listing_date,
        "listed_days": listed_days,
        "current_close": current_close,
        "high_52w_close": high_52w,
        "high_52w_date": high_date,
        "pct_from_52w_high": pct_from_high,
        "days_since_52w_high": days_since_high,
        "avg_turnover_42d": avg_turnover_42d,
        "median_turnover_42d": median_turnover_42d,
        "return_12m": ret_12m,
        "return_6m": ret_6m,
        "return_3m": ret_3m,
        "return_1d": ret_1d,
        "reset_date_used": reset_date_used,
        "reset_low": reset_low,
        "reset_recovery": reset_recovery,
        "index_return_12m": index_ret_12m,
        "index_return_6m": index_ret_6m,
        "index_return_3m": index_ret_3m,
        "rs_12m": rs_12m,
        "rs_6m": rs_6m,
        "rs_3m": rs_3m,
        "rs_line_at_52w_high": rs_line_at_high,
        "rs_line_slope_21d": rs_line_slope_21d,
        "atr_percent_21d": atr_percent,
        "ema8_value": ema8_value,
        "ema21_value": ema21_value,
        "above_50dma": above_50dma,
        "above_21ema": above_21ema,
        "above_8ema": above_8ema,
        "ema8_slope_5d": ema8_slope_5d,
        "ema21_slope_5d": ema21_slope_5d,
        "short_trend_bearish": short_trend_bearish,
        "short_trend_breakdown": short_trend_breakdown,
        "uptrend_consistency_pct": uptrend_consistency_pct,
        "uptrend_consistency_lookback": uptrend_lookback,
        "ema21_uptrend_since": ema21_uptrend_since,
        "ema21_uptrend_days": ema21_uptrend_days,
        "ema21_uptrend_pct": ema21_uptrend_pct,
        "green_candle_count": green_candle_count,
        "daysbelow50dma": days_below_50dma,
        "spike_date": spike_date,
        "spike_volume": spike_volume,
        "spike_price_change_pct": spike_price_change_pct / 100.0 if not pd.isna(spike_price_change_pct) else math.nan,
        "spike_window_days": spike_window_days,
        "spike_top1_threshold": spike_top1_threshold,
        "spike_label": spike_label,
        "score_spike_base": spike_base_points,
        "score_spike_bonus": spike_bonus_points,
        "score_spike_total": spike_total_points,
        "gapup_date": gapup_date,
        "gapup_volume": gapup_volume,
        "gapup_pct": gapup_pct / 100.0 if not pd.isna(gapup_pct) else math.nan,
        "gapup_close_pct": gapup_close_pct / 100.0 if not pd.isna(gapup_close_pct) else math.nan,
        "gapup_top1_threshold": gapup_top1_threshold,
        "score_gapup": score_gapup,
        "score_new_listing": score_new_listing,
        "issue_price": issue_price if issue_price and issue_price > 0 else math.nan,
        "ipo_gain": (
            (current_close - issue_price) / issue_price
            if issue_price and issue_price > 0 else math.nan
        ),
        "score_liquid_leaders_bonus": liquid_leader_bonus,
    }


def top_n_threshold(series: pd.Series, pct: float) -> float:
    clean = series.dropna()
    if clean.empty:
        return math.nan
    rank = max(1, math.ceil(len(clean) * (pct / 100.0)))
    return float(clean.sort_values(ascending=False).iloc[rank - 1])


def bottom_n_threshold(series: pd.Series, pct: float) -> float:
    clean = series.dropna()
    if clean.empty:
        return math.nan
    rank = max(1, math.ceil(len(clean) * (pct / 100.0)))
    return float(clean.sort_values(ascending=True).iloc[rank - 1])


def apply_scoring(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()

    thresholds = {
        "turnover_bottom10": bottom_n_threshold(work["avg_turnover_42d"], 10),
        "turnover_bottom20": bottom_n_threshold(work["avg_turnover_42d"], 20),
        "turnover_bottom30": bottom_n_threshold(work["avg_turnover_42d"], 30),
        "median_turnover_bottom10": bottom_n_threshold(work["median_turnover_42d"], 10),
        "median_turnover_bottom20": bottom_n_threshold(work["median_turnover_42d"], 20),
        "ret12_top10": top_n_threshold(work["return_12m"], 10),
        "ret12_top20": top_n_threshold(work["return_12m"], 20),
        "ret12_top30": top_n_threshold(work["return_12m"], 30),
        "ret12_bottom10": bottom_n_threshold(work["return_12m"], 10),
        "ret12_bottom20": bottom_n_threshold(work["return_12m"], 20),
        "ret6_top10": top_n_threshold(work["return_6m"], 10),
        "ret6_top20": top_n_threshold(work["return_6m"], 20),
        "ret6_top30": top_n_threshold(work["return_6m"], 30),
        "ret6_bottom10": bottom_n_threshold(work["return_6m"], 10),
        "ret3_top10": top_n_threshold(work["return_3m"], 10),
        "ret3_top20": top_n_threshold(work["return_3m"], 20),
        "ret3_top30": top_n_threshold(work["return_3m"], 30),
        "ret3_bottom10": bottom_n_threshold(work["return_3m"], 10),
        "ret1_top10": top_n_threshold(work["return_1d"], 10),
        "ret1_top20": top_n_threshold(work["return_1d"], 20),
        "ret1_bottom20": bottom_n_threshold(work["return_1d"], 20),
        "rs12_top10": top_n_threshold(work["rs_12m"], 10),
        "rs12_top30": top_n_threshold(work["rs_12m"], 30),
        "rs6_top10": top_n_threshold(work["rs_6m"], 10),
        "rs6_top30": top_n_threshold(work["rs_6m"], 30),
        "rs3_top10": top_n_threshold(work["rs_3m"], 10),
        "rs3_top30": top_n_threshold(work["rs_3m"], 30),
        "atr_bottom10": bottom_n_threshold(work["atr_percent_21d"], 10),
        "uptrend_consistency_top30": top_n_threshold(work["uptrend_consistency_pct"], 30),
        "green_candle_top10": top_n_threshold(work["green_candle_count"], 10),
        "green_candle_top20": top_n_threshold(work["green_candle_count"], 20),
    }

    eligible_52w = work["listed_days"] > 1
    within_10 = eligible_52w & (work["pct_from_52w_high"] <= 0.10)
    within_15 = eligible_52w & (work["pct_from_52w_high"] > 0.10) & (work["pct_from_52w_high"] <= 0.15)
    within_20 = eligible_52w & (work["pct_from_52w_high"] > 0.15) & (work["pct_from_52w_high"] <= 0.20)
    off_20 = eligible_52w & (work["pct_from_52w_high"] > 0.20) & (work["pct_from_52w_high"] < 0.25)
    off_25 = eligible_52w & (work["pct_from_52w_high"] >= 0.25) & (work["pct_from_52w_high"] < 0.30)
    off_30 = eligible_52w & (work["pct_from_52w_high"] >= 0.30)
    recent_10 = eligible_52w & (work["days_since_52w_high"] <= 10)

    # Distance is the base 52W score and also carries the negative penalty once a stock
    # rolls over too far from the high.
    work["score_52w_price"] = np.select(
        [
            within_10,
            within_15,
            within_20,
            off_20,
            off_25,
            off_30,
        ],
        [10, 6, 2, -4, -10, -16],
        default=0,
    )
    # Recency only matters if the stock is still within 20% of the 52W high, which
    # avoids the old stacked recency/bonus behavior for names already well off the top.
    work["score_52w_recency"] = np.select(
        [
            recent_10 & within_10,
            recent_10 & within_15,
            recent_10 & within_20,
        ],
        [2, 2, 2],
        default=0,
    )
    work["score_52w_bonus"] = 0
    work["score_52w_total"] = (
        work["score_52w_price"] + work["score_52w_recency"] + work["score_52w_bonus"]
    )
    work["score_above_50dma"] = np.where(work["above_50dma"], 4, -4)
    work["score_above_21ema"] = np.where(work["above_21ema"], 2, 0)
    work["score_above_8ema"] = np.where(work["above_8ema"], 2, 0)

    _liq_low = work["avg_turnover_42d"] < 40.0   # penalty only applies when avg TO < 40 Cr
    work["score_liquidity"] = np.select(
        [
            _liq_low & (work["avg_turnover_42d"] <= thresholds["turnover_bottom10"]),
            _liq_low & (work["avg_turnover_42d"] <= thresholds["turnover_bottom20"]),
            _liq_low & (work["avg_turnover_42d"] <= thresholds["turnover_bottom30"]),
        ],
        [-8, -6, -4],
        default=0,
    )
    work["score_median_turnover"] = np.select(
        [
            (work["median_turnover_42d"] <= thresholds["median_turnover_bottom10"])
            & (work["median_turnover_42d"] < MEDIAN_TURNOVER_LOW_THRESHOLD_CRORES),
            (work["median_turnover_42d"] <= thresholds["median_turnover_bottom20"])
            & (work["median_turnover_42d"] < MEDIAN_TURNOVER_LOW_THRESHOLD_CRORES),
        ],
        [-8, -6],
        default=0,
    )

    top12_10 = work["return_12m"] >= thresholds["ret12_top10"]
    top12_20 = work["return_12m"] >= thresholds["ret12_top20"]
    top12_30 = work["return_12m"] >= thresholds["ret12_top30"]
    top6_10 = work["return_6m"] >= thresholds["ret6_top10"]
    top6_20 = work["return_6m"] >= thresholds["ret6_top20"]
    top6_30 = work["return_6m"] >= thresholds["ret6_top30"]
    top3_10 = work["return_3m"] >= thresholds["ret3_top10"]
    top3_20 = work["return_3m"] >= thresholds["ret3_top20"]
    top3_30 = work["return_3m"] >= thresholds["ret3_top30"]
    top1d_10 = work["return_1d"] >= thresholds["ret1_top10"]
    top1d_20 = work["return_1d"] >= thresholds["ret1_top20"]
    mature_listing = work["listed_days"] >= LOOKBACK_3M
    full_12m_history = work["listed_days"] >= LOOKBACK_12M   # must have full year to be penalised
    bottom12_10 = full_12m_history & (work["return_12m"] <= thresholds["ret12_bottom10"])
    bottom12_20 = full_12m_history & (work["return_12m"] <= thresholds["ret12_bottom20"])
    bottom6_10 = mature_listing & (work["return_6m"] <= thresholds["ret6_bottom10"])
    bottom3_10 = mature_listing & (work["return_3m"] <= thresholds["ret3_bottom10"])
    bottom1d_20 = (work["return_1d"] <= thresholds["ret1_bottom20"]) & (work["return_1d"] < -0.03)

    # Per-stock flag used later for the sector 12M+52W concentration bonus:
    # True when a stock is in the top-30% of 12M return AND within 20% of its 52W high.
    work["top30_12m_within20_52w"] = (
        top12_30 & eligible_52w & (work["pct_from_52w_high"] <= 0.20)
    )

    work["score_perf_12m"] = np.select([top12_10, top12_20, top12_30], [8, 6, 4], default=0)
    work["score_perf_6m"] = np.select([top6_10, top6_20, top6_30], [6, 4, 2], default=0)
    work["score_perf_3m"] = np.select([top3_10, top3_20, top3_30], [6, 4, 2], default=0)
    work["score_perf_1d"] = np.select([top1d_10, top1d_20], [4, 2], default=0)
    work["score_perf_12m_penalty"] = np.select(
        [bottom12_10, bottom12_20],
        [-6, -4],
        default=0,
    )
    work["score_perf_6m_penalty"] = np.where(bottom6_10, -2, 0)
    work["score_perf_3m_penalty"] = np.where(bottom3_10, -4, 0)
    work["score_perf_1d_penalty"] = np.where(bottom1d_20, -4, 0)
    work["score_perf_bonus"] = np.where(
        top12_10
        & (work["days_since_52w_high"] <= 10)
        & (work["pct_from_52w_high"] <= 0.20),
        2,
        0,
    )
    # 12M top-30% AND within 15% of 52W high bonus
    work["score_12m_52w_bonus"] = np.where(
        top12_30 & (work["pct_from_52w_high"] <= 0.15),
        4,
        0,
    )
    reset_rank = work["reset_recovery"].rank(method="min", ascending=False)
    work["score_reset_recovery"] = np.select(
        [
            reset_rank <= 10,
            reset_rank <= 20,
        ],
        [4, 2],
        default=0,
    )
    recent_listing_mask = work["listed_days"] < 60
    work.loc[recent_listing_mask, "score_perf_3m"] = np.maximum(work.loc[recent_listing_mask, "score_perf_3m"], 2)
    work.loc[recent_listing_mask, "score_reset_recovery"] = np.maximum(work.loc[recent_listing_mask, "score_reset_recovery"], 1)

    def median_score(series: pd.Series) -> float:
        clean = pd.to_numeric(series, errors="coerce").dropna()
        if clean.empty:
            return 0.0
        return float(clean.median())

    missing_12m_history_mask = work["listed_days"] < LOOKBACK_12M
    missing_6m_history_mask = work["listed_days"] < LOOKBACK_6M
    median_12m_score = median_score(work.loc[~missing_12m_history_mask, "score_perf_12m"])
    median_6m_score = median_score(work.loc[~missing_6m_history_mask, "score_perf_6m"])

    # Median substitution for insufficient history — but preserve top-tier points:
    # a stock with < full 12M history that still ranks top 10/20/30% keeps its earned score.
    apply_12m_median = missing_12m_history_mask & ~(top12_10 | top12_20 | top12_30)
    work.loc[apply_12m_median, "score_perf_12m"] = median_12m_score
    work.loc[missing_6m_history_mask, "score_perf_6m"] = median_6m_score
    work.loc[missing_12m_history_mask, "score_perf_bonus"] = 0
    work.loc[missing_12m_history_mask, "score_12m_52w_bonus"] = 0
    work.loc[missing_6m_history_mask, "score_perf_6m_penalty"] = 0
    work["score_performance_total"] = (
        work["score_perf_12m"]
        + work["score_perf_12m_penalty"]
        + work["score_perf_6m"]
        + work["score_perf_3m"]
        + work["score_perf_1d"]
        + work["score_perf_6m_penalty"]
        + work["score_perf_3m_penalty"]
        + work["score_perf_1d_penalty"]
        + work["score_perf_bonus"]
        + work["score_12m_52w_bonus"]
        + work["score_reset_recovery"]
    )

    top_rs12_10 = work["rs_12m"] >= thresholds["rs12_top10"]
    top_rs12_30 = work["rs_12m"] >= thresholds["rs12_top30"]
    top_rs6_10 = work["rs_6m"] >= thresholds["rs6_top10"]
    top_rs6_30 = work["rs_6m"] >= thresholds["rs6_top30"]
    top_rs3_10 = work["rs_3m"] >= thresholds["rs3_top10"]
    top_rs3_30 = work["rs_3m"] >= thresholds["rs3_top30"]

    work["score_rs_12m"] = np.select([top_rs12_10, top_rs12_30], [6, 3], default=0)
    work["score_rs_6m"] = np.select([top_rs6_10, top_rs6_30], [4, 2], default=0)
    work["score_rs_3m"] = np.select([top_rs3_10, top_rs3_30], [4, 2], default=0)
    work["score_rs_line_high"] = np.where(work["rs_line_at_52w_high"], 2, 0)
    work["score_rs_slope_penalty"] = np.where(work["rs_line_slope_21d"] < 0, -2, 0)
    work["score_rs_total"] = (
        work["score_rs_12m"]
        + work["score_rs_6m"]
        + work["score_rs_3m"]
        + work["score_rs_line_high"]
        + work["score_rs_slope_penalty"]
    )

    work["score_volatility"] = np.select(
        [
            (work["atr_percent_21d"] <= thresholds["atr_bottom10"]) & (work["atr_percent_21d"] < 3.0),
            (work["atr_percent_21d"] <= thresholds["atr_bottom10"]) & (work["atr_percent_21d"] < 4.0),
        ],
        [-6, -3],
        default=0,
    )
    work["score_uptrend_consistency"] = np.where(
        work["uptrend_consistency_pct"] >= thresholds["uptrend_consistency_top30"],
        4,
        0,
    )
    work.loc[recent_listing_mask, "score_uptrend_consistency"] = np.maximum(
        work.loc[recent_listing_mask, "score_uptrend_consistency"], 1
    )
    work["score_green_candles"] = np.select(
        [
            work["green_candle_count"] >= thresholds["green_candle_top10"],
            work["green_candle_count"] >= thresholds["green_candle_top20"],
        ],
        [4, 2],
        default=0,
    )
    work["score_ema21_uptrend"] = np.select(
        [
            work["ema21_uptrend_days"] >= 20,
            work["ema21_uptrend_days"] >= 10,
            work["ema21_uptrend_pct"] >= 0.90,
            work["ema21_uptrend_pct"] >= 0.80,
        ],
        [4, 2, 2, 1],
        default=0,
    )
    work["score_daysbelow50dma_penalty"] = 0

    def median_or_default(series: pd.Series, default_value: float = 0.0) -> float:
        clean = pd.to_numeric(series, errors="coerce").dropna()
        if clean.empty:
            return default_value
        return float(clean.median())

    recent_listing_trend_mask = work["listed_days"] < 60
    mature_trend_mask = ~recent_listing_trend_mask
    median_score_50dma = median_or_default(work.loc[mature_trend_mask, "score_above_50dma"], 0.0)
    median_score_21ema = median_or_default(work.loc[mature_trend_mask, "score_above_21ema"], 0.0)
    median_score_8ema = median_or_default(work.loc[mature_trend_mask, "score_above_8ema"], 0.0)
    median_score_uptrend = median_or_default(work.loc[mature_trend_mask, "score_uptrend_consistency"], 0.0)
    median_score_ema21_uptrend = median_or_default(work.loc[mature_trend_mask, "score_ema21_uptrend"], 0.0)
    median_score_green = median_or_default(work.loc[mature_trend_mask, "score_green_candles"], 0.0)

    work.loc[recent_listing_trend_mask, "score_above_50dma"] = median_score_50dma
    work.loc[recent_listing_trend_mask, "score_above_21ema"] = median_score_21ema
    work.loc[recent_listing_trend_mask, "score_above_8ema"] = median_score_8ema
    work.loc[recent_listing_trend_mask, "score_uptrend_consistency"] = median_score_uptrend
    work.loc[recent_listing_trend_mask, "score_ema21_uptrend"] = median_score_ema21_uptrend
    work.loc[recent_listing_trend_mask, "score_green_candles"] = median_score_green
    # IPO performance bonus — stocks listed < 6M with CMP above issue price.
    # Keep the original top-10 / top-20 scoring and add an extra +3
    # for names in the top 40 percentile with at least 20% IPO gain.
    ipo_sub = (work["listed_days"] < LOOKBACK_6M) & (work["ipo_gain"] > 0)
    if ipo_sub.sum() >= 2:
        ipo_gains = work.loc[ipo_sub, "ipo_gain"]
        ipo_top10_t = top_n_threshold(ipo_gains, 10)
        ipo_top20_t = top_n_threshold(ipo_gains, 20)
        ipo_top40_t = top_n_threshold(ipo_gains, 40)
        base_ipo_score = np.select(
            [
                ipo_sub & (work["ipo_gain"] >= ipo_top10_t),
                ipo_sub & (work["ipo_gain"] >= ipo_top20_t),
            ],
            [IPO_TOP10_POINTS, IPO_TOP20_POINTS],
            default=0,
        )
        extra_ipo_bonus = np.where(
            ipo_sub & (work["ipo_gain"] >= ipo_top40_t) & (work["ipo_gain"] >= IPO_MIN_GAIN),
            IPO_TOP40_POINTS,
            0,
        )
        work["score_ipo_perf"] = base_ipo_score + extra_ipo_bonus
    else:
        work["score_ipo_perf"] = 0

    work["pre_sector_score"] = (
        work["score_52w_total"]
        + work["score_liquidity"]
        + work["score_median_turnover"]
        + work["score_performance_total"]
        + work["score_rs_total"]
        + work["score_volatility"]
        + work["score_above_50dma"]
        + work["score_above_21ema"]
        + work["score_above_8ema"]
        + work["score_uptrend_consistency"]
        + work["score_ema21_uptrend"]
        + work["score_green_candles"]
        + work["score_spike_total"]
        + work["score_gapup"]
        + work["score_new_listing"]
        + work["score_liquid_leaders_bonus"]
        + work["score_ipo_perf"]
    )
    top_quartile_threshold = top_n_threshold(work["pre_sector_score"], 25)
    work["is_top_quartile_pre_sector"] = work["pre_sector_score"] >= top_quartile_threshold
    work["pre_sector_rank"] = work["pre_sector_score"].rank(method="min", ascending=False)

    sector_rows: List[Dict[str, object]] = []
    for sector_name, grp in work.groupby("sector"):
        ordered = grp.sort_values(
            ["pre_sector_score", "score_rs_total", "score_performance_total", "symbol"],
            ascending=[False, False, False, True],
        ).reset_index(drop=True)
        stock_count = int(len(ordered))
        positive = ordered[ordered["pre_sector_score"] > 0].copy()
        positive_count = int(len(positive))
        top_quartile_count = int(ordered["is_top_quartile_pre_sector"].sum())
        strength_ratio = (top_quartile_count / stock_count) if stock_count > 0 else 0.0
        weighted_pre = float(positive["pre_sector_score"].mean()) if positive_count > 0 else math.nan
        weighted_rs = float(positive["score_rs_total"].mean()) if positive_count > 0 else math.nan
        weighted_perf = float(positive["score_performance_total"].mean()) if positive_count > 0 else math.nan
        top20_hits = int((positive["pre_sector_rank"] <= 20).sum()) if positive_count > 0 else 0
        # Count stocks that qualify for the 12M+52W proximity concentration bonus.
        qualifying_12m_52w = int(ordered["top30_12m_within20_52w"].sum())
        sector_rows.append(
            {
                "sector": sector_name,
                "sector_stock_count": stock_count,
                "sector_positive_score_count": positive_count,
                "sector_top_quartile_count": top_quartile_count,
                "sector_strength_ratio": strength_ratio if stock_count > 1 else -1.0,
                "sector_weighted_pre_score": weighted_pre,
                "sector_weighted_rs_score": weighted_rs,
                "sector_weighted_perf_score": weighted_perf,
                "sector_top20_hits": top20_hits,
                "sector_12m_52w_qualifying_count": qualifying_12m_52w,
            }
        )

    sector_stats = pd.DataFrame(sector_rows)
    eligible_sector_mask = sector_stats["sector_positive_score_count"] > 1

    # -- Top-20 concentration bonuses ----------------------------------------
    # Breadth bonus: how many of this sector's stocks landed in the Top-20.
    def _breadth_bonus(n: int) -> int:
        if n >= 3:  return SECTOR_TOP20_BREADTH_BONUS_3PLUS
        if n == 2:  return SECTOR_TOP20_BREADTH_BONUS_2
        if n == 1:  return SECTOR_TOP20_BREADTH_BONUS_1
        return 0

    # Penetration bonus: fraction of sector's total stocks that are in Top-20.
    def _penetration_bonus(hits: int, total: int) -> int:
        if total == 0 or hits == 0:
            return 0
        pct = hits / total
        if pct >= 1.0 and total >= 2:  return SECTOR_TOP20_PENETRATION_BONUS_100
        if pct >= 0.67:                return SECTOR_TOP20_PENETRATION_BONUS_67
        if pct >= 0.50:                return SECTOR_TOP20_PENETRATION_BONUS_50
        return 0

    sector_stats["sector_top20_breadth_bonus"] = sector_stats["sector_top20_hits"].apply(
        lambda n: _breadth_bonus(int(n))
    )
    sector_stats["sector_top20_penetration_bonus"] = sector_stats.apply(
        lambda r: _penetration_bonus(int(r["sector_top20_hits"]), int(r["sector_stock_count"])),
        axis=1,
    )

    # -- 12M return + 52W high proximity concentration bonus -----------------
    # Only eligible for sectors with >= 2 stocks (enforced via eligible_sector_mask).
    # >3 qualifying stocks → base 6 + 1 per qualifying stock
    # >2 qualifying stocks → base 4 + 1 per qualifying stock
    def _12m_52w_bonus(count: int) -> int:
        if count > 3:
            return SECTOR_12M_52W_GT3_BASE + count
        if count > 2:
            return SECTOR_12M_52W_GT2_BASE + count
        return 0

    sector_stats["sector_12m_52w_bonus"] = sector_stats["sector_12m_52w_qualifying_count"].apply(
        lambda n: _12m_52w_bonus(int(n))
    )

    if eligible_sector_mask.any():
        sector_stats["sector_leadership_score"] = np.where(
            eligible_sector_mask,
            sector_stats["sector_weighted_pre_score"]
            + sector_stats["sector_top20_breadth_bonus"]
            + sector_stats["sector_top20_penetration_bonus"]
            + sector_stats["sector_12m_52w_bonus"],
            -1.0,
        )
    else:
        sector_stats["sector_leadership_score"] = -1.0

    sector_top10 = top_n_threshold(sector_stats.loc[eligible_sector_mask, "sector_leadership_score"], 10)
    sector_top20 = top_n_threshold(sector_stats.loc[eligible_sector_mask, "sector_leadership_score"], 20)
    sector_top30 = top_n_threshold(sector_stats.loc[eligible_sector_mask, "sector_leadership_score"], 30)
    sector_stats["score_sector"] = np.select(
        [
            eligible_sector_mask & (sector_stats["sector_leadership_score"] >= sector_top10),
            eligible_sector_mask & (sector_stats["sector_leadership_score"] >= sector_top20),
            eligible_sector_mask & (sector_stats["sector_leadership_score"] >= sector_top30),
        ],
        [SECTOR_TOP10_POINTS, SECTOR_TOP20_POINTS, SECTOR_TOP30_POINTS],
        default=0,
    )
    work = work.merge(sector_stats, on="sector", how="left")
    work["composite_score"] = work["pre_sector_score"] + work["score_sector"]
    work["rank"] = work["composite_score"].rank(method="min", ascending=False).astype(int)
    work = work.sort_values(["composite_score", "symbol"], ascending=[False, True]).reset_index(drop=True)

    for name, value in thresholds.items():
        work[name] = value
    work["sector_top10_threshold"] = sector_top10
    work["sector_top20_threshold"] = sector_top20
    work["sector_top30_threshold"] = sector_top30
    work["top_quartile_score_threshold"] = top_quartile_threshold
    return work


def percent_or_blank(value: object) -> object:
    if pd.isna(value):
        return ""
    return round(float(value) * 100.0, 2)


def date_or_blank(value: object) -> object:
    if pd.isna(value) or value is None:
        return ""
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def round_or_blank(value: object, digits: int = 2) -> object:
    if pd.isna(value):
        return ""
    if isinstance(value, (int, float, np.integer, np.floating)):
        return round(float(value), digits)
    return value


def score_or_blank(value: object) -> object:
    if pd.isna(value):
        return ""
    if isinstance(value, (int, float, np.integer, np.floating)):
        return int(round(float(value)))
    return value


def smart_num_fmt(decimals: int = 2) -> str:
    return "#,##0" if decimals <= 0 else "#,##0." + ("#" * decimals)


def force_close_excel_workbook(path: Path) -> None:
    """
    Deletes path.  If the file is locked by Excel (PermissionError):
      1. Tries win32com to close just that workbook.
      2. Falls back to PowerShell COM automation.
      3. Retries the delete.
      4. Hard-exits with a clear message if still locked.
    """
    fname = str(path)
    if not path.exists():
        return
    try:
        os.remove(fname)
        print(f"  Deleted old file: {path.name}")
        return
    except PermissionError:
        pass

    print(f"  File is open in Excel — attempting to close it: {path.name}")

    # ── Attempt 1: win32com ───────────────────────────────────────
    closed = False
    try:
        import win32com.client  # type: ignore
        xl = win32com.client.GetActiveObject("Excel.Application")
        for wb in list(xl.Workbooks):
            if os.path.normcase(wb.FullName) == os.path.normcase(fname):
                wb.Close(SaveChanges=False)
                closed = True
                print("  Closed workbook via Excel COM.")
                break
    except Exception:
        pass

    # ── Attempt 2: PowerShell COM fallback ────────────────────────
    if not closed:
        try:
            ps = (
                "$xl = [Runtime.InteropServices.Marshal]"
                "::GetActiveObject('Excel.Application'); "
                f"$xl.Workbooks | Where-Object {{ $_.FullName -eq '{fname}' }}"
                " | ForEach-Object { $_.Close($false) }"
            )
            subprocess.run(["powershell", "-Command", ps],
                           capture_output=True, timeout=10)
            closed = True
            print("  Closed workbook via PowerShell.")
        except Exception:
            pass

    # ── Retry delete ──────────────────────────────────────────────
    time.sleep(0.5)
    try:
        os.remove(fname)
        print(f"  Deleted old file: {path.name}")
    except PermissionError:
        print(f"\n  ERROR: Still cannot delete {fname}.")
        print("  Please close Excel completely and re-run.")
        sys.exit(1)


def export_tradingview_dayone(scored: pd.DataFrame, reports_dir: Path, cutoff_date: date) -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"rated_list_{cutoff_date.strftime('%d%b%Y').lower()}.txt"

    top_rows = scored.head(20).copy()
    if "sector" not in top_rows.columns:
        top_rows["sector"] = "Unknown"
    top_rows["sector"] = (
        top_rows["sector"]
        .astype(str)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
        .fillna("Unknown")
    )

    sectors: Dict[str, List[str]] = {}
    for _, row in top_rows.iterrows():
        sector = str(row.get("sector", "Unknown")).strip() or "Unknown"
        symbol = str(row.get("symbol", "")).strip().upper().replace("NSE:", "")
        if not symbol:
            continue
        sectors.setdefault(sector, []).append(f"NSE:{symbol}")

    lines = [
        f"### Rated List - {cutoff_date.strftime('%d %b %Y')} (Top {len(top_rows)})",
        "",
    ]
    for sector in sorted(sectors.keys()):
        lines.append(f"###{sector}")
        lines.extend(sectors[sector])
        lines.append("")

    out_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return out_path


def robust_force_close_excel_workbook(path: Path) -> None:
    fname = str(path)
    if not path.exists():
        return

    try:
        os.remove(fname)
        print(f"  Deleted old file: {path.name}")
        return
    except PermissionError:
        pass

    print(f"  File is open in Excel - force closing it: {path.name}")
    resolved_target = os.path.normcase(str(path.resolve()))

    try:
        import win32com.client  # type: ignore

        xl = win32com.client.GetActiveObject("Excel.Application")
        for wb in list(xl.Workbooks):
            try:
                wb_fullname = os.path.normcase(str(Path(str(wb.FullName)).resolve()))
            except Exception:
                wb_fullname = os.path.normcase(str(wb.FullName))
            if wb_fullname == resolved_target:
                wb.Close(SaveChanges=False)
                print("  Closed workbook via Excel COM.")
                break
        try:
            if xl.Workbooks.Count == 0:
                xl.Quit()
        except Exception:
            pass
    except Exception:
        pass

    try:
        escaped = fname.replace("'", "''")
        ps = (
            f"$target = [System.IO.Path]::GetFullPath('{escaped}'); "
            "$xl = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application'); "
            "$matched = @($xl.Workbooks | Where-Object { "
            "[System.IO.Path]::GetFullPath($_.FullName) -eq $target }); "
            "$matched | ForEach-Object { $_.Close($false) }; "
            "if ($xl.Workbooks.Count -eq 0) { $xl.Quit() }"
        )
        subprocess.run(
            ["powershell", "-Command", ps],
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
        print("  Closed workbook via PowerShell.")
    except Exception:
        pass

    for _ in range(6):
        time.sleep(0.5)
        try:
            os.remove(fname)
            print(f"  Deleted old file: {path.name}")
            return
        except PermissionError:
            continue

    print("  Workbook still locked - force closing Excel.")
    try:
        subprocess.run(
            ["taskkill", "/F", "/IM", "EXCEL.EXE"],
            capture_output=True,
            text=True,
            timeout=15,
            check=False,
        )
    except Exception:
        pass

    for _ in range(8):
        time.sleep(0.5)
        try:
            os.remove(fname)
            print(f"  Deleted old file: {path.name}")
            return
        except PermissionError:
            continue

    print(f"\n  ERROR: Still cannot delete {fname}.")
    print("  Please close Excel completely and re-run.")
    sys.exit(1)


def export_workbook(
    scored: pd.DataFrame,
    warnings: List[WarningLog],
    output_dir: Path,
    cutoff_date: date,
    reset_date: date,
    symbol_file: Path,
) -> Path:  # noqa: C901
    # ── imports needed only here ──────────────────────────────────────
    from openpyxl.styles import Border, Side
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"stock_rating_{cutoff_date.strftime('%d%b%Y')}.xlsx"
    robust_force_close_excel_workbook(out_path)

    # ── palette ───────────────────────────────────────────────────────
    T_FILL   = PatternFill("solid", fgColor="1F3864")   # title bar — deep navy
    H_FILL   = PatternFill("solid", fgColor="2E4B8F")   # column header — cobalt
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")   # alternating row — pale blue
    WHT_FILL = PatternFill("solid", fgColor="FFFFFF")

    # Score-band row fills + fonts  (fill, font-colour)
    # All bands use white text on saturated backgrounds — maximum readability
    SCORE_BAND = [
        (28, PatternFill("solid", fgColor="C6E0B4"), "1B4332"),
        (22, PatternFill("solid", fgColor="D9EAD3"), "1B4332"),
        (14, PatternFill("solid", fgColor="FFF2CC"), "7F6000"),
        ( 7, PatternFill("solid", fgColor="FCE4D6"), "9E480E"),
        ( 0, PatternFill("solid", fgColor="F4CCCC"), "9C0006"),
        (-99,PatternFill("solid", fgColor="F4CCCC"), "9C0006"),
    ]
    def score_band(score):
        for threshold, fill, fc in SCORE_BAND:
            if score >= threshold:
                return fill, fc
        return SCORE_BAND[-1][1], SCORE_BAND[-1][2]

    # Section header fills — one per scoring group
    SEC_FILLS = {
        "id":       PatternFill("solid", fgColor="1F3864"),
        "52w":      PatternFill("solid", fgColor="1A5276"),
        "liq":      PatternFill("solid", fgColor="145A32"),
        "perf":     PatternFill("solid", fgColor="784212"),
        "rs":       PatternFill("solid", fgColor="6C3483"),
        "trend":    PatternFill("solid", fgColor="0E6655"),
        "spike":    PatternFill("solid", fgColor="7D6608"),
        "gapup":    PatternFill("solid", fgColor="7D6608"),
        "vol":      PatternFill("solid", fgColor="922B21"),
        "sector":   PatternFill("solid", fgColor="2C3E50"),
        "total":    PatternFill("solid", fgColor="1F3864"),
    }
    # Column header fills (lighter tints)
    COL_FILLS = {
        "id":       PatternFill("solid", fgColor="2E4B8F"),
        "52w":      PatternFill("solid", fgColor="1F618D"),
        "liq":      PatternFill("solid", fgColor="1E8449"),
        "perf":     PatternFill("solid", fgColor="A04000"),
        "rs":       PatternFill("solid", fgColor="884EA0"),
        "trend":    PatternFill("solid", fgColor="117A65"),
        "spike":    PatternFill("solid", fgColor="9A7D0A"),
        "gapup":    PatternFill("solid", fgColor="9A7D0A"),
        "vol":      PatternFill("solid", fgColor="B03A2E"),
        "sector":   PatternFill("solid", fgColor="2C3E50"),
        "total":    PatternFill("solid", fgColor="1F3864"),
    }

    POS_FILL  = PatternFill("solid", fgColor="D5F5E3")   # score cell — positive
    NEG_FILL  = PatternFill("solid", fgColor="FADBD8")   # score cell — negative
    ZERO_FILL = PatternFill("solid", fgColor="F4F6F7")   # score cell — zero
    YES_FILL  = PatternFill("solid", fgColor="D5F5E3")
    NO_FILL   = PatternFill("solid", fgColor="FADBD8")
    GOLD_FILL = PatternFill("solid", fgColor="FFD700")
    SILV_FILL = PatternFill("solid", fgColor="C0C0C0")
    BRNZ_FILL = PatternFill("solid", fgColor="CD7F32")

    thin  = Side(style="thin",   color="CCCCCC")
    med   = Side(style="medium", color="888888")
    BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
    BDR_M = Border(left=med,  right=med,  top=med,  bottom=med)

    def sc(cell, bold=False, italic=False, color="000000", fill=None,
           align="center", size=11, wrap=False, border=BDR, num_fmt=None):
        cell.font      = Font(bold=bold, italic=italic, color=color,
                              name="Arial", size=size)
        if fill:
            cell.fill  = fill
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   wrap_text=wrap)
        cell.border    = border
        if num_fmt:
            cell.number_format = num_fmt

    def score_cell(cell, val, bold=True, size=10):
        """Colour a numeric score cell green/red/grey and write value."""
        cell.value = val
        if isinstance(val, (int, float, np.integer, np.floating)):
            f = POS_FILL if val > 0 else NEG_FILL if val < 0 else ZERO_FILL
            c = "1F6B3A" if val > 0 else "8B0000" if val < 0 else "555555"
        else:
            f, c = ZERO_FILL, "555555"
        sc(cell, bold=bold, color=c, fill=f, size=size)

    def pct_fmt(v):
        if pd.isna(v):
            return ""
        return round(float(v), 4)
    def dt_fmt(v):  return date_or_blank(v)
    def r2(v):      return round_or_blank(v, 2)

    wb = Workbook()
    wb.remove(wb.active)   # sheets added in order below

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 1 ── 🏆 DASHBOARD  (Top 20)
    # ══════════════════════════════════════════════════════════════════
    dash = wb.create_sheet("🏆 Dashboard")
    top20 = scored.head(20).reset_index(drop=True)
    ncols_d = 14

    dash.merge_cells(f"A1:{get_column_letter(ncols_d)}1")
    dash["A1"] = (
        f"NSE STOCK RATING  —  TOP 20 DASHBOARD   |   "
        f"As of {cutoff_date.strftime('%d-%b-%Y')}   |   "
        f"Reset: {reset_date.strftime('%d-%b-%Y')}   |   "
        f"{len(scored)} stocks rated"
    )
    sc(dash["A1"], bold=True, color="FFFFFF", fill=T_FILL, size=13, align="center")
    dash.row_dimensions[1].height = 28

    d_hdrs = [
        "#", "Symbol", "Sector", "Score", "Rating",
        f"Close\n{cutoff_date.strftime('%d-%b')}",
        "Avg TO\n42D (Cr)", "Med TO\n42D (Cr)",
        "21EMA\nUp Sc", "Uptrend\nCon", "Spike\nScore", "Gap-Up\nScore",
        "Sector\nScore", "Criteria Met",
    ]
    for ci, h in enumerate(d_hdrs, 1):
        cell = dash.cell(row=2, column=ci, value=h)
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=11, wrap=True)
    dash.row_dimensions[2].height = 30

    medal_fills = [GOLD_FILL, SILV_FILL, BRNZ_FILL]
    cmap_keys = [
        ("score_52w_total",          "52W High Proximity/Recency"),
        ("score_liquidity",          "Liquidity"),
        ("score_median_turnover",    "Median Turnover"),
        ("score_performance_total",  "Performance (12M/6M/3M)"),
        ("score_rs_total",           "Relative Strength"),
        ("score_above_50dma",        "Above 50DMA"),
        ("score_above_21ema",        "Above 21EMA"),
        ("score_above_8ema",         "Above 8EMA"),
        ("score_uptrend_consistency","Uptrend Consistency"),
        ("score_ema21_uptrend",      "21EMA Uptrend"),
        ("score_green_candles",      "Green Candles"),
        ("score_spike_total",        "Volume Spike"),
        ("score_gapup",              "Gap-Up"),
        ("score_volatility",         "Low Volatility"),
        ("score_sector",             "Sector Strength"),
        ("score_new_listing",        "New Listing"),
        ("score_ipo_perf",           "IPO Performance"),
        ("score_liquid_leaders_bonus","Liquid Leaders"),
    ]
    for ri, (_, row) in enumerate(top20.iterrows(), start=3):
        rank  = ri - 2
        score = score_or_blank(row["composite_score"])
        s_fill, s_fc = score_band(row["composite_score"])
        row_bg = medal_fills[rank - 1] if rank <= 3 else (
            ALT_FILL if rank % 2 == 0 else WHT_FILL)

        rating = row.get("rating", "")
        met    = [lbl for key, lbl in cmap_keys if (row.get(key, 0) or 0) > 0]

        vals = [
            rank,
            row["symbol"],
            row.get("sector", ""),
            score,
            rating,
            r2(row["current_close"]),
            r2(row["avg_turnover_42d"]),
            r2(row["median_turnover_42d"]),
            round_or_blank(row["score_ema21_uptrend"]),
            round_or_blank(row["score_uptrend_consistency"]),
            round_or_blank(row["score_spike_total"]),
            round_or_blank(row["score_gapup"]),
            round_or_blank(row["score_sector"]),
            " | ".join(met) if met else "—",
        ]
        for ci, val in enumerate(vals, 1):
            cell = dash.cell(row=ri, column=ci, value=val)
            if ci == 4:    # Score
                sc(cell, bold=True, color=s_fc, fill=s_fill, size=11, border=BDR_M)
            elif ci == 5:  # Rating
                sc(cell, bold=True, color=s_fc, fill=s_fill, size=11)
            elif ci == 14: # Criteria Met
                sc(cell, fill=row_bg, align="left", size=10, wrap=False)
            elif ci in (2, 3):
                sc(cell, fill=row_bg, align="left", size=11,
                   bold=(ci == 2))
            else:
                sc(cell, fill=row_bg, size=11)
                if ci == 6:
                    cell.number_format = "#,##0.00"
                elif ci in (7, 8):
                    cell.number_format = "#,##0.0"
        dash.row_dimensions[ri].height = 15

    dash.freeze_panes = "D3"
    dash.auto_filter.ref = f"A2:{get_column_letter(ncols_d)}{max(2, len(top20)+2)}"
    d_col_w = [5, 18, 22, 7, 14, 9, 10, 10, 8, 8, 7, 8, 7, 55]
    for ci, w in enumerate(d_col_w, 1):
        dash.column_dimensions[get_column_letter(ci)].width = w

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 2 ── 📊 FULL RATINGS
    # ══════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("📊 Full Ratings")

    # Column spec:  (data_key,  display_name,  group,  width,  num_fmt)
    COL_SPEC = [
        # ── Identity
        ("rank",                    "#",               "id",    4,  None),
        ("symbol",                  "Symbol",          "id",   16,  None),
        ("sector",                  "Sector",          "id",   22,  None),
        ("composite_score",         "Total\nScore",    "total", 9,  "0"),
        ("pre_sector_score",        "Pre-Sec\nScore",  "total", 9,  "0"),
        # ── 52W
        ("current_close",           "Close",           "52w",   9,  "#,##0.##"),
        ("high_52w_close",          "52W High",        "52w",   9,  "#,##0.##"),
        ("high_52w_date",           "52W High\nDate",  "52w",  11,  None),
        ("pct_from_52w_high",       "% from\n52W Hi",  "52w",   8,  "0.0%"),
        ("days_since_52w_high",     "Days\n52W Hi",    "52w",   7,  "0"),
        ("score_52w_price",         "Sc\nDist",        "52w",   5,  "0"),
        ("score_52w_recency",       "Sc\nRec",         "52w",   5,  "0"),
        ("score_52w_bonus",         "Sc\nAdj",         "52w",   5,  "0"),
        ("score_52w_total",         "52W\nTotal",      "52w",   6,  "0"),
        # ── Liquidity
        ("listing_date",            "Listed",          "liq",  11,  None),
        ("listed_days",             "Listed\nDays",    "liq",   7,  "0"),
        ("avg_turnover_42d",        "Avg TO\n42D(Cr)", "liq",  10,  "#,##0.0"),
        ("median_turnover_42d",     "Med TO\n42D(Cr)", "liq",  10,  "#,##0.0"),
        ("score_liquidity",         "Sc\nLiq",         "liq",   5,  "0"),
        ("score_median_turnover",   "Sc\nTO",          "liq",   5,  "0"),
        ("score_new_listing",       "Sc\nNew\nList",   "liq",   5,  "0"),
        ("score_ipo_perf",         "Sc\nIPO",         "liq",   6,  "0"),
        ("score_liquid_leaders_bonus","Sc\nLiq\nLead", "liq",   6,  "0"),
        # ── Performance
        ("return_12m",              "Ret\n12M",        "perf",  7,  "0.0%"),
        ("return_6m",               "Ret\n6M",         "perf",  7,  "0.0%"),
        ("return_3m",               "Ret\n3M",         "perf",  7,  "0.0%"),
        ("return_1d",               "Ret\n1D",         "perf",  7,  "0.0%"),
        ("reset_low",               "Reset\nLow",      "perf",  9,  "#,##0.##"),
        ("reset_recovery",          "Reset\nRecov",    "perf",  8,  "0.0%"),
        ("score_reset_recovery",    "Sc\nRecov",       "perf",  5,  "0"),
        ("score_perf_12m",          "Sc\n12M",         "perf",  5,  "0"),
        ("score_perf_12m_penalty",   "Pen\n12M",        "perf",  5,  "0"),
        ("score_perf_6m",           "Sc\n6M",          "perf",  5,  "0"),
        ("score_perf_3m",           "Sc\n3M",          "perf",  5,  "0"),
        ("score_perf_1d",           "Sc\n1D",          "perf",  5,  "0"),
        ("score_perf_6m_penalty",   "Pen\n6M",         "perf",  5,  "0"),
        ("score_perf_3m_penalty",   "Pen\n3M",         "perf",  5,  "0"),
        ("score_perf_1d_penalty",   "Pen\n1D",         "perf",  5,  "0"),
        ("score_12m_52w_bonus",     "Sc\n12M\n52W",   "perf",  6,  "0"),
        ("score_perf_bonus",        "Sc\nBonus",       "perf",  5,  "0"),
        ("score_performance_total", "Perf\nTotal",     "perf",  6,  "0"),
        # ── Relative Strength
        ("rs_12m",                  "RS\n12M",         "rs",    7,  "0.0%"),
        ("rs_6m",                   "RS\n6M",          "rs",    7,  "0.0%"),
        ("rs_3m",                   "RS\n3M",          "rs",    7,  "0.0%"),
        ("rs_line_at_52w_high",     "RS@\n52W Hi",     "rs",    7,  None),
        ("rs_line_slope_21d",       "RS\nSlope",       "rs",    7,  "0.000"),
        ("score_rs_12m",            "Sc\n12M",         "rs",    5,  "0"),
        ("score_rs_6m",             "Sc\n6M",          "rs",    5,  "0"),
        ("score_rs_3m",             "Sc\n3M",          "rs",    5,  "0"),
        ("score_rs_line_high",      "Sc\nHi",          "rs",    5,  "0"),
        ("score_rs_slope_penalty",  "Pen\nSlope",      "rs",    5,  "0"),
        ("score_rs_total",          "RS\nTotal",       "rs",    6,  "0"),
        # ── Uptrend
        ("above_50dma",             "Above\n50DMA",    "trend", 8,  None),
        ("score_above_50dma",       "Sc\n50DMA",       "trend", 6,  "0"),
        ("above_21ema",             "Above\n21EMA",    "trend", 8,  None),
        ("score_above_21ema",       "Sc\n21EMA",       "trend", 6,  "0"),
        ("above_8ema",              "Above\n8EMA",     "trend", 8,  None),
        ("score_above_8ema",        "Sc\n8EMA",        "trend", 6,  "0"),
        ("uptrend_consistency_pct", "Uptrnd\nCon%",    "trend", 9,  "0.0%"),
        ("score_uptrend_consistency","Sc\nTrend",      "trend", 6,  "0"),
        ("green_candle_count",      "Green\nCandles",  "trend", 8,  "0"),
        ("score_green_candles",     "Sc\nGreen",       "trend", 6,  "0"),
        ("daysbelow50dma",          "Days<\n50DMA",    "trend", 7,  "0"),
        # ── Volume Spike
        ("spike_date",              "Spike\nDate",     "spike",11,  None),
        ("spike_volume",            "Spike\nVol",      "spike",10,  "#,##0"),
        ("spike_price_change_pct",  "Spike\nChg%",     "spike", 8,  "0.0%"),
        ("spike_window_days",       "Spike\nWin",      "spike", 6,  "0"),
        ("spike_label",             "Spike\nLabel",    "spike", 9,  None),
        ("score_spike_base",        "Sc\nBase",        "spike", 5,  "0"),
        ("score_spike_bonus",       "Sc\nBonus",       "spike", 5,  "0"),
        ("score_spike_total",       "Spike\nTotal",    "spike", 6,  "0"),
        # ── Gap-Up
        ("gapup_date",              "GapUp\nDate",     "gapup",11,  None),
        ("gapup_volume",            "GapUp\nVol",      "gapup",10,  "#,##0"),
        ("gapup_pct",               "GapUp\n%",        "gapup", 8,  "0.0%"),
        ("gapup_close_pct",         "GapUp\nClose%",   "gapup", 8,  "0.0%"),
        ("score_gapup",             "Sc\nGapUp",       "gapup", 6,  "0"),
        # ── Volatility
        ("atr_percent_21d",         "ATR%\n21D",       "vol",   9,  "0.00"),
        ("score_volatility",        "Sc\nATR",         "vol",   5,  "0"),
        # ── Sector
        ("sector_stock_count",      "Sec\n#Stk",       "sector",5,  "0"),
        ("sector_top_quartile_count","Sec\nTop25",     "sector",6,  "0"),
        ("sector_strength_ratio",   "Sec\nStr%",       "sector",9,  "0.0%"),
        ("sector_leadership_score", "Sec\nScore",      "sector",8,  "0.0"),
        ("score_sector",            "Sc\nSect",        "sector",7,  "0"),
    ]

    # Keep the full dataset in the sheet, but open with a cleaner default view.
    HIDDEN_KEYS = {
        "high_52w_date",
        "days_since_52w_high",
        "score_52w_price",
        "score_52w_recency",
        "score_52w_bonus",
        "listing_date",
        "listed_days",
        "score_liquidity",
        "score_median_turnover",
        "reset_low",
        "score_reset_recovery",
        "score_perf_12m",
        "score_perf_6m",
        "score_perf_3m",
        "score_perf_6m_penalty",
        "score_perf_3m_penalty",
        "score_perf_bonus",
        "rs_line_at_52w_high",
        "rs_line_slope_21d",
        "score_rs_12m",
        "score_rs_3m",
        "score_rs_line_high",
        "score_rs_slope_penalty",
        "spike_date",
        "spike_volume",
        "spike_window_days",
        "spike_label",
        "score_spike_base",
        "score_spike_bonus",
        "gapup_date",
        "gapup_volume",
        "score_volatility",
        "sector_stock_count",
        "sector_top_quartile_count",
    }

    PCT_COLS  = {"pct_from_52w_high","return_12m","return_6m","return_3m",
                 "reset_recovery","rs_12m","rs_6m","rs_3m",
                 "uptrend_consistency_pct","spike_price_change_pct",
                 "gapup_pct","gapup_close_pct","sector_strength_ratio"}
    DATE_COLS = {"listing_date","high_52w_date","spike_date","gapup_date"}
    SCORE_KEYS= {k for k, *_ in COL_SPEC if k.startswith("score_") or k.startswith("pre_")}
    INT_COLS  = {"rank","listed_days","days_since_52w_high",
                 "sector_stock_count","sector_top_quartile_count","spike_window_days",
                 "daysbelow50dma","green_candle_count"}

    ncols_r = len(COL_SPEC)
    groups   = [g for _, _, g, _, _ in COL_SPEC]

    # ── Row 1: title ──────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(ncols_r)}1")
    ws["A1"] = (
        f"NSE STOCK TECHNICAL RATING  —  Full Data   |   "
        f"As of {cutoff_date.strftime('%d-%b-%Y')}   |   "
        f"Reset: {reset_date.strftime('%d-%b-%Y')}   |   "
        f"Source: {symbol_file.name}   |   {len(scored)} stocks"
    )
    sc(ws["A1"], bold=True, color="FFFFFF", fill=T_FILL, size=12, align="center")
    ws.row_dimensions[1].height = 26

    # ── Row 2: section header groups ─────────────────────────────────
    GROUP_LABELS = {
        "id":    "IDENTITY & SCORE",
        "total": "COMPOSITE",
        "52w":   "52-WEEK HIGH",
        "liq":   "LIQUIDITY & TURNOVER",
        "perf":  "PERFORMANCE",
        "rs":    "RELATIVE STRENGTH",
        "trend": "UPTREND",
        "spike": "VOLUME SPIKE",
        "gapup": "GAP-UP",
        "vol":   "VOLATILITY (ATR)",
        "sector":"SECTOR",
    }
    prev_group = None
    sec_start  = 1
    for ci, g in enumerate(groups, 1):
        if g != prev_group:
            if prev_group is not None:
                ws.merge_cells(start_row=2, start_column=sec_start,
                               end_row=2,   end_column=ci - 1)
            sec_start  = ci
            prev_group = g
    ws.merge_cells(start_row=2, start_column=sec_start,
                   end_row=2,   end_column=ncols_r)

    ci_cursor = 1
    prev_group = None
    for ci, (_, _, g, _, _) in enumerate(COL_SPEC, 1):
        if g != prev_group:
            cell = ws.cell(row=2, column=ci, value=GROUP_LABELS.get(g, g.upper()))
            sc(cell, bold=True, color="FFFFFF", fill=SEC_FILLS.get(g, H_FILL),
               size=11, align="center")
            prev_group = g
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[2].hidden = True

    # ── Row 3: column headers ─────────────────────────────────────────
    for ci, (_, hdr, grp, _, _) in enumerate(COL_SPEC, 1):
        cell = ws.cell(row=3, column=ci, value=hdr)
        sc(cell, bold=True, color="FFFFFF",
           fill=COL_FILLS.get(grp, H_FILL), size=10, wrap=True)
    ws.row_dimensions[3].height = 40

    # ── Data rows ────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(scored.iterrows(), start=4):
        s_fill, s_fc = score_band(row["composite_score"])
        afill = ALT_FILL if (ri % 2 == 0) else WHT_FILL

        for ci, (key, _, grp, _, nfmt) in enumerate(COL_SPEC, 1):
            raw = row.get(key)
            if key in PCT_COLS:
                val = pct_fmt(raw)
            elif key in DATE_COLS:
                val = dt_fmt(raw)
            elif key in INT_COLS:
                val = int(raw) if (raw is not None and not (isinstance(raw, float) and math.isnan(raw))) else ""
            elif isinstance(raw, (float, np.floating)) and not math.isnan(raw):
                val = round(float(raw), 4)
            else:
                val = raw

            cell = ws.cell(row=ri, column=ci, value=val)

            if key == "composite_score":
                sc(cell, bold=True, color=s_fc, fill=s_fill, size=11, border=BDR_M)
                if nfmt: cell.number_format = nfmt
            elif key in SCORE_KEYS:
                score_cell(cell, val, size=10)
                if nfmt: cell.number_format = nfmt
            elif key == "symbol":
                sc(cell, bold=True, fill=afill, align="left", size=11)
            elif key == "sector":
                sc(cell, fill=afill, align="left", size=10)
            elif key == "rank":
                sc(cell, bold=True, fill=s_fill, color=s_fc, size=11)
            else:
                sc(cell, fill=afill, size=10)
                if nfmt: cell.number_format = nfmt

        ws.row_dimensions[ri].height = 16

    # ── Column widths ─────────────────────────────────────────────────
    for ci, (key, _, _, w, _) in enumerate(COL_SPEC, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        if key in HIDDEN_KEYS:
            ws.column_dimensions[get_column_letter(ci)].hidden = True

    ws.freeze_panes = "D4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols_r)}{max(3, len(scored)+3)}"

    # ── Color scale on composite_score column ─────────────────────────
    sc_col = get_column_letter(4)
    ws.conditional_formatting.add(
        f"{sc_col}4:{sc_col}{len(scored)+3}",
        ColorScaleRule(
            start_type="min", start_color="F1948A",
            mid_type="num",   mid_value=14, mid_color="FAD7A0",
            end_type="max",   end_color="82E0AA",
        )
    )

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 3 ── 🎯 SCORECARD
    # ══════════════════════════════════════════════════════════════════
    sc_ws = wb.create_sheet("🎯 Scorecard")
    SC_COLS = [
        ("rank",                    "#",              4),
        ("symbol",                  "Symbol",        16),
        ("sector",                  "Sector",        20),
        ("score_new_listing",       "New\nList",      8),
        ("score_ipo_perf",          "IPO\nPerf",      8),
        ("score_52w_total",         "52W\nScore",     8),
        ("score_performance_total", "Perf\nTotal",    9),
        ("score_perf_1d",           "1D\nScore",      8),
        ("score_perf_12m_penalty",  "12M\nPen",       7),
        ("score_perf_1d_penalty",   "1D\nPen",        7),
        ("score_reset_recovery",    "Reset\nScore",   8),
        ("score_perf_6m_penalty",   "6M\nPen",        7),
        ("score_perf_3m_penalty",   "3M\nPen",        7),
        ("score_rs_total",          "RS\nTotal",      8),
        ("score_rs_6m",             "RS\n6M",         7),
        ("score_above_50dma",       "50DMA\nScore",   8),
        ("score_above_21ema",       "21EMA\nScore",   8),
        ("score_ema21_uptrend",     "21EMA\nUp",      8),
        ("score_above_8ema",        "8EMA\nScore",    8),
        ("score_uptrend_consistency","Trend\nScore",  8),
        ("score_green_candles",     "Green\nScore",   8),
        ("score_spike_total",       "Spike",          7),
        ("score_gapup",             "GapUp",          7),
        ("score_volatility",        "ATR\nPen",       7),
        ("score_liquidity",         "Liquidity",      8),
        ("score_median_turnover",   "Median\nTO",     8),
        ("score_liquid_leaders_bonus","Liq\nLead",    8),
        ("sector_leadership_score", "Sector\nScore", 10),
        ("score_sector",            "Sector\nPts",    7),
        ("composite_score",         "TOTAL\nSCORE",  10),
    ]
    ncols_sc = len(SC_COLS)

    sc_ws.merge_cells(f"A1:{get_column_letter(ncols_sc)}1")
    sc_ws["A1"] = (
        f"SCORE BREAKDOWN — ALL COMPONENTS   |   "
        f"As of {cutoff_date.strftime('%d-%b-%Y')}   |   {len(scored)} stocks"
    )
    sc(sc_ws["A1"], bold=True, color="FFFFFF", fill=T_FILL, size=12, align="center")
    sc_ws.row_dimensions[1].height = 26

    for ci, (_, hdr, _) in enumerate(SC_COLS, 1):
        cell = sc_ws.cell(row=2, column=ci, value=hdr)
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=11, wrap=True)
    sc_ws.row_dimensions[2].height = 36

    for ri, (_, row) in enumerate(scored.iterrows(), start=3):
        s_fill, s_fc = score_band(row["composite_score"])
        afill = ALT_FILL if ri % 2 == 1 else WHT_FILL

        for ci, (key, _, w) in enumerate(SC_COLS, 1):
            val = row.get(key)
            cell = sc_ws.cell(row=ri, column=ci)
            if key == "composite_score":
                cell.value = score_or_blank(val)
                sc(cell, bold=True, color=s_fc, fill=s_fill, size=11, border=BDR_M)
                cell.number_format = "0"
            elif key == "rank":
                cell.value = int(val) if val else ""
                sc(cell, bold=True, fill=afill, size=11)
            elif key == "symbol":
                cell.value = val
                sc(cell, bold=True, fill=afill, align="left", size=11)
            elif key == "sector":
                cell.value = val
                sc(cell, fill=afill, align="left", size=10)
            else:
                score_cell(cell, score_or_blank(val), size=11)
                cell.number_format = "0"
        sc_ws.row_dimensions[ri].height = 14

    for ci, (_, _, w) in enumerate(SC_COLS, 1):
        sc_ws.column_dimensions[get_column_letter(ci)].width = w
    sc_ws.freeze_panes = "D3"
    sc_ws.auto_filter.ref = f"A2:{get_column_letter(ncols_sc)}{max(2, len(scored)+2)}"

    # Color scale on Total Score column
    tot_col = get_column_letter(ncols_sc)
    sc_ws.conditional_formatting.add(
        f"{tot_col}3:{tot_col}{len(scored)+2}",
        ColorScaleRule(
            start_type="min", start_color="F1948A",
            mid_type="num",   mid_value=14, mid_color="FAD7A0",
            end_type="max",   end_color="82E0AA",
        )
    )

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 4 ── 🏦 INSTITUTIONAL PICKS
    #  Top 50 by composite score → weighted institutional score:
    #    40% sector strength (count × avg score in top 50)
    #    30% liquidity      (median_turnover_42d)
    #    30% stock rating   (composite_score)
    #  → Top 20 by institutional score
    # ══════════════════════════════════════════════════════════════════
    ip_ws = wb.create_sheet("🏦 Institutional Picks")

    top50 = scored.head(50).copy()

    # Sector strength from the top-50 pool
    ip_sec_agg = (
        top50.groupby("sector", dropna=False)
        .agg(_ip_sec_count=("composite_score", "count"),
             _ip_sec_avg   =("composite_score", "mean"))
        .reset_index()
    )
    ip_sec_agg["_ip_sec_strength"] = ip_sec_agg["_ip_sec_count"] * ip_sec_agg["_ip_sec_avg"]
    top50 = top50.merge(
        ip_sec_agg[["sector", "_ip_sec_count", "_ip_sec_avg", "_ip_sec_strength"]],
        on="sector", how="left"
    )

    # Percentile-rank each dimension within the top-50 pool
    top50["_ip_sector_pct"]    = top50["_ip_sec_strength"].rank(pct=True, method="average") * 100
    top50["_ip_liquidity_pct"] = top50["median_turnover_42d"].rank(pct=True, method="average") * 100
    top50["_ip_rating_pct"]    = top50["composite_score"].rank(pct=True, method="average") * 100

    # Weighted institutional score: 40% sector + 30% liquidity + 30% rating
    IP_WT_SECTOR    = 0.40
    IP_WT_LIQUIDITY = 0.30
    IP_WT_RATING    = 0.30
    top50["inst_score"] = (
        IP_WT_SECTOR    * top50["_ip_sector_pct"] +
        IP_WT_LIQUIDITY * top50["_ip_liquidity_pct"] +
        IP_WT_RATING    * top50["_ip_rating_pct"]
    )

    # Additional liquidity gate for institutional picks: exclude stocks where
    # median 42D turnover < INST_PICKS_MIN_MEDIAN_TURNOVER_CRORES (10 Cr).
    # This is stricter than the universe gate (15 Cr AND logic) — a stock can pass
    # the universe filter via avg turnover but still be too illiquid for inst picks.
    _inst_eligible = top50["median_turnover_42d"] >= INST_PICKS_MIN_MEDIAN_TURNOVER_CRORES
    inst_picks = (
        top50[_inst_eligible]
        .sort_values("inst_score", ascending=False)
        .head(20)
        .reset_index(drop=True)
    )
    inst_picks["inst_rank"] = range(1, len(inst_picks) + 1)

    IP_FILL  = PatternFill("solid", fgColor="0B3D91")   # deep institutional blue
    IP_H     = PatternFill("solid", fgColor="1A56A0")   # column header
    IP_ALT   = PatternFill("solid", fgColor="EAF2FB")
    IP_GOLD  = PatternFill("solid", fgColor="FFD700")
    IP_SILV  = PatternFill("solid", fgColor="C0C0C0")
    IP_BRNZ  = PatternFill("solid", fgColor="CD7F32")

    IP_COLS = [
        ("inst_rank",          "#",                   5),
        ("symbol",             "Symbol",             16),
        ("sector",             "Sector",             22),
        ("inst_score",         "Inst\nScore",        10),
        ("_ip_sector_pct",     "Sector\nStr% (40%)", 13),
        ("_ip_liquidity_pct",  "Liquidity\n%ile (30%)",13),
        ("_ip_rating_pct",     "Rating\n%ile (30%)", 13),
        ("composite_score",    "Rating\nScore",      11),
        ("rank",               "Rating\nRank",       10),
        ("median_turnover_42d","Median TO\n42d (Cr)",14),
        ("_ip_sec_count",      "Sector\nCount",      10),
        ("_ip_sec_avg",        "Sector\nAvg Score",  13),
        ("current_close",      "Close",              10),
        ("perf_12m_pct",       "12M\nReturn%",       12),
        ("perf_3m_pct",        "3M\nReturn%",        12),
    ]
    ncols_ip = len(IP_COLS)

    # ── Title bar ─────────────────────────────────────────────────────
    ip_ws.merge_cells(f"A1:{get_column_letter(ncols_ip)}1")
    ip_ws["A1"] = (
        f"🏦 INSTITUTIONAL PICKS  —  "
        f"Top 50 by Score → Weighted: 40% Sector + 30% Liquidity + 30% Rating  |  "
        f"As of {cutoff_date.strftime('%d-%b-%Y')}"
    )
    sc(ip_ws["A1"], bold=True, color="FFFFFF", fill=IP_FILL, size=13, align="center")
    ip_ws.row_dimensions[1].height = 28

    # ── Sub-title ─────────────────────────────────────────────────────
    ip_ws.merge_cells(f"A2:{get_column_letter(ncols_ip)}2")
    ip_ws["A2"] = (
        "Institutional Score = 40% Sector Strength + 30% Liquidity (Median Turnover) + 30% Stock Rating — "
        "all components percentile-ranked within top 50"
    )
    sc(ip_ws["A2"], italic=True, color="FFFFFF", fill=IP_H, size=10, align="center")
    ip_ws.row_dimensions[2].height = 18

    # ── Column headers — colour-coded by component ────────────────────
    IP_SECTOR_KEYS   = {"_ip_sector_pct", "_ip_sec_count", "_ip_sec_avg"}
    IP_LIQUIDITY_KEY = {"median_turnover_42d", "_ip_liquidity_pct"}
    IP_RATING_KEYS   = {"composite_score", "rank", "_ip_rating_pct"}
    IP_SEC_HDR  = PatternFill("solid", fgColor="1A5276")   # blue   — sector cols
    IP_LIQ_HDR  = PatternFill("solid", fgColor="1E8449")   # green  — liquidity cols
    IP_RAT_HDR  = PatternFill("solid", fgColor="6E2F8A")   # purple — rating cols

    for ci, (key, hdr, _) in enumerate(IP_COLS, 1):
        cell = ip_ws.cell(row=3, column=ci, value=hdr)
        if key in IP_SECTOR_KEYS:
            hfill = IP_SEC_HDR
        elif key in IP_LIQUIDITY_KEY:
            hfill = IP_LIQ_HDR
        elif key in IP_RATING_KEYS:
            hfill = IP_RAT_HDR
        elif key == "inst_score":
            hfill = IP_FILL
        else:
            hfill = IP_H
        sc(cell, bold=True, color="FFFFFF", fill=hfill, size=10, wrap=True)
    ip_ws.row_dimensions[3].height = 36

    # ── Data rows ─────────────────────────────────────────────────────
    IP_SEC_CELL = PatternFill("solid", fgColor="D6EAF8")   # light blue  — sector data
    IP_LIQ_CELL = PatternFill("solid", fgColor="D5F5E3")   # light green — liquidity data
    IP_RAT_CELL = PatternFill("solid", fgColor="EAD9F5")   # light purple — rating data

    for ri, row in inst_picks.iterrows():
        er       = ri + 4
        alt      = ri % 2 == 1
        row_fill = IP_ALT if alt else WHT_FILL
        s_fill, s_fc = score_band(row["composite_score"])

        if row["inst_rank"] == 1:
            rank_fill = IP_GOLD
        elif row["inst_rank"] == 2:
            rank_fill = IP_SILV
        elif row["inst_rank"] == 3:
            rank_fill = IP_BRNZ
        else:
            rank_fill = row_fill

        for ci, (key, _, _) in enumerate(IP_COLS, 1):
            val  = row.get(key)
            cell = ip_ws.cell(row=er, column=ci)

            if key == "inst_rank":
                cell.value = int(val) if pd.notna(val) else ""
                sc(cell, bold=True, color="1F3864", fill=rank_fill, size=12, align="center")

            elif key == "symbol":
                cell.value = val
                sc(cell, bold=True, color="0B3D91", fill=row_fill, align="left", size=12)

            elif key == "sector":
                cell.value = val
                sc(cell, color="1A5276", fill=IP_SEC_CELL, align="left", size=10)

            elif key == "inst_score":
                cell.value = round_or_blank(val, 1)
                sc(cell, bold=True, color="0B1F33", fill=IP_FILL, size=12, border=BDR_M)
                cell.number_format = "0.0"

            elif key == "_ip_sector_pct":
                cell.value = round_or_blank(val, 1)
                sc(cell, bold=True, color="1A5276", fill=IP_SEC_CELL, size=11)
                cell.number_format = "0.0"

            elif key in ("_ip_sec_count", "_ip_sec_avg"):
                cell.value = round_or_blank(val, 1) if key == "_ip_sec_avg" else (int(val) if pd.notna(val) else "")
                sc(cell, color="1A5276", fill=IP_SEC_CELL, size=10)
                if key == "_ip_sec_avg":
                    cell.number_format = "0.0"

            elif key == "_ip_liquidity_pct":
                cell.value = round_or_blank(val, 1)
                sc(cell, bold=True, color="145A32", fill=IP_LIQ_CELL, size=11)
                cell.number_format = "0.0"

            elif key == "median_turnover_42d":
                cell.value = round_or_blank(val, 2)
                sc(cell, color="145A32", fill=IP_LIQ_CELL, size=11)
                cell.number_format = "#,##0.00"

            elif key == "_ip_rating_pct":
                cell.value = round_or_blank(val, 1)
                sc(cell, bold=True, color="6E2F8A", fill=IP_RAT_CELL, size=11)
                cell.number_format = "0.0"

            elif key == "composite_score":
                cell.value = score_or_blank(val)
                sc(cell, bold=True, color=s_fc, fill=s_fill, size=11, border=BDR_M)
                cell.number_format = "0"

            elif key == "rank":
                cell.value = int(val) if pd.notna(val) else ""
                sc(cell, color="555555", fill=IP_RAT_CELL, size=11)

            elif key in ("perf_12m_pct", "perf_3m_pct"):
                v = round_or_blank(val, 4)
                cell.value = v
                if isinstance(v, float):
                    pfill = PatternFill("solid", fgColor="D5F5E3") if v >= 0 else PatternFill("solid", fgColor="FADBD8")
                    pclr  = "1F6B3A" if v >= 0 else "8B0000"
                    sc(cell, color=pclr, fill=pfill, size=11)
                else:
                    sc(cell, fill=row_fill, size=11)
                cell.number_format = "0.00%"

            else:
                cell.value = round_or_blank(val, 2) if isinstance(val, float) else val
                sc(cell, fill=row_fill, size=11)

        ip_ws.row_dimensions[er].height = 16

    # ── Column widths ─────────────────────────────────────────────────
    for ci, (_, _, w) in enumerate(IP_COLS, 1):
        ip_ws.column_dimensions[get_column_letter(ci)].width = w

    ip_ws.freeze_panes = "C4"

    # Color scale on Institutional Score column (col 4)
    last_ip_row = 3 + len(inst_picks)
    inst_score_col = get_column_letter(4)
    if last_ip_row > 4:
        ip_ws.conditional_formatting.add(
            f"{inst_score_col}4:{inst_score_col}{last_ip_row}",
            ColorScaleRule(
                start_type="min", start_color="FADBD8",
                mid_type="percentile", mid_value=50, mid_color="FAD7A0",
                end_type="max", end_color="D5F5E3",
            )
        )

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 5 ── 🎨 THEMATIC PICKS
    #  Top 50 by composite score → weighted score:
    #    65% sector strength (count × avg score in top 50)
    #    35% liquidity (median_turnover_42d)
    #  → Top 20 by thematic score
    # ══════════════════════════════════════════════════════════════════
    tp_ws = wb.create_sheet("🎨 Thematic Picks")

    # ── Build thematic score ───────────────────────────────────────────
    top50_tp = scored.head(50).copy()

    # Sector strength from top-50 universe:
    #   count  = how many top-50 stocks belong to this sector (breadth)
    #   avg_cs = average composite_score of those stocks (quality)
    #   strength = count × avg_cs  →  rewards sectors with both breadth and score
    sector_agg = (
        top50_tp.groupby("sector", dropna=False)
        .agg(
            _sec_count=("composite_score", "count"),
            _sec_avg  =("composite_score", "mean"),
        )
        .reset_index()
    )
    sector_agg["_sec_strength"] = sector_agg["_sec_count"] * sector_agg["_sec_avg"]
    top50_tp = top50_tp.merge(
        sector_agg[["sector", "_sec_count", "_sec_avg", "_sec_strength"]],
        on="sector", how="left"
    )

    # Percentile-rank both dimensions within the top-50 pool
    top50_tp["_sector_pct"]    = top50_tp["_sec_strength"].rank(pct=True, method="average") * 100
    top50_tp["_liquidity_pct"] = top50_tp["median_turnover_42d"].rank(pct=True, method="average") * 100

    # Weighted thematic score: 65% sector, 35% liquidity
    SECTOR_WT   = 0.65
    LIQUIDITY_WT = 0.35
    top50_tp["thematic_score"] = (
        SECTOR_WT   * top50_tp["_sector_pct"] +
        LIQUIDITY_WT * top50_tp["_liquidity_pct"]
    )

    thematic_picks = (
        top50_tp
        .sort_values("thematic_score", ascending=False)
        .head(20)
        .reset_index(drop=True)
    )
    thematic_picks["thematic_rank"] = range(1, len(thematic_picks) + 1)

    # ── Palette ───────────────────────────────────────────────────────
    TP_FILL  = PatternFill("solid", fgColor="4A235A")   # deep purple title
    TP_H     = PatternFill("solid", fgColor="6C3483")   # header purple
    TP_ALT   = PatternFill("solid", fgColor="F5EEF8")   # alt row lavender
    TP_SEC   = PatternFill("solid", fgColor="EAF4FB")   # sector highlight (blue tint)
    TP_LIQ   = PatternFill("solid", fgColor="EAFAF1")   # liquidity highlight (green tint)
    TP_GOLD  = PatternFill("solid", fgColor="FFD700")
    TP_SILV  = PatternFill("solid", fgColor="C0C0C0")
    TP_BRNZ  = PatternFill("solid", fgColor="CD7F32")

    TP_COLS = [
        ("thematic_rank",        "#",                   5),
        ("symbol",               "Symbol",             16),
        ("sector",               "Sector",             22),
        ("_sec_count",           "Sector\nCount",      10),   # stocks from sector in top 50
        ("_sec_avg",             "Sector\nAvg Score",  12),
        ("_sector_pct",          "Sector\nStrength%",  12),   # 65% weight input
        ("median_turnover_42d",  "Median TO\n42d (Cr)",14),
        ("_liquidity_pct",       "Liquidity\n%ile",    12),   # 35% weight input
        ("thematic_score",       "Thematic\nScore",    12),   # final weighted score
        ("composite_score",      "Rating\nScore",      10),
        ("rank",                 "Rating\nRank",       10),
        ("current_close",        "Close",              10),
        ("perf_12m_pct",         "12M\nReturn%",       12),
        ("perf_3m_pct",          "3M\nReturn%",        12),
    ]
    ncols_tp = len(TP_COLS)

    # ── Title bar ─────────────────────────────────────────────────────
    tp_ws.merge_cells(f"A1:{get_column_letter(ncols_tp)}1")
    tp_ws["A1"] = (
        f"🎨 THEMATIC PICKS  —  "
        f"Top 50 by Score → Weighted: 65% Leading Sector + 35% Liquidity  |  "
        f"As of {cutoff_date.strftime('%d-%b-%Y')}"
    )
    sc(tp_ws["A1"], bold=True, color="FFFFFF", fill=TP_FILL, size=13, align="center")
    tp_ws.row_dimensions[1].height = 28

    # ── Sub-title ─────────────────────────────────────────────────────
    tp_ws.merge_cells(f"A2:{get_column_letter(ncols_tp)}2")
    tp_ws["A2"] = (
        "Sector Strength = (stocks in sector from top 50) × (avg composite score)   |   "
        "Stocks riding leading sectors with good liquidity float to the top"
    )
    sc(tp_ws["A2"], italic=True, color="FFFFFF", fill=TP_H, size=10, align="center")
    tp_ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────────────────
    # Colour-code header groups: sector cols → blue tint, liquidity col → green tint
    SECTOR_COLS  = {"_sec_count", "_sec_avg", "_sector_pct"}
    LIQUIDITY_C  = {"median_turnover_42d", "_liquidity_pct"}
    for ci, (key, hdr, _) in enumerate(TP_COLS, 1):
        cell = tp_ws.cell(row=3, column=ci, value=hdr)
        if key in SECTOR_COLS:
            sc(cell, bold=True, color="FFFFFF", fill=PatternFill("solid", fgColor="1A5276"), size=10, wrap=True)
        elif key in LIQUIDITY_C:
            sc(cell, bold=True, color="FFFFFF", fill=PatternFill("solid", fgColor="1E8449"), size=10, wrap=True)
        elif key == "thematic_score":
            sc(cell, bold=True, color="FFFFFF", fill=TP_FILL, size=10, wrap=True)
        else:
            sc(cell, bold=True, color="FFFFFF", fill=TP_H, size=10, wrap=True)
    tp_ws.row_dimensions[3].height = 36

    # ── Data rows ─────────────────────────────────────────────────────
    for ri, row in thematic_picks.iterrows():
        er       = ri + 4
        alt      = ri % 2 == 1
        row_fill = TP_ALT if alt else WHT_FILL
        s_fill, s_fc = score_band(row["composite_score"])

        if row["thematic_rank"] == 1:
            rank_fill = TP_GOLD
        elif row["thematic_rank"] == 2:
            rank_fill = TP_SILV
        elif row["thematic_rank"] == 3:
            rank_fill = TP_BRNZ
        else:
            rank_fill = row_fill

        for ci, (key, _, _) in enumerate(TP_COLS, 1):
            val  = row.get(key)
            cell = tp_ws.cell(row=er, column=ci)

            if key == "thematic_rank":
                cell.value = int(val) if pd.notna(val) else ""
                sc(cell, bold=True, color="4A235A", fill=rank_fill, size=12, align="center")

            elif key == "symbol":
                cell.value = val
                sc(cell, bold=True, color="4A235A", fill=row_fill, align="left", size=12)

            elif key == "sector":
                cell.value = val
                sc(cell, bold=False, color="1A5276", fill=TP_SEC, align="left", size=10)

            elif key == "_sec_count":
                cell.value = int(val) if pd.notna(val) else ""
                sc(cell, bold=True, color="1A5276", fill=TP_SEC, size=11)

            elif key == "_sec_avg":
                cell.value = round_or_blank(val, 1)
                sc(cell, color="1A5276", fill=TP_SEC, size=11)
                cell.number_format = "0.0"

            elif key == "_sector_pct":
                cell.value = round_or_blank(val, 1)
                sc(cell, bold=True, color="1A5276", fill=TP_SEC, size=11)
                cell.number_format = "0.0"

            elif key == "median_turnover_42d":
                cell.value = round_or_blank(val, 2)
                sc(cell, bold=True, color="145A32", fill=TP_LIQ, size=11)
                cell.number_format = "#,##0.00"

            elif key == "_liquidity_pct":
                cell.value = round_or_blank(val, 1)
                sc(cell, color="145A32", fill=TP_LIQ, size=11)
                cell.number_format = "0.0"

            elif key == "thematic_score":
                cell.value = round_or_blank(val, 1)
                sc(cell, bold=True, color="FFFFFF", fill=TP_FILL, size=12, border=BDR_M)
                cell.number_format = "0.0"

            elif key == "composite_score":
                cell.value = score_or_blank(val)
                sc(cell, bold=True, color=s_fc, fill=s_fill, size=11)
                cell.number_format = "0"

            elif key == "rank":
                cell.value = int(val) if pd.notna(val) else ""
                sc(cell, color="555555", fill=row_fill, size=11)

            elif key in ("perf_12m_pct", "perf_3m_pct"):
                v = round_or_blank(val, 4)
                cell.value = v
                if isinstance(v, float):
                    pfill = PatternFill("solid", fgColor="D5F5E3") if v >= 0 else PatternFill("solid", fgColor="FADBD8")
                    pclr  = "1F6B3A" if v >= 0 else "8B0000"
                    sc(cell, color=pclr, fill=pfill, size=11)
                else:
                    sc(cell, fill=row_fill, size=11)
                cell.number_format = "0.00%"

            else:
                cell.value = round_or_blank(val, 2) if isinstance(val, float) else val
                sc(cell, fill=row_fill, size=11)

        tp_ws.row_dimensions[er].height = 16

    # ── Column widths ─────────────────────────────────────────────────
    for ci, (_, _, w) in enumerate(TP_COLS, 1):
        tp_ws.column_dimensions[get_column_letter(ci)].width = w

    tp_ws.freeze_panes = "C4"

    # Color scale on Thematic Score column (col 9)
    th_col    = get_column_letter(9)
    last_tp_r = 3 + len(thematic_picks)
    if last_tp_r > 4:
        tp_ws.conditional_formatting.add(
            f"{th_col}4:{th_col}{last_tp_r}",
            ColorScaleRule(
                start_type="min", start_color="F5CBA7",
                mid_type="percentile", mid_value=50, mid_color="D7BDE2",
                end_type="max", end_color="A9CCE3",
            )
        )
    # Color scale on Sector Strength% (col 6)
    sec_col = get_column_letter(6)
    if last_tp_r > 4:
        tp_ws.conditional_formatting.add(
            f"{sec_col}4:{sec_col}{last_tp_r}",
            ColorScaleRule(
                start_type="min", start_color="D6EAF8",
                end_type="max", end_color="1A5276",
            )
        )

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 6 (was 4) ── 21EMA UPTREND
    # ══════════════════════════════════════════════════════════════════
    ema_ws = wb.create_sheet("21EMA Uptrend")
    EMA_COLS = [
        ("rank", "#", 4),
        ("symbol", "Symbol", 16),
        ("sector", "Sector", 20),
        ("current_close", "Close", 10),
        ("ema21_value", "21EMA", 10),
        ("above_21ema", "Above\n21EMA", 9),
        ("ema21_slope_5d", "21EMA\nSlope 5D", 11),
        ("uptrend_consistency_lookback", "Window\nDays", 9),
        ("ema21_uptrend_since", "Uptrend\nSince", 12),
        ("ema21_uptrend_days", "Uptrend\nDays", 10),
        ("ema21_uptrend_pct", "Uptrend\nPct", 10),
        ("score_ema21_uptrend", "21EMA Up\nScore", 10),
    ]
    ncols_ema = len(EMA_COLS)
    ema_ws.merge_cells(f"A1:{get_column_letter(ncols_ema)}1")
    ema_ws["A1"] = (
        f"21EMA UPTREND BREAKDOWN   |   "
        f"As of {cutoff_date.strftime('%d-%b-%Y')}   |   "
        f"Window = max({UPTREND_CONSISTENCY_MIN_LOOKBACK}, days since reset)"
    )
    sc(ema_ws["A1"], bold=True, color="FFFFFF", fill=T_FILL, size=12, align="center")
    ema_ws.row_dimensions[1].height = 26
    for ci, (_, hdr, width) in enumerate(EMA_COLS, 1):
        cell = ema_ws.cell(row=2, column=ci, value=hdr)
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=11, wrap=True)
        ema_ws.column_dimensions[get_column_letter(ci)].width = width
    ema_ws.row_dimensions[2].height = 32

    for ri, (_, row) in enumerate(scored.iterrows(), start=3):
        afill = ALT_FILL if ri % 2 == 1 else WHT_FILL
        values = [
            row.get("rank"),
            row.get("symbol"),
            row.get("sector"),
            row.get("current_close"),
            row.get("ema21_value"),
            row.get("above_21ema"),
            row.get("ema21_slope_5d"),
            row.get("uptrend_consistency_lookback"),
            row.get("ema21_uptrend_since"),
            row.get("ema21_uptrend_days"),
            row.get("ema21_uptrend_pct"),
            row.get("score_ema21_uptrend"),
        ]
        for ci, value in enumerate(values, 1):
            cell = ema_ws.cell(row=ri, column=ci, value=value)
            if ci == 12:
                score_cell(cell, score_or_blank(value), size=11)
                cell.number_format = "0"
            elif ci in (2, 3):
                sc(cell, fill=afill, align="left", size=10, bold=(ci == 2))
            else:
                sc(cell, fill=afill, size=10)
                if ci in (4, 5):
                    cell.number_format = "#,##0.##"
                elif ci == 6:
                    cell.value = "Yes" if value else "No"
                elif ci == 7:
                    cell.number_format = "0.00"
                elif ci in (8, 10):
                    cell.number_format = "0"
                elif ci == 9:
                    cell.value = date_or_blank(value)
                elif ci == 11:
                    cell.number_format = "0.0%"
        ema_ws.row_dimensions[ri].height = 15
    ema_ws.freeze_panes = "C3"
    ema_ws.auto_filter.ref = f"A2:{get_column_letter(ncols_ema)}{max(2, len(scored)+2)}"

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 5 ── 🏭 SECTOR VIEW
    # ══════════════════════════════════════════════════════════════════
    sv = wb.create_sheet("🏭 Sector View")
    sv_hdrs = ["Symbol","Close","Score","Rating",
               "Ret from Low%","1D Ret%","Uptrend Con%"]
    sv.merge_cells("A1:G1")
    sv["A1"] = "SECTOR VIEW — stocks ranked by score within sector"
    sc(sv["A1"], bold=True, color="FFFFFF", fill=T_FILL, size=12, align="center")
    sv.row_dimensions[1].height = 24

    from collections import defaultdict
    sector_groups: Dict[str, list] = defaultdict(list)
    for _, row in scored.iterrows():
        sector_groups[row.get("sector", "Unknown")].append(row)

    sv_row = 2
    for sector_name in sorted(sector_groups.keys()):
        stocks = sorted(sector_groups[sector_name],
                        key=lambda r: -(r["composite_score"] or 0))
        sv.merge_cells(f"A{sv_row}:G{sv_row}")
        cell = sv.cell(row=sv_row, column=1,
                       value=f"  {sector_name.upper()}  ({len(stocks)} stocks)")
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=11, align="left")
        sv.row_dimensions[sv_row].height = 18
        sv_row += 1

        for hi, hdr in enumerate(sv_hdrs, 1):
            cell = sv.cell(row=sv_row, column=hi, value=hdr)
            sc(cell, bold=True, color="FFFFFF",
               fill=PatternFill("solid", fgColor="3A5F8A"), size=10)
        sv_row += 1

        for si, row in enumerate(stocks):
            score = row["composite_score"] or 0
            s_fill, s_fc = score_band(score)
            afill = ALT_FILL if si % 2 == 0 else WHT_FILL
            rating = row.get("rating", "")
            vals = [
                row["symbol"],
                r2(row["current_close"]),
                round_or_blank(score),
                rating,
                pct_fmt(row.get("reset_recovery")),
                pct_fmt(row.get("return_3m")),
                pct_fmt(row.get("uptrend_consistency_pct")),
            ]
            for ci2, val in enumerate(vals, 1):
                cell = sv.cell(row=sv_row, column=ci2, value=val)
                if ci2 in (3, 4):
                    sc(cell, bold=True, color=s_fc, fill=s_fill, size=11)
                    if ci2 == 3: cell.number_format = "0.0"
                else:
                    sc(cell, fill=afill, align="left" if ci2==1 else "center",
                       size=11)
                    if ci2 in (2, 5, 6, 7):
                        cell.number_format = ("#,##0.00" if ci2 == 2 else "0.0%")
            sv.row_dimensions[sv_row].height = 14
            sv_row += 1

        sv_row += 1  # blank separator between sectors

    for ci2, w in enumerate([18, 9, 8, 14, 12, 10, 11], 1):
        sv.column_dimensions[get_column_letter(ci2)].width = w

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 5 ── SECTOR LEADERS
    # ══════════════════════════════════════════════════════════════════
    sector_summary_base = (
        scored.sort_values(["sector", "pre_sector_score", "score_rs_total", "composite_score", "symbol"],
                           ascending=[True, False, False, False, True])
        .groupby("sector", as_index=False)
        .agg(
            sector_stock_count=("sector_stock_count", "max"),
            sector_positive_score_count=("sector_positive_score_count", "max"),
            sector_top_quartile_count=("sector_top_quartile_count", "max"),
            sector_strength_ratio=("sector_strength_ratio", "max"),
            sector_weighted_pre_score=("sector_weighted_pre_score", "max"),
            sector_weighted_rs_score=("sector_weighted_rs_score", "max"),
            sector_weighted_perf_score=("sector_weighted_perf_score", "max"),
            sector_top20_hits=("sector_top20_hits", "max"),
            sector_top20_breadth_bonus=("sector_top20_breadth_bonus", "max"),
            sector_top20_penetration_bonus=("sector_top20_penetration_bonus", "max"),
            sector_leadership_score=("sector_leadership_score", "max"),
            score_sector=("score_sector", "max"),
            best_symbol=("symbol", "first"),
            best_composite_score=("composite_score", "max"),
            avg_composite_score=("composite_score", "mean"),
            median_sector_score=("composite_score", "median"),
            avg_pre_sector_score=("pre_sector_score", "mean"),
            avg_rs_total=("score_rs_total", "mean"),
            avg_perf_total=("score_performance_total", "mean"),
            count_above_20_score=("composite_score", lambda s: int((s >= 20).sum())),
            count_top_20=("symbol", lambda s: int(sum(1 for sym in s if sym in top20["symbol"].values))),
        )
    )

    leader_rows: List[Dict[str, object]] = []
    for sector_name, grp in (
        scored.sort_values(["sector", "pre_sector_score", "score_rs_total", "composite_score", "symbol"],
                           ascending=[True, False, False, False, True])
        .groupby("sector")
    ):
        if len(grp) < 2:
            continue
        top2 = grp.head(2)
        leader_rows.append(
            {
                "sector": sector_name,
                "sector_stock_count": int(len(grp)),
                "sector_positive_score_count": int(grp["sector_positive_score_count"].iloc[0]),
                "leader_1": top2.iloc[0]["symbol"],
                "leader_2": top2.iloc[1]["symbol"],
                "leader_avg_pre_sector": float(grp["sector_weighted_pre_score"].iloc[0]),
                "leader_avg_rs_total": float(grp["sector_weighted_rs_score"].iloc[0]),
                "leader_avg_perf_total": float(grp["sector_weighted_perf_score"].iloc[0]),
                "leader_strength_ratio": float(grp["sector_strength_ratio"].iloc[0]),
                "leader_best_score": float(grp["composite_score"].max()),
                "sector_top20_hits": int(grp["sector_top20_hits"].iloc[0]),
                "leadership_score": float(grp["sector_leadership_score"].iloc[0]),
                "score_sector": int(grp["score_sector"].iloc[0]),
            }
        )

    sector_leaders = pd.DataFrame(leader_rows)
    if not sector_leaders.empty:
        sector_leaders = sector_leaders.sort_values(
            ["leadership_score", "score_sector", "sector_positive_score_count", "sector"],
            ascending=[False, False, False, True],
        ).reset_index(drop=True)

    sl_ws = wb.create_sheet("Sector Leaders")
    SL_COLS = [
        ("Rank", 5), ("Sector", 22), ("# Stocks", 7), ("# Pos", 6),
        ("Sector Score", 11), ("Sector Pts", 9),
        ("Leader 1", 14), ("Leader 2", 14), ("Avg RS", 8), ("Avg Pre-Sec", 10),
        ("Avg Perf", 9), ("Str%", 8), ("Best Score", 9),
    ]
    ncols_sl = len(SL_COLS)
    sl_ws.merge_cells(f"A1:{get_column_letter(ncols_sl)}1")
    sl_ws["A1"] = (
        f"SECTOR LEADERSHIP BOARD   |   As of {cutoff_date.strftime('%d-%b-%Y')}   "
        f"|   Sorted by actual sector score first, then sector points"
    )
    sc(sl_ws["A1"], bold=True, color="FFFFFF", fill=T_FILL, size=12, align="center")
    sl_ws.row_dimensions[1].height = 24
    for ci, (hdr, w) in enumerate(SL_COLS, 1):
        cell = sl_ws.cell(row=2, column=ci, value=hdr)
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=10)
        sl_ws.column_dimensions[get_column_letter(ci)].width = w
    qualifying_leaders = (
        sector_leaders[sector_leaders["score_sector"] > 0].reset_index(drop=True)
        if not sector_leaders.empty else sector_leaders
    )
    for ri, (_, lrow) in enumerate(qualifying_leaders.iterrows(), start=3):
        rank2 = ri - 2
        row_bg = medal_fills[rank2 - 1] if rank2 <= 3 else (ALT_FILL if rank2 % 2 == 0 else WHT_FILL)
        pts_val = int(lrow.get("score_sector", 0) or 0)
        if pts_val >= 8:
            pts_fill = PatternFill("solid", fgColor="FFD700"); pts_fc = "000000"
        elif pts_val >= 6:
            pts_fill = PatternFill("solid", fgColor="2E8B57"); pts_fc = "FFFFFF"
        elif pts_val >= 4:
            pts_fill = PatternFill("solid", fgColor="90EE90"); pts_fc = "000000"
        else:
            pts_fill = WHT_FILL; pts_fc = "000000"
        vals = [
            rank2,
            lrow.get("sector", ""),
            lrow.get("sector_stock_count", ""),
            lrow.get("sector_positive_score_count", ""),
            round_or_blank(lrow.get("leadership_score"), 1),
            pts_val,
            lrow.get("leader_1", ""),
            lrow.get("leader_2", ""),
            round_or_blank(lrow.get("leader_avg_rs_total"), 1),
            round_or_blank(lrow.get("leader_avg_pre_sector"), 1),
            round_or_blank(lrow.get("leader_avg_perf_total"), 1),
            pct_fmt(lrow.get("leader_strength_ratio")),
            round_or_blank(lrow.get("leader_best_score"), 1),
        ]
        for ci, (val, _) in enumerate(zip(vals, SL_COLS), 1):
            cell = sl_ws.cell(row=ri, column=ci, value=val)
            if ci == 5:  # Sector Score
                s_fill, s_fc = score_band(float(lrow.get("leadership_score", 0) or 0))
                sc(cell, bold=True, color=s_fc, fill=s_fill, size=11)
                cell.number_format = "0.0"
            elif ci == 6:  # Sector Pts
                sc(cell, bold=True, color=pts_fc, fill=pts_fill, size=11)
                cell.number_format = "0"
            else:
                sc(cell, fill=row_bg, align="left" if ci in (2, 7, 8) else "center", size=11)
            if ci == 12:
                cell.number_format = "0.0%"
    sl_ws.freeze_panes = "A3"
    sl_ws.auto_filter.ref = f"A2:{get_column_letter(ncols_sl)}{max(2, len(qualifying_leaders)+2)}"

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 6 ── 📋 SECTOR SUMMARY  (aggregated)
    # ══════════════════════════════════════════════════════════════════
    ss_ws = wb.create_sheet("📋 Sector Summary")

    sector_summary = sector_summary_base.merge(
        sector_leaders[[
            "sector",
            "leadership_score",
            "leader_1",
            "leader_2",
            "leader_avg_rs_total",
            "leader_avg_pre_sector",
            "leader_avg_perf_total",
        ]] if not sector_leaders.empty else pd.DataFrame(columns=[
            "sector", "leadership_score", "leader_1", "leader_2",
            "leader_avg_rs_total", "leader_avg_pre_sector", "leader_avg_perf_total",
        ]),
        on="sector",
        how="left",
    )
    sector_summary["leadership_score"] = sector_summary["leadership_score"].fillna(-1.0)
    sector_summary.sort_values(
        ["leadership_score", "score_sector", "sector_positive_score_count", "avg_composite_score", "avg_pre_sector_score", "avg_rs_total"],
        ascending=[False, False, False, False, False, False],
        inplace=True,
    )
    sector_summary.reset_index(drop=True, inplace=True)

    SS_COLS = [
        ("Rank",               5), ("Sector",            22),
        ("# Stocks",           7), ("# Pos",             6), ("Sector Score",     10),
        ("Leader 1",          14), ("Leader 2",         14),
        ("Best Symbol",       14), ("Best Score",        8),
        ("Avg Score",          8), ("Median Score",      8),
        ("Avg RS",             8), ("Avg Pre-Sec",      10),
        ("# Top-20",           7), ("Breadth Pts",       8), ("Pen Pts",          7),
        ("# Score≥20",         7), ("Top-Qrt Count",     9), ("Strength Ratio",   9),
        ("Sector Pts",         8),
    ]
    ncols_ss = len(SS_COLS)

    ss_ws.merge_cells(f"A1:{get_column_letter(ncols_ss)}1")
    ss_ws["A1"] = (
        f"SECTOR SUMMARY   |   As of {cutoff_date.strftime('%d-%b-%Y')}   "
        f"|   Ranked by actual sector score first "
        f"(AvgPreSec + Top-20 Breadth Bonus + Penetration Bonus + 12M/52W Bonus)"
    )
    sc(ss_ws["A1"], bold=True, color="FFFFFF", fill=T_FILL, size=12, align="center")
    ss_ws.row_dimensions[1].height = 24

    for ci, (hdr, _) in enumerate(SS_COLS, 1):
        cell = ss_ws.cell(row=2, column=ci, value=hdr)
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=11, wrap=True)
    ss_ws.row_dimensions[2].height = 28

    for ri, (_, srow) in enumerate(sector_summary.iterrows(), start=3):
        rank2 = ri - 2
        row_bg = (medal_fills[rank2-1] if rank2 <= 3
                  else ALT_FILL if rank2 % 2 == 0 else WHT_FILL)
        avg_sc = round(float(srow.get("avg_composite_score", 0) or 0), 1)
        s_fill2, s_fc2 = score_band(avg_sc)

        vals = [
            rank2,
            srow.get("sector", ""),
            srow.get("sector_stock_count", ""),
            srow.get("sector_positive_score_count", ""),
            round_or_blank(srow.get("leadership_score"), 1) if float(srow.get("leadership_score", -1) or -1) >= 0 else "",
            srow.get("leader_1", ""),
            srow.get("leader_2", ""),
            srow.get("best_symbol", ""),
            round_or_blank(srow.get("best_composite_score")),
            round(avg_sc, 1),
            round_or_blank(srow.get("median_sector_score"), 1),
            round_or_blank(srow.get("avg_rs_total"), 1),
            round_or_blank(srow.get("avg_pre_sector_score"), 1),
            srow.get("count_top_20", ""),
            round_or_blank(srow.get("sector_top20_breadth_bonus")),
            round_or_blank(srow.get("sector_top20_penetration_bonus")),
            srow.get("count_above_20_score", ""),
            srow.get("sector_top_quartile_count", ""),
            pct_fmt(srow.get("sector_strength_ratio")),
            round_or_blank(srow.get("score_sector")),
        ]
        for ci, (val, (_, w)) in enumerate(zip(vals, SS_COLS), 1):
            cell = ss_ws.cell(row=ri, column=ci, value=val)
            if ci == 5:   # Sector Score — colour-coded
                lead_val = float(srow.get("leadership_score", 0) or 0)
                lead_fill, lead_fc = score_band(lead_val)
                sc(cell, bold=True, color=lead_fc, fill=lead_fill, size=11, border=BDR_M)
                cell.number_format = "0.0"
            elif ci == 10:   # Avg Score — colour-coded
                sc(cell, bold=True, color=s_fc2, fill=s_fill2, size=11,
                   border=BDR_M)
                cell.number_format = "0.0"
            elif ci == 2:
                sc(cell, bold=(rank2<=3), fill=row_bg, align="left", size=11)
            elif ci == 1:
                sc(cell, bold=True, fill=row_bg, size=11)
            elif ci == ncols_ss:  # Sector Pts — colour by tier
                pts = val if isinstance(val, (int, float)) else 0
                if pts >= 8:
                    sp_fill = PatternFill("solid", fgColor="FFD700")
                    sp_fc = "000000"
                elif pts >= 6:
                    sp_fill = PatternFill("solid", fgColor="2E8B57")
                    sp_fc = "FFFFFF"
                elif pts >= 4:
                    sp_fill = PatternFill("solid", fgColor="90EE90")
                    sp_fc = "000000"
                else:
                    sp_fill = row_bg
                    sp_fc = "808080"
                sc(cell, bold=(pts >= 4), color=sp_fc, fill=sp_fill, size=11)
                cell.number_format = "0"
            else:
                sc(cell, fill=row_bg, size=11)
                if ci in (5, 7, 12):
                    cell.number_format = "0.0"
        ss_ws.row_dimensions[ri].height = 14

    for ci, (_, w) in enumerate(SS_COLS, 1):
        ss_ws.column_dimensions[get_column_letter(ci)].width = w
    ss_ws.freeze_panes = "B3"

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 6 ── ℹ️ SUMMARY  (run metadata)
    # ══════════════════════════════════════════════════════════════════
    summ = wb.create_sheet("ℹ️ Summary")
    rr_ws = wb.create_sheet("📘 Ready Reckoner")

    def rr_value(key: str, formatter=None) -> object:
        if scored.empty or key not in scored.columns:
            return ""
        value = scored.iloc[0][key]
        if pd.isna(value):
            return ""
        return formatter(value) if formatter else round_or_blank(value, 2)

    RR_COLS = [
        ("Category", 18),
        ("Metric", 26),
        ("Rule", 56),
        ("Points", 12),
        ("Current Threshold / Notes", 34),
    ]
    rr_ws.merge_cells(f"A1:{get_column_letter(len(RR_COLS))}1")
    rr_ws["A1"] = (
        f"READY RECKONER   |   As of {cutoff_date.strftime('%d-%b-%Y')}   "
        f"|   Summary of points and scoring criteria used by this run"
    )
    sc(rr_ws["A1"], bold=True, color="FFFFFF", fill=T_FILL, size=12, align="center")
    rr_ws.row_dimensions[1].height = 24

    for ci, (hdr, width) in enumerate(RR_COLS, 1):
        cell = rr_ws.cell(row=2, column=ci, value=hdr)
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=11, wrap=True)
        rr_ws.column_dimensions[get_column_letter(ci)].width = width
    rr_ws.row_dimensions[2].height = 28

    rr_rows = [
        ("Liquidity", "Universe inclusion", "Average 42D turnover must be available", "Required", "Missing value => excluded"),
        ("Liquidity", "Universe inclusion", f"Excluded if median < {MIN_MEDIAN_TURNOVER_CRORES:.0f} Cr AND avg < {MIN_AVG_TURNOVER_42D_CRORES:.0f} Cr (42D)", "Required", f"Both below threshold => excluded from universe"),
        ("Liquidity", "Inst Picks gate", f"Excluded from inst picks if median 42D turnover < {INST_PICKS_MIN_MEDIAN_TURNOVER_CRORES:.0f} Cr", "Required", f"Median < {INST_PICKS_MIN_MEDIAN_TURNOVER_CRORES:.0f} Cr => excluded from institutional picks"),
        ("Liquidity", "Avg turnover 42D", "Bottom 10% of rated universe AND avg TO < 40 Cr", "-8", rr_value("turnover_bottom10")),
        ("Liquidity", "Avg turnover 42D", "Bottom 20% of rated universe AND avg TO < 40 Cr", "-6", rr_value("turnover_bottom20")),
        ("Liquidity", "Avg turnover 42D", "Bottom 30% of rated universe AND avg TO < 40 Cr", "-4", rr_value("turnover_bottom30")),
        ("Liquidity", "Median turnover 42D", f"Bottom 10% of universe AND < {MEDIAN_TURNOVER_LOW_THRESHOLD_CRORES:.0f} Cr", "-8", rr_value("median_turnover_bottom10")),
        ("Liquidity", "Median turnover 42D", f"Bottom 20% of universe AND < {MEDIAN_TURNOVER_LOW_THRESHOLD_CRORES:.0f} Cr", "-6", rr_value("median_turnover_bottom20")),
        ("Liquidity", "Liquid leaders bonus", "Top 10 symbols in liquid leaders source order", "+6", "Based on updated_gmlist order"),
        ("Liquidity", "Liquid leaders bonus", "Top 20 symbols in liquid leaders source order", "+3", "Ranks 11-20"),
        ("52W High", "Distance from 52W high", "<= 10% / 15% / 20% from 52W high", "+10 / +6 / +2", "Daily high basis"),
        ("52W High", "Distance penalty", "> 20% and < 25% from 52W high", "-4", "Applied even if 52W high is recent"),
        ("52W High", "Distance penalty", ">= 25% and < 30% from 52W high", "-10", "Applied even if 52W high is recent"),
        ("52W High", "Distance penalty", ">= 30% from 52W high", "-16", "Applied even if 52W high is recent"),
        ("52W High", "Recency of 52W high", "Within 10 trading days and still within 20% of 52W high", "+2", "No recency points beyond 20% away"),
        ("52W High", "Adjustment", "Old 52W bonus stacking removed", "0", "No duplicate 52W points"),
        ("Trend", "50DMA", "Close above 50DMA", "+4", ""),
        ("Trend", "50DMA", "Close below 50DMA", "-4", ""),
        ("Trend", "21EMA", "Close above 21EMA", "+2", ""),
        ("Trend", "21EMA uptrend", f"Consecutive cutoff-ending uptrend days >= 20 / 10 over max({UPTREND_CONSISTENCY_MIN_LOOKBACK}, reset-window)", "+4 / +2", "Needs close > 21EMA and rising 21EMA"),
        ("Trend", "21EMA uptrend consistency", f"Window consistency >= 90% / 80% over max({UPTREND_CONSISTENCY_MIN_LOOKBACK}, reset-window)", "+2 / +1", "Fallback credit if cutoff-ending streak is shorter"),
        ("Trend", "8EMA", "Close above 8EMA", "+2", ""),
        ("Trend", "Uptrend consistency", f"Top 30% of rated universe using max({UPTREND_CONSISTENCY_MIN_LOOKBACK}, reset-window) trading days", "+4", rr_value("uptrend_consistency_top30", percent_or_blank)),
        ("Trend", "Green candle count", "Top 10% of rated universe", "+4", rr_value("green_candle_top10")),
        ("Trend", "Green candle count", "Top 20% of rated universe", "+2", rr_value("green_candle_top20")),
        ("Trend", "Recent listings", "Listed < 60 days", "Median trend scores; no negative trend penalty", "Applied to 50DMA/21EMA/8EMA/uptrend/green"),
        ("Performance", "12M return", "Top 10% of rated universe", "+8", rr_value("ret12_top10", percent_or_blank)),
        ("Performance", "12M return", "Top 20% of rated universe", "+6", rr_value("ret12_top20", percent_or_blank)),
        ("Performance", "12M return", "Top 30% of rated universe", "+4", rr_value("ret12_top30", percent_or_blank)),
        ("Performance", "12M return", "Bottom 10% of mature listings", "-6", rr_value("ret12_bottom10", percent_or_blank)),
        ("Performance", "12M return", "Bottom 20% of mature listings", "-4", rr_value("ret12_bottom20", percent_or_blank)),
        ("Performance", "12M return", "Insufficient listing history (< 250 trading days)", "Median score", "Uses universe median 12M score"),
        ("Performance", "6M return", "Top 10% of rated universe", "+6", rr_value("ret6_top10", percent_or_blank)),
        ("Performance", "6M return", "Top 20% of rated universe", "+4", rr_value("ret6_top20", percent_or_blank)),
        ("Performance", "6M return", "Top 30% of rated universe", "+2", rr_value("ret6_top30", percent_or_blank)),
        ("Performance", "6M return", "Bottom 10% of mature listings", "-2", rr_value("ret6_bottom10", percent_or_blank)),
        ("Performance", "6M return", "Insufficient listing history (< 125 trading days)", "Median score", "Uses universe median 6M score and removes 6M penalty"),
        ("Performance", "3M return", "Top 10% of rated universe", "+6", rr_value("ret3_top10", percent_or_blank)),
        ("Performance", "3M return", "Top 20% of rated universe", "+4", rr_value("ret3_top20", percent_or_blank)),
        ("Performance", "3M return", "Top 30% of rated universe", "+2", rr_value("ret3_top30", percent_or_blank)),
        ("Performance", "3M return", "Bottom 10% of mature listings", "-4", rr_value("ret3_bottom10", percent_or_blank)),
        ("Performance", "1D return", "Top 10% of rated universe", "+4", rr_value("ret1_top10", percent_or_blank)),
        ("Performance", "1D return", "Top 20% of rated universe", "+2", rr_value("ret1_top20", percent_or_blank)),
        ("Performance", "1D return", "Bottom 20% of universe AND return < -3%", "-4", rr_value("ret1_bottom20", percent_or_blank)),
        ("Performance", "Bonus", "12M top-10 stock with 52W high in last 10 days", "+2", ""),
        ("Performance", "Reset recovery rank", "Top 10 by reset recovery", "+4", "Rank based"),
        ("Performance", "Reset recovery rank", "Top 20 by reset recovery", "+2", "Rank based"),
        ("Relative Strength", "RS 12M", "Top 10% of rated universe", "+6", rr_value("rs12_top10", percent_or_blank)),
        ("Relative Strength", "RS 12M", "Top 30% of rated universe", "+3", rr_value("rs12_top30", percent_or_blank)),
        ("Relative Strength", "RS 6M", "Top 10% of rated universe", "+4", rr_value("rs6_top10", percent_or_blank)),
        ("Relative Strength", "RS 6M", "Top 30% of rated universe", "+2", rr_value("rs6_top30", percent_or_blank)),
        ("Relative Strength", "RS 3M", "Top 10% of rated universe", "+4", rr_value("rs3_top10", percent_or_blank)),
        ("Relative Strength", "RS 3M", "Top 30% of rated universe", "+2", rr_value("rs3_top30", percent_or_blank)),
        ("Relative Strength", "RS line", "RS line at 52W high", "+2", ""),
        ("Relative Strength", "RS slope 21D", "Negative RS slope", "-2", ""),
        ("Volatility", "ATR% 21D", "Bottom 10% and ATR% < 3.0", "-6", rr_value("atr_bottom10")),
        ("Volatility", "ATR% 21D", "Bottom 10% and ATR% < 4.0", "-3", rr_value("atr_bottom10")),
        ("Event", "Volume spike", "Top 1% volume in lookback, within 10D / 30D / 60D", "+10 / +8 / +6", "Bonus +4 if price change > 6%; +2 if > 3%"),
        ("Event", "Gap-up", "Gap-up > 3%, close gain > 5%, and either close gain >= 9% or top-1% volume", "+6", "Recent gap-up window"),
        ("Listing", "New listing", "Listed < 30 days", "+4", ""),
        ("Performance", "12M return + 52W proximity", "Top 30% 12M return AND within 15% of 52W high", "+4", "Bonus on top of regular 12M score"),
        ("IPO Performance", "IPO gain vs issue price", "Top 10% of IPOs listed < 6M with CMP > issue price", f"+{IPO_TOP10_POINTS}", "Percentile within IPO sub-universe"),
        ("IPO Performance", "IPO gain vs issue price", "Top 20% of IPOs listed < 6M with CMP > issue price", f"+{IPO_TOP20_POINTS}", "Percentile within IPO sub-universe"),
        ("IPO Performance", "IPO gain vs issue price", "Additional: Top 40% of IPOs listed < 6M with CMP > issue price and IPO gain >= 20%", f"+{IPO_TOP40_POINTS}", "Added on top of the regular IPO percentile score"),
        ("Listing", "New listing", "Listed < 60 days", "+2", ""),
        ("Sector", "Sector strength", "Top 10% eligible sectors by leadership score", f"+{SECTOR_TOP10_POINTS}", rr_value("sector_top10_threshold")),
        ("Sector", "Sector strength", "Top 20% eligible sectors by leadership score", f"+{SECTOR_TOP20_POINTS}", rr_value("sector_top20_threshold")),
        ("Sector", "Sector strength", "Top 30% eligible sectors by leadership score", f"+{SECTOR_TOP30_POINTS}", rr_value("sector_top30_threshold")),
        ("Sector", "Eligibility", "Sector must have more than 1 positive-score stock", "Required", ""),
        ("Sector", "Top-20 Breadth Bonus", "3+ sector stocks in Top-20 dashboard", f"+{SECTOR_TOP20_BREADTH_BONUS_3PLUS} to Leader Score", "Added before top-10%/30% threshold"),
        ("Sector", "Top-20 Breadth Bonus", "2 sector stocks in Top-20 dashboard", f"+{SECTOR_TOP20_BREADTH_BONUS_2} to Leader Score", "Added before top-10%/30% threshold"),
        ("Sector", "Top-20 Breadth Bonus", "1 sector stock in Top-20 dashboard", f"+{SECTOR_TOP20_BREADTH_BONUS_1} to Leader Score", "Added before top-10%/30% threshold"),
        ("Sector", "Top-20 Penetration Bonus", "100% of sector stocks in Top-20 (min 2 stocks)", f"+{SECTOR_TOP20_PENETRATION_BONUS_100} to Leader Score", "Added before top-10%/30% threshold"),
        ("Sector", "Top-20 Penetration Bonus", ">=67% of sector stocks in Top-20 (min 2 stocks)", f"+{SECTOR_TOP20_PENETRATION_BONUS_67} to Leader Score", "Added before top-10%/30% threshold"),
        ("Sector", "Top-20 Penetration Bonus", ">=50% of sector stocks in Top-20 (min 2 stocks)", f"+{SECTOR_TOP20_PENETRATION_BONUS_50} to Leader Score", "Added before top-10%/30% threshold"),
        ("Sector", "Top quartile reference", "Pre-sector score top quartile", "Reference", rr_value("top_quartile_score_threshold")),
    ]

    for ri, row_vals in enumerate(rr_rows, start=3):
        row_fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        for ci, value in enumerate(row_vals, start=1):
            cell = rr_ws.cell(row=ri, column=ci, value=value)
            sc(cell, fill=row_fill, align="left", size=11, wrap=(ci in (2, 3, 5)))
    rr_ws.freeze_panes = "A3"
    rr_ws.auto_filter.ref = f"A2:E{len(rr_rows) + 2}"

    summary_rows = [
        ("Metric", "Value"),
        ("Cutoff Date",                   cutoff_date.isoformat()),
        ("Reset Date",                    reset_date.isoformat()),
        ("Symbol File",                   symbol_file.name),
        ("Symbols Rated",                 int(len(scored))),
        ("─── Scoring Windows ───",        ""),
        ("12M / 52W Window",              "250 trading candles"),
        ("6M Window",                     "125 trading candles"),
        ("3M Window",                      "60 trading candles"),
        ("Uptrend Consistency Window",    f"max({UPTREND_CONSISTENCY_MIN_LOOKBACK} trading candles, days since reset)"),
        ("─── Turnover ───",               ""),
        ("Turnover Unit",                  "Crores"),
        ("Universe exclusion (Cr)",        f"Median < {MIN_MEDIAN_TURNOVER_CRORES:.0f} Cr AND Avg < {MIN_AVG_TURNOVER_42D_CRORES:.0f} Cr (42D)"),
        ("Inst Picks median gate (Cr)",   f"< {INST_PICKS_MIN_MEDIAN_TURNOVER_CRORES:.0f} Cr median 42D => excluded from inst picks"),
        ("─── Thresholds ───",             ""),
        ("Top Score",                     round_or_blank(scored["composite_score"].max()) if not scored.empty else ""),
        ("Bottom Score",                  round_or_blank(scored["composite_score"].min()) if not scored.empty else ""),
        ("Liquidity Bottom-10 Threshold", round_or_blank(scored["turnover_bottom10"].iloc[0])          if not scored.empty else ""),
        ("Liquidity Bottom-20 Threshold", round_or_blank(scored["turnover_bottom20"].iloc[0])          if not scored.empty else ""),
        ("Liquidity Bottom-30 Threshold", round_or_blank(scored["turnover_bottom30"].iloc[0])          if not scored.empty else ""),
        ("Med TO Bottom-10% Threshold",   round_or_blank(scored["median_turnover_bottom10"].iloc[0])    if not scored.empty else ""),
        ("Med TO Bottom-20% Threshold",   round_or_blank(scored["median_turnover_bottom20"].iloc[0])    if not scored.empty else ""),
        ("ATR% Bottom-10 Threshold",      round_or_blank(scored["atr_bottom10"].iloc[0])               if not scored.empty else ""),
        ("Index Return 12M",              percent_or_blank(scored["index_return_12m"].iloc[0])          if not scored.empty else ""),
        ("Index Return 6M",               percent_or_blank(scored["index_return_6m"].iloc[0])           if not scored.empty else ""),
        ("Index Return 3M",               percent_or_blank(scored["index_return_3m"].iloc[0])           if not scored.empty else ""),
        ("12M Bottom-10 Threshold",        percent_or_blank(scored["ret12_bottom10"].iloc[0])             if not scored.empty else ""),
        ("12M Bottom-20 Threshold",        percent_or_blank(scored["ret12_bottom20"].iloc[0])             if not scored.empty else ""),
        ("12M Top-10 Threshold",          percent_or_blank(scored["ret12_top10"].iloc[0])               if not scored.empty else ""),
        ("12M Top-20 Threshold",          percent_or_blank(scored["ret12_top20"].iloc[0])               if not scored.empty else ""),
        ("12M Top-30 Threshold",          percent_or_blank(scored["ret12_top30"].iloc[0])               if not scored.empty else ""),
        ("6M Top-10 Threshold",           percent_or_blank(scored["ret6_top10"].iloc[0])                if not scored.empty else ""),
        ("6M Top-20 Threshold",           percent_or_blank(scored["ret6_top20"].iloc[0])                if not scored.empty else ""),
        ("6M Top-30 Threshold",           percent_or_blank(scored["ret6_top30"].iloc[0])                if not scored.empty else ""),
        ("3M Top-10 Threshold",           percent_or_blank(scored["ret3_top10"].iloc[0])                if not scored.empty else ""),
        ("3M Top-20 Threshold",           percent_or_blank(scored["ret3_top20"].iloc[0])                if not scored.empty else ""),
        ("3M Top-30 Threshold",           percent_or_blank(scored["ret3_top30"].iloc[0])                if not scored.empty else ""),
        ("Green Candles Top-10 Threshold",round_or_blank(scored["green_candle_top10"].iloc[0])         if not scored.empty else ""),
        ("Green Candles Top-20 Threshold",round_or_blank(scored["green_candle_top20"].iloc[0])         if not scored.empty else ""),
        ("RS 12M Top-10 Threshold",       percent_or_blank(scored["rs12_top10"].iloc[0])                if not scored.empty else ""),
        ("RS 12M Top-30 Threshold",       percent_or_blank(scored["rs12_top30"].iloc[0])                if not scored.empty else ""),
        ("RS 6M Top-10 Threshold",        percent_or_blank(scored["rs6_top10"].iloc[0])                 if not scored.empty else ""),
        ("RS 6M Top-30 Threshold",        percent_or_blank(scored["rs6_top30"].iloc[0])                 if not scored.empty else ""),
        ("RS 3M Top-10 Threshold",        percent_or_blank(scored["rs3_top10"].iloc[0])                 if not scored.empty else ""),
        ("RS 3M Top-30 Threshold",        percent_or_blank(scored["rs3_top30"].iloc[0])                 if not scored.empty else ""),
        ("Uptrend Top-30 Threshold",      percent_or_blank(scored["uptrend_consistency_top30"].iloc[0]) if not scored.empty else ""),
        ("Pre-Sector Top-Qtrl Threshold", round_or_blank(scored["top_quartile_score_threshold"].iloc[0])if not scored.empty else ""),
        ("Sector Leadership Top-10",      round_or_blank(scored["sector_top10_threshold"].iloc[0])      if not scored.empty else ""),
        ("Sector Leadership Top-20",      round_or_blank(scored["sector_top20_threshold"].iloc[0])      if not scored.empty else ""),
        ("Sector Leadership Top-30",      round_or_blank(scored["sector_top30_threshold"].iloc[0])      if not scored.empty else ""),
        ("─── Assumptions ───",            ""),
        ("Sector Scoring",                "Top 10% sectors = +4pts;  Top 30% = +2pts"),
        ("Sector Leaders",                "Top 5 sectors, min 2 positive-score stocks; higher positive-stock count takes precedence, then average stock score"),
        ("52W High Basis",                "Daily HIGH (not close)"),
    ]
    for ri2, (label, val) in enumerate(summary_rows, start=1):
        cell_a = summ.cell(row=ri2, column=1, value=label)
        cell_b = summ.cell(row=ri2, column=2, value=val)
        if ri2 == 1:
            sc(cell_a, bold=True, color="FFFFFF", fill=H_FILL, size=11)
            sc(cell_b, bold=True, color="FFFFFF", fill=H_FILL, size=11)
        elif str(label).startswith("───"):
            sc(cell_a, bold=True, italic=True, color="FFFFFF",
               fill=PatternFill("solid", fgColor="3A5F8A"), size=11, align="left")
            sc(cell_b, bold=False, color="FFFFFF",
               fill=PatternFill("solid", fgColor="3A5F8A"), size=11)
        else:
            sc(cell_a, fill=ALT_FILL if ri2 % 2 == 0 else WHT_FILL,
               align="left", size=11)
            sc(cell_b, fill=ALT_FILL if ri2 % 2 == 0 else WHT_FILL,
               align="left", size=11)
    summ.column_dimensions["A"].width = 34
    summ.column_dimensions["B"].width = 42

    # ══════════════════════════════════════════════════════════════════
    #  SHEET 7 ── ⚠️ WARNINGS
    # ══════════════════════════════════════════════════════════════════
    warn_ws = wb.create_sheet("⚠️ Warnings")
    exclusions = [item for item in warnings if "excluded from rating universe" in item.message.lower()]
    other_warnings = [item for item in warnings if item not in exclusions]

    excl_ws = wb.create_sheet("Exclusions")
    for ci, hdr in enumerate(["Symbol", "Reason", "Avg TO 21D (Cr)", "Avg TO 42D (Cr)", "Median TO 42D (Cr)"], 1):
        cell = excl_ws.cell(row=1, column=ci, value=hdr)
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=11)
    for ri2, item in enumerate(exclusions, start=2):
        row_fill = ALT_FILL if ri2 % 2 == 0 else WHT_FILL
        sc(excl_ws.cell(row=ri2, column=1, value=item.symbol), fill=row_fill, align="left", size=11)
        sc(excl_ws.cell(row=ri2, column=2, value=item.message), fill=row_fill, align="left", size=11)
        avg21_cell = excl_ws.cell(row=ri2, column=3, value=item.avg_turnover_21d)
        avg42_cell = excl_ws.cell(row=ri2, column=4, value=item.avg_turnover_42d)
        med_cell = excl_ws.cell(row=ri2, column=5, value=item.median_turnover_42d)
        sc(avg21_cell, fill=row_fill, size=11, num_fmt="#,##0.00")
        sc(avg42_cell, fill=row_fill, size=11, num_fmt="#,##0.00")
        sc(med_cell, fill=row_fill, size=11, num_fmt="#,##0.00")
    excl_ws.auto_filter.ref = f"A1:E{max(1, len(exclusions)+1)}"
    excl_ws.freeze_panes = "A2"
    excl_ws.column_dimensions["A"].width = 16
    excl_ws.column_dimensions["B"].width = 90
    excl_ws.column_dimensions["C"].width = 18
    excl_ws.column_dimensions["D"].width = 18
    excl_ws.column_dimensions["E"].width = 20

    for ci, hdr in enumerate(["Symbol", "Warning Message"], 1):
        cell = warn_ws.cell(row=1, column=ci, value=hdr)
        sc(cell, bold=True, color="FFFFFF", fill=H_FILL, size=11)
    for ri2, item in enumerate(other_warnings, start=2):
        sc(warn_ws.cell(row=ri2, column=1, value=item.symbol),
           fill=ALT_FILL if ri2 % 2 == 0 else WHT_FILL,
           align="left", size=11)
        sc(warn_ws.cell(row=ri2, column=2, value=item.message),
           fill=ALT_FILL if ri2 % 2 == 0 else WHT_FILL,
           align="left", size=11)
    warn_ws.auto_filter.ref = f"A1:B{max(1, len(other_warnings)+1)}"
    warn_ws.freeze_panes = "A2"
    warn_ws.column_dimensions["A"].width = 16
    warn_ws.column_dimensions["B"].width = 90

    # ── Set active sheet to Dashboard ────────────────────────────────
    wb.active = dash
    wb.save(out_path)
    return out_path


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    args = parse_args()
    cutoff_date: date = args.cutoff_date
    reset_date:  date = args.reset_date
    output_dir        = Path(args.output_dir)
    token_file        = Path(args.token)
    index_symbol: str = args.index_symbol

    print("=" * 70)
    print("  NSE STOCK TECHNICAL RATING")
    print(f"  Cutoff date  : {cutoff_date.strftime('%d-%b-%Y')}")
    print(f"  Reset date   : {reset_date.strftime('%d-%b-%Y')}")
    print(f"  Token file   : {token_file}")
    print(f"  Output dir   : {output_dir}")
    print(f"  Index symbol : {index_symbol}")
    print("=" * 70)

    # ── Resolve symbol file ───────────────────────────────────────────────
    symbol_file = locate_symbol_file(cutoff_date, args.symbols, args.symbols_dir)
    symbols     = read_symbols(symbol_file)
    print(f"\n  Symbol file  : {symbol_file}  ({len(symbols)} symbols)")

    # ── DB connection ─────────────────────────────────────────────────────
    conn = get_db_connection()
    print("  MySQL bhav DB : connected")
    ensure_inactive_symbols_table(conn)

    warnings: List[WarningLog] = []
    symbol_replacements = load_symbol_replacements(conn)
    remapped_pairs = [
        (symbol, symbol_replacements[symbol.upper()])
        for symbol in symbols
        if symbol.upper() in symbol_replacements and symbol_replacements[symbol.upper()] != symbol.upper()
    ]
    liquid_leader_bonus_map = parse_liquid_leader_bonus_map(args.liquid_leader_map, symbol_replacements)
    if liquid_leader_bonus_map:
        print(f"  Liquid leader bonus source : inline neo map ({len(liquid_leader_bonus_map)} mapped keys)")
    else:
        print("  Liquid leader bonus source : none provided")
    if remapped_pairs:
        print("  Using mapped replacement symbols from bhav.inactive_symbols for Kite lookups:")
        for old_symbol, new_symbol in remapped_pairs:
            print(f"    {old_symbol} -> {new_symbol}  (turnover keeps old bhav symbol)")
            warnings.append(
                WarningLog(old_symbol, f"Using active Kite symbol {new_symbol}; bhav turnover remains on original symbol.")
            )
    effective_kite_symbols = remap_symbols(symbols, symbol_replacements)

    # ── Sector map ────────────────────────────────────────────────────────
    sector_map = load_sector_map(conn, list(dict.fromkeys(symbols + effective_kite_symbols)))

    # ── Date window for history — exactly 1 year back from cutoff ────────
    fetch_start = cutoff_date - timedelta(days=365)
    fetch_start = max(fetch_start, date(2000, 1, 1))

    # ── Index history (for RS) ────────────────────────────────────────────
    index_df = load_index_history(conn, fetch_start, cutoff_date, index_symbol)

    # ── Turnover map (liquidity filter) ───────────────────────────────────
    turnover_map = load_turnover_map(conn, symbols, fetch_start, cutoff_date)
    ipo_issue_price_map = load_ipo_issue_price_map(conn, symbols, cutoff_date)
    if ipo_issue_price_map:
        print(f"  IPO issue prices loaded for: {list(ipo_issue_price_map.keys())}")
    eligible = filter_symbols_by_turnover(
        symbols,
        turnover_map,
        MIN_MEDIAN_TURNOVER_CRORES,
        MIN_AVG_TURNOVER_42D_CRORES,
        warnings,
    )
    if len(eligible) < len(symbols):
        excluded = len(symbols) - len(eligible)
        print(f"\n  Turnover filter: {excluded} symbol(s) excluded; {len(eligible)} eligible")
    symbols = eligible

    # ── Kite session ──────────────────────────────────────────────────────
    kite = get_kite_client(token_file)
    instrument_map = build_instrument_map(kite, effective_kite_symbols)

    # ── Per-symbol fetch + metrics ────────────────────────────────────────
    total   = len(symbols)
    metrics_list: List[Dict] = []
    skipped: List[str] = []

    print(f"\n  Fetching OHLCV history + computing metrics ({total} symbols)...\n")

    for idx, symbol in enumerate(symbols, 1):
        kite_symbol = symbol_replacements.get(symbol.upper(), symbol.upper())
        info = instrument_map.get(kite_symbol)
        if info is None:
            record_inactive_symbol(conn, symbol)
            warnings.append(WarningLog(symbol, "Not found in Kite NSE instruments; skipped."))
            skipped.append(symbol)
            print(f"  [{idx:>3}/{total}] {symbol:<18} — not in Kite instruments, skipped")
            continue

        instrument_token = int(info["instrument_token"])
        listing_date     = info.get("listing_date")

        display_symbol = f"{symbol}->{kite_symbol}" if kite_symbol != symbol.upper() else symbol
        print(f"  [{idx:>3}/{total}] {display_symbol:<18}", end=" ", flush=True)
        df = fetch_history_with_retry(
            kite, instrument_token, kite_symbol, fetch_start, cutoff_date
        )
        time.sleep(0.35)

        if df is None or df.empty:
            warnings.append(WarningLog(symbol, "No OHLCV data returned by Kite; skipped."))
            skipped.append(symbol)
            print("NO DATA — skipped")
            continue

        turnover_override = turnover_map.get(symbol.upper()) or turnover_map.get(kite_symbol.upper())
        result = compute_stock_metrics(
            kite_symbol, df, index_df, sector_map.get(symbol, sector_map.get(kite_symbol, "Unknown")),
            cutoff_date, reset_date, listing_date,
            turnover_override,
            liquid_leader_bonus_map.get(symbol.upper(), liquid_leader_bonus_map.get(kite_symbol.upper(), 0)),
            ipo_issue_price_map.get(symbol.upper()) or ipo_issue_price_map.get(kite_symbol.upper()),
            warnings,
        )
        if result is None:
            skipped.append(symbol)
            print("metrics=None — skipped")
            continue

        metrics_list.append(result)
        cs   = result.get("composite_score", "?")
        rs   = result.get("rs_1m", None)
        rs_s = f"{rs:+.1f}%" if rs is not None and not (isinstance(rs, float) and math.isnan(rs)) else "N/A"
        print(f"ok  score={cs}  RS1m={rs_s}")

    # ── Build DataFrame + scoring ─────────────────────────────────────────
    if not metrics_list:
        print("\n  No stocks could be rated. Check symbol file, Kite token, and DB.")
        conn.close()
        return

    raw_df = pd.DataFrame(metrics_list)
    print(f"\n  Applying scoring model to {len(raw_df)} stocks...")
    scored = apply_scoring(raw_df)
    scored = scored.sort_values("composite_score", ascending=False).reset_index(drop=True)

    # ── Console summary ───────────────────────────────────────────────────
    print(f"\n{'='*70}")
    print(f"  Rated: {len(scored)}  |  Skipped: {len(skipped)}")
    if skipped:
        print(f"  Skipped: {', '.join(skipped)}")
    print(f"{'='*70}")
    top_n = min(20, len(scored))
    print(f"\n  {'#':<4} {'SYMBOL':<18} {'SCORE':>6}  {'SECTOR'}")
    print(f"  {'─'*60}")
    for i, row in scored.head(top_n).iterrows():
        print(f"  {i+1:<4} {str(row.get('symbol','')):<18} "
              f"{row.get('composite_score', 0):>6.1f}  "
              f"{row.get('sector', '')}")

    # ── Export workbook ───────────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    wb_path = export_workbook(scored, warnings, output_dir, cutoff_date, reset_date, symbol_file)
    print(f"\n  Excel  → {wb_path}")

    # ── Export TradingView day-one list ───────────────────────────────────
    reports_dir = Path(__file__).resolve().parent / "reports"
    tv_path = export_tradingview_dayone(scored, reports_dir, cutoff_date)
    print(f"  Day-1  → {tv_path}")

    # ── Export Institutional Picks TXT (40% sector + 30% liquidity + 30% rating) ──
    top50_ip = scored.head(50).copy()
    ip_sec = (
        top50_ip.groupby("sector", dropna=False)
        .agg(_c=("composite_score", "count"), _a=("composite_score", "mean"))
        .reset_index()
    )
    ip_sec["_ss"] = ip_sec["_c"] * ip_sec["_a"]
    top50_ip = top50_ip.merge(ip_sec[["sector", "_ss"]], on="sector", how="left")
    top50_ip["_sp"] = top50_ip["_ss"].rank(pct=True, method="average") * 100
    top50_ip["_lp"] = top50_ip["median_turnover_42d"].rank(pct=True, method="average") * 100
    top50_ip["_rp"] = top50_ip["composite_score"].rank(pct=True, method="average") * 100
    top50_ip["_is"] = 0.40 * top50_ip["_sp"] + 0.30 * top50_ip["_lp"] + 0.30 * top50_ip["_rp"]
    _inst_eligible_txt = top50_ip["median_turnover_42d"] >= INST_PICKS_MIN_MEDIAN_TURNOVER_CRORES
    inst_picks_df = (
        top50_ip[_inst_eligible_txt]
        .sort_values("_is", ascending=False)
        .head(20)
        .reset_index(drop=True)
    )

    reports_dir.mkdir(parents=True, exist_ok=True)
    ip_txt_path = reports_dir / f"institutional_picks_{cutoff_date.strftime('%d%b%Y').lower()}.txt"

    ip_lines = [
        f"### Institutional Picks — {cutoff_date.strftime('%d %b %Y')}",
        f"### Selection: Top 50 by score → 40% Sector Strength + 30% Liquidity + 30% Rating",
        f"### Additional gate: median 42D turnover >= {INST_PICKS_MIN_MEDIAN_TURNOVER_CRORES:.0f} Cr",
        "",
    ]
    for _, row in inst_picks_df.iterrows():
        symbol = str(row.get("symbol", "")).strip().upper().replace("NSE:", "")
        if symbol:
            ip_lines.append(f"NSE:{symbol}")

    ip_txt_path.write_text("\n".join(ip_lines).rstrip() + "\n", encoding="utf-8")
    print(f"  Inst.  → {ip_txt_path}")

    # ── Export Thematic Picks TXT (Top 50 → 65% sector + 35% liquidity) ─
    top50_main = scored.head(50).copy()
    sec_agg_main = (
        top50_main.groupby("sector", dropna=False)
        .agg(_sc=("composite_score", "count"), _sa=("composite_score", "mean"))
        .reset_index()
    )
    sec_agg_main["_ss"] = sec_agg_main["_sc"] * sec_agg_main["_sa"]
    top50_main = top50_main.merge(sec_agg_main[["sector", "_ss"]], on="sector", how="left")
    top50_main["_sp"] = top50_main["_ss"].rank(pct=True, method="average") * 100
    top50_main["_lp"] = top50_main["median_turnover_42d"].rank(pct=True, method="average") * 100
    top50_main["_ts"] = 0.65 * top50_main["_sp"] + 0.35 * top50_main["_lp"]
    thematic_df = (
        top50_main.sort_values("_ts", ascending=False).head(20).reset_index(drop=True)
    )

    tp_txt_path = reports_dir / f"thematic_picks_{cutoff_date.strftime('%d%b%Y').lower()}.txt"
    tp_lines = [
        f"### Thematic Picks — {cutoff_date.strftime('%d %b %Y')}",
        f"### Selection: Top 50 by score → 65% Leading Sector + 35% Liquidity",
        "",
    ]
    for _, row in thematic_df.iterrows():
        symbol = str(row.get("symbol", "")).strip().upper().replace("NSE:", "")
        if symbol:
            tp_lines.append(f"NSE:{symbol}")

    tp_txt_path.write_text("\n".join(tp_lines).rstrip() + "\n", encoding="utf-8")
    print(f"  Themat → {tp_txt_path}")

    # ── Auto-run sector_discovery2.py if present ──────────────────────────
    sd2_script = Path(__file__).resolve().parent / "sector_discovery2.py"
    if sd2_script.exists():
        print("\n" + "=" * 70)
        print("  AUTO-RUNNING SECTOR DISCOVERY 2 ...")
        print("=" * 70)
        sd2_cmd = [
            sys.executable, str(sd2_script),
            "--as-of",   cutoff_date.strftime("%Y-%m-%d"),
            "--reset",   reset_date.strftime("%Y-%m-%d"),
            "--symbols", str(symbol_file),
            "--out",     str(output_dir),
            "--token",   str(token_file),
        ]
        subprocess.run(sd2_cmd)
    else:
        print(f"\n  [i] sector_discovery2.py not found — skipping sector discovery run.")

    conn.close()
    print("\n  Done!")


if __name__ == "__main__":
    main()
