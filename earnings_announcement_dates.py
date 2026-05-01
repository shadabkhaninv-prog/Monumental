#!/usr/bin/env python3
"""
Build an Excel sheet of current-quarter earnings announcement dates for a symbol list.

Default behavior:
- Reads symbols from a plain text file, one symbol per line.
- Looks up current-quarter board-meeting and financial-results data from NSE APIs.
- Writes an Excel workbook with a summary sheet and a detailed results sheet.

Example:
    python earnings_announcement_dates.py --input gmlist\\updated_gmlist_30Apr2026.txt --output output\\earnings_dates.xlsx
"""

from __future__ import annotations

import argparse
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path

import pandas as pd
import requests
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_INPUT = Path(__file__).resolve().parent / "gmlist" / "updated_gmlist_30Apr2026.txt"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "output" / "current_quarter_earnings_announcement_dates.xlsx"
DEFAULT_TABLE = "earnings_announcement_dates"
DB_CONFIG = {
    "host": "localhost",
    "port": 3306,
    "user": "root",
    "password": "root",
    "database": "bhav",
}
CURRENT_Q_END = date(2026, 3, 31)
CURRENT_Q_LABEL = "FY2025-26 Q4 (quarter ended 31-Mar-2026)"

CURRENT_PATTERNS = [
    r"March\s+2026",
    r"March\s+31,\s*2026",
    r"March\s+31\s+2026",
    r"quarter\s+ended\s+March\s+2026",
    r"year\s+ended\s+March\s+2026",
    r"period\s+ended\s+March\s+2026",
    r"ended\s+March\s+2026",
]
CURRENT_RE = re.compile("|".join(CURRENT_PATTERNS), re.I)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Find current-quarter earnings announcement dates from NSE.")
    parser.add_argument(
        "--input",
        default=str(DEFAULT_INPUT),
        help=f"Text file with one symbol per line (default: {DEFAULT_INPUT}).",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help=f"Output workbook path (default: {DEFAULT_OUTPUT}).",
    )
    parser.add_argument(
        "--max-workers",
        type=int,
        default=8,
        help="Parallel workers for exchange lookups (default: 8).",
    )
    parser.add_argument("--host", default=DB_CONFIG["host"], help="MySQL host (default: localhost).")
    parser.add_argument("--port", type=int, default=DB_CONFIG["port"], help="MySQL port (default: 3306).")
    parser.add_argument("--user", default=DB_CONFIG["user"], help="MySQL user (default: root).")
    parser.add_argument("--password", default=DB_CONFIG["password"], help="MySQL password (default: root).")
    parser.add_argument("--database", default=DB_CONFIG["database"], help="MySQL database (default: bhav).")
    parser.add_argument(
        "--table",
        default=DEFAULT_TABLE,
        help=f"MySQL table to upsert into (default: {DEFAULT_TABLE}).",
    )
    return parser.parse_args()


def load_symbols(path: Path) -> list[str]:
    symbols: list[str] = []
    seen: set[str] = set()
    for raw in path.read_text(encoding="utf-8").splitlines():
        sym = raw.strip().upper()
        if not sym or sym.startswith("#") or sym in seen:
            continue
        symbols.append(sym)
        seen.add(sym)
    if not symbols:
        raise SystemExit(f"No symbols found in {path}")
    return symbols


def parse_dmy(value: str | None):
    text = (value or "").strip()
    if not text:
        return None
    for fmt in ("%d-%b-%Y", "%d-%b-%Y %H:%M", "%d-%b-%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json, text/plain, */*",
        }
    )
    return session


def make_mysql_connection(args: argparse.Namespace) -> mysql.connector.MySQLConnection:
    return mysql.connector.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        database=args.database,
    )


def ensure_earnings_table(conn: mysql.connector.MySQLConnection, table_name: str) -> None:
    sql = f"""
        DROP TABLE IF EXISTS `{table_name}`;
        CREATE TABLE `{table_name}` (
            symbol VARCHAR(32) NOT NULL,
            quarter_end DATE NOT NULL,
            announcement_date DATE NULL,
            status VARCHAR(16) NOT NULL DEFAULT 'Not found',
            PRIMARY KEY (symbol)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """
    cur = conn.cursor()
    for stmt in sql.strip().split(";"):
        statement = stmt.strip()
        if statement:
            cur.execute(statement)
    conn.commit()
    cur.close()


def upsert_earnings_rows(
    conn: mysql.connector.MySQLConnection,
    table_name: str,
    rows: list[dict],
) -> int:
    if not rows:
        return 0

    sql = f"""
        INSERT INTO `{table_name}` (
            symbol,
            quarter_end,
            announcement_date,
            status
        )
        VALUES (
            %(Symbol)s,
            %(Quarter End)s,
            %(Announcement Date)s,
            %(Status)s
        )
        ON DUPLICATE KEY UPDATE
            quarter_end = VALUES(quarter_end),
            announcement_date = VALUES(announcement_date),
            status = VALUES(status)
    """
    payload = []
    for row in rows:
        payload.append(
            {
                "Symbol": row["Symbol"],
                "Quarter End": row["Quarter End"],
                "Announcement Date": row["Announcement Date"],
                "Status": row["Status"],
            }
        )

    cur = conn.cursor()
    cur.executemany(sql, payload)
    affected = cur.rowcount
    conn.commit()
    cur.close()
    return affected


@lru_cache(maxsize=None)
def get_meta(session_key: int, sym: str) -> dict:
    session = _SESSION_REGISTRY[session_key]
    resp = session.get(
        "https://www.nseindia.com/api/equity-meta-info",
        params={"symbol": sym},
        headers={"Referer": "https://www.nseindia.com/companies-listing/corporate-filings-application"},
        timeout=25,
    )
    resp.raise_for_status()
    return resp.json()


@lru_cache(maxsize=None)
def get_board(session_key: int, sym: str) -> list[dict]:
    session = _SESSION_REGISTRY[session_key]
    resp = session.get(
        "https://www.nseindia.com/api/corporate-board-meetings",
        params={"index": "equities", "symbol": sym},
        headers={"Referer": "https://www.nseindia.com/companies-listing/corporate-filings-board-meetings"},
        timeout=25,
    )
    resp.raise_for_status()
    return resp.json()


@lru_cache(maxsize=None)
def get_results(session_key: int, sym: str) -> list[dict]:
    session = _SESSION_REGISTRY[session_key]
    resp = session.get(
        "https://www.nseindia.com/api/corporates-financial-results",
        params={"index": "equities", "symbol": sym, "period": "Quarterly"},
        headers={"Referer": "https://www.nseindia.com/companies-listing/corporate-filings-financial-results"},
        timeout=25,
    )
    resp.raise_for_status()
    return resp.json()


def score_board(row: dict) -> tuple[int, int, datetime]:
    purpose = (row.get("bm_purpose") or "").lower()
    desc = (row.get("bm_desc") or "").lower()
    text = f"{purpose} {desc}"
    if "financial results" in purpose or "financial results" in desc:
        purpose_score = 0
    elif "board meeting intimation" in purpose:
        purpose_score = 1
    elif "board meeting" in purpose:
        purpose_score = 2
    else:
        purpose_score = 3
    return (
        purpose_score,
        0 if CURRENT_RE.search(text) else 1,
        parse_dmy(row.get("bm_timestamp")) or datetime.max,
    )


def pick_board_row(rows: list[dict] | None) -> dict | None:
    candidates = [row for row in rows or [] if CURRENT_RE.search(" ".join(str(row.get(k, "")) for k in ("bm_purpose", "bm_desc", "bm_date")))]
    if not candidates:
        return None
    candidates.sort(key=score_board)
    return candidates[0]


def score_result(row: dict) -> tuple[int, datetime]:
    consolidated = (row.get("consolidated") or "").lower()
    return (0 if "consolidated" in consolidated else 1, parse_dmy(row.get("filingDate")) or datetime.max)


def pick_result_row(rows: list[dict] | None) -> dict | None:
    candidates = []
    for row in rows or []:
        fy = row.get("financialYear") or ""
        to_date = row.get("toDate") or ""
        if "31-Mar-2026" in fy or to_date == "31-Mar-2026" or "Mar 2026" in fy:
            candidates.append(row)
    if not candidates:
        return None
    candidates.sort(key=score_result)
    return candidates[0]


def extract_timestamp_from_result(row: dict) -> str:
    for key in ("broadCastDate", "filingDate", "exchdisstime"):
        val = (row.get(key) or "").strip()
        if val:
            return val
    return ""


def build_row(session_key: int, sym: str) -> dict:
    meta = get_meta(session_key, sym)
    company = meta.get("companyName") or ""

    board_row = pick_board_row(get_board(session_key, sym))
    if board_row:
        bm_date = parse_dmy(board_row.get("bm_date"))
        return {
            "Symbol": sym,
            "Company Name": company or board_row.get("sm_name") or "",
            "Status": "Found",
            "Announcement Date": bm_date.date() if bm_date else None,
            "Announcement Timestamp": board_row.get("bm_timestamp") or "",
            "Source": "NSE corporate-board-meetings",
            "Basis": board_row.get("bm_purpose") or "",
            "Evidence": board_row.get("bm_desc") or "",
            "Quarter End": CURRENT_Q_END,
            "Quarter Label": CURRENT_Q_LABEL,
            "Comments": "Selected the row that mentions March 2026 / year ended March 2026.",
        }

    result_row = pick_result_row(get_results(session_key, sym))
    if result_row:
        filing = extract_timestamp_from_result(result_row)
        filing_dt = parse_dmy(filing.split(" ")[0]) if filing else None
        return {
            "Symbol": sym,
            "Company Name": company or result_row.get("companyName") or "",
            "Status": "Found",
            "Announcement Date": filing_dt.date() if filing_dt else None,
            "Announcement Timestamp": filing,
            "Source": "NSE corporates-financial-results",
            "Basis": f"{result_row.get('relatingTo') or ''} | {result_row.get('consolidated') or ''}",
            "Evidence": f"{result_row.get('financialYear') or ''} | from {result_row.get('fromDate') or ''} to {result_row.get('toDate') or ''}",
            "Quarter End": CURRENT_Q_END,
            "Quarter Label": CURRENT_Q_LABEL,
            "Comments": "Used the latest filed financial-result row for the current quarter.",
        }

    return {
        "Symbol": sym,
        "Company Name": company,
        "Status": "Not found",
        "Announcement Date": None,
        "Announcement Timestamp": "",
        "Source": "",
        "Basis": "",
        "Evidence": "No March-2026 board-meeting or financial-results row found in NSE APIs as of 30-Apr-2026.",
        "Quarter End": CURRENT_Q_END,
        "Quarter Label": CURRENT_Q_LABEL,
        "Comments": "Exchange posting not yet available in the queried NSE feeds.",
    }


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for sheet_name in ("Summary", "EarningsDates"):
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if sheet_name == "EarningsDates":
            widths = {
                "A": 14,
                "B": 16,
                "C": 32,
                "D": 12,
                "E": 18,
                "F": 24,
                "G": 26,
                "H": 34,
                "I": 70,
                "J": 14,
                "K": 34,
                "L": 24,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            for cell in ws["D"][1:]:
                if cell.value:
                    cell.number_format = "yyyy-mm-dd"
            for cell in ws["E"][1:]:
                if cell.value:
                    cell.number_format = "dd-mmm-yyyy"
            for cell in ws["J"][1:]:
                if cell.value:
                    cell.number_format = "dd-mmm-yyyy"
        else:
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 12
    wb.save(path)


_SESSION_REGISTRY: dict[int, requests.Session] = {}


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    symbols = load_symbols(input_path)
    order_map = {sym: idx for idx, sym in enumerate(symbols)}
    session = make_session()
    session_key = id(session)
    _SESSION_REGISTRY[session_key] = session

    mysql_conn = make_mysql_connection(args)
    try:
        rows = []
        with ThreadPoolExecutor(max_workers=args.max_workers) as pool:
            futures = {pool.submit(build_row, session_key, sym): sym for sym in symbols}
            for future in as_completed(futures):
                rows.append(future.result())

        rows.sort(key=lambda r: order_map.get(r["Symbol"], 10**9))
        df = pd.DataFrame(rows)

        summary = pd.DataFrame(
            [
                {"Metric": "Total symbols", "Value": len(df)},
                {"Metric": "Found", "Value": int((df["Status"] == "Found").sum())},
                {"Metric": "Not found", "Value": int((df["Status"] != "Found").sum())},
                {"Metric": "Reference date", "Value": "30-Apr-2026"},
                {"Metric": "Current quarter", "Value": CURRENT_Q_LABEL},
            ]
        )

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="Summary", index=False)
            df.to_excel(writer, sheet_name="EarningsDates", index=False)

        style_workbook(output_path)
        ensure_earnings_table(mysql_conn, args.table)
        affected = upsert_earnings_rows(mysql_conn, args.table, rows)
        print(f"Wrote {output_path}")
        print(f"Upserted {len(rows)} rows into {args.database}.{args.table} (affected rows: {affected}).")
        print(df["Status"].value_counts(dropna=False).to_string())
    finally:
        mysql_conn.close()
        _SESSION_REGISTRY.pop(session_key, None)


if __name__ == "__main__":
    main()
