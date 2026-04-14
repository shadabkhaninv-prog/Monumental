#!/usr/bin/env python3
"""
Run the neo liquid momentum scanner logic on the top turnover universe,
restricted to the top 5 sectors identified by stock_rating.py.

Usage:
    python neo_top_sector_scanner.py 2026-04-02 2026-03-04
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import neo_liquid_momentum_scanner as neo


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


DEFAULT_TOP_TURNOVER = 1000
DEFAULT_TOP_REPORT = 30
ROOT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT_DIR / "output"
REPORTS_DIR = ROOT_DIR / "reports"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Neo liquid momentum scan restricted to stock-rating top sectors."
    )
    parser.add_argument("as_of_date", help="As-of date (YYYY-MM-DD)")
    parser.add_argument("reset_date", help="Reset date (YYYY-MM-DD)")
    parser.add_argument("--source", choices=["kite", "bhav"], default="kite")
    parser.add_argument("--rating-file", help="Explicit stock-rating workbook path")
    parser.add_argument("--output-dir", default=str(REPORTS_DIR), help="Folder for report outputs")
    parser.add_argument("--top-turnover", type=int, default=DEFAULT_TOP_TURNOVER)
    parser.add_argument("--top-report", type=int, default=DEFAULT_TOP_REPORT)
    return parser.parse_args()


def locate_rating_workbook(cutoff_date: datetime, explicit_path: str | None) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if not path.exists():
            raise SystemExit(f"Stock rating workbook not found: {path}")
        return path

    path = OUTPUT_DIR / f"stock_rating_{cutoff_date.strftime('%d%b%Y')}.xlsx"
    if not path.exists():
        raise SystemExit(f"Stock rating workbook not found: {path}")
    return path.resolve()


def read_top_rating_sectors(rating_path: Path) -> list[str]:
    excel = pd.ExcelFile(rating_path, engine="openpyxl")
    sheet_name = next((name for name in excel.sheet_names if "Sector Leaders" in name), None)
    if not sheet_name:
        raise SystemExit(f"Sector Leaders sheet not found in {rating_path}")
    df = pd.read_excel(rating_path, sheet_name=sheet_name, header=1, engine="openpyxl")
    if "Sector" not in df.columns:
        raise SystemExit(f"Sector column not found in Sector Leaders sheet: {rating_path}")
    sectors = (
        df["Sector"]
        .dropna()
        .astype(str)
        .str.strip()
        .replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
        .dropna()
        .head(5)
        .tolist()
    )
    if not sectors:
        raise SystemExit(f"No top sectors found in {rating_path}")
    return sectors


def fetch_sector_map_with_theme(conn) -> pd.DataFrame:
    try:
        df = pd.read_sql("SELECT symbol, sector1 AS sector, sector2 FROM sectors", conn)
    except Exception as exc:
        raise SystemExit(f"Could not read sectors table: {exc}") from exc
    df["symbol"] = df["symbol"].astype(str).str.strip().str.upper()
    df["sector"] = df["sector"].fillna("").astype(str).str.strip()
    df["sector2"] = df["sector2"].fillna("").astype(str).str.strip()
    return df


def build_sector_filtered_universe(
    turnover_df: pd.DataFrame,
    sectors_df: pd.DataFrame,
    selected_sectors: list[str],
    top_turnover: int,
) -> tuple[pd.DataFrame, list[str]]:
    selected = {str(s).strip() for s in selected_sectors if str(s).strip()}
    merged = turnover_df.merge(sectors_df, on="symbol", how="left")
    merged["sector"] = merged["sector"].fillna("Unknown")
    merged["sector2"] = merged["sector2"].fillna("")
    merged["matched_sector"] = merged.apply(
        lambda row: (
            str(row.get("sector", "")).strip()
            if str(row.get("sector", "")).strip() in selected
            else str(row.get("sector2", "")).strip()
            if str(row.get("sector2", "")).strip() in selected
            else ""
        ),
        axis=1,
    )
    merged = merged[merged["matched_sector"] != ""].copy()
    merged["sector"] = merged["matched_sector"]
    merged = merged.sort_values("avg_turnover_cr", ascending=False).head(top_turnover).reset_index(drop=True)
    return merged, merged["symbol"].astype(str).tolist()


def write_report(
    top_report: pd.DataFrame,
    all_scored: pd.DataFrame,
    top5_sectors: pd.DataFrame,
    eligible_universe: pd.DataFrame,
    selected_sectors: list[str],
    summary: dict,
    out_path: Path,
) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    neo._force_delete_xlsx(out_path)

    wb = Workbook()
    title_fill = PatternFill("solid", fgColor="1F3864")
    header_fill = PatternFill("solid", fgColor="2E4B8F")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    def style_header(ws, row_num: int, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def populate_sheet(ws, df: pd.DataFrame, title: str) -> None:
        if df.empty and len(df.columns) == 0:
            df = pd.DataFrame([{"Message": "No data available"}])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))
        ws.cell(row=1, column=1, value=title).font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        ws.cell(row=1, column=1).fill = title_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        ws.append(list(df.columns))
        style_header(ws, 2, len(df.columns))
        for idx, (_, row) in enumerate(df.iterrows(), start=3):
            ws.append(row.tolist())
            fill = alt_fill if idx % 2 == 1 else white_fill
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=idx, column=col)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(24, len(str(col_name)) + 2))
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}{max(2, len(df) + 2)}"

    top_cols = [
        "rank", "symbol", "sector", "source", "total_score", "rating",
        "avg_turnover_cr", "median_turnover_cr", "ret_12m", "ret_6m", "ret_3m",
        "ret_reset", "rs_composite", "atr_pct", "vol_period", "vol_day_move_pct",
    ]
    top_view = top_report[[c for c in top_cols if c in top_report.columns]].copy()
    all_view = all_scored[[c for c in top_cols if c in all_scored.columns]].copy()
    sector_view = top5_sectors.copy()
    eligible_cols = ["symbol", "sector", "avg_turnover_cr", "median_turnover_cr", "latest_close"]
    eligible_view = eligible_universe[[c for c in eligible_cols if c in eligible_universe.columns]].copy()

    ws1 = wb.active
    ws1.title = "Top Candidates"
    populate_sheet(ws1, top_view, "Top Candidates — sector-filtered neo scan")

    ws2 = wb.create_sheet("All Scored")
    populate_sheet(ws2, all_view, "All Scored — sector-filtered neo scan")

    ws3 = wb.create_sheet("Eligible Universe")
    populate_sheet(ws3, eligible_view, "Eligible after sector filter - debug universe")

    ws4 = wb.create_sheet("Sector Summary")
    populate_sheet(ws4, sector_view, "Sector Summary - filtered universe")

    ws5 = wb.create_sheet("Summary")
    summary_rows = [
        ("Metric", "Value"),
        ("Selected sectors", ", ".join(selected_sectors)),
        ("Top-turnover universe", summary.get("top_turnover_count", "")),
        ("Eligible after sector filter", summary.get("after_sector_filter", "")),
        ("12M return gate", summary.get("return_gate_percentile", "")),
        ("Scored symbols", summary.get("total_scored", "")),
        ("Top report size", summary.get("top_report_count", "")),
        ("Stock-rating workbook", summary.get("rating_file", "")),
    ]
    for row in summary_rows:
        ws5.append(list(row))
    style_header(ws5, 1, 2)
    for row in range(2, len(summary_rows) + 1):
        fill = alt_fill if row % 2 == 0 else white_fill
        for col in range(1, 3):
            cell = ws5.cell(row=row, column=col)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center", wrap_text=True)
    ws5.column_dimensions["A"].width = 24
    ws5.column_dimensions["B"].width = 80
    ws5.freeze_panes = "A2"

    wb.save(out_path)
    return out_path


def main() -> None:
    args = parse_args()
    as_of = datetime.strptime(args.as_of_date, "%Y-%m-%d")
    reset = datetime.strptime(args.reset_date, "%Y-%m-%d")
    ohlcv_from = min(as_of - timedelta(days=375), reset - timedelta(days=5))

    rating_path = locate_rating_workbook(as_of, args.rating_file)
    selected_sectors = read_top_rating_sectors(rating_path)
    log.info(f"Using stock-rating sectors: {', '.join(selected_sectors)}")

    conn = neo.get_db_connection()
    kite = neo.get_kite_client() if args.source == "kite" else None

    turnover_df = neo.fetch_bhav_turnover(conn, as_of)
    if turnover_df.empty or "symbol" not in turnover_df.columns:
        raise SystemExit("No turnover data available from bhav DB.")
    turnover_df = turnover_df[turnover_df["latest_close"] >= neo.MIN_CLOSE].copy()
    turnover_df = turnover_df[
        (turnover_df["avg_turnover_cr"] >= neo.MIN_EXTENDED_TURNOVER_CR)
        & (turnover_df["median_turnover_cr"] >= neo.MIN_EXTENDED_TURNOVER_CR)
    ].copy()

    sectors_df = fetch_sector_map_with_theme(conn)
    sector_universe_df, sector_symbols = build_sector_filtered_universe(
        turnover_df, sectors_df, selected_sectors, args.top_turnover
    )
    log.info(f"Sector-filtered turnover universe: {len(sector_symbols)} symbols")

    if kite:
        valid_eq = neo.get_nse_eq_symbols(kite)
        sector_universe_df = sector_universe_df[sector_universe_df["symbol"].isin(valid_eq)].copy()
        sector_symbols = sector_universe_df["symbol"].astype(str).tolist()
        token_map = neo.build_nse_token_map(kite, sector_symbols)
    else:
        token_map = {}

    index_df = pd.DataFrame()
    if kite:
        index_df = neo.fetch_ohlcv_kite(kite, neo.NIFTY_SMALLCAP_250_SYMBOL, ohlcv_from, as_of, is_index=True)
        if index_df.empty:
            index_df = neo.fetch_ohlcv_kite(kite, neo.NIFTY_50_SYMBOL, ohlcv_from, as_of, is_index=True)

    failed_symbols: list[str] = []
    ohlcv_map: dict[str, pd.DataFrame] = {}
    for i, sym in enumerate(sector_symbols):
        if i > 0 and i % 100 == 0:
            log.info(f"Fetched {i}/{len(sector_symbols)} symbols...")
        ohlcv_map[sym] = neo.fetch_ohlcv(
            kite, conn, sym, ohlcv_from, as_of, args.source, failed_symbols, token_map=token_map
        )

    metrics_df = neo.compute_all_metrics(sector_symbols, ohlcv_map, index_df, as_of, reset)
    metrics_df = metrics_df.merge(
        sector_universe_df[["symbol", "avg_turnover_cr", "median_turnover_cr", "sector"]],
        on="symbol",
        how="left",
    )
    scored = neo.score_universe(metrics_df)
    if not scored.empty:
        scored = scored.merge(sectors_df[["symbol", "sector"]], on="symbol", how="left", suffixes=("", "_db"))
        if "sector_db" in scored.columns:
            scored["sector"] = scored["sector"].fillna(scored["sector_db"])
            scored.drop(columns=["sector_db"], inplace=True)
        scored["source"] = "Top1000TopSectors"
        scored = scored.sort_values(["total_score", "ret_12m"], ascending=[False, False]).reset_index(drop=True)
        scored["rank"] = range(1, len(scored) + 1)

    top_report = scored.head(args.top_report).copy()
    top5_sectors = (
        neo.compute_sector_summary(scored.drop(columns=["sector"], errors="ignore"), sectors_df)
        if not scored.empty else pd.DataFrame()
    )

    output_dir = Path(args.output_dir).expanduser().resolve()
    report_path = output_dir / (
        f"TopSectorLiquidCandidates_{as_of.strftime('%d%b%Y').upper()}_Reset{reset.strftime('%d%b%Y').upper()}.xlsx"
    )
    summary = {
        "top_turnover_count": min(args.top_turnover, len(turnover_df)),
        "after_sector_filter": len(sector_symbols),
        "return_gate_percentile": "Not applied in this report",
        "total_scored": len(scored),
        "top_report_count": len(top_report),
        "rating_file": rating_path.name,
    }
    write_report(top_report, scored, top5_sectors, sector_universe_df, selected_sectors, summary, report_path)

    watchlist_path = output_dir / f"top_sector_wl_{as_of.strftime('%d%b%Y')}.txt"
    neo.write_tradingview_watchlist(top_report, sectors_df, watchlist_path, as_of)

    conn.close()
    log.info(f"Report saved: {report_path}")
    log.info(f"Watchlist saved: {watchlist_path}")
    if failed_symbols:
        log.warning(f"Failed API symbols: {failed_symbols[:10]}{'...' if len(failed_symbols) > 10 else ''}")


if __name__ == "__main__":
    main()
